"""Gate Access & Attendance Tracking - City College of Bayawan."""
import base64
import csv
import io
import os
import platform
import zipfile
import datetime
import json
import logging
import calendar

logger = logging.getLogger(__name__)
import re
from types import SimpleNamespace
from urllib.parse import urlencode
from django.shortcuts import render, redirect, get_object_or_404
from django.template.loader import render_to_string
from django.urls import reverse
from django.conf import settings
from django.templatetags.static import static
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse, HttpResponse, HttpResponseForbidden
from django.views.decorators.http import require_GET, require_POST
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.contrib import messages
from django.db.models import Count, Q, Case, When, IntegerField, Sum, F
from django.db.models.functions import ExtractYear, ExtractMonth
from django.db import transaction
from django.core.cache import cache
from django.core.files.storage import default_storage
from django.contrib.staticfiles import finders

from collections import Counter, defaultdict

from .notifications import notify_student_status_change

from gate_analytics.roles import role_required, has_role, get_user_role
from .models import (
    Student,
    StaffPersonnelProfile,
    GateEntry,
    GateIncident,
    Event,
    EventAgenda,
    EventAttendance,
    EventRegistration,
    AttendanceLog,
    ScannerDevice,
    GeneratedReport,
    VisitorEntry,
    VisitorPass,
    VisitorVisit,
    CAMPUS_DEPARTMENT_CHOICES,
    SiteTheme,
    GateShift,
    AuditLog,
    AdminNotification,
    BlockedIP,
)
from .forms import (
    StudentForm,
    StudentModalForm,
    StudentStudentAffairsForm,
    StaffPersonnelCreateForm,
    SiteThemeEidSignatoryForm,
    SiteThemeReportSignatoryForm,
)
from .policy import (
    get_student_current_state,
    evaluate_scan,
    get_gate_policy,
    daily_gate_repeat_cooldown,
    format_cooldown_wait_remaining,
    _entry_is_in_direction,
)
from .page_loader_session import pop_post_login_loader


def _register_or_touch_scanner_device(device_id):
    """Register a scanner device on first use and update last_seen_at. Idempotent."""
    if not (device_id and isinstance(device_id, str)):
        return None
    base_id = device_id.replace('-OFFLINE', '').strip()
    if not base_id:
        return None
    ScannerDevice.objects.get_or_create(
        device_id=base_id,
        defaults={'name': f'Scanner {base_id[:8]}', 'is_active': True},
    )
    ScannerDevice.objects.filter(device_id=base_id).update(last_seen_at=timezone.now())
    return base_id


def _resolve_student_from_gate_input(raw: str):
    """
    Resolve a scanner/manual string to a Student (by ID or name).
    Returns (student, err_kind) where err_kind is None, 'multiple', or 'none'.
    """
    q = (raw or '').strip()
    if not q:
        return None, 'none'

    st = Student.objects.filter(student_id__iexact=q).first()
    if st:
        return st, None

    if re.match(r'^[A-Za-z0-9\-_.]+$', q) and len(q) >= 3:
        qs = Student.objects.filter(student_id__icontains=q)
        n = qs.count()
        if n == 1:
            return qs.first(), None
        if n > 1:
            return None, 'multiple'

    parts = q.split()
    base = Student.objects.all()
    if len(parts) >= 2:
        first_t, last_t = parts[0], parts[-1]
        qs = base.filter(
            (
                (Q(first_name__icontains=first_t) | Q(middle_name__icontains=first_t))
                & Q(last_name__icontains=last_t)
            )
            | (
                (Q(first_name__icontains=last_t) | Q(middle_name__icontains=last_t))
                & Q(last_name__icontains=first_t)
            )
        ).distinct()
        n = qs.count()
        if n == 1:
            return qs.first(), None
        if n > 1:
            return None, 'multiple'
    if len(parts) == 1 and len(parts[0]) >= 2:
        tok = parts[0]
        qs = base.filter(
            Q(first_name__icontains=tok)
            | Q(last_name__icontains=tok)
            | Q(middle_name__icontains=tok)
        )
        n = qs.count()
        if n == 1:
            return qs.first(), None
        if n > 1:
            return None, 'multiple'
    return None, 'none'


MANUAL_OFFICE_REASON_LABELS = {
    # Guard-facing: scenario first, then who gets notified / what happens next.
    'SAS_ID_CONCERN': 'ID or identity issue — notify Student Affairs (SAS)',
    'GATE_RESOLVED_OFFICE_ENDORSEMENT': 'Others',
}


def _normalize_manual_office_reason(raw_reason):
    code = (raw_reason or '').strip().upper()
    if not code:
        return '', ''
    if code in MANUAL_OFFICE_REASON_LABELS:
        return code, MANUAL_OFFICE_REASON_LABELS[code]
    # Backward compatibility: keep legacy free-text reasons as-is.
    return '', (raw_reason or '').strip()[:200]


def _notify_manual_office_referral(student, actor, entry, office_code, office_label, selected_event=None):
    """
    When guard uses manual entry with office routing: GateIncident + Admin/SAS incident alerts,
    plus admin-only in-app rows with a direct link to edit the student (inactive if SAS unresolved).
    """
    if not office_code or office_code not in MANUAL_OFFICE_REASON_LABELS:
        return
    label = (office_label or MANUAL_OFFICE_REASON_LABELS.get(office_code, office_code) or '').strip()
    if office_code == 'SAS_ID_CONCERN':
        detail_prefix = 'Manual gate referral to SAS for ID concern'
    elif office_code == 'GATE_RESOLVED_OFFICE_ENDORSEMENT':
        detail_prefix = f'Manual gate referral (Others): {label}'
    else:
        detail_prefix = 'Manual gate referral'
    try:
        actor_name = (actor.get_full_name() if actor else '') or (actor.username if actor else 'Gate staff')
        event_label = f' | Event: {selected_event.name}' if selected_event else ''
        msg = (
            f'{detail_prefix}.\n'
            f'{student.get_full_name()} ({student.student_id}) • {actor_name}{event_label}'
        )

        # Create a GateIncident so it appears in /gate/incidents/ and in the navbar incident dropdown.
        # De-dupe: don't create multiple identical referrals within a short window (per routing type).
        try:
            recent_cutoff = timezone.now() - timezone.timedelta(minutes=10)
            exists_recent = GateIncident.objects.filter(
                student=student,
                reason='other',
                details__icontains=detail_prefix,
                timestamp__gte=recent_cutoff,
            ).exists()
            incident = None
            if not exists_recent:
                incident = GateIncident.objects.create(
                    student=student,
                    scanned_id=student.student_id,
                    reason='other',
                    details=msg[:1000],
                    staff_alerted=True,
                    sas_review_status='to_check',
                )
        except Exception:
            incident = None

        try:
            from .admin_notification_service import AdminNotificationService
            if incident:
                AdminNotificationService.notify_incident(incident, priority='high')
            else:
                AdminNotificationService.create_notification(
                    notification_type='incident',
                    title=f'Office referral: {label[:72]}',
                    message=msg[:1000],
                    priority='high',
                    broadcast=True,
                    related_student=student,
                    related_entry=entry,
                    related_event=selected_event,
                )
            AdminNotificationService.notify_admins_gate_manual_referral(
                student=student,
                actor=actor,
                office_label=label,
                related_incident=incident,
            )
        except Exception:
            logger.exception(
                'Manual office referral: admin notification failed (student_id=%s)',
                getattr(student, 'student_id', student.pk),
            )
    except Exception:
        logger.exception(
            'Manual office referral: unexpected error (student_id=%s)',
            getattr(student, 'student_id', getattr(student, 'pk', '')),
        )


def _apply_student_office_hold(student, note):
    if not student:
        return
    try:
        update_fields = []
        if not student.office_clearance_hold:
            student.office_clearance_hold = True
            update_fields.append('office_clearance_hold')
        if note:
            student.office_clearance_note = (note or '')[:255]
            update_fields.append('office_clearance_note')
        if update_fields:
            student.save(update_fields=update_fields)
    except Exception:
        pass


def _clear_student_office_hold_if_no_pending(student):
    if not student:
        return
    pending_exists = GateIncident.objects.filter(
        student=student,
        sas_review_status='to_check',
    ).exists()
    if not pending_exists and student.office_clearance_hold:
        student.office_clearance_hold = False
        student.office_clearance_note = ''
        student.save(update_fields=['office_clearance_hold', 'office_clearance_note'])


def _get_unresolved_office_incident(student):
    """Latest unresolved incident that should block IN until office resolution."""
    if not student:
        return None
    return (
        GateIncident.objects.filter(
            student=student,
            sas_review_status='to_check',
        )
        .order_by('-timestamp')
        .first()
    )


def _audit_kwargs_for_gate_entry(request, device_id=None):
    """Return device_id and ip_address for GateEntry audit trail. Pass device_id if already in scope (e.g. save_scan)."""
    if device_id is None:
        device_id = (request.POST.get('device_id') or (getattr(request, 'GET', None) and request.GET.get('device_id')) or '').strip()
    device_id = (device_id or '')[:128]
    xff = (request.META.get('HTTP_X_FORWARDED_FOR') or '').split(',')[0].strip()
    raw_ip = xff or request.META.get('REMOTE_ADDR') or ''
    ip = raw_ip[:45] if raw_ip else None
    return {'device_id': device_id or '', 'ip_address': ip}


def _guard_embed_query_token_ok(request):
    """GET ?embed=1&guard_token=... matches GATE_GUARD_DISPLAY_TOKEN (gate entry list iframe)."""
    embed = (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes')
    token = (request.GET.get('guard_token') or '').strip()
    expected = getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or ''
    return bool(embed and expected and token == expected)


def _guard_embed_get_token_ok(request):
    """GET ?guard_token=... only (e.g. event attendees iframe — no embed=1 in URL)."""
    token = (request.GET.get('guard_token') or '').strip()
    expected = getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or ''
    return bool(expected and token == expected)


def _resolve_save_scan_actor(request):
    """
    Who is recording this scan: logged-in gate staff, or guard monitor (GATE_GUARD_DISPLAY_TOKEN).
    Returns User for audit, None when token is valid but no GATE_GUARD_EMBED_RECORDED_BY_USER_ID, or False if unauthorized.
    """
    token = (request.POST.get('guard_token') or request.headers.get('X-Gate-Guard-Token') or '').strip()
    expected = getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or ''
    if token and expected and token == expected:
        uid = getattr(settings, 'GATE_GUARD_EMBED_RECORDED_BY_USER_ID', None)
        if uid:
            from django.contrib.auth import get_user_model
            User = get_user_model()
            try:
                return User.objects.get(pk=int(uid))
            except (ValueError, User.DoesNotExist):
                pass
        return None
    if request.user.is_authenticated:
        role = get_user_role(request.user)
        if role in ('admin', 'staff', 'faculty'):
            return request.user
    return False


def _scan_ui_photo_url(request, student):
    """
    Photo URL for scan JSON and gate popup <img>. Prefer relative /media/... paths so the
    browser resolves against the current page origin (fixes 127.0.0.1 vs localhost and
    similar host mismatches). Remote storage may return an absolute https URL — keep as-is.
    """
    if not student or not getattr(student, 'photo', None) or not student.photo:
        return None
    try:
        url = student.photo.url
    except Exception:
        return None
    if not url:
        return None
    url = str(url).strip()
    if url.startswith('http://') or url.startswith('https://'):
        return url
    if url.startswith('/'):
        return url
    try:
        return request.build_absolute_uri(url)
    except Exception:
        return None


def _fmt_student_name(student):
    """Return name in "Last, First Middle" order for exports and preview.
    Falls back to get_full_name() if fields are missing to avoid blank output.
    """
    if not student:
        return ''
    try:
        lname = (student.last_name or '').strip()
        fname = (student.first_name or '').strip()
        mname = (student.middle_name or '').strip()
    except AttributeError:
        return student.get_full_name() if hasattr(student, 'get_full_name') else ''
    if not lname and not fname:
        return student.get_full_name() if hasattr(student, 'get_full_name') else ''
    full = fname
    if mname:
        full += ' ' + mname
    if lname:
        return f"{lname}, {full}".strip(', ')
    return full


def _student_sex_display(student):
    """Human-readable sex/gender from Student.sex. Empty field → 'Not set' (data not captured on student profile)."""
    if not student:
        return '—'
    raw = (getattr(student, 'sex', None) or '').strip()
    if not raw:
        return 'Not set'
    return student.get_sex_display()


PER_PAGE_OPTIONS = [10, 20, 30, 40, 50, 100]


def _get_per_page_and_query(request, param_name='page'):
    """Return (per_page, query_extra, query_extra_base). Default per_page 10."""
    raw = request.GET.get('per_page', '10')
    try:
        per_page = int(raw)
        if per_page not in PER_PAGE_OPTIONS:
            per_page = 10
    except (ValueError, TypeError):
        per_page = 10
    q_extra = request.GET.copy()
    q_extra.pop(param_name, None)
    query_extra = q_extra.urlencode()
    q_base = request.GET.copy()
    q_base.pop(param_name, None)
    q_base.pop('per_page', None)
    query_extra_base = q_base.urlencode()
    return per_page, query_extra, query_extra_base


def _gate_entries_to_visits(entries):
    """
    Group gate entries into visits: pair IN with the immediately following OUT
    (same student, same local day) so we show one row per visit instead of two.
    Returns list of (in_entry, out_entry) with out_entry None if no OUT.

    IN/OUT: Prefer scan_type; fallback to notes (legacy). Matches policy/get_student_current_state.
    Grouping: Uses local-date (timezone.localtime) so grouping is consistent with
    _local_day_bounds filtering; avoids splitting one local day across UTC date boundary.
    """
    def _is_out(entry):
        st = getattr(entry, 'scan_type', None)
        if st:
            return (st or '').strip().upper() == 'OUT'
        notes = (entry.notes or '').strip().upper()
        return notes == 'OUT' or notes.startswith('OUT')

    by_student_date = defaultdict(list)
    for e in entries:
        # Group by local date so entries filtered by local day stay in one bucket
        local_ts = timezone.localtime(e.timestamp) if timezone.is_aware(e.timestamp) else e.timestamp
        local_date = local_ts.date()
        by_student_date[(e.student_id, local_date)].append(e)

    visits = []
    for (_student_id, _date), lst in by_student_date.items():
        lst = sorted(lst, key=lambda x: x.timestamp)
        i = 0
        while i < len(lst):
            e = lst[i]
            is_out = _is_out(e)
            if is_out:
                visits.append((None, e))
                i += 1
            else:
                if i + 1 < len(lst):
                    next_e = lst[i + 1]
                    if _is_out(next_e):
                        visits.append((e, next_e))
                        i += 2
                        continue
                visits.append((e, None))
                i += 1

    # sort visits by timestamp ascending so earliest scan appears first
    visits.sort(key=lambda v: (v[0].timestamp if v[0] else v[1].timestamp))
    return visits


def _hydrate_gate_entry_students(entries):
    """Attach Student to GateEntry rows where student_id is set but .student is missing (ORM join edge cases)."""
    if not entries:
        return
    from .models import Student
    need_ids = []
    for e in entries:
        sid = getattr(e, 'student_id', None)
        if sid and getattr(e, 'student', None) is None:
            need_ids.append(sid)
    if not need_ids:
        return
    by_pk = {s.pk: s for s in Student.objects.filter(pk__in=list(set(need_ids)))}
    for e in entries:
        sid = getattr(e, 'student_id', None)
        if sid and getattr(e, 'student', None) is None and sid in by_pk:
            e.student = by_pk[sid]


def _create_event_log_single_duplicate(event, student, scan_type, result, **extra_fields):
    """
    Create an AttendanceLog row, but cap DUPLICATE entries to a single row per
    (event, student, scan_type). Further duplicate scans still notify in UI but
    do not spam the log table.
    """
    if result == 'DUPLICATE' and student is not None:
        if AttendanceLog.objects.filter(
            event=event,
            student=student,
            scan_type=scan_type,
            result='DUPLICATE',
            voided=False,
        ).exists():
            return None
    return AttendanceLog.objects.create(
        event=event,
        student=student,
        scan_type=scan_type,
        result=result,
        **extra_fields,
    )


def _granted_visits_count_for_date(date, daily_gate_only=False):
    """Count granted visits (IN+OUT grouped) for a single date.
    This is a convenience wrapper around `_granted_visits_count_for_bounds` which
    accepts arbitrary start/end datetimes.

    When daily_gate_only=True, only entries with event=None and visitor_visit=None
    are counted (student gate visits)."""
    day_start, day_end = _local_day_bounds(date)
    return _granted_visits_count_for_bounds(day_start, day_end, daily_gate_only=daily_gate_only)


def _granted_visits_count_for_bounds(day_start, day_end, daily_gate_only=False, report_timestamp_q=None):
    """Count granted visits between two aware datetimes (inclusive start, exclusive end).
    Entries are grouped into visits (IN/OUT pairs) before counting.

    The caller is responsible for constructing day_start/day_end from filters, which
    may span multiple days when a date range is applied.

    If daily_gate_only=True only entries with event=None and visitor_visit=None are
    considered.

    When report_timestamp_q is set (multi-day + time-of-day window), use it instead of day bounds.
    """
    qs = GateEntry.objects.filter(granted=True)
    qs = _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end)
    if daily_gate_only:
        qs = qs.filter(event__isnull=True, visitor_visit__isnull=True)
    entries = list(qs.order_by('-timestamp')[:500])
    return len(_gate_entries_to_visits(entries))


def _currently_inside_count(date=None):
    """Count students currently inside (last daily-gate scan today is IN). Uses local day bounds."""
    if date is None:
        date = timezone.localdate()
    day_start, day_end = _local_day_bounds(date)
    entries = list(
        GateEntry.objects.filter(
            timestamp__gte=day_start, timestamp__lt=day_end,
            granted=True, event__isnull=True, visitor_visit__isnull=True,
            student_id__isnull=False,
        ).order_by('student_id', '-timestamp')
    )
    # Last entry per student: if OUT, they're outside; else IN (or legacy notes)
    last_by_student = {}
    for e in entries:
        if e.student_id not in last_by_student:
            last_by_student[e.student_id] = e
    def _is_out(ent):
        st = getattr(ent, 'scan_type', None)
        if st:
            return (st or '').strip().upper() == 'OUT'
        n = (ent.notes or '').strip().upper()
        return n == 'OUT' or n.startswith('OUT')
    return sum(1 for e in last_by_student.values() if not _is_out(e))


def _currently_inside_list(date=None):
    """List of GateEntry (IN) for students currently inside on the given date. Same logic as _currently_inside_count. Uses local day bounds."""
    if date is None:
        date = timezone.localdate()
    day_start, day_end = _local_day_bounds(date)
    entries = list(
        GateEntry.objects.filter(
            timestamp__gte=day_start, timestamp__lt=day_end,
            granted=True, event__isnull=True, visitor_visit__isnull=True,
            student_id__isnull=False,
        ).select_related('student').order_by('student_id', '-timestamp')
    )
    last_by_student = {}
    for e in entries:
        if e.student_id not in last_by_student:
            last_by_student[e.student_id] = e
    def _is_out(ent):
        st = getattr(ent, 'scan_type', None)
        if st:
            return (st or '').strip().upper() == 'OUT'
        n = (ent.notes or '').strip().upper()
        return n == 'OUT' or n.startswith('OUT')
    inside_entries = [e for e in last_by_student.values() if not _is_out(e)]
    inside_entries.sort(key=lambda e: e.timestamp, reverse=True)
    return inside_entries


def _local_day_bounds(date):
    """Return (day_start, day_end) in app timezone for the given date. Use for all 'today' or date filtering to avoid UTC vs local mismatch."""
    tz = timezone.get_current_timezone()
    day_start = timezone.make_aware(datetime.datetime.combine(date, datetime.time.min), tz)
    day_end = day_start + datetime.timedelta(days=1)
    return day_start, day_end


def _local_month_bounds(year, month):
    """Return (month_start, month_end) in app timezone for the given year/month (1-12)."""
    tz = timezone.get_current_timezone()
    month_start = timezone.make_aware(datetime.datetime(year, month, 1, 0, 0, 0), tz)
    if month == 12:
        month_end = timezone.make_aware(datetime.datetime(year + 1, 1, 1, 0, 0, 0), tz)
    else:
        month_end = timezone.make_aware(datetime.datetime(year, month + 1, 1, 0, 0, 0), tz)
    return month_start, month_end


def _local_year_bounds(year):
    """Return (year_start, year_end) in app timezone for the given year."""
    tz = timezone.get_current_timezone()
    year_start = timezone.make_aware(datetime.datetime(year, 1, 1, 0, 0, 0), tz)
    year_end = timezone.make_aware(datetime.datetime(year + 1, 1, 1, 0, 0, 0), tz)
    return year_start, year_end


def _calendar_weeks_overlapping_month(year, month):
    """
    Monday–Sunday weeks that overlap the given calendar month.
    Returns list of (week_index, monday, sunday, clip_start, clip_end) where:
    - week_index is 1-based within this month (first overlapping week = 1),
    - monday/sunday are the full calendar week (Mon–Sun),
    - clip_start/clip_end are dates in [year, month] used for counting entries.
    """
    first = datetime.date(year, month, 1)
    if month == 12:
        last = datetime.date(year + 1, 1, 1) - datetime.timedelta(days=1)
    else:
        last = datetime.date(year, month + 1, 1) - datetime.timedelta(days=1)
    cur = first
    while cur.weekday() != 0:  # Monday = 0
        cur -= datetime.timedelta(days=1)
    weeks = []
    week_index = 0
    while cur <= last:
        monday = cur
        sunday = cur + datetime.timedelta(days=6)
        clip_start = max(monday, first)
        clip_end = min(sunday, last)
        if clip_start <= clip_end:
            week_index += 1
            weeks.append((week_index, monday, sunday, clip_start, clip_end))
        cur += datetime.timedelta(days=7)
    return weeks


# Old name removed from weekly analytics; keep alias so cached bytecode / stale edits don't 500.
_calendar_weeks_fully_inside_month = _calendar_weeks_overlapping_month


def _report_per_day_time_window_q(start_d, end_d, t_from, t_to, field_name, tz):
    """OR of field in [d+t_from, d+t_to) for each calendar day from start_d through end_d (inclusive)."""
    q = Q()
    d = start_d
    while d <= end_d:
        ds = timezone.make_aware(datetime.datetime.combine(d, t_from), tz)
        de = timezone.make_aware(datetime.datetime.combine(d, t_to), tz)
        q |= Q(**{f'{field_name}__gte': ds, f'{field_name}__lt': de})
        d += datetime.timedelta(days=1)
    return q


def _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end):
    if report_timestamp_q is not None:
        return qs.filter(report_timestamp_q)
    return qs.filter(timestamp__gte=day_start, timestamp__lt=day_end)


def _apply_report_checked_in_at_filter(qs, report_checked_in_at_q, day_start, day_end):
    if report_checked_in_at_q is not None:
        return qs.filter(report_checked_in_at_q)
    return qs.filter(checked_in_at__gte=day_start, checked_in_at__lt=day_end)


def _apply_report_checked_out_at_filter(qs, report_checked_out_at_q, day_start, day_end):
    if report_checked_out_at_q is not None:
        return qs.filter(report_checked_out_at_q)
    return qs.filter(checked_out_at__gte=day_start, checked_out_at__lt=day_end)


def _apply_report_recorded_at_filter(qs, report_recorded_at_q, day_start, day_end):
    if report_recorded_at_q is not None:
        return qs.filter(report_recorded_at_q)
    return qs.filter(recorded_at__gte=day_start, recorded_at__lt=day_end)


def _report_filter_from_request(request):
    """
    Parse report filter from GET. Default: Today (local day).
    Supports: date_range (today, yesterday, this_week, last_7_days, last_30_days, custom),
    from_date, to_date (for custom range), all_day (ignore time when set),
    from_time, to_time, search, event_id.
    Returns: (filter_date, day_start, day_end, from_time, to_time, search_q, event_id,
            date_range_label, from_date_str, to_date_str, time_error,
            report_timestamp_q, report_checked_in_at_q, report_checked_out_at_q, report_recorded_at_q).
    When the selected calendar range spans multiple days and both from_time and to_time are set,
    the four *_q values are OR-of-daily windows (same clock on each day); otherwise they are None
    and callers use day_start/day_end. Single-day ranges still use a single [day_start, day_end) pair.
    """
    tz = timezone.get_current_timezone()
    today = timezone.localdate()
    date_range = (request.GET.get('date_range') or 'today').strip().lower()
    from_date = (request.GET.get('from_date') or '').strip()
    to_date = (request.GET.get('to_date') or '').strip()
    all_day = request.GET.get('all_day', '').strip().lower() in ('1', 'true', 'on')
    from_time = (request.GET.get('from_time') or '').strip() if not all_day else ''
    to_time = (request.GET.get('to_time') or '').strip() if not all_day else ''
    search_q = (request.GET.get('search') or request.GET.get('q') or '').strip()
    event_id = request.GET.get('event_id')
    try:
        event_id = int(event_id) if event_id else None
    except (TypeError, ValueError):
        event_id = None

    filter_date = today
    from_date_str = from_date or today.isoformat()
    to_date_str = to_date or from_date_str

    if date_range == 'yesterday':
        filter_date = today - datetime.timedelta(days=1)
        from_date_str = filter_date.isoformat()
        to_date_str = from_date_str
    elif date_range == 'this_week':
        filter_date = today - datetime.timedelta(days=today.weekday())
        from_date_str = filter_date.isoformat()
        to_date_str = today.isoformat()
    elif date_range == 'last_7_days':
        filter_date = today - datetime.timedelta(days=6)
        from_date_str = filter_date.isoformat()
        to_date_str = today.isoformat()
    elif date_range == 'last_30_days':
        filter_date = today - datetime.timedelta(days=29)
        from_date_str = filter_date.isoformat()
        to_date_str = today.isoformat()
    elif date_range == 'custom' and from_date:
        try:
            filter_date = datetime.date.fromisoformat(from_date)
            from_date_str = from_date
            if to_date:
                try:
                    end_d = datetime.date.fromisoformat(to_date)
                    if end_d >= filter_date:
                        to_date_str = to_date
                    else:
                        to_date_str = from_date_str
                except ValueError:
                    to_date_str = from_date_str
            else:
                to_date_str = from_date_str
        except ValueError:
            from_date_str = today.isoformat()
            to_date_str = from_date_str

    # Bounds: single day or range
    day_start, day_end = _local_day_bounds(filter_date)
    if date_range == 'this_week':
        day_end = timezone.make_aware(
            datetime.datetime.combine(today + datetime.timedelta(days=1), datetime.time.min), tz
        )
    elif date_range == 'last_7_days':
        day_end = timezone.make_aware(
            datetime.datetime.combine(today + datetime.timedelta(days=1), datetime.time.min), tz
        )
    elif date_range == 'last_30_days':
        day_end = timezone.make_aware(
            datetime.datetime.combine(today + datetime.timedelta(days=1), datetime.time.min), tz
        )
    elif date_range == 'custom' and to_date_str != from_date_str:
        try:
            end_d = datetime.date.fromisoformat(to_date_str)
            day_end = timezone.make_aware(
                datetime.datetime.combine(end_d + datetime.timedelta(days=1), datetime.time.min), tz
            )
        except ValueError:
            pass

    report_timestamp_q = None
    report_checked_in_at_q = None
    report_checked_out_at_q = None
    report_recorded_at_q = None

    # Time window (only when not all_day)
    time_error = False
    if not all_day and (from_time or to_time):
        if from_time and to_time:
            try:
                t_from = datetime.time.fromisoformat(from_time)
                t_to = datetime.time.fromisoformat(to_time)
                if t_to <= t_from:
                    time_error = True
                    to_time = ''
                elif from_date_str != to_date_str:
                    try:
                        start_d = datetime.date.fromisoformat(from_date_str)
                        end_d = datetime.date.fromisoformat(to_date_str)
                        if end_d >= start_d:
                            report_timestamp_q = _report_per_day_time_window_q(start_d, end_d, t_from, t_to, 'timestamp', tz)
                            report_checked_in_at_q = _report_per_day_time_window_q(start_d, end_d, t_from, t_to, 'checked_in_at', tz)
                            report_checked_out_at_q = _report_per_day_time_window_q(start_d, end_d, t_from, t_to, 'checked_out_at', tz)
                            report_recorded_at_q = _report_per_day_time_window_q(start_d, end_d, t_from, t_to, 'recorded_at', tz)
                        else:
                            time_error = True
                            to_time = ''
                    except ValueError:
                        pass
                else:
                    dt_from = timezone.make_aware(datetime.datetime.combine(filter_date, t_from), tz)
                    dt_to = timezone.make_aware(datetime.datetime.combine(filter_date, t_to), tz)
                    day_start = dt_from
                    day_end = dt_to
            except ValueError:
                pass
        elif from_time:
            try:
                t = datetime.time.fromisoformat(from_time)
                day_start = timezone.make_aware(datetime.datetime.combine(filter_date, t), tz)
            except ValueError:
                pass
        elif to_time:
            try:
                t = datetime.time.fromisoformat(to_time)
                day_end = timezone.make_aware(datetime.datetime.combine(filter_date, t), tz)
                if day_end < day_start:
                    day_end = day_start + datetime.timedelta(days=1)
            except ValueError:
                pass

    # Label for display
    if date_range == 'custom':
        if to_date_str != from_date_str:
            date_range_label = f"{filter_date.strftime('%b %d')} – {datetime.date.fromisoformat(to_date_str).strftime('%b %d, %Y')}"
        else:
            date_range_label = f"Custom ({filter_date.strftime('%b %d, %Y')})"
    else:
        labels = {
            'today': 'Today',
            'yesterday': 'Yesterday',
            'this_week': 'This week',
            'last_7_days': 'Last 7 days',
            'last_30_days': 'Last 30 days',
        }
        date_range_label = labels.get(date_range, str(filter_date))

    return (
        filter_date, day_start, day_end, from_time, to_time, search_q, event_id,
        date_range_label, from_date_str, to_date_str, time_error,
        report_timestamp_q, report_checked_in_at_q, report_checked_out_at_q, report_recorded_at_q,
    )


def _report_applied_filter_chips(request, date_range_label, from_time, to_time, search_q, event_id, active_events=None):
    """Build list of applied filter chips with remove URLs for the report filter bar.
    Only include a chip when that filter is explicitly in the request, so "remove" actually clears it.
    """
    from django.http import QueryDict
    q = request.GET.copy()
    path = request.path
    chips = []
    # Date: only show chip when user has set date params (so remove clears to "no date" and chip disappears)
    if request.GET.get('date_range') or request.GET.get('from_date'):
        q_no_date = q.copy()
        for k in ('date_range', 'from_date', 'to_date'):
            q_no_date.pop(k, None)
        remove_url = path + ('?' + q_no_date.urlencode() if q_no_date else '')
        chips.append({'label': date_range_label, 'remove_url': remove_url})
    # Time (when not all day)
    if from_time or to_time:
        q_no_time = q.copy()
        for k in ('from_time', 'to_time', 'all_day'):
            q_no_time.pop(k, None)
        time_label = (from_time or '00:00') + ' – ' + (to_time or '23:59')
        chips.append({'label': time_label, 'remove_url': path + ('?' + q_no_time.urlencode() if q_no_time else '')})
    # Search
    if search_q:
        q_no_search = q.copy()
        q_no_search.pop('search', None)
        q_no_search.pop('q', None)
        chips.append({'label': 'Search: ' + search_q[:30] + ('…' if len(search_q) > 30 else ''), 'remove_url': path + ('?' + q_no_search.urlencode() if q_no_search else '')})
    # Event (scope = specific event)
    if event_id and active_events:
        event_name = None
        for ev in active_events:
            if ev.id == event_id:
                event_name = ev.name
                break
        if event_name:
            q_no_event = q.copy()
            q_no_event.pop('event_id', None)
            chips.append({'label': 'Event: ' + event_name[:25] + ('…' if len(event_name) > 25 else ''), 'remove_url': path + ('?' + q_no_event.urlencode() if q_no_event else '')})
    return chips


def _top_departments_visitors(month_start, day_end, selected_year):
    """
    Top departments/offices visited from both VisitorEntry (who_to_visit) and VisitorVisit (department),
    so counts match the visitor list page. Returns (monthly_list, annual_list), each a list of
    dicts [{'who_to_visit': str, 'count': int}, ...] sorted by count desc, max 10.
    """
    from collections import Counter
    # Monthly: same bounds as visitors_this_month
    entry_month = VisitorEntry.objects.filter(
        timestamp__gte=month_start, timestamp__lt=day_end,
    ).values_list('who_to_visit', flat=True)
    visit_month = VisitorVisit.objects.filter(
        checked_in_at__gte=month_start, checked_in_at__lt=day_end,
    ).values_list('department', flat=True)
    monthly = Counter()
    for name in entry_month:
        monthly[(name or '').strip() or '—'] += 1
    for name in visit_month:
        monthly[(name or '').strip() or '—'] += 1
    monthly_list = [{'who_to_visit': k, 'count': c} for k, c in monthly.most_common(10)]

    # Annual: selected year
    entry_year = VisitorEntry.objects.filter(
        timestamp__year=selected_year,
    ).values_list('who_to_visit', flat=True)
    visit_year = VisitorVisit.objects.filter(
        checked_in_at__year=selected_year,
    ).values_list('department', flat=True)
    annual = Counter()
    for name in entry_year:
        annual[(name or '').strip() or '—'] += 1
    for name in visit_year:
        annual[(name or '').strip() or '—'] += 1
    annual_list = [{'who_to_visit': k, 'count': c} for k, c in annual.most_common(10)]
    return monthly_list, annual_list


def _get_active_events():
    """Events that are active or scheduled for today and on-campus (for gate scan "Tracking event" dropdown)."""
    today = timezone.localdate()
    return Event.objects.filter(
        status__in=('active', 'scheduled'),
        start_date__lte=today,
        end_date__gte=today,
        event_location='on_campus',
    ).order_by('start_date', 'name')


def _event_audience_students_qs(event):
    """
    Return active students included by event audience scope.
    For specific_students scope, uses EventRegistration list.
    """
    scope = (getattr(event, 'audience_scope', 'all') or 'all').strip().lower()
    qs = Student.objects.filter(is_active=True)

    if scope == 'all':
        return qs
    if scope == 'specific_students':
        student_ids = EventRegistration.objects.filter(
            event=event, status='active'
        ).values_list('student_id', flat=True)
        return qs.filter(id__in=student_ids)

    course = (getattr(event, 'audience_course', '') or '').strip()
    year = (getattr(event, 'audience_year_level', '') or '').strip()
    section = (getattr(event, 'audience_section', '') or '').strip()

    if scope in ('course', 'course_year', 'course_section', 'course_section_year'):
        if not course:
            return qs.none()
        qs = qs.filter(course__iexact=course)
    if scope in ('year_level', 'course_year', 'course_section_year'):
        if not year:
            return qs.none()
        qs = qs.filter(year_level=year)
    if scope in ('course_section', 'course_section_year'):
        if not section:
            return qs.none()
        qs = qs.filter(section__iexact=section)
    if scope == 'year_level' and not year:
        return qs.none()
    return qs


def _is_student_allowed_for_event(event, student):
    """Check if student is in event audience (all/course/year/course+year/course+section/specific_students)."""
    if student is None:
        return False
    scope = (getattr(event, 'audience_scope', 'all') or 'all').strip().lower()
    if scope == 'specific_students':
        return EventRegistration.objects.filter(event=event, student=student, status='active').exists()
    if hasattr(event, 'audience_matches_student'):
        return event.audience_matches_student(student)
    return True


def _can_override_audience(user):
    """Only admin/faculty may override event audience rules."""
    try:
        return has_role(user, 'admin', 'faculty')
    except Exception:
        return False


def _should_override_audience(request):
    """Return True when override checkbox/flag is set in the request."""
    val = (request.POST.get('override_audience') or request.GET.get('override_audience') or '').strip().lower()
    return val in ('1', 'true', 'on', 'yes', 'y')


def _get_past_events_with_stats(limit=20):
    """Events that have ended (end_date < today), with attendance summary for analytics."""
    today = timezone.localdate()
    past_events = list(
        Event.objects.filter(end_date__lt=today)
        .order_by('-end_date', '-start_date')[:limit]
    )
    if not past_events:
        return []
    event_ids = [e.pk for e in past_events]
    stats = (
        EventAttendance.objects.filter(event_id__in=event_ids)
        .values('event_id')
        .annotate(
            total=Count('id'),
            participated_count=Count(Case(When(participated=True, then=1), output_field=IntegerField())),
            checked_in_count=Count(Case(When(checked_in_at__isnull=False, then=1), output_field=IntegerField())),
        )
    )
    by_event = {s['event_id']: s for s in stats}
    result = []
    for event in past_events:
        s = by_event.get(event.pk, {})
        result.append({
            'event': event,
            'total_tracked': s.get('total', 0),
            'participated': s.get('participated_count', 0),
            'checked_in': s.get('checked_in_count', 0),
        })
    return result


def _format_event_time_12h(dt):
    """Format datetime for event exports: standard 12-hour time e.g. '2/25/2026 9:39 PM'."""
    if not dt:
        return ''
    local = timezone.localtime(dt)
    h = local.hour
    if h == 0:
        h12, am_pm = 12, 'AM'
    elif h < 12:
        h12, am_pm = h, 'AM'
    else:
        h12, am_pm = (12 if h == 12 else h - 12), 'PM'
    return local.strftime('%m/%d/%Y') + ' {}:{:02d} {}'.format(h12, local.minute, am_pm)


# Lunch break: 11:59 AM–12:59 PM — relaxed OUT handling during lunch
LUNCH_EXIT_START = datetime.time(11, 59, 0)   # 11:59 AM
LUNCH_EXIT_END = datetime.time(12, 59, 59)    # 12:59:59 PM (inclusive)


def _is_lunch_exit_window(now=None):
    """True if current time is in lunch break (11:59 AM–12:59 PM)."""
    if now is None:
        now = timezone.localtime(timezone.now())
    now_time = now.time() if hasattr(now, 'time') else now
    return LUNCH_EXIT_START <= now_time <= LUNCH_EXIT_END


def _http_request_wants_json(request):
    """
    True when the client expects JSON (not an HTML redirect).
    Some browsers/proxies omit X-Requested-With; jQuery still sends Accept: application/json for dataType=json.
    """
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return True
    accept = (request.headers.get('Accept') or '').lower()
    if 'application/json' in accept:
        return True
    if (request.POST.get('ajax') or '').strip().lower() in ('1', 'true', 'yes', 'on'):
        return True
    if (request.POST.get('json') or '').strip().lower() in ('1', 'true', 'yes', 'on'):
        return True
    return False


def _is_within_event_window(event, now=None, grace_minutes=30):
    """Check if current time is within event's time window (with grace period)."""
    if now is None:
        now = timezone.now()
    
    # Convert dates to timezone-aware datetime for comparison
    # start_date and end_date are DateFields, so we convert to datetime
    tz = timezone.get_current_timezone()
    start_dt = timezone.make_aware(
        datetime.datetime.combine(event.start_date, datetime.time.min),
        timezone=tz
    )
    end_dt = timezone.make_aware(
        datetime.datetime.combine(event.end_date, datetime.time.max),
        timezone=tz
    )
    
    # Add grace period
    start_grace = start_dt - timezone.timedelta(minutes=grace_minutes)
    end_grace = end_dt + timezone.timedelta(minutes=grace_minutes)
    
    return start_grace <= now <= end_grace


# Whole-day events: return IN only in this window (after lunch OUT)
AFTERNOON_RETURN_IN_START = datetime.time(12, 45, 0)
AFTERNOON_RETURN_IN_END = datetime.time(13, 0, 0)


def _event_agenda_rows_with_start_end(event):
    """Agenda rows that define a timed session (both start and end). Empty => treat as whole-day."""
    return list(
        EventAgenda.objects.filter(event=event)
        .exclude(start_time__isnull=True)
        .exclude(end_time__isnull=True)
    )


def _is_whole_day_event_schedule(event):
    """No session times on agenda => whole-day rules (lunch OUT, afternoon return IN)."""
    return len(_event_agenda_rows_with_start_end(event)) == 0


def _time_in_afternoon_return_window(t):
    """12:45 PM–1:00 PM inclusive (for whole-day return IN after lunch)."""
    return AFTERNOON_RETURN_IN_START <= t <= AFTERNOON_RETURN_IN_END


def _time_in_session_out_window(now_local, end_time, minutes_before=15, grace_after_minutes=5):
    """
    OUT allowed in the last `minutes_before` before end_time, plus optional grace after end.
    Compares clock times on the same calendar day (no overnight sessions).
    """
    if not end_time:
        return False
    now_t = now_local.time() if hasattr(now_local, 'time') else now_local
    end_m = end_time.hour * 60 + end_time.minute
    start_m = max(0, end_m - minutes_before)
    now_m = now_t.hour * 60 + now_t.minute
    grace_end = end_m + grace_after_minutes
    return start_m <= now_m <= grace_end


def _format_time_ampm(t):
    """Format datetime.time for messages (12-hour, portable)."""
    if not t:
        return ''
    h = t.hour % 12 or 12
    ampm = 'AM' if t.hour < 12 else 'PM'
    return f'{h}:{t.minute:02d} {ampm}'


def _scan_event_schedule_allowed(event, scan_type, now_local, attendance, reg, is_token_based):
    """
    Event attendance scanner only: enforce IN/OUT windows from agenda or whole-day rules.
    Returns (ok: bool, message: str).
    """
    whole_day = _is_whole_day_event_schedule(event)
    now_t = now_local.time() if hasattr(now_local, 'time') else now_local

    if is_token_based and reg:
        cin, cout = reg.checked_in_at, reg.checked_out_at
    else:
        cin, cout = attendance.checked_in_at, attendance.checked_out_at

    if whole_day:
        if scan_type == 'OUT':
            if not _is_lunch_exit_window(now_local):
                return False, (
                    'For whole-day events, check-out is only allowed during lunch break '
                    f'({_format_time_ampm(LUNCH_EXIT_START)}–{_format_time_ampm(LUNCH_EXIT_END)}).'
                )
            return True, ''
        # IN
        first_in = cin is None
        if first_in:
            if _time_in_afternoon_return_window(now_t):
                return False, (
                    'First check-in for whole-day events is not allowed between '
                    f'{_format_time_ampm(AFTERNOON_RETURN_IN_START)} and '
                    f'{_format_time_ampm(AFTERNOON_RETURN_IN_END)} '
                    '(reserved for return after lunch).'
                )
            return True, ''
        # Return IN after lunch OUT
        if cin is not None and cout is not None and cout > cin:
            if not _time_in_afternoon_return_window(now_t):
                return False, (
                    'After lunch, check-in again only between '
                    f'{_format_time_ampm(AFTERNOON_RETURN_IN_START)} and '
                    f'{_format_time_ampm(AFTERNOON_RETURN_IN_END)}.'
                )
            return True, ''
        return True, ''

    # Session(s) from agenda: OUT only in the last 15 minutes before each session end (per row).
    rows = _event_agenda_rows_with_start_end(event)
    if scan_type == 'OUT':
        allowed = any(
            _time_in_session_out_window(now_local, row.end_time)
            for row in rows
        )
        if not allowed:
            ends = sorted({_format_time_ampm(row.end_time) for row in rows})
            return False, (
                'Check-out is only allowed in the last 15 minutes before a scheduled session end '
                f'(sessions end at {", ".join(ends)}).'
            )
        return True, ''
    return True, ''


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
@ensure_csrf_cookie
def gate_scan(request):
    """Gate kiosk: assigned staff logs in and scans; physical security may assist (no separate login)."""
    from gate_analytics.roles import get_user_role
    event_raw = (request.GET.get('event') or '').strip()
    if event_raw.isdigit():
        gd_url = _guard_display_scanner_url_for_event(request, event_raw)
        if gd_url:
            return redirect(gd_url)
    user_role = get_user_role(request.user)
    active_events = _get_active_events()
    kiosk_mode = (request.GET.get('kiosk') or '').strip().lower() in ('1', 'true', 'yes', 'on')
    return render(request, 'gate/gate_scan.html', {
        'site_name': 'City College of Bayawan',
        'page_title': 'Gate Entry - Scan Student ID',
        'active_events': active_events,
        'campus_departments': CAMPUS_DEPARTMENT_CHOICES,
        'user_role': user_role,
        'kiosk_mode': kiosk_mode,
        'guard_student_popup_style': getattr(settings, 'GATE_GUARD_STUDENT_POPUP_STYLE', 'split'),
    })


@require_GET
def gate_scan_sw(request):
    """Serve the gate scan service worker (for offline support)."""
    static_dirs = getattr(settings, 'STATICFILES_DIRS', None) or []
    path = os.path.join(static_dirs[0], 'js', 'sw-gate-scan.js') if static_dirs else None
    if not path or not os.path.isfile(path):
        return HttpResponse('/* service worker not found */', content_type='application/javascript')
    with open(path, 'r', encoding='utf-8') as f:
        return HttpResponse(f.read(), content_type='application/javascript')


@require_GET
@ensure_csrf_cookie
def guard_scanner_dashboard(request):
    """Guard wall: same QR scanner UX as /gate/ (token auth, no login). Uses GATE_GUARD_DISPLAY_TOKEN."""
    expected = getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or ''
    print(f"[DEBUG Guard Dashboard] Expected token: '{expected}'")
    if not expected:
        return HttpResponseForbidden(
            'Guard display is not configured. Set GATE_GUARD_DISPLAY_TOKEN in the server environment.'
        )
    token = (request.GET.get('token') or '').strip()
    print(f"[DEBUG Guard Dashboard] Received token: '{token}'")
    if token != expected:
        print(f"[DEBUG Guard Dashboard] Token mismatch! Expected '{expected}', got '{token}'")
        return HttpResponseForbidden('Invalid or missing token.')
    active_events = _get_active_events()
    # Block the guard-wall until staff clicks Start on the dashboard; stays open until Stop (even if they log out).
    from gate.gate_personnel_views import gate_scanner_session_armed

    scanner_active = gate_scanner_session_armed()
    print(f"[DEBUG Guard Dashboard] Scanner active from cache: {scanner_active}")
    print(f"[DEBUG Guard Dashboard] Rendering with guard_wall=True, guard_embed_token='{token}'")
    event_attendance_scanner_event = None
    event_raw = (request.GET.get('event') or '').strip()
    if event_raw.isdigit():
        event_attendance_scanner_event = Event.objects.filter(pk=int(event_raw)).first()
    page_title = 'Gate Entry - Scan Student ID'
    if event_attendance_scanner_event:
        page_title = f'Event attendance – {event_attendance_scanner_event.name}'
    resp = render(request, 'gate/gate_scan.html', {
        'site_name': getattr(settings, 'SITE_NAME', 'City College of Bayawan'),
        'page_title': page_title,
        'active_events': active_events,
        'campus_departments': CAMPUS_DEPARTMENT_CHOICES,
        'user_role': 'staff',
        'kiosk_mode': True,
        'guard_embed': False,
        'guard_wall': True,
        'guard_embed_token': token,
        'guard_student_popup_style': getattr(settings, 'GATE_GUARD_STUDENT_POPUP_STYLE', 'split'),
        'guard_scanner_active': scanner_active,
        'event_attendance_scanner_event': event_attendance_scanner_event,
        # Visible build/debug id for field testing
        'guard_build_id': getattr(settings, 'GATE_GUARD_BUILD_ID', '') or 'sw-v35',
    })
    # Prevent any intermediate caching (ngrok/browser/proxy).
    resp['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    resp['Pragma'] = 'no-cache'
    resp['Expires'] = '0'
    return resp


@require_GET
@ensure_csrf_cookie
def gate_scan_guard_embed(request):
    """Token-only gate scanner for guard monitors (camera + IN/OUT; no staff login)."""
    expected = getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or ''
    if not expected:
        return HttpResponseForbidden(
            'Guard embed is not configured. Set GATE_GUARD_DISPLAY_TOKEN in the server environment.'
        )
    token = (request.GET.get('token') or '').strip()
    if token != expected:
        return HttpResponseForbidden('Invalid or missing token.')
    active_events = _get_active_events()
    return render(request, 'gate/gate_scan.html', {
        'site_name': getattr(settings, 'SITE_NAME', 'City College of Bayawan'),
        'page_title': 'Gate Entry - Scan Student ID',
        'active_events': active_events,
        'campus_departments': CAMPUS_DEPARTMENT_CHOICES,
        'user_role': 'staff',
        'kiosk_mode': True,
        'guard_embed': True,
        'guard_embed_token': token,
        'guard_student_popup_style': getattr(settings, 'GATE_GUARD_STUDENT_POPUP_STYLE', 'split'),
    })


@require_GET
@login_required(login_url='/login/')
def lookup_student(request):
    """Lookup student by ID (from QR scan). Returns JSON for AJAX or redirect for form."""
    student_id = (request.GET.get('student_id') or '').strip()
    if not student_id:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'found': False, 'error': 'No student ID provided'})
        return redirect('gate-scan')
    try:
        student = Student.objects.get(student_id=student_id, is_active=True)
    except Student.DoesNotExist:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({
                'found': False,
                'student_id': student_id,
                'error': 'Student not found or inactive',
            })
        return render(request, 'gate/gate_scan.html', {
            'site_name': 'City College of Bayawan',
            'error': 'Student not found or inactive.',
            'scanned_id': student_id,
        })
    active_events = _get_active_events()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({
            'found': True,
            'student': {
                'id': student.pk,
                'student_id': student.student_id,
                'full_name': student.get_full_name(),
                'email': student.email or '',
                'photo_url': student.photo.url if student.photo else None,
            },
            'active_events': [{'id': e.id, 'name': e.name} for e in active_events],
        })
    return render(request, 'gate/gate_verify.html', {
        'site_name': 'City College of Bayawan',
        'student': student,
        'active_events': active_events,
    })


@require_POST
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
@transaction.atomic
def record_entry(request):
    """Record gate entry: grant or deny. Creates GateEntry, optionally GateIncident and EventAttendance."""
    student_id = request.POST.get('student_id')
    granted = request.POST.get('granted', 'true').lower() in ('1', 'true', 'yes')
    reason = request.POST.get('reason', 'identity_mismatch')
    notes = request.POST.get('notes', '')
    student_pk = request.POST.get('student_pk')

    if granted and student_pk:
        try:
            student = Student.objects.get(pk=student_pk, student_id=student_id, is_active=True)
        except Student.DoesNotExist:
            return JsonResponse({'ok': False, 'error': 'Student not found'}, status=400)
        # Access control uses account_status/is_active only: APPROVED=active, INACTIVE=frozen.
        # Grant gate entry only. Event attendance is recorded separately via Events → Attendance Scanner (instructor).
        GateEntry.objects.create(
            student=student, granted=True, notes=notes, recorded_by=request.user,
            **_audit_kwargs_for_gate_entry(request),
        )
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'message': 'Entry granted'})
        return redirect('gate-scan')
    elif not granted:
        valid_incident_reasons = {c[0] for c in GateIncident.REASON_CHOICES}
        inc_reason = (reason or 'identity_mismatch').strip()
        if inc_reason not in valid_incident_reasons:
            inc_reason = 'identity_mismatch'
        student = None
        if student_pk:
            try:
                student = Student.objects.get(pk=student_pk)
            except Student.DoesNotExist:
                pass
        if student:
            incident = GateIncident.objects.create(
                student=student,
                reason=inc_reason,
                details=(notes or '')[:2000],
                scanned_id=(student.student_id or '')[:100],
                staff_alerted=True,
            )
            _apply_student_office_hold(student, 'Blocked: unresolved gate incident. Resolve at office first.')
            GateEntry.objects.create(
                student=student,
                granted=False,
                incident=incident,
                notes=notes,
                recorded_by=request.user,
                result='DENIED',
                **_audit_kwargs_for_gate_entry(request),
            )
            try:
                from .admin_notification_service import AdminNotificationService
                AdminNotificationService.notify_incident(incident)
            except Exception:
                pass
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return JsonResponse({'ok': True, 'message': 'Entry denied'})
        return redirect('gate-scan')

    return JsonResponse({'ok': False, 'error': 'Invalid request'}, status=400)


def _save_scan_incident_and_entry(
    *,
    student,
    result,
    notes,
    incident_reason,
    incident_detail,
    scanned_id,
    actor,
    request,
    device_id,
    event=None,
):
    """Log a denied scan with GateIncident + GateEntry and notify admins / Student Affairs (same as record_entry denies)."""
    valid_incident_reasons = {c[0] for c in GateIncident.REASON_CHOICES}
    inc_reason = (incident_reason or 'other').strip()
    if inc_reason not in valid_incident_reasons:
        inc_reason = 'other'
    incident = GateIncident.objects.create(
        student=student,
        reason=inc_reason,
        details=(incident_detail or '')[:2000],
        scanned_id=(scanned_id or '')[:100],
        staff_alerted=True,
    )
    if student is not None:
        _apply_student_office_hold(student, incident_detail or 'Blocked: unresolved gate incident. Resolve at office first.')
    kwargs = {
        'student': student,
        'granted': False,
        'incident': incident,
        'notes': notes,
        'result': result,
        'recorded_by': actor,
        **_audit_kwargs_for_gate_entry(request, device_id=device_id),
    }
    if event is not None:
        kwargs['event'] = event
    GateEntry.objects.create(**kwargs)
    try:
        from .admin_notification_service import AdminNotificationService

        AdminNotificationService.notify_incident(incident)
    except Exception:
        pass


@require_POST
@transaction.atomic
def save_scan(request):
    """Record a scan from the QR scanner or manual entry. Returns JSON for the scanner UI."""
    actor = _resolve_save_scan_actor(request)
    if actor is False:
        return JsonResponse({
            'success': False,
            'message': 'Unauthorized',
            'color': 'error',
        }, status=403)
    student_id = (request.POST.get('student_id') or '').strip()
    manual_status = (request.POST.get('manual_status') or '').strip().upper()  # IN or OUT
    gate_manual_entry = (request.POST.get('gate_manual_entry') or '').strip().lower()  # 'in' | 'out' | ''
    manual_gate_form = request.POST.get('manual_gate_form') == '1'  # gate_scan.html Manual Entry (Student) only
    manual_entry_reason = (request.POST.get('manual_entry_reason') or '').strip()[:200]
    manual_entry_other_detail = (request.POST.get('manual_entry_other_detail') or '').strip()[:500]
    manual_reason_code, manual_reason_label = _normalize_manual_office_reason(manual_entry_reason)
    if manual_gate_form and gate_manual_entry == 'in' and manual_reason_code == 'GATE_RESOLVED_OFFICE_ENDORSEMENT':
        if len(manual_entry_other_detail) < 2:
            return JsonResponse({
                'success': False,
                'message': (
                    'When "Others" is selected, type a short reason (e.g. which office, or why manual check-in).'
                ),
                'color': 'warning',
            })
        manual_reason_label = f'Others — {manual_entry_other_detail}'
    local_time = request.POST.get('local_time', '').strip()
    event_id = request.POST.get('event_id', '').strip()
    device_id = (request.POST.get('device_id') or '').strip()
    # Auto-register scanner device so Reports Hub shows "Scanner devices" in use
    _register_or_touch_scanner_device(device_id)

    if not student_id:
        return JsonResponse({
            'success': False,
            'message': 'No student ID provided.',
            'color': 'error',
        })

    visitor_match = re.search(r'(VIS-\d{1,6}|VISITOR-[A-Za-z0-9_-]{6,64})', student_id, flags=re.IGNORECASE)
    if visitor_match:
        student_id = visitor_match.group(1).strip()
        if student_id[:8].upper() == 'VISITOR-':
            student_id = 'VISITOR-' + student_id[8:]
        elif student_id[:4].upper() == 'VIS-':
            student_id = 'VIS-' + student_id[4:]

    # Reusable visitor pass (VIS-001 style): AVAILABLE → check-in modal; IN_USE → check-out
    if student_id[:4].upper() == 'VIS-' and len(student_id) <= 16:
        now = timezone.now()
        pass_obj = VisitorPass.objects.filter(code=student_id).select_related('current_visit').first()
        if not pass_obj:
            return JsonResponse({
                'success': False,
                'message': 'Unknown visitor pass code.',
                'color': 'error',
            })
        if pass_obj.status == VisitorPass.STATUS_DISABLED:
            return JsonResponse({
                'success': False,
                'message': 'This pass is disabled. Contact staff at the gate.',
                'color': 'error',
            })
        if pass_obj.status == VisitorPass.STATUS_AVAILABLE:
            return JsonResponse({
                'success': False,
                'need_visitor_checkin': True,
                'pass_code': pass_obj.code,
                'message': 'Visitor check-in: enter details below.',
                'color': 'info',
            })
        if pass_obj.status == VisitorPass.STATUS_IN_USE and pass_obj.current_visit_id:
            visit = pass_obj.current_visit
            # Double-scan guard: if checked in within last 30 seconds, treat as duplicate
            from datetime import timedelta
            if visit.checked_in_at and (now - visit.checked_in_at) < timedelta(seconds=30):
                return JsonResponse({
                    'success': False,
                    'message': 'Duplicate scan. Pass was just used for check-in.',
                    'color': 'warning',
                    'duplicate': True,
                })
            time_str = timezone.localtime(visit.checked_in_at).strftime('%I:%M %p')
            return JsonResponse({
                'success': False,
                'need_visitor_checkout': True,
                'pass_code': pass_obj.code,
                'visit_id': visit.id,
                'visit': {
                    'full_name': visit.full_name,
                    'department': visit.department or '',
                    'purpose': visit.purpose or '',
                    'checked_in_at': visit.checked_in_at.isoformat(),
                    'checked_in_time': time_str,
                },
                'message': f'{visit.full_name} – Check out?',
                'color': 'info',
            })
        # IN_USE but no current_visit (data inconsistency): release pass
        pass_obj.status = VisitorPass.STATUS_AVAILABLE
        pass_obj.current_visit_id = None
        pass_obj.save(update_fields=['status', 'current_visit_id'])
        return JsonResponse({
            'success': False,
            'need_visitor_checkin': True,
            'pass_code': pass_obj.code,
            'message': 'Visitor check-in: enter details below.',
            'color': 'info',
        })

    # Legacy one-time visitor pass (VISITOR-xxx)
    if student_id[:8].upper() == 'VISITOR-':
        from django.utils import timezone as tz
        now = tz.now()
        pass_obj = VisitorPass.objects.filter(code=student_id, used_at__isnull=True).first()
        if not pass_obj:
            return JsonResponse({
                'success': False,
                'message': 'Invalid or already used visitor pass.',
                'color': 'error',
            })
        if pass_obj.valid_from and pass_obj.valid_until and (now < pass_obj.valid_from or now > pass_obj.valid_until):
            return JsonResponse({
                'success': False,
                'message': f'Visitor pass valid {pass_obj.valid_from.date()}–{pass_obj.valid_until.date()} only.',
                'color': 'warning',
            })
        guest, _ = Student.objects.get_or_create(
            student_id='GUEST',
            defaults={'first_name': 'Guest', 'last_name': 'Visitor', 'email': '', 'is_active': True},
        )
        GateEntry.objects.create(
            student=guest,
            granted=True,
            notes=f"Visitor: {pass_obj.guest_name or 'Guest'} ({pass_obj.code})",
            recorded_by=actor,
            **_audit_kwargs_for_gate_entry(request),
        )
        pass_obj.used_at = now
        pass_obj.save(update_fields=['used_at'])
        time_str = tz.localtime(now).strftime('%I:%M %p')
        return JsonResponse({
            'success': True,
            'message': f'{pass_obj.guest_name or "Visitor"} (visitor) checked in.',
            'color': 'success',
            'status': 'IN',
            'student_name': pass_obj.guest_name or 'Visitor',
            'time': time_str,
        })

    lookup_raw = student_id
    student, resolve_err = _resolve_student_from_gate_input(student_id)
    if resolve_err == 'multiple':
        _save_scan_incident_and_entry(
            student=None,
            result='NOT_FOUND',
            notes='Ambiguous lookup (multiple matches)',
            incident_reason='other',
            incident_detail='Several students match that name or ID; full student ID required.',
            scanned_id=lookup_raw,
            actor=actor,
            request=request,
            device_id=device_id,
        )
        return JsonResponse({
            'success': False,
            'message': 'Several students match that name or ID. Enter the full student ID.',
            'color': 'warning',
        })
    if not student:
        # Do not create GateEntry, GateIncident, or other DB rows for unknown IDs — no attendance recorded.
        return JsonResponse({
            'success': False,
            'not_registered': True,
            'student_id': lookup_raw,
            'message': 'This student ID is not recognized. No attendance was recorded.',
            'color': 'warning',
        })
    student_id = student.student_id
    if not student.is_active or student.account_status != Student.ACCOUNT_STATUS_APPROVED:
        _save_scan_incident_and_entry(
            student=student,
            result='NOT_APPROVED',
            notes='Account inactive',
            incident_reason='other',
            incident_detail='Student account is inactive; QR / gate scan not allowed until active again.',
            scanned_id=student_id,
            actor=actor,
            request=request,
            device_id=device_id,
        )
        photo_url = _scan_ui_photo_url(request, student)
        return JsonResponse({
            'success': False,
            'inactive': True,
            'student_id': student_id,
            'message': 'This student account is inactive. The QR code cannot be used until the account is active again.',
            'color': 'warning',
            'student': {
                'student_id': student.student_id,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'email': student.email or '',
                'photo_url': photo_url,
                'course_or_section': getattr(student, 'course_or_section', '') or '',
                'year_level': getattr(student, 'year_level', '') or '',
            },
        })

    # Student block / allowlist (temporary block or allow-only window)
    today_block = timezone.localdate()
    from .models import StudentBlock
    blocking = StudentBlock.objects.filter(student=student, block_from__lte=today_block, block_until__gte=today_block)
    if blocking.filter(is_allowlist=False).exists():
        _save_scan_incident_and_entry(
            student=student,
            result='BLOCKED',
            notes='StudentBlock: access temporarily blocked',
            incident_reason='other',
            incident_detail='Access temporarily blocked (StudentBlock). Contact admin.',
            scanned_id=student_id,
            actor=actor,
            request=request,
            device_id=device_id,
        )
        return JsonResponse({
            'success': False,
            'student_id': student_id,
            'message': 'Access temporarily blocked. Contact admin.',
            'color': 'error',
        })
    if blocking.filter(is_allowlist=True).exists():
        pass  # explicit allow in this window
    elif StudentBlock.objects.filter(student=student, is_allowlist=True).exists():
        _save_scan_incident_and_entry(
            student=student,
            result='BLOCKED',
            notes='Allowlist: outside allowed date window',
            incident_reason='other',
            incident_detail='Access allowed only in a specific date window. Today is outside that window.',
            scanned_id=student_id,
            actor=actor,
            request=request,
            device_id=device_id,
        )
        return JsonResponse({
            'success': False,
            'student_id': student_id,
            'message': 'Access allowed only in a specific date window. Today is outside that window.',
            'color': 'warning',
        })

    selected_event = None
    if event_id:
        try:
            # Match scan_event_qr: any active/scheduled event (on-campus + field trip).
            # Requiring event_location='on_campus' here caused guard / locked-event scans to
            # fall through to daily gate when the event was mis-tagged or when event_id was missing.
            selected_event = Event.objects.filter(
                id=int(event_id),
                status__in=('active', 'scheduled'),
            ).first()
        except ValueError:
            pass

    if gate_manual_entry == 'in' and not manual_entry_reason:
        return JsonResponse({
            'success': False,
            'message': 'Select an office routing for manual check-in (for ID concern, choose Student Affairs).',
            'color': 'warning',
        })

    # Manual Entry (Student): check-in must use office routing (gate_manual_entry=in + reason above).
    # Blocks forged POSTs that send manual_status=IN without the manual gate path.
    if manual_gate_form and manual_status == 'IN' and gate_manual_entry != 'in':
        return JsonResponse({
            'success': False,
            'message': 'Select an office routing for manual check-in before recording entry.',
            'color': 'warning',
        })

    today = timezone.localdate()
    gate_eval_result = None  # set in daily-gate path for schedule_hint / out_reason_code


    # Event tracking: separate from daily gate. Only check event-specific attendance.
    if selected_event:
        # Must NOT use get_student_current_state() here — that mixes daily gate + all events and
        # suggests OUT when the student is already "inside" from morning gate, breaking first
        # event scan. Mirror scan_event_qr: IN until checked in for this event, then OUT.
        att0, _ = EventAttendance.objects.get_or_create(
            student=student,
            event=selected_event,
            defaults={'participated': False},
        )
        suggested = 'IN' if att0.checked_in_at is None else 'OUT'
        if gate_manual_entry == 'out':
            status = 'OUT'
        elif gate_manual_entry == 'in':
            status = 'IN'
        else:
            status = manual_status if manual_status in ('IN', 'OUT') else suggested
        day_start, day_end = _local_day_bounds(today)
        if not _is_student_allowed_for_event(selected_event, student):
            aud = selected_event.audience_summary()
            _save_scan_incident_and_entry(
                student=student,
                result='DENIED',
                notes=f'Event audience: {aud}',
                incident_reason='other',
                incident_detail=(
                    f'{student.get_full_name()} is not included in this event audience ({aud}).'
                ),
                scanned_id=student_id,
                actor=actor,
                request=request,
                device_id=device_id,
                event=selected_event,
            )
            photo_url = _scan_ui_photo_url(request, student)
            return JsonResponse({
                'success': False,
                'message': f'{student.get_full_name()} is not included in this event audience ({selected_event.audience_summary()}).',
                'color': 'warning',
                'student_name': student.get_full_name(),
                'student': {
                    'first_name': student.first_name,
                    'middle_name': student.middle_name or '',
                    'last_name': student.last_name,
                    'student_id': student.student_id,
                    'email': student.email or '',
                    'photo_url': photo_url or '',
                    'course_or_section': getattr(student, 'course_or_section', '') or '',
                    'year_level': getattr(student, 'year_level', '') or '',
                },
            })
        event_entry_today = GateEntry.objects.filter(
            student=student, event=selected_event, granted=True,
            timestamp__gte=day_start, timestamp__lt=day_end,
        ).order_by('-timestamp').first()
        already_in_event = event_entry_today is not None and (event_entry_today.notes or '').strip().upper() != 'OUT'
        if already_in_event and status == 'IN':
            first_today = GateEntry.objects.filter(
                student=student, event=selected_event, granted=True,
                timestamp__gte=day_start, timestamp__lt=day_end,
            ).order_by('timestamp').first()
            first_time = timezone.localtime(first_today.timestamp).strftime('%I:%M %p') if first_today else ''
            photo_url = _scan_ui_photo_url(request, student)
            return JsonResponse({
                'success': False,
                'already_scanned': True,
                'message': f'{student.get_full_name()} already checked in to this event.',
                'color': 'warning',
                'student_name': student.get_full_name(),
                'student': {
                    'first_name': student.first_name,
                    'middle_name': student.middle_name or '',
                    'last_name': student.last_name,
                    'student_id': student.student_id,
                    'email': student.email or '',
                    'photo_url': photo_url,
                    'course_or_section': getattr(student, 'course_or_section', '') or '',
                    'year_level': getattr(student, 'year_level', '') or '',
                },
                'first_scan_time': first_time,
            })
    else:
        # Daily gate (no event): IN/OUT from evaluate_scan() — college-style, no class schedule.
        state = get_student_current_state(student, today, daily_gate_only=True)
        suggested = 'OUT' if state == 'INSIDE' else 'IN'
        if gate_manual_entry == 'out':
            status = 'OUT'
        elif gate_manual_entry == 'in':
            status = 'IN'
        else:
            status = manual_status if manual_status in ('IN', 'OUT') else suggested
        # Duplicate = already scanned at gate today (entry with no event).
        # Use local calendar day bounds so "today" matches the gate entries list (avoids UTC vs local date mismatch).
        day_start, day_end = _local_day_bounds(today)
        today_entries = GateEntry.objects.filter(
            student=student, granted=True, event__isnull=True,
            timestamp__gte=day_start, timestamp__lt=day_end,
        )
        latest_today = today_entries.order_by('-timestamp').first()
        already_inside = latest_today is not None and _entry_is_in_direction(latest_today)
        now_dt = timezone.localtime(timezone.now())
        # Lunch: staff tries IN while already inside during lunch — auto OUT (before repeat cooldown).
        if already_inside and status == 'IN':
            in_lunch_window = _is_lunch_exit_window(now_dt)
            if in_lunch_window:
                # During lunch (11:59 AM–12:59 PM): automatically record OUT and return success — no modal
                eval_out = evaluate_scan(
                    student, 'OUT', now_dt,
                    personnel_override_reason=None,
                    daily_gate_only=True,
                )
                if not eval_out['allowed']:
                    eval_out = {
                        'allowed': True,
                        'result': 'SUCCESS',
                        'message': 'Lunch break exit allowed.',
                        'out_reason_code': 'LUNCH',
                        'out_reason_text': 'Lunch break',
                        'schedule_hint': '',
                        'next_suggested': 'IN',
                        'schedule_based': False,
                    }
                GateEntry.objects.create(
                    student=student,
                    event=None,
                    granted=True,
                    result='SUCCESS',
                    scan_type='OUT',
                    notes='OUT',
                    out_reason=eval_out.get('out_reason_text') or 'Lunch break',
                    out_reason_code=eval_out.get('out_reason_code') or 'LUNCH',
                    recorded_by=actor,
                    **_audit_kwargs_for_gate_entry(request, device_id=device_id),
                )
                time_str = now_dt.strftime('%I:%M %p')
                photo_url = _scan_ui_photo_url(request, student)
                return JsonResponse({
                    'success': True,
                    'message': f'{student.get_full_name()} checked out (lunch break).',
                    'color': 'success',
                    'status': 'OUT',
                    'student_name': student.get_full_name(),
                    'time': time_str,
                    'result': 'ALLOWED',
                    'student': {
                        'first_name': student.first_name,
                        'middle_name': student.middle_name or '',
                        'last_name': student.last_name,
                        'student_id': student.student_id,
                        'email': student.email or '',
                        'photo_url': photo_url or '',
                        'course_or_section': getattr(student, 'course_or_section', '') or '',
                        'year_level': getattr(student, 'year_level', '') or '',
                    },
                    'schedule_hint': 'Lunch break (11:59 AM–12:59 PM): exit recorded.',
                    'next_suggested': 'IN',
                    'all_classes_done': False,
                })
        # Global repeat cooldown: minimum time after ANY daily gate scan before the next (default).
        # Without this, auto IN/OUT suggestion always alternates, so "same-direction" cooldown never triggers on QR scans.
        _cool_scope = getattr(settings, 'GATE_SCAN_REPEAT_COOLDOWN_SCOPE', 'global')
        if _cool_scope == 'global' and latest_today:
            cooldown_td = daily_gate_repeat_cooldown()
            elapsed_since_last = now_dt - timezone.localtime(latest_today.timestamp)
            if elapsed_since_last < cooldown_td:
                secs_left = max(0.0, (cooldown_td - elapsed_since_last).total_seconds())
                wait_human = format_cooldown_wait_remaining(secs_left)
                first_today = today_entries.order_by('timestamp').first()
                first_time = timezone.localtime(first_today.timestamp).strftime('%I:%M %p') if first_today else ''
                photo_url = _scan_ui_photo_url(request, student)
                return JsonResponse({
                    'success': False,
                    'already_scanned': True,
                    'repeat_cooldown': True,
                    'message': (
                        f'{student.get_full_name()} was scanned recently. '
                        f'Please wait {wait_human} before the next gate scan.'
                    ),
                    'color': 'warning',
                    'student_name': student.get_full_name(),
                    'student': {
                        'first_name': student.first_name,
                        'middle_name': student.middle_name or '',
                        'last_name': student.last_name,
                        'student_id': student.student_id,
                        'email': student.email or '',
                        'photo_url': photo_url or '',
                        'course_or_section': getattr(student, 'course_or_section', '') or '',
                        'year_level': getattr(student, 'year_level', '') or '',
                    },
                    'first_scan_time': first_time,
                    'in_lunch_window': _is_lunch_exit_window(now_dt),
                    'in_lunch_exit_window': _is_lunch_exit_window(now_dt),
                })

        # Daily gate: college-style IN/OUT (evaluate_scan — no class schedule)
        override_reason = (request.POST.get('out_reason') or request.POST.get('reason') or request.POST.get('note') or '').strip() or None
        eval_result = evaluate_scan(
            student, status, now_dt,
            personnel_override_reason=override_reason,  # required for forced OUT when still OUTSIDE
            daily_gate_only=True,
        )
        # Safety net: during lunch window (11:59 AM–12:59 PM), allow both IN and OUT
        if not eval_result['allowed'] and _is_lunch_exit_window(now_dt):
            if status == 'OUT':
                eval_result = {
                    'allowed': True,
                    'result': 'SUCCESS',
                    'message': 'Lunch break exit allowed.',
                    'out_reason_code': 'LUNCH',
                    'out_reason_text': override_reason or 'Lunch break',
                    'schedule_hint': eval_result.get('schedule_hint', ''),
                    'next_suggested': 'IN',
                    'deny_detail': '',
                    'schedule_based': False,
                }
            elif status == 'IN':
                eval_result = {
                    'allowed': True,
                    'result': 'SUCCESS',
                    'message': 'Return from lunch allowed (lunch window).',
                    'out_reason_code': '',
                    'out_reason_text': '',
                    'schedule_hint': eval_result.get('schedule_hint', ''),
                    'next_suggested': 'OUT',
                    'deny_detail': '',
                    'schedule_based': False,
                }
        if not eval_result['allowed']:
            photo_url = _scan_ui_photo_url(request, student)
            now_str = now_dt.strftime('%I:%M %p')
            in_class = False
            all_classes_done = True
            all_classes_done_based_on_schedule = False
            in_lunch_window = _is_lunch_exit_window(now_dt)
            return JsonResponse({
                'success': False,
                'message': eval_result['message'],
                'color': 'warning',
                'schedule_hint': eval_result.get('schedule_hint', ''),
                'next_suggested': eval_result.get('next_suggested', suggested),
                'deny_detail': eval_result.get('deny_detail', eval_result.get('deny_reason', '')),
                'require_note': eval_result.get('result') == 'REQUIRE_REASON',
                'status': status,
                'time': now_str,
                'denied': True,
                'student_name': student.get_full_name(),
                'student': {
                    'first_name': student.first_name,
                    'middle_name': student.middle_name or '',
                    'last_name': student.last_name,
                    'student_id': student.student_id,
                    'email': student.email or '',
                    'photo_url': photo_url or '',
                    'course_or_section': getattr(student, 'course_or_section', '') or '',
                    'year_level': getattr(student, 'year_level', '') or '',
                },
                'all_classes_done': all_classes_done,
                'all_classes_done_based_on_schedule': all_classes_done_based_on_schedule,
                'in_class_now': in_class,
                'class_until': None,
                'schedule_based': eval_result.get('schedule_based', False),
                'in_lunch_exit_window': in_lunch_window,
            })
        gate_eval_result = eval_result

    notes = status
    if gate_manual_entry == 'in' and manual_entry_reason:
        notes = f'{status} | Manual office: {manual_reason_label or manual_entry_reason}'
    elif gate_manual_entry == 'out':
        notes = f'{status} | Manual (no ID scan)'
    out_reason_code = (gate_eval_result.get('out_reason_code') or '') if gate_eval_result else ''
    out_reason_text = (gate_eval_result.get('out_reason_text') or '') if gate_eval_result else ''

    entry = GateEntry.objects.create(
        student=student,
        event=selected_event,
        granted=True,
        result='SUCCESS',
        scan_type=status,
        notes=notes,
        out_reason=out_reason_text,
        out_reason_code=out_reason_code or '',
        recorded_by=actor,
        **_audit_kwargs_for_gate_entry(request, device_id=device_id),
    )
    if gate_manual_entry == 'in' and (manual_reason_code or manual_entry_reason):
        _notify_manual_office_referral(
            student=student,
            actor=actor,
            entry=entry,
            office_code=manual_reason_code,
            office_label=manual_reason_label or manual_entry_reason,
            selected_event=selected_event,
        )
    
    # Log gate activity for scan (staff/faculty/admin)
    if actor is not None and get_user_role(actor) in ('admin', 'staff', 'faculty'):
        from .gate_personnel_services import GateActivityLogger
        GateActivityLogger.log_scan(
            guard=actor,
            entry=entry,
            device_id=device_id,
            ip_address=request.META.get('REMOTE_ADDR')
        )
    
    # If event is selected (gate scan with "Tracking event"), update EventAttendance check-in/out times
    if selected_event:
        att, _ = EventAttendance.objects.get_or_create(
            student=student,
            event=selected_event,
            defaults={'participated': True},
        )
        att.participated = True
        now_att = timezone.now()
        update_fields = ['participated']
        if status == 'IN' and att.checked_in_at is None:
            att.checked_in_at = now_att
            update_fields.append('checked_in_at')
        elif status == 'OUT' and att.checked_in_at is not None and att.checked_out_at is None:
            att.checked_out_at = now_att
            update_fields.append('checked_out_at')
        att.save(update_fields=update_fields)
    
    now = timezone.localtime(timezone.now())
    time_str = now.strftime('%I:%M %p')

    if status == 'IN':
        message = f'{student.get_full_name()} checked in.'
    else:
        message = f'{student.get_full_name()} checked out.'

    photo_url = _scan_ui_photo_url(request, student)

    resp = {
        'success': True,
        'message': message,
        'color': 'success',
        'status': status,
        'student_name': student.get_full_name(),
        'time': time_str,
        'result': 'ALLOWED',
        'student': {
            'first_name': student.first_name,
            'middle_name': student.middle_name or '',
            'last_name': student.last_name,
            'student_id': student.student_id,
            'email': student.email or '',
            'photo_url': photo_url,
            'course_or_section': getattr(student, 'course_or_section', '') or '',
            'year_level': getattr(student, 'year_level', '') or '',
        },
    }
    if selected_event:
        resp['event_attendance'] = True
    if gate_eval_result:
        resp['schedule_hint'] = gate_eval_result.get('schedule_hint', '')
        resp['next_suggested'] = gate_eval_result.get('next_suggested', 'OUT' if state == 'INSIDE' else 'IN')
        resp['schedule_based'] = gate_eval_result.get('schedule_based', False)
        if status == 'OUT' and (gate_eval_result.get('out_reason_code') == 'ALL_CLASSES_DONE'):
            resp['all_classes_done'] = True
        if gate_eval_result.get('forced_out_no_in'):
            resp['forced_out_no_in'] = True
            resp['message'] = gate_eval_result.get('message', resp.get('message', ''))
    return JsonResponse(resp)


@require_POST
@csrf_exempt
@transaction.atomic
def save_scan_guard(request):
    """
    Guard-display/embed scanner submit endpoint.
    CSRF is bypassed because these clients are token-authenticated and often not logged in.
    Still requires a valid guard token (GATE_GUARD_DISPLAY_TOKEN) in POST or header.
    """
    token = (request.POST.get('guard_token') or request.headers.get('X-Gate-Guard-Token') or '').strip()
    expected = getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or ''
    if not (token and expected and token == expected):
        return JsonResponse({'success': False, 'message': 'Unauthorized', 'color': 'error'}, status=403)
    return save_scan(request)


@require_POST
@transaction.atomic
def scan_event_qr(request):
    """
    Hybrid event attendance scanning - accepts BOTH:
    1. Permanent student QR (student_id) - for regular events
    2. Token-based QR (EVT:<event_id>:<token>) - for secure events
    
    Expects: {event_id, qr, scan_type, device_id, client_scan_time}
    """
    actor = _resolve_save_scan_actor(request)
    if actor is False:
        return JsonResponse({
            'ok': False,
            'result': 'DENIED',
            'message': 'Unauthorized',
            'color': 'error',
        }, status=403)
    import datetime as py_datetime
    
    event_id = request.POST.get('event_id', '').strip()
    qr = (request.POST.get('qr') or '').strip()
    # Client may send scan_type; we replace it below from registration/attendance state.
    scan_type = request.POST.get('scan_type', 'IN').strip().upper()
    device_id = request.POST.get('device_id', '').strip()
    client_scan_time_str = request.POST.get('client_scan_time', '').strip()
    
    # Parse client scan time if provided (for offline scans)
    client_scan_time = None
    if client_scan_time_str:
        try:
            s = client_scan_time_str.replace('Z', '+00:00')
            dt = py_datetime.datetime.fromisoformat(s)
            client_scan_time = dt if timezone.is_aware(dt) else timezone.make_aware(dt, timezone.get_current_timezone())
        except (ValueError, AttributeError):
            client_scan_time = None
    
    # Validate event exists
    try:
        event = Event.objects.get(id=int(event_id))
    except (Event.DoesNotExist, ValueError):
        return JsonResponse({
            'ok': False,
            'result': 'INVALID',
            'message': 'Event not found.',
            'color': 'error',
        }, status=400)

    recorded_by = actor

    # Auto-register scanner device on first use so dashboard shows "current scanner devices"
    _register_or_touch_scanner_device(device_id)
    
    # Determine QR type: token-based (EVT:event_id:token) or student ID
    is_token_based = qr.startswith('EVT:')
    
    if is_token_based:
        # TOKEN-BASED QR (for high-security events)
        parts = qr.split(':')
        if len(parts) != 3:
            AttendanceLog.objects.create(
                event=event,
                scan_type=scan_type,
                result='INVALID',
                token='',
                device_id=device_id,
                recorded_by=recorded_by,
                client_scan_time=client_scan_time,
                remarks=f'Invalid token QR format: {qr[:255]}'
            )
            return JsonResponse({
                'ok': False,
                'result': 'INVALID',
                'message': 'Invalid QR code format.',
                'color': 'error',
            }, status=400)
        
        qr_event_id = parts[1]
        token = parts[2]
        
        # Validate QR is for correct event
        if str(event_id) != str(qr_event_id):
            AttendanceLog.objects.create(
                event=event,
                scan_type=scan_type,
                result='WRONG_EVENT',
                token=token,
                device_id=device_id,
                recorded_by=recorded_by,
                client_scan_time=client_scan_time,
                remarks=f'Token QR is for event {qr_event_id}, not {event_id}. Raw: {qr[:255]}'
            )
            return JsonResponse({
                'ok': False,
                'result': 'WRONG_EVENT',
                'message': 'This QR code is for a different event.',
                'color': 'error',
            }, status=400)
        
        # Token lookup with lock
        reg = EventRegistration.objects.select_for_update().filter(
            token=token
        ).select_related('student', 'event').first()
        
        if not reg:
            AttendanceLog.objects.create(
                event=event,
                scan_type=scan_type,
                result='INVALID',
                token=token,
                device_id=device_id,
                recorded_by=recorded_by,
                client_scan_time=client_scan_time,
                remarks=f'Token not found. Raw: {qr[:255]}'
            )
            return JsonResponse({
                'ok': False,
                'result': 'INVALID',
                'message': 'Invalid QR code or token.',
                'color': 'error',
            }, status=400)
        
        # CRITICAL: Verify token belongs to the selected event
        if reg.event_id != event.id:
            AttendanceLog.objects.create(
                event=event,
                student=reg.student,
                registration=reg,
                scan_type=scan_type,
                result='WRONG_EVENT',
                token=token,
                device_id=device_id,
                recorded_by=recorded_by,
                client_scan_time=client_scan_time,
                remarks=f'Token belongs to event {reg.event_id}, selected {event.id}. Raw: {qr[:255]}'
            )
            return JsonResponse({
                'ok': False,
                'result': 'WRONG_EVENT',
                'message': 'This token is for a different event.',
                'color': 'error',
            }, status=400)
        
        # Check if token is revoked
        if reg.status != 'active':
            AttendanceLog.objects.create(
                event=event,
                student=reg.student,
                registration=reg,
                scan_type=scan_type,
                result='REVOKED',
                token=token,
                device_id=device_id,
                recorded_by=recorded_by,
                client_scan_time=client_scan_time,
                remarks='Token has been revoked'
            )
            return JsonResponse({
                'ok': False,
                'result': 'REVOKED',
                'message': 'This QR code has been revoked.',
                'color': 'error',
            }, status=400)
        
        student = reg.student
        
    else:
        # STUDENT ID QR (permanent eEID) – primary flow: always accept for event attendance
        # Parse: "STU:2022-00123" or just "2022-00123"
        student_id = qr
        if student_id.startswith('STU:'):
            student_id = student_id.split('STU:', 1)[1].strip()
        student_id = student_id.strip()
        
        # Lookup student
        student = Student.objects.filter(student_id=student_id, is_active=True).first()
        if not student:
            AttendanceLog.objects.create(
                event=event,
                scan_type=scan_type,
                result='INVALID',
                token='',
                device_id=device_id,
                recorded_by=recorded_by,
                client_scan_time=client_scan_time,
                remarks=f'Student not found or inactive: {student_id}. Raw: {qr[:255]}'
            )
            return JsonResponse({
                'ok': False,
                'result': 'INVALID',
                'message': f'Student ID {student_id} not found or inactive.',
                'color': 'error',
                'not_registered': True,
                'student_id': student_id,
            }, status=400)
        
        reg = None  # No token registration for student ID scans

    # Inactive accounts cannot use event QR. Student-ID branch already filters is_active=True;
    # token-based flow must reject registrations tied to a frozen/inactive student.
    if not student.is_active or student.account_status != Student.ACCOUNT_STATUS_APPROVED:
        AttendanceLog.objects.create(
            event=event,
            student=student,
            registration=reg,
            scan_type=scan_type,
            result='INVALID',
            token=reg.token if (is_token_based and reg) else '',
            device_id=device_id,
            recorded_by=recorded_by,
            client_scan_time=client_scan_time,
            remarks='Student account inactive; event QR not allowed.',
        )
        return JsonResponse({
            'ok': False,
            'result': 'INVALID',
            'message': 'This student account is inactive. The QR code cannot be used until the account is active again.',
            'color': 'error',
            'inactive': True,
            'student_id': student.student_id,
        }, status=400)

    # Audience eligibility validation (applies to token + student ID scans)
    if not _is_student_allowed_for_event(event, student):
        audience_summary = event.audience_summary() if hasattr(event, 'audience_summary') else event.get_audience_scope_display()
        AttendanceLog.objects.create(
            event=event,
            student=student,
            registration=reg,
            scan_type=scan_type,
            result='INVALID',
            token=reg.token if (is_token_based and reg) else '',
            device_id=device_id,
            recorded_by=recorded_by,
            client_scan_time=client_scan_time,
            remarks=f'Not in target audience ({audience_summary}). Raw: {qr[:200]}'
        )
        return JsonResponse({
            'ok': False,
            'result': 'INVALID',
            'message': f'{student.get_full_name()} is not included in this event audience ({audience_summary}).',
            'color': 'warning',
            'student': {
                'student_id': student.student_id,
                'name': student.get_full_name(),
            },
        }, status=400)
    
    # Time window validation (applies to both types)
    if not _is_within_event_window(event, grace_minutes=30):
        AttendanceLog.objects.create(
            event=event,
            student=student,
            registration=reg,
            scan_type=scan_type,
            result='OUTSIDE_WINDOW',
            token=reg.token if reg else '',
            device_id=device_id,
                recorded_by=recorded_by,
            client_scan_time=client_scan_time,
            remarks=f'Event: {event.start_date} to {event.end_date}. Raw: {qr[:200]}'
        )
        return JsonResponse({
            'ok': False,
            'result': 'OUTSIDE_WINDOW',
            'message': f'Event is scheduled for {event.start_date} - {event.end_date}.',
            'color': 'warning',
        }, status=400)
    
    # Get or create EventAttendance record for duplicate checking
    attendance, att_created = EventAttendance.objects.select_for_update().get_or_create(
        event=event,
        student=student,
        defaults={'participated': False}
    )
    
    # Auto IN/OUT from registration/attendance (no IN/OUT toggle on scanner UI).
    if is_token_based and reg:
        if reg.checked_in_at is None:
            scan_type = 'IN'
        elif reg.checked_out_at is None:
            scan_type = 'OUT'
        else:
            if reg.checked_out_at > reg.checked_in_at and _is_whole_day_event_schedule(event):
                scan_type = 'IN'
            else:
                scan_type = 'OUT'
    else:
        if attendance.checked_in_at is None:
            scan_type = 'IN'
        elif attendance.checked_out_at is None:
            scan_type = 'OUT'
        else:
            if attendance.checked_out_at > attendance.checked_in_at and _is_whole_day_event_schedule(event):
                scan_type = 'IN'
            else:
                scan_type = 'OUT'
    
    now = timezone.now()
    now_local = timezone.localtime(now)
    sched_ok, sched_msg = _scan_event_schedule_allowed(
        event, scan_type, now_local, attendance, reg, is_token_based
    )
    if not sched_ok:
        AttendanceLog.objects.create(
            event=event,
            student=student,
            registration=reg,
            scan_type=scan_type,
            result='OUTSIDE_WINDOW',
            token=reg.token if reg else '',
            device_id=device_id,
            recorded_by=recorded_by,
            client_scan_time=client_scan_time,
            remarks=sched_msg[:255],
        )
        return JsonResponse({
            'ok': False,
            'result': 'OUTSIDE_WINDOW',
            'message': sched_msg,
            'color': 'warning',
            'student': {
                'student_id': student.student_id,
                'name': student.get_full_name(),
            },
        }, status=400)
    
    # For token-based: use EventRegistration timestamps
    # For student ID: use EventAttendance (we'll add fields) or check AttendanceLog
    if is_token_based and reg:
        # Use token registration timestamps
        if scan_type == 'IN':
            if reg.checked_in_at is not None:
                if reg.checked_out_at is None or reg.checked_out_at < reg.checked_in_at:
                    AttendanceLog.objects.create(
                        event=event,
                        student=student,
                        registration=reg,
                        scan_type='IN',
                        result='DUPLICATE',
                        token=reg.token,
                        device_id=device_id,
                        recorded_by=recorded_by,
                        client_scan_time=client_scan_time,
                        remarks=f'Already checked in at {timezone.localtime(reg.checked_in_at).strftime("%Y-%m-%d %I:%M %p")}'
                    )
                    return JsonResponse({
                        'ok': False,
                        'result': 'DUPLICATE',
                        'message': f'{student.get_full_name()} already checked in at {timezone.localtime(reg.checked_in_at).strftime("%I:%M %p")}.',
                        'color': 'warning',
                        'student': {
                            'student_id': student.student_id,
                            'name': student.get_full_name(),
                        },
                        'checked_in_at': timezone.localtime(reg.checked_in_at).strftime('%Y-%m-%d %I:%M %p'),
                    }, status=200)
            return_afternoon = (
                _is_whole_day_event_schedule(event)
                and reg.checked_in_at is not None
                and reg.checked_out_at is not None
                and reg.checked_out_at > reg.checked_in_at
            )
            reg.checked_in_at = now
            if return_afternoon:
                reg.checked_out_at = None
            reg.save(update_fields=['checked_in_at', 'checked_out_at'] if return_afternoon else ['checked_in_at'])

        elif scan_type == 'OUT':
            if reg.checked_in_at is None:
                AttendanceLog.objects.create(
                    event=event,
                    student=student,
                    registration=reg,
                    scan_type='OUT',
                    result='NOT_CHECKED_IN',
                    token=reg.token,
                    device_id=device_id,
                    recorded_by=recorded_by,
                    client_scan_time=client_scan_time,
                    remarks='Cannot check out before checking in'
                )
                return JsonResponse({
                    'ok': False,
                    'result': 'NOT_CHECKED_IN',
                    'message': f'{student.get_full_name()} has not checked in yet.',
                    'color': 'warning',
                    'student': {
                        'student_id': student.student_id,
                        'name': student.get_full_name(),
                    },
                }, status=200)
            
            if reg.checked_out_at is not None and reg.checked_out_at > reg.checked_in_at:
                AttendanceLog.objects.create(
                    event=event,
                    student=student,
                    registration=reg,
                    scan_type='OUT',
                    result='DUPLICATE',
                    token=reg.token,
                    device_id=device_id,
                    recorded_by=recorded_by,
                    client_scan_time=client_scan_time,
                    remarks=f'Already checked out at {timezone.localtime(reg.checked_out_at).strftime("%Y-%m-%d %I:%M %p")}'
                )
                return JsonResponse({
                    'ok': False,
                    'result': 'DUPLICATE',
                    'message': f'{student.get_full_name()} already checked out at {timezone.localtime(reg.checked_out_at).strftime("%I:%M %p")}.',
                    'color': 'warning',
                    'student': {
                        'student_id': student.student_id,
                        'name': student.get_full_name(),
                    },
                    'checked_out_at': timezone.localtime(reg.checked_out_at).strftime('%Y-%m-%d %I:%M %p'),
                }, status=200)
            
            reg.checked_out_at = now
            reg.save(update_fields=['checked_out_at'])
    
    else:
        # ====================================================================
        # STUDENT ID QR: Use EventAttendance timestamps for duplicate checking
        # ====================================================================
        if scan_type == 'IN':
            # Check if already checked in (no checkout after)
            if attendance.checked_in_at is not None:
                if attendance.checked_out_at is None or attendance.checked_out_at < attendance.checked_in_at:
                    # Already checked in, no checkout after
                    AttendanceLog.objects.create(
                        event=event,
                        student=student,
                        scan_type='IN',
                        result='DUPLICATE',
                        token='',
                        device_id=device_id,
                    recorded_by=recorded_by,
                        client_scan_time=client_scan_time,
                        remarks=f'Already checked in at {timezone.localtime(attendance.checked_in_at).strftime("%Y-%m-%d %I:%M %p")}'
                    )
                    return JsonResponse({
                        'ok': False,
                        'result': 'DUPLICATE',
                        'message': f'{student.get_full_name()} already checked in at {timezone.localtime(attendance.checked_in_at).strftime("%I:%M %p")}.',
                        'color': 'warning',
                        'student': {
                            'student_id': student.student_id,
                            'name': student.get_full_name(),
                        },
                        'checked_in_at': timezone.localtime(attendance.checked_in_at).strftime('%Y-%m-%d %I:%M %p'),
                    }, status=200)
            
            # Record check-in (whole-day afternoon return: clear lunch OUT so end-of-day OUT works)
            return_afternoon = (
                _is_whole_day_event_schedule(event)
                and attendance.checked_in_at is not None
                and attendance.checked_out_at is not None
                and attendance.checked_out_at > attendance.checked_in_at
            )
            attendance.checked_in_at = now
            if return_afternoon:
                attendance.checked_out_at = None
            attendance.save(
                update_fields=['checked_in_at', 'checked_out_at'] if return_afternoon else ['checked_in_at']
            )

        elif scan_type == 'OUT':
            if attendance.checked_in_at is None:
                AttendanceLog.objects.create(
                    event=event,
                    student=student,
                    scan_type='OUT',
                    result='NOT_CHECKED_IN',
                    token='',
                    device_id=device_id,
                    recorded_by=recorded_by,
                    client_scan_time=client_scan_time,
                    remarks='Cannot check out before checking in'
                )
                return JsonResponse({
                    'ok': False,
                    'result': 'NOT_CHECKED_IN',
                    'message': f'{student.get_full_name()} has not checked in yet.',
                    'color': 'warning',
                    'student': {
                        'student_id': student.student_id,
                        'name': student.get_full_name(),
                    },
                }, status=200)
            
            # Check for duplicate OUT
            if attendance.checked_out_at is not None and attendance.checked_out_at > attendance.checked_in_at:
                AttendanceLog.objects.create(
                    event=event,
                    student=student,
                    scan_type='OUT',
                    result='DUPLICATE',
                    token='',
                    device_id=device_id,
                    recorded_by=recorded_by,
                    client_scan_time=client_scan_time,
                    remarks=f'Already checked out at {timezone.localtime(attendance.checked_out_at).strftime("%Y-%m-%d %I:%M %p")}'
                )
                return JsonResponse({
                    'ok': False,
                    'result': 'DUPLICATE',
                    'message': f'{student.get_full_name()} already checked out at {timezone.localtime(attendance.checked_out_at).strftime("%I:%M %p")}.',
                    'color': 'warning',
                    'student': {
                        'student_id': student.student_id,
                        'name': student.get_full_name(),
                    },
                    'checked_out_at': timezone.localtime(attendance.checked_out_at).strftime('%Y-%m-%d %I:%M %p'),
                }, status=200)
            
            # Check if leaving event early (before event end time)
            event_end_datetime = timezone.make_aware(
                datetime.datetime.combine(event.end_date, datetime.time(23, 59, 59))
            )
            is_early_checkout = now < event_end_datetime
            
            early_out_reason = request.POST.get('early_out_reason', '').strip()
            
            # Record check-out (optional early-out note for reports)
            attendance.checked_out_at = now
            if is_early_checkout and early_out_reason:
                attendance.early_out_reason = early_out_reason
                attendance.save(update_fields=['checked_out_at', 'early_out_reason'])
            else:
                attendance.save(update_fields=['checked_out_at'])
    
    # Mark as participated
    attendance.participated = True
    attendance.save(update_fields=['participated'])
    
    # Create GateEntry so "Event attendees" modal (which lists GateEntry with event_id) shows this scan
    GateEntry.objects.create(
        student=student,
        event=event,
        granted=True,
        notes=scan_type,
        recorded_by=recorded_by,
        timestamp=now,
        **_audit_kwargs_for_gate_entry(request, device_id=device_id),
    )
    
    # Log successful scan
    AttendanceLog.objects.create(
        event=event,
        student=student,
        registration=reg if is_token_based else None,
        scan_type=scan_type,
        result='SUCCESS',
        token=reg.token if (is_token_based and reg) else '',
        device_id=device_id,
        recorded_by=recorded_by,
        client_scan_time=client_scan_time,
        remarks='OK'
    )
    
    # Get student photo URL
    photo_url = _scan_ui_photo_url(request, student)

    time_str = timezone.localtime(now).strftime('%I:%M %p')
    
    # Build response with check-in/out times
    checked_in_time = None
    checked_out_time = None
    
    if is_token_based and reg:
        if reg.checked_in_at:
            checked_in_time = timezone.localtime(reg.checked_in_at).strftime('%Y-%m-%d %I:%M %p')
        if reg.checked_out_at:
            checked_out_time = timezone.localtime(reg.checked_out_at).strftime('%Y-%m-%d %I:%M %p')
    else:
        # For student ID scans, use EventAttendance timestamps
        if attendance.checked_in_at:
            checked_in_time = timezone.localtime(attendance.checked_in_at).strftime('%Y-%m-%d %I:%M %p')
        if attendance.checked_out_at:
            checked_out_time = timezone.localtime(attendance.checked_out_at).strftime('%Y-%m-%d %I:%M %p')
    
    return JsonResponse({
        'ok': True,
        'result': 'SUCCESS',
        'message': f'{student.get_full_name()} checked {scan_type.lower()} successfully.',
        'color': 'success',
        'scan_type': scan_type,
        'status': scan_type,
        'event_attendance': True,
        'time': time_str,
        'student_name': student.get_full_name(),
        'qr_type': 'token' if is_token_based else 'student_id',
        'student': {
            'student_id': student.student_id,
            'name': student.get_full_name(),
            'first_name': student.first_name,
            'middle_name': student.middle_name or '',
            'last_name': student.last_name,
            'email': student.email or '',
            'photo_url': photo_url,
            'course_or_section': getattr(student, 'course_or_section', '') or '',
            'year_level': getattr(student, 'year_level', '') or '',
        },
        'checked_in_at': checked_in_time,
        'checked_out_at': checked_out_time,
    }, status=200)


@require_POST
@csrf_exempt
@transaction.atomic
def scan_event_qr_guard(request):
    """
    Same as scan_event_qr but CSRF-exempt; requires GATE_GUARD_DISPLAY_TOKEN in POST/header.
    Used by /gate/guard-display/ and embed scanner (token-only clients).
    """
    token = (request.POST.get('guard_token') or request.headers.get('X-Gate-Guard-Token') or '').strip()
    expected = getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or ''
    if not (token and expected and token == expected):
        return JsonResponse({
            'ok': False,
            'result': 'DENIED',
            'message': 'Unauthorized',
            'color': 'error',
        }, status=403)
    return scan_event_qr(request)


@require_POST
def register_student_from_scan(request):
    if _resolve_save_scan_actor(request) is False:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    student_id = (request.POST.get('student_id') or '').strip()
    first_name = (request.POST.get('first_name') or '').strip()
    middle_name = (request.POST.get('middle_name') or '').strip()
    last_name = (request.POST.get('last_name') or '').strip()
    sex = (request.POST.get('sex') or '').strip().upper()
    email = (request.POST.get('email') or '').strip()

    valid_sex_values = {choice[0] for choice in Student.SEX_CHOICES}
    if sex and sex not in valid_sex_values:
        sex = ''

    if not student_id:
        return JsonResponse({'success': False, 'message': 'Missing student ID.'}, status=400)
    if not first_name or not last_name:
        return JsonResponse({'success': False, 'message': 'Please provide first name and last name.'}, status=400)
    existing = Student.objects.filter(student_id=student_id).first()
    if existing and existing.is_active:
        return JsonResponse({'success': False, 'message': 'This Student ID is already registered.'}, status=400)

    if existing and not existing.is_active:
        existing.first_name = first_name
        existing.middle_name = middle_name
        existing.last_name = last_name
        if sex:
            existing.sex = sex
        if email:
            existing.email = email
        existing.account_status = Student.ACCOUNT_STATUS_APPROVED
        existing.is_active = True
        existing.save()
        student = existing
        created_msg = f'{student.get_full_name()} registration updated.'
    else:
        student = Student.objects.create(
            student_id=student_id,
            first_name=first_name,
            middle_name=middle_name,
            last_name=last_name,
            sex=sex or '',
            email=email or '',
            account_status=Student.ACCOUNT_STATUS_APPROVED,
            is_active=True,
        )
        created_msg = f'{student.get_full_name()} registration submitted.'

    now = timezone.localtime(timezone.now())
    time_str = now.strftime('%I:%M %p')

    return JsonResponse({
        'success': False,
        'registered': True,
        'message': created_msg,
        'color': 'info',
        'student_name': student.get_full_name(),
        'time': time_str,
        'student_id': student.student_id,
        'student': {
            'first_name': student.first_name,
            'middle_name': student.middle_name or '',
            'last_name': student.last_name,
            'student_id': student.student_id,
            'sex': student.get_sex_display() if student.sex else '',
            'email': student.email or '',
            'photo_url': None,
        },
    })


def _resolve_visitor_department(department_value, who_to_visit_other):
    """Resolve department dropdown + optional 'other' text into a single who_to_visit label."""
    if not department_value or department_value == 'other':
        return (who_to_visit_other or '').strip() or 'Other'
    for value, label in CAMPUS_DEPARTMENT_CHOICES:
        if value == department_value and value:
            return label
    return (who_to_visit_other or '').strip() or department_value


@require_POST
def record_visitor_entry(request):
    """Record a manual visitor entry: staff allows a visitor to enter (name, purpose, department/office). Optional face photo as proof."""
    actor = _resolve_save_scan_actor(request)
    if actor is False:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    visitor_name = (request.POST.get('visitor_name') or '').strip()
    purpose = (request.POST.get('purpose') or '').strip()
    department_value = (request.POST.get('department') or '').strip()
    who_to_visit_other = (request.POST.get('who_to_visit_other') or '').strip()
    who_to_visit = _resolve_visitor_department(department_value, who_to_visit_other)
    photo_file = request.FILES.get('photo')

    if not visitor_name:
        return JsonResponse({'success': False, 'message': 'Please enter the visitor\'s name.'}, status=400)
    if not purpose:
        return JsonResponse({'success': False, 'message': 'Please enter the purpose of visit.'}, status=400)
    if not who_to_visit:
        return JsonResponse({'success': False, 'message': 'Please select a department/office or specify where they are going.'}, status=400)

    entry_kwargs = dict(
        visitor_name=visitor_name,
        purpose=purpose,
        who_to_visit=who_to_visit,
        recorded_by=actor,
    )
    if photo_file:
        entry_kwargs['photo'] = photo_file
    VisitorEntry.objects.create(**entry_kwargs)
    now = timezone.localtime(timezone.now())
    time_str = now.strftime('%I:%M %p')
    return JsonResponse({
        'success': True,
        'message': f'{visitor_name} entered. Purpose: {purpose}. Visiting: {who_to_visit}.',
        'time': time_str,
    })


EARLY_OUT_REASON_MIN_LENGTH = 5  # legacy; event early-out no longer requires a minimum-length note


@require_POST
def record_early_out(request):
    """Record OUT when student was already scanned today (daily gate). Uses same policy as gate scan."""
    actor = _resolve_save_scan_actor(request)
    if actor is False:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    student_id = (request.POST.get('student_id') or '').strip()
    reason = (request.POST.get('note') or request.POST.get('reason') or '').strip()

    if not student_id:
        return JsonResponse({
            'success': False,
            'message': 'No student ID provided.',
        }, status=400)

    try:
        student = Student.objects.get(student_id=student_id, is_active=True)
    except Student.DoesNotExist:
        return JsonResponse({
            'success': False,
            'message': 'Student not found or inactive.',
        }, status=400)

    # Use same local calendar day as save_scan so "today" matches (avoids UTC vs local date mismatch).
    today = timezone.localdate()
    day_start, day_end = _local_day_bounds(today)
    already_today = GateEntry.objects.filter(
        student=student, granted=True, event__isnull=True,
        timestamp__gte=day_start, timestamp__lt=day_end,
    ).exists()
    if not already_today:
        return JsonResponse({
            'success': False,
            'message': 'Student has not been scanned in today. Use normal scan to record OUT.',
        }, status=400)

    now_dt = timezone.localtime(timezone.now())
    override_reason = (reason or '').strip() or None
    eval_result = evaluate_scan(student, 'OUT', now_dt, personnel_override_reason=override_reason, daily_gate_only=True)
    if not eval_result['allowed']:
        return JsonResponse({
            'success': False,
            'blocked_during_class': eval_result.get('result') in ('DENIED', 'REQUIRE_REASON'),
            'message': eval_result['message'],
            'schedule_hint': eval_result.get('schedule_hint', ''),
            'require_note': eval_result.get('result') == 'REQUIRE_REASON',
        }, status=400)

    out_reason_code = eval_result.get('out_reason_code') or ''
    out_reason_text = eval_result.get('out_reason_text') or reason or ''

    GateEntry.objects.create(
        student=student,
        granted=True,
        result='SUCCESS',
        scan_type='OUT',
        notes='OUT',
        out_reason=out_reason_text,
        out_reason_code=out_reason_code,
        recorded_by=actor,
        **_audit_kwargs_for_gate_entry(request),
    )

    now = timezone.localtime(timezone.now())
    time_str = now.strftime('%I:%M %p')
    photo_url = _scan_ui_photo_url(request, student)

    return JsonResponse({
        'success': True,
        'message': f'{student.get_full_name()} checked out (early out).',
        'color': 'success',
        'status': 'OUT',
        'student_name': student.get_full_name(),
        'time': time_str,
        'out_note': reason,
        'student': {
            'first_name': student.first_name,
            'middle_name': student.middle_name or '',
            'last_name': student.last_name,
            'student_id': student.student_id,
            'email': student.email or '',
            'photo_url': photo_url,
        },
    })


# ---------- Reusable visitor pass: check-in / check-out lifecycle ----------

@require_POST
@transaction.atomic
def visitor_checkin_submit(request):
    """Submit visitor check-in form (reusable pass). Creates VisitorVisit, sets pass IN_USE, GateEntry(IN)."""
    actor = _resolve_save_scan_actor(request)
    if actor is False:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    pass_code = (request.POST.get('pass_code') or '').strip()
    full_name = (request.POST.get('full_name') or '').strip()
    purpose = (request.POST.get('purpose') or '').strip()
    department = (request.POST.get('department') or '').strip()
    notes = (request.POST.get('notes') or '').strip()

    if not pass_code or not full_name:
        return JsonResponse({'success': False, 'message': 'Pass code and full name are required.'}, status=400)
    if not purpose:
        return JsonResponse({'success': False, 'message': 'Purpose is required.'}, status=400)
    if not department:
        return JsonResponse({'success': False, 'message': 'Department / office is required.'}, status=400)

    pass_obj = VisitorPass.objects.filter(code=pass_code).first()
    if not pass_obj:
        return JsonResponse({'success': False, 'message': 'Invalid pass code.'}, status=400)
    if pass_obj.status != VisitorPass.STATUS_AVAILABLE:
        return JsonResponse({'success': False, 'message': 'Pass is not available (in use or disabled).'}, status=400)

    now = timezone.now()
    visit = VisitorVisit.objects.create(
        pass_obj=pass_obj,
        full_name=full_name,
        purpose=purpose,
        department=department,
        checked_in_at=now,
        checked_in_by=actor,
        status=VisitorVisit.STATUS_INSIDE,
        notes=notes,
    )
    if request.FILES.get('photo_in'):
        visit.photo_in = request.FILES['photo_in']
        visit.save(update_fields=['photo_in'])

    pass_obj.status = VisitorPass.STATUS_IN_USE
    pass_obj.current_visit = visit
    pass_obj.save(update_fields=['status', 'current_visit_id'])

    GateEntry.objects.create(
        student=None,
        visitor_visit=visit,
        granted=True,
        result='SUCCESS',
        scan_type='IN',
        notes='IN',
        recorded_by=actor,
        **_audit_kwargs_for_gate_entry(request),
    )

    time_str = timezone.localtime(now).strftime('%I:%M %p')
    return JsonResponse({
        'success': True,
        'message': f'{full_name} checked in ({pass_code}).',
        'color': 'success',
        'status': 'IN',
        'visit_id': visit.id,
        'pass_code': pass_code,
        'time': time_str,
    })


@require_POST
@transaction.atomic
def visitor_checkout_submit(request):
    """Check out visitor (same pass scanned). Sets visit OUTSIDE, pass AVAILABLE, GateEntry(OUT)."""
    actor = _resolve_save_scan_actor(request)
    if actor is False:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    pass_code = (request.POST.get('pass_code') or '').strip()
    visit_id = request.POST.get('visit_id')
    if not pass_code and not visit_id:
        return JsonResponse({'success': False, 'message': 'Pass code or visit ID required.'}, status=400)

    if visit_id:
        visit = VisitorVisit.objects.filter(id=visit_id, status=VisitorVisit.STATUS_INSIDE).select_related('pass_obj').first()
    else:
        pass_obj = VisitorPass.objects.filter(code=pass_code).select_related('current_visit').first()
        if not pass_obj or pass_obj.status != VisitorPass.STATUS_IN_USE or not pass_obj.current_visit_id:
            return JsonResponse({'success': False, 'message': 'Pass not in use or no active visit.'}, status=400)
        visit = pass_obj.current_visit

    if not visit or visit.status != VisitorVisit.STATUS_INSIDE:
        return JsonResponse({'success': False, 'message': 'No active visit found.'}, status=400)

    now = timezone.now()
    # Minimum 1 minute before visitor can check out (prevents accidental double-scan)
    min_stay = datetime.timedelta(minutes=1)
    if now - visit.checked_in_at < min_stay:
        remaining = min_stay - (now - visit.checked_in_at)
        secs = max(0, int(remaining.total_seconds()))
        return JsonResponse({
            'success': False,
            'message': f'Visitor must stay at least 1 minute before check-out. Please try again in {secs} second(s).',
            'color': 'warning',
        }, status=400)

    visit.checked_out_at = now
    visit.checked_out_by = actor
    visit.status = VisitorVisit.STATUS_OUTSIDE
    visit.save(update_fields=['checked_out_at', 'checked_out_by', 'status'])
    if request.FILES.get('photo_out'):
        visit.photo_out = request.FILES['photo_out']
        visit.save(update_fields=['photo_out'])

    pass_obj = visit.pass_obj
    pass_obj.status = VisitorPass.STATUS_AVAILABLE
    pass_obj.current_visit = None
    pass_obj.last_used_at = now
    pass_obj.save(update_fields=['status', 'current_visit_id', 'last_used_at'])

    GateEntry.objects.create(
        student=None,
        visitor_visit=visit,
        granted=True,
        result='SUCCESS',
        scan_type='OUT',
        notes='OUT',
        recorded_by=actor,
        **_audit_kwargs_for_gate_entry(request),
    )

    time_str = timezone.localtime(now).strftime('%I:%M %p')
    return JsonResponse({
        'success': True,
        'message': f'{visit.full_name} checked out. Pass {pass_obj.code} released.',
        'color': 'success',
        'status': 'OUT',
        'time': time_str,
    })


@require_POST
@transaction.atomic
def visitor_force_checkout(request):
    """Force checkout a visitor (e.g. lost pass, no return). Requires a short staff note."""
    actor = _resolve_save_scan_actor(request)
    if actor is False:
        return JsonResponse({'success': False, 'message': 'Unauthorized'}, status=403)
    visit_id = request.POST.get('visit_id')
    reason = (request.POST.get('note') or request.POST.get('reason') or '').strip()
    if not visit_id or not reason or len(reason) < 3:
        return JsonResponse({'success': False, 'message': 'Visit ID and a short note (min 3 characters) required.'}, status=400)

    visit = VisitorVisit.objects.filter(id=visit_id, status=VisitorVisit.STATUS_INSIDE).select_related('pass_obj').first()
    if not visit:
        return JsonResponse({'success': False, 'message': 'Active visit not found.'}, status=404)

    now = timezone.now()
    visit.checked_out_at = now
    visit.checked_out_by = actor
    visit.status = VisitorVisit.STATUS_OUTSIDE
    visit.notes = (visit.notes or '') + f'\n[Force checkout] {reason}'
    visit.save(update_fields=['checked_out_at', 'checked_out_by', 'status', 'notes'])

    pass_obj = visit.pass_obj
    pass_obj.status = VisitorPass.STATUS_AVAILABLE
    pass_obj.current_visit = None
    pass_obj.last_used_at = now
    pass_obj.save(update_fields=['status', 'current_visit_id', 'last_used_at'])

    GateEntry.objects.create(
        student=None,
        visitor_visit=visit,
        granted=True,
        result='SUCCESS',
        scan_type='OUT',
        notes=f'OUT (force: {reason[:100]})',
        recorded_by=actor,
        **_audit_kwargs_for_gate_entry(request),
    )
    return JsonResponse({'success': True, 'message': f'Visit force-checked out. Pass {pass_obj.code} released.'})


@require_POST
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def visitor_disable_pass(request):
    """Disable a reusable pass (e.g. lost). If IN_USE, force checkout first."""
    pass_code = (request.POST.get('pass_code') or '').strip()
    if not pass_code:
        return JsonResponse({'success': False, 'message': 'Pass code required.'}, status=400)

    pass_obj = VisitorPass.objects.filter(code=pass_code).first()
    if not pass_obj:
        return JsonResponse({'success': False, 'message': 'Pass not found.'}, status=404)

    if pass_obj.status == VisitorPass.STATUS_IN_USE and pass_obj.current_visit_id:
        return JsonResponse({
            'success': False,
            'message': 'Pass is in use. Force checkout the visit first, then disable.',
            'visit_id': pass_obj.current_visit_id,
        }, status=400)

    pass_obj.status = VisitorPass.STATUS_DISABLED
    pass_obj.current_visit = None
    pass_obj.save(update_fields=['status', 'current_visit_id'])
    return JsonResponse({'success': True, 'message': f'Pass {pass_code} disabled.'})


@require_POST
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def mark_participation(request):
    """Mark student as participated or non-participant for an event."""
    student_pk = request.POST.get('student_pk')
    event_pk = request.POST.get('event_pk')
    participated = request.POST.get('participated', 'true').lower() in ('1', 'true', 'yes')
    student = get_object_or_404(Student, pk=student_pk)
    event = get_object_or_404(Event, pk=event_pk)
    att, _ = EventAttendance.objects.get_or_create(student=student, event=event, defaults={'participated': False})
    att.participated = participated
    att.save()
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return JsonResponse({'ok': True})
    return redirect('gate-scan')


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def analytics_dashboard(request):
    """Analytics: gate stats, event participation (incidents feature removed)."""
    today = timezone.localdate()
    day_start, day_end = _local_day_bounds(today)

    selected_year_raw = (request.GET.get('report_year') or request.GET.get('year') or '').strip()
    selected_year = int(selected_year_raw) if selected_year_raw.isdigit() else today.year

    entries_today = GateEntry.objects.filter(timestamp__gte=day_start, timestamp__lt=day_end)
    # Student gate visits only (daily gate, no event/visitor) so count matches Gate entry list
    granted_today = _granted_visits_count_for_date(today, daily_gate_only=True)
    denied_entries_count = entries_today.filter(granted=False).count()
    denied_today = denied_entries_count
    total_students = Student.objects.filter(is_active=True).count()
    inside_now = _currently_inside_count(today)
    active_events = _get_active_events()
    report_event_choices = Event.objects.order_by('-start_date', 'name')[:150]
    past_events_with_stats = _get_past_events_with_stats(limit=20)
    recent_incidents = []
    participation_stats = EventAttendance.objects.values('event__name').annotate(
        total=Count('id'),
        participated_count=Count(Case(When(participated=True, then=1), output_field=IntegerField())),
        non_participant_count=Count(Case(When(participated=False, then=1), output_field=IntegerField())),
    )

    years_entries = GateEntry.objects.annotate(y=ExtractYear('timestamp')).values_list('y', flat=True).distinct()
    data_years = {y for y in list(years_entries) if y}

    default_years = {today.year - i for i in range(0, 6)}
    available_years = sorted(data_years.union(default_years).union({selected_year}))

    reason_labels = []
    reason_counts = []
    reason_colors = []

    # Visitor stats (monthly and annually) – use local datetime bounds for timezone correctness
    # Include both VisitorEntry (manual log) and VisitorVisit (reusable pass check-ins) to match visitor list page
    start_of_month = today.replace(day=1)
    month_start, _ = _local_day_bounds(start_of_month)
    visitors_this_month = (
        VisitorEntry.objects.filter(
            timestamp__gte=month_start, timestamp__lt=day_end,
        ).count()
        + VisitorVisit.objects.filter(
            checked_in_at__gte=month_start, checked_in_at__lt=day_end,
        ).count()
    )
    visitors_this_year = (
        VisitorEntry.objects.filter(timestamp__year=selected_year).count()
        + VisitorVisit.objects.filter(checked_in_at__year=selected_year).count()
    )

    # Student daily gate (granted IN/OUT scans only; excludes events and visitor rows) — matches gate entry list + analytics links
    first_of_month = today.replace(day=1)
    month_start_bounds, _ = _local_day_bounds(first_of_month)
    student_entries_this_month = GateEntry.objects.filter(
        timestamp__gte=month_start_bounds,
        timestamp__lt=day_end,
        granted=True,
        student_id__isnull=False,
        event__isnull=True,
        visitor_visit__isnull=True,
    )
    student_entries_this_month = student_entries_this_month.count()
    ys, ye = _local_year_bounds(selected_year)
    student_entries_this_year = GateEntry.objects.filter(
        timestamp__gte=ys,
        timestamp__lt=ye,
        granted=True,
        student_id__isnull=False,
        event__isnull=True,
        visitor_visit__isnull=True,
    )
    student_entries_this_year = student_entries_this_year.count()
    year_end_for_link = min(datetime.date(selected_year, 12, 31), today)

    # Top departments/offices visited (who_to_visit) – monthly and annually; include VisitorEntry + VisitorVisit to match visitor list
    top_departments_monthly, top_departments_annually = _top_departments_visitors(month_start, day_end, selected_year)

    theme = SiteTheme.objects.first()
    default_first_signatory_name = (getattr(theme, 'default_first_signatory_name', '') or '').strip()
    default_first_signatory_title = (getattr(theme, 'default_first_signatory_title', '') or '').strip()
    default_second_signatory_name = (getattr(theme, 'default_second_signatory_name', '') or '').strip()
    default_second_signatory_title = (getattr(theme, 'default_second_signatory_title', '') or '').strip()
    student_section_choices = [
        s for s in Student.objects.exclude(section='').values_list('section', flat=True).distinct().order_by('section')
    ]
    student_reason_choices = [('EARLY_OUT', 'Early out')] + [
        (code, label) for code, label in GateEntry.OUT_REASON_CODE_CHOICES if code
    ]
    visitor_department_choices = [(code, label) for code, label in CAMPUS_DEPARTMENT_CHOICES if code]

    # --- Analytics overview (date range + optional gate / scanner) ---
    from django.utils.dateparse import parse_date

    def _overview_student_gate_qs(start_d, end_d, gate_device_id):
        rs, _ = _local_day_bounds(start_d)
        _, re_excl = _local_day_bounds(end_d)
        qs = GateEntry.objects.filter(
            timestamp__gte=rs,
            timestamp__lt=re_excl,
            granted=True,
            student_id__isnull=False,
            event__isnull=True,
            visitor_visit__isnull=True,
        )
        if gate_device_id:
            qs = qs.filter(device_id=gate_device_id)
        return qs

    def _overview_visitor_count(start_d, end_d, gate_device_id):
        rs, _ = _local_day_bounds(start_d)
        _, re_excl = _local_day_bounds(end_d)
        if not gate_device_id:
            return (
                VisitorEntry.objects.filter(timestamp__gte=rs, timestamp__lt=re_excl).count()
                + VisitorVisit.objects.filter(checked_in_at__gte=rs, checked_in_at__lt=re_excl).count()
            )
        return GateEntry.objects.filter(
            timestamp__gte=rs,
            timestamp__lt=re_excl,
            granted=True,
            visitor_visit__isnull=False,
            device_id=gate_device_id,
        ).count()

    def _overview_peak_rush_label(hour):
        if hour is None:
            return '—'
        if 5 <= hour < 12:
            return 'Morning Rush'
        if 12 <= hour < 14:
            return 'Lunch'
        if 14 <= hour < 18:
            return 'Afternoon'
        if 18 <= hour < 22:
            return 'Evening'
        return 'Off-peak'

    def _pct_change(curr, prev):
        if prev is None:
            return None, None
        if prev == 0:
            if curr == 0:
                return 0.0, 'flat'
            return 100.0, 'up'
        delta = ((curr - prev) / prev) * 100.0
        if delta > 0.5:
            return delta, 'up'
        if delta < -0.5:
            return delta, 'down'
        return delta, 'flat'

    overview_preset = (request.GET.get('overview_preset') or '').strip().lower()
    _preset_allowed = ('today', 'yesterday', 'last_7_days', 'custom')
    if overview_preset not in _preset_allowed:
        overview_preset = ''

    overview_from_raw = (request.GET.get('overview_from') or '').strip()
    overview_to_raw = (request.GET.get('overview_to') or '').strip()
    overview_from_parsed = parse_date(overview_from_raw) if overview_from_raw else None
    overview_to_parsed = parse_date(overview_to_raw) if overview_to_raw else None

    # Legacy URLs: both dates set without overview_preset → custom range
    if not overview_preset:
        if overview_from_parsed and overview_to_parsed:
            overview_preset = 'custom'
        else:
            overview_preset = 'today'

    if overview_preset == 'custom':
        if overview_from_parsed and overview_to_parsed:
            overview_start = overview_from_parsed
            overview_end = overview_to_parsed
            if overview_start > overview_end:
                overview_start, overview_end = overview_end, overview_start
        else:
            overview_start = overview_end = today
    else:
        if overview_preset == 'today':
            overview_start = overview_end = today
        elif overview_preset == 'yesterday':
            _yd = today - datetime.timedelta(days=1)
            overview_start = overview_end = _yd
        elif overview_preset == 'last_7_days':
            overview_start = today - datetime.timedelta(days=6)
            overview_end = today
        else:
            overview_start = overview_end = today

    # Do not allow future end date past today for consistency with operational data
    if overview_end > today:
        overview_end = today
    if overview_start > overview_end:
        overview_start = overview_end

    overview_gate = ''

    overview_num_days = (overview_end - overview_start).days + 1
    span_days = overview_num_days
    prev_overview_end = overview_start - datetime.timedelta(days=1)
    prev_overview_start = prev_overview_end - datetime.timedelta(days=span_days - 1)

    ov_qs = _overview_student_gate_qs(overview_start, overview_end, overview_gate)
    overview_total_scans = ov_qs.count()
    prev_total_scans = _overview_student_gate_qs(prev_overview_start, prev_overview_end, overview_gate).count()

    overview_total_pct, overview_total_trend = _pct_change(overview_total_scans, prev_total_scans)

    overview_avg_daily = (overview_total_scans / overview_num_days) if overview_num_days else 0
    prev_avg_den = overview_num_days if overview_num_days else 1
    prev_avg_daily = (prev_total_scans / prev_avg_den) if prev_avg_den else 0
    if prev_avg_daily == 0:
        overview_avg_status = 'normal' if overview_avg_daily == 0 else 'above'
    else:
        rel_avg = (overview_avg_daily - prev_avg_daily) / prev_avg_daily
        if abs(rel_avg) < 0.12:
            overview_avg_status = 'normal'
        elif overview_avg_daily > prev_avg_daily:
            overview_avg_status = 'above'
        else:
            overview_avg_status = 'below'

    overview_visitors = _overview_visitor_count(overview_start, overview_end, overview_gate)
    prev_visitors = _overview_visitor_count(prev_overview_start, prev_overview_end, overview_gate)
    overview_visitors_pct, overview_visitors_trend = _pct_change(overview_visitors, prev_visitors)

    hour_counts = Counter()
    for ts in ov_qs.values_list('timestamp', flat=True):
        hour_counts[timezone.localtime(ts).hour] += 1
    if hour_counts:
        peak_hour = max(hour_counts.keys(), key=lambda h: hour_counts[h])
        peak_ampm = datetime.datetime(2000, 1, 1, peak_hour, 0).strftime('%I:%M %p').lstrip('0')
        overview_peak_display = peak_ampm.replace(' 0', ' ')
        overview_peak_rush = _overview_peak_rush_label(peak_hour)
    else:
        peak_hour = None
        overview_peak_display = '—'
        overview_peak_rush = 'No data'

    # Hourly traffic + entry-by-role: single calendar day vs full overview range (no separate UI control).
    traffic_hour_scope = 'range' if overview_start < overview_end else 'today'

    def _traffic_flow_student_gate_qs(range_start, range_end_exclusive, gate_device_id):
        qs = GateEntry.objects.filter(
            timestamp__gte=range_start,
            timestamp__lt=range_end_exclusive,
            granted=True,
            student_id__isnull=False,
            event__isnull=True,
            visitor_visit__isnull=True,
        )
        if gate_device_id:
            qs = qs.filter(device_id=gate_device_id)
        return qs

    def _fmt_hour_ampm(h):
        if h == 0:
            return '12 AM'
        if h < 12:
            return '%d AM' % h
        if h == 12:
            return '12 PM'
        return '%d PM' % (h - 12)

    tf_range_start, _ = _local_day_bounds(overview_start)
    _, tf_range_end_excl = _local_day_bounds(overview_end)

    # Campus hours: 6 AM – 5 PM (hours 6–17), line/area chart
    traffic_flow_labels = [_fmt_hour_ampm(h) for h in range(6, 18)]

    if traffic_hour_scope == 'today':
        day_start, day_end = _local_day_bounds(overview_end)
        qs_tf = _traffic_flow_student_gate_qs(day_start, day_end, overview_gate)
        traffic_flow_subtitle = overview_end.strftime('%b %d, %Y')
    else:
        qs_tf = _traffic_flow_student_gate_qs(tf_range_start, tf_range_end_excl, overview_gate)
        traffic_flow_subtitle = (
            f'{overview_start.strftime("%b %d, %Y")} – {overview_end.strftime("%b %d, %Y")}'
        )

    tf_rows = list(qs_tf.values_list('timestamp', 'scan_type'))
    in_counts = Counter()
    out_counts = Counter()
    for ts, st in tf_rows:
        h = timezone.localtime(ts).hour
        if h < 6 or h > 17:
            continue
        stu = (st or '').upper()
        if stu == 'IN':
            in_counts[h] += 1
        elif stu == 'OUT':
            out_counts[h] += 1
    traffic_flow_in = [in_counts[h] for h in range(6, 18)]
    traffic_flow_out = [out_counts[h] for h in range(6, 18)]
    traffic_flow_title = 'Hourly Traffic Flow'

    # Entry by role (IN scans, same time window + gate as hourly chart above) — donut chart
    # Students vs visitors only (no separate staff/contractor slices — not tracked in this deployment).
    if traffic_hour_scope == 'today':
        _role_start, _role_end_excl = _local_day_bounds(overview_end)
        entry_role_range_note = overview_end.strftime('%b %d, %Y')
    else:
        _role_start, _role_end_excl = tf_range_start, tf_range_end_excl
        entry_role_range_note = (
            f'{overview_start.strftime("%b %d")} – {overview_end.strftime("%b %d, %Y")}'
        )
    role_qs = GateEntry.objects.filter(
        timestamp__gte=_role_start,
        timestamp__lt=_role_end_excl,
        granted=True,
        scan_type='IN',
    ).select_related('student')

    entry_role_counts = {'students': 0, 'visitors': 0}
    for e in role_qs.iterator(chunk_size=500):
        if e.visitor_visit_id:
            entry_role_counts['visitors'] += 1
            continue
        st = e.student
        if not st:
            entry_role_counts['visitors'] += 1
            continue
        sid_u = (st.student_id or '').strip().upper()
        if sid_u == 'GUEST':
            entry_role_counts['visitors'] += 1
            continue
        entry_role_counts['students'] += 1

    entry_role_labels = ['Students', 'Visitors']
    entry_role_values = [
        entry_role_counts['students'],
        entry_role_counts['visitors'],
    ]
    entry_role_total = sum(entry_role_values)

    # Weekly peak hours heatmap (3 time bands × Mon–Sun), full overview range
    hm_qs = _traffic_flow_student_gate_qs(tf_range_start, tf_range_end_excl, overview_gate)
    hm_matrix = [[0] * 7 for _ in range(3)]

    def _heatmap_band(h):
        if 6 <= h <= 10:
            return 0
        if 11 <= h <= 14:
            return 1
        if 15 <= h <= 17:
            return 2
        return None

    for ts in hm_qs.values_list('timestamp', flat=True):
        lt = timezone.localtime(ts)
        b = _heatmap_band(lt.hour)
        if b is None:
            continue
        hm_matrix[b][lt.weekday()] += 1

    hm_flat = [hm_matrix[r][c] for r in range(3) for c in range(7)]
    hm_min = min(hm_flat) if hm_flat else 0
    hm_max = max(hm_flat) if hm_flat else 0

    def _heatmap_tier(v):
        if hm_max == hm_min:
            return 'low' if v == 0 else 'peak'
        ratio = (v - hm_min) / (hm_max - hm_min)
        if ratio < 0.25:
            return 'low'
        if ratio < 0.5:
            return 'medium'
        if ratio < 0.75:
            return 'high'
        return 'peak'

    heatmap_row_labels = ['6–10 AM', '11 AM–2 PM', '3–5 PM']
    heatmap_weekday_labels = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
    heatmap_rows = []
    for r in range(3):
        cells = [{'value': hm_matrix[r][c], 'tier': _heatmap_tier(hm_matrix[r][c])} for c in range(7)]
        heatmap_rows.append({'label': heatmap_row_labels[r], 'cells': cells})
    heatmap_subtitle = (
        f'{overview_start.strftime("%b %d, %Y")} – {overview_end.strftime("%b %d, %Y")} · IN+OUT scans'
    )

    # Bar charts: student daily gate scans by program, year level, gender (same window as overview / heatmap)
    _bar_student_qs = GateEntry.objects.filter(
        timestamp__gte=tf_range_start,
        timestamp__lt=tf_range_end_excl,
        granted=True,
        student_id__isnull=False,
        event__isnull=True,
        visitor_visit__isnull=True,
    )
    _course_lbl = dict(Student.COURSE_CHOICES)
    _year_lbl = dict(Student.YEAR_LEVEL_CHOICES)
    _sex_lbl = dict(Student.SEX_CHOICES)

    def _bar_course_label(code):
        code = (code or '').strip()
        return _course_lbl.get(code, code) if code else 'Unspecified'

    def _bar_year_label(code):
        code = (code or '').strip()
        return _year_lbl.get(code, code) if code else 'Unspecified'

    def _bar_sex_label(code):
        code = (code or '').strip()
        return _sex_lbl.get(code, code) if code else ''

    # Clear Model.Meta.ordering so GROUP BY is only student__course (otherwise -timestamp splits every row).
    _by_course = list(
        _bar_student_qs.order_by().values('student__course').annotate(c=Count('id'))
    )

    def _course_sort_key(row):
        v = row.get('student__course') or ''
        return ({'BST': 0, 'BSE': 1}.get(v, 2), v)

    _by_course.sort(key=_course_sort_key)

    _by_year = list(
        _bar_student_qs.order_by().values('student__year_level').annotate(c=Count('id'))
    )

    def _year_sort_key(row):
        v = row.get('student__year_level') or ''
        if v in ('1', '2', '3', '4'):
            return (0, int(v))
        return (1, v)

    _by_year.sort(key=_year_sort_key)

    _by_sex = list(
        _bar_student_qs.filter(student__sex__in=[Student.SEX_MALE, Student.SEX_FEMALE])
        .order_by()
        .values('student__sex')
        .annotate(c=Count('id'))
    )
    _sex_order = {'MALE': 0, 'FEMALE': 1}

    def _sex_sort_key(row):
        v = row.get('student__sex') or ''
        return (_sex_order.get(v, 99), v)

    _by_sex.sort(key=_sex_sort_key)

    student_bar_subtitle = (
        f'{overview_start.strftime("%b %d, %Y")} – {overview_end.strftime("%b %d, %Y")} · daily gate IN+OUT scans'
    )
    student_bar_any_data = _bar_student_qs.exists()

    def _gate_student_list_url(**kwargs):
        from urllib.parse import urlencode
        q = {k: v for k, v in kwargs.items() if v is not None and str(v).strip() != ''}
        base = reverse('gate-student-list')
        return base + ('?' + urlencode(q) if q else '')

    profile_kpi_by_program = []
    for r in _by_course[:10]:
        code = (r.get('student__course') or '').strip()
        profile_kpi_by_program.append({
            'label': _bar_course_label(code),
            'count': r['c'],
            'key': code,
            'link': _gate_student_list_url(course=code) if code else reverse('gate-student-list'),
        })
    profile_kpi_by_year = []
    for r in _by_year[:10]:
        yl = (r.get('student__year_level') or '').strip()
        profile_kpi_by_year.append({
            'label': _bar_year_label(yl),
            'count': r['c'],
            'key': yl,
            'link': _gate_student_list_url(year_level=yl) if yl else reverse('gate-student-list'),
        })
    profile_kpi_by_sex = []
    for r in _by_sex[:6]:
        sx = (r.get('student__sex') or '').strip()
        profile_kpi_by_sex.append({
            'label': _bar_sex_label(sx) or '—',
            'count': r['c'],
            'key': sx,
            'link': _gate_student_list_url(sex=sx) if sx else reverse('gate-student-list'),
        })

    # Legacy key: empty dict (student chart + meta removed). Keeps partials/templates safe if still referenced.
    chart_meta = {}

    ctx = {
        'site_name': 'City College of Bayawan',
        'today_iso': today.isoformat(),
        'granted_today': granted_today,
        'denied_today': denied_today,
        'incidents_today': 0,
        'total_students': total_students,
        'inside_now': inside_now,
        'active_events': active_events,
        'recent_incidents': recent_incidents,
        'participation_stats': participation_stats,
        'past_events_with_stats': past_events_with_stats,
        'selected_year': selected_year,
        'available_years': available_years,
        'visitors_this_month': visitors_this_month,
        'visitors_this_year': visitors_this_year,
        'student_entries_this_month': student_entries_this_month,
        'student_entries_this_year': student_entries_this_year,
        'student_entries_month_from': first_of_month.isoformat(),
        'student_entries_month_to': today.isoformat(),
        'student_entries_year_from': datetime.date(selected_year, 1, 1).isoformat(),
        'student_entries_year_to': year_end_for_link.isoformat(),
        'top_departments_monthly': top_departments_monthly,
        'top_departments_annually': top_departments_annually,
        'chart_reason_labels_json': json.dumps(reason_labels),
        'chart_reason_counts_json': json.dumps(reason_counts),
        'chart_reason_colors_json': json.dumps(reason_colors),
        'default_first_signatory_name': default_first_signatory_name,
        'default_first_signatory_title': default_first_signatory_title,
        'default_second_signatory_name': default_second_signatory_name,
        'default_second_signatory_title': default_second_signatory_title,
        'student_program_choices': Student.COURSE_CHOICES,
        'student_section_choices': student_section_choices,
        'student_reason_choices': student_reason_choices,
        'visitor_department_choices': visitor_department_choices,
        'report_event_choices': report_event_choices,
        'overview_start': overview_start,
        'overview_end': overview_end,
        'overview_from_iso': overview_start.isoformat(),
        'overview_to_iso': overview_end.isoformat(),
        'overview_preset': overview_preset,
        'overview_total_scans': overview_total_scans,
        'overview_total_pct': overview_total_pct,
        'overview_total_trend': overview_total_trend,
        'overview_avg_daily': round(overview_avg_daily, 1),
        'overview_avg_status': overview_avg_status,
        'overview_visitors': overview_visitors,
        'overview_visitors_pct': overview_visitors_pct,
        'overview_visitors_trend': overview_visitors_trend,
        'overview_peak_display': overview_peak_display,
        'overview_peak_rush': overview_peak_rush,
        'overview_num_days': overview_num_days,
        'traffic_hour_scope': traffic_hour_scope,
        'traffic_flow_title': traffic_flow_title,
        'traffic_flow_subtitle': traffic_flow_subtitle,
        'traffic_flow_labels': traffic_flow_labels,
        'traffic_flow_in': traffic_flow_in,
        'traffic_flow_out': traffic_flow_out,
        'entry_role_labels': entry_role_labels,
        'entry_role_values': entry_role_values,
        'entry_role_total': entry_role_total,
        'entry_role_range_note': entry_role_range_note,
        'heatmap_rows': heatmap_rows,
        'heatmap_weekday_labels': heatmap_weekday_labels,
        'heatmap_subtitle': heatmap_subtitle,
        'student_bar_subtitle': student_bar_subtitle,
        'student_bar_any_data': student_bar_any_data,
        'profile_kpi_by_program': profile_kpi_by_program,
        'profile_kpi_by_year': profile_kpi_by_year,
        'profile_kpi_by_sex': profile_kpi_by_sex,
        'chart_meta': chart_meta,
    }
    ctx['ccb_post_login_loader'] = pop_post_login_loader(request)
    if request.GET.get('partial') == '1':
        return render(request, 'gate/analytics_inner.html', ctx)
    return render(request, 'gate/analytics.html', ctx)


def _parse_analytics_overview_gate_date_range(request):
    """Same overview date range as analytics_dashboard; returns (tf_start, tf_end_exclusive, overview_start, overview_end)."""
    from django.utils.dateparse import parse_date

    today = timezone.localdate()
    overview_preset = (request.GET.get('overview_preset') or '').strip().lower()
    _preset_allowed = ('today', 'yesterday', 'last_7_days', 'custom')
    if overview_preset not in _preset_allowed:
        overview_preset = ''
    overview_from_raw = (request.GET.get('overview_from') or '').strip()
    overview_to_raw = (request.GET.get('overview_to') or '').strip()
    overview_from_parsed = parse_date(overview_from_raw) if overview_from_raw else None
    overview_to_parsed = parse_date(overview_to_raw) if overview_to_raw else None
    if not overview_preset:
        if overview_from_parsed and overview_to_parsed:
            overview_preset = 'custom'
        else:
            overview_preset = 'today'
    if overview_preset == 'custom':
        if overview_from_parsed and overview_to_parsed:
            overview_start = overview_from_parsed
            overview_end = overview_to_parsed
            if overview_start > overview_end:
                overview_start, overview_end = overview_end, overview_start
        else:
            overview_start = overview_end = today
    else:
        if overview_preset == 'today':
            overview_start = overview_end = today
        elif overview_preset == 'yesterday':
            _yd = today - datetime.timedelta(days=1)
            overview_start = overview_end = _yd
        elif overview_preset == 'last_7_days':
            overview_start = today - datetime.timedelta(days=6)
            overview_end = today
        else:
            overview_start = overview_end = today
    if overview_end > today:
        overview_end = today
    if overview_start > overview_end:
        overview_start = overview_end
    tf_range_start, _ = _local_day_bounds(overview_start)
    _, tf_range_end_excl = _local_day_bounds(overview_end)
    return tf_range_start, tf_range_end_excl, overview_start, overview_end


@require_GET
def analytics_profile_entries_json(request):
    """JSON list of student daily gate IN/OUT scans for a profile slice (program, year level, gender).

    Always returns JSON (including auth errors) so fetch() can parse the body — HTML 403/redirect
    from decorators would break r.json() in the browser and show a generic error.
    """
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Sign in required. Refresh the page and try again.'}, status=401)
    role = get_user_role(request.user)
    if role is None:
        return JsonResponse({'error': 'Your account has no role. Contact an administrator.'}, status=403)
    if role not in ('admin', 'staff', 'faculty'):
        return JsonResponse({'error': 'You do not have permission to load gate entries.'}, status=403)

    dim = (request.GET.get('dimension') or '').strip().lower()
    raw_val = request.GET.get('value')
    val = (raw_val or '').strip() if raw_val is not None else ''

    if dim not in ('course', 'year_level', 'sex'):
        return JsonResponse({'error': 'Invalid dimension'}, status=400)

    try:
        tf_start, tf_end_excl, overview_start, overview_end = _parse_analytics_overview_gate_date_range(request)

        qs = GateEntry.objects.filter(
            timestamp__gte=tf_start,
            timestamp__lt=tf_end_excl,
            granted=True,
            student_id__isnull=False,
            event__isnull=True,
            visitor_visit__isnull=True,
        ).select_related('student')

        if dim == 'course':
            qs = qs.filter(student__course=val)
        elif dim == 'year_level':
            qs = qs.filter(student__year_level=val)
        elif dim == 'sex':
            if not val:
                return JsonResponse({'error': 'Missing value'}, status=400)
            qs = qs.filter(student__sex=val)

        agg = qs.aggregate(
            total=Count('id'),
            in_count=Count('id', filter=Q(scan_type='IN')),
            out_count=Count('id', filter=Q(scan_type='OUT')),
        )
        total = agg['total'] or 0
        in_count = agg['in_count'] or 0
        out_count = agg['out_count'] or 0

        LIMIT = 500
        out_code_lbl = dict(GateEntry.OUT_REASON_CODE_CHOICES)
        rows = []
        for e in qs.order_by('-timestamp')[:LIMIT]:
            st = e.student
            sid = (st.student_id or '').strip() if st else ''
            fn = (getattr(st, 'first_name', None) or '') if st else ''
            ln = (getattr(st, 'last_name', None) or '') if st else ''
            name = (fn + ' ' + ln).strip() or '—'
            sex_disp = _student_sex_display(st)
            yl_disp = '—'
            if st and getattr(st, 'year_level', None):
                yl_disp = st.get_year_level_display()
            sec_disp = '—'
            if st and (st.section or '').strip():
                sec_disp = (st.section or '').strip()
            lt = timezone.localtime(e.timestamp)
            local_time = lt.strftime('%b %d, %Y %I:%M %p').replace(' 0', ' ')
            code = (e.out_reason_code or '').strip()
            note = (e.out_reason or '').strip()
            if code and note:
                out_disp = f'{out_code_lbl.get(code, code)} · {note}'
            elif note:
                out_disp = note
            elif code:
                out_disp = out_code_lbl.get(code, code)
            else:
                out_disp = ''
            if len(out_disp) > 120:
                out_disp = out_disp[:117] + '…'
            rows.append(
                {
                    'entry_id': e.pk,
                    'timestamp': e.timestamp.isoformat(),
                    'local_time': local_time,
                    'scan_type': (e.scan_type or '').upper(),
                    'student_id': sid,
                    'student_name': name,
                    'gender': sex_disp,
                    'course': (st.course or '') if st else '',
                    'year_level': yl_disp,
                    'section': sec_disp,
                    'out_note': out_disp,
                }
            )

        range_label = (
            f'{overview_start.strftime("%b %d, %Y")} – {overview_end.strftime("%b %d, %Y")}'
            if overview_start != overview_end
            else overview_start.strftime('%b %d, %Y')
        )

        gate_log_q = urlencode(
            {
                'from_date': overview_start.isoformat(),
                'to_date': overview_end.isoformat(),
                'student_gate_only': '1',
            }
        )
        gate_log_url = reverse('gate-entry-list') + '?' + gate_log_q

        return JsonResponse(
            {
                'dimension': dim,
                'value': val,
                'range_label': range_label,
                'total': total,
                'in_count': in_count,
                'out_count': out_count,
                'entries': rows,
                'truncated': total > LIMIT,
                'source': 'gate_entries',
                'gate_log_url': gate_log_url,
            }
        )
    except Exception as e:
        logging.getLogger(__name__).exception('analytics_profile_entries_json failed: %s', e)
        return JsonResponse({'error': 'Server error while loading gate entries.'}, status=500)


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def analytics_report(request):
    """Redirect: analytics is view-only per panel; print report removed."""
    return redirect('gate-analytics')


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_qr_image(request, pk):
    """Generate QR code image for student (encodes student_id for gate scan). On-demand, no DB change."""
    student = get_object_or_404(Student, pk=pk)
    payload = (student.student_id or '').strip() or str(student.pk)
    payload = str(payload)
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=8, border=3)
        qr.add_data(payload)
        qr.make(fit=True)
        # No fill_color/back_color for maximum compatibility across qrcode/PIL versions
        img = qr.make_image()
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        buffer.seek(0)
        return HttpResponse(buffer.getvalue(), content_type='image/png')
    except Exception as e:
        import logging
        logging.getLogger(__name__).warning('student_qr_image failed for pk=%s student_id=%r: %s', pk, getattr(student, 'student_id', None), e)
        try:
            from PIL import Image
            buf = io.BytesIO()
            placeholder = Image.new('RGB', (56, 56), color=(240, 240, 240))
            placeholder.save(buf, format='PNG')
            buf.seek(0)
            return HttpResponse(buf.getvalue(), content_type='image/png')
        except Exception:
            return HttpResponse(
                b'\x89PNG\r\n\x1a\n\x00\x00\x00\rIHDR\x00\x00\x00\x01\x00\x00\x00\x01\x08\x02\x00\x00\x00\x90wS\xde\x00\x00\x00\x0cIDATx\x9cc\xf8\x0f\x00\x00\x01\x01\x00\x05\x18\xd8N\x00\x00\x00\x00IEND\xaeB`\x82',
                content_type='image/png'
            )


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_eid_card(request, pk):
    """Printable student e-ID card (front + back) with permanent QR. View in modal or print/download."""
    student = get_object_or_404(Student, pk=pk)
    qr_url = request.build_absolute_uri(reverse('gate-student-qr', kwargs={'pk': student.pk}))
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')
    logo_url = None
    try:
        theme = SiteTheme.objects.first()
        if theme and theme.logo:
            logo_url = request.build_absolute_uri(theme.logo.url)
    except Exception:
        pass
    if not logo_url:
        try:
            logo_url = request.build_absolute_uri(static('gate/images/university-logo.png'))
        except Exception:
            pass
    photo_url = None
    if getattr(student, 'photo', None) and student.photo:
        try:
            url = student.photo.url
            if url.startswith('http://') or url.startswith('https://'):
                photo_url = url
            else:
                photo_url = request.build_absolute_uri(url)
        except Exception:
            pass
    signature_url = None
    if getattr(student, 'signature', None) and student.signature:
        try:
            url = student.signature.url
            if url.startswith('http://') or url.startswith('https://'):
                signature_url = url
            else:
                signature_url = request.build_absolute_uri(url)
        except Exception:
            pass
    return render(request, 'gate/student_eid_card.html', {
        'student': student,
        'qr_url': qr_url,
        'site_name': site_name,
        'logo_url': logo_url,
        'photo_url': photo_url,
        'signature_url': signature_url,
    })


@login_required(login_url='/login/')
@role_required('student affairs')
def eid_signatory_settings(request):
    """
    Legacy / bookmark URL: opens the Student Affairs e-ID signatory modal on the student list.
    Saving is done via POST on the student list (same form).
    """
    from django.urls import reverse
    from urllib.parse import urlencode

    return redirect(reverse('gate-student-list') + '?' + urlencode({'eid_sig': '1'}))


def _student_qr_png_bytes(student, box_size=8):
    """Return QR code PNG bytes for a student (for embedding in PDF / print-all)."""
    payload = (getattr(student, 'student_id', None) or '').strip() or str(student.pk)
    payload = str(payload)
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=box_size, border=3)
        qr.add_data(payload)
        qr.make(fit=True)
        img = qr.make_image()
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None


def _build_student_eid_export_pdf_template_context(request, students):
    """Build template context for gate/student_eid_export_pdf.html (data-URI assets, no external URLs)."""
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')

    def _file_to_data_uri(filepath, mime='image/png'):
        try:
            with open(filepath, 'rb') as f:
                return f'data:{mime};base64,{base64.b64encode(f.read()).decode("ascii")}'
        except Exception:
            return None

    def _model_file_to_data_uri(field, mime='image/jpeg'):
        try:
            if field and field.path and os.path.isfile(field.path):
                return _file_to_data_uri(field.path, mime)
        except Exception:
            pass
        return None

    def _model_file_to_resized_jpeg_data_uri(field, max_size=(480, 618), quality=88):
        if not (field and getattr(field, "path", None) and os.path.isfile(field.path)):
            return None
        try:
            from PIL import Image
            img = Image.open(field.path)
            img = img.convert("RGB")
            img.thumbnail(max_size, Image.Resampling.LANCZOS)
            buf = io.BytesIO()
            img.save(buf, format="JPEG", quality=quality, optimize=True)
            buf.seek(0)
            b = buf.getvalue()
            return f"data:image/jpeg;base64,{base64.b64encode(b).decode('ascii')}"
        except Exception:
            try:
                return _model_file_to_data_uri(field, mime="image/jpeg")
            except Exception:
                return None

    logo_url = None
    try:
        theme = SiteTheme.objects.first()
        if theme and theme.logo:
            logo_url = _model_file_to_data_uri(theme.logo, 'image/png')
    except Exception:
        pass
    if not logo_url:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'gate', 'images', 'CBB.png')
        logo_url = _file_to_data_uri(logo_path, 'image/png')

    bg_path = os.path.join(settings.BASE_DIR, 'static', 'gate', 'images', 'CCB.jpg')
    bg_data_uri = _file_to_data_uri(bg_path, 'image/jpeg')

    def _mime_for_image_field(field):
        if not field:
            return 'image/png'
        try:
            path = field.path
        except Exception:
            return 'image/png'
        ext = os.path.splitext(path)[1].lower()
        if ext in ('.jpg', '.jpeg'):
            return 'image/jpeg'
        if ext == '.webp':
            return 'image/webp'
        return 'image/png'

    pdf_theme = SiteTheme.objects.first()
    pdf_fs_name = pdf_fs_title = pdf_ss_name = pdf_ss_title = ''
    pdf_first_sig_data_uri = pdf_second_sig_data_uri = ''
    if pdf_theme:
        pdf_fs_name = (pdf_theme.default_first_signatory_name or '').strip()
        pdf_fs_title = (pdf_theme.default_first_signatory_title or '').strip()
        pdf_ss_name = (pdf_theme.default_second_signatory_name or '').strip()
        pdf_ss_title = (pdf_theme.default_second_signatory_title or '').strip()
        pdf_first_sig_data_uri = (
            _model_file_to_data_uri(
                pdf_theme.first_signatory_signature,
                _mime_for_image_field(pdf_theme.first_signatory_signature),
            )
            or ''
        )
        pdf_second_sig_data_uri = (
            _model_file_to_data_uri(
                pdf_theme.second_signatory_signature,
                _mime_for_image_field(pdf_theme.second_signatory_signature),
            )
            or ''
        )

    student_cards = []
    for student in students:
        qr_bytes = _student_qr_png_bytes(student, box_size=12)
        if qr_bytes:
            qr_url = f'data:image/png;base64,{base64.b64encode(qr_bytes).decode("ascii")}'
        else:
            qr_url = ''

        photo_url = _model_file_to_resized_jpeg_data_uri(getattr(student, 'photo', None))
        if not photo_url:
            try:
                if getattr(student, 'photo', None) and student.photo:
                    photo_url = request.build_absolute_uri(student.photo.url)
            except Exception:
                photo_url = ''
        student_cards.append({'student': student, 'qr_url': qr_url, 'photo_url': photo_url})

    return {
        'student_cards': student_cards,
        'site_name': site_name,
        'logo_url': logo_url,
        'bg_data_uri': bg_data_uri,
        'first_signatory_name': pdf_fs_name,
        'first_signatory_title': pdf_fs_title,
        'second_signatory_name': pdf_ss_name,
        'second_signatory_title': pdf_ss_title,
        'first_signatory_signature_data_uri': pdf_first_sig_data_uri,
        'second_signatory_signature_data_uri': pdf_second_sig_data_uri,
    }


def _eid_student_cards_zip_response(request, template_ctx, image_format):
    """
    ZIP of per-card raster images (front/back .id-card elements) via Playwright screenshots.
    image_format: 'png' or 'jpeg'.
    """
    import time

    _log = logging.getLogger(__name__)
    html = render_to_string('gate/student_eid_export_pdf.html', template_ctx, request=request)
    started_at = time.monotonic()
    cards_meta = template_ctx.get('student_cards') or []
    try:
        from playwright.sync_api import sync_playwright
        with sync_playwright() as p:
            browser = p.chromium.launch()
            context = None
            n_students = max(1, len(cards_meta))
            vp_h = min(12000, 800 + n_students * 520)
            try:
                context = browser.new_context(
                    device_scale_factor=2,
                    viewport={'width': 1200, 'height': vp_h},
                )
                page = context.new_page()
                n = max(120000, 15000 + len(cards_meta) * 400)
                page.set_default_timeout(n)
                try:
                    page.emulate_media(media='print')
                except Exception:
                    pass
                page.set_content(html, wait_until='domcontentloaded', timeout=n)
                page.wait_for_timeout(min(4000, 800 + len(cards_meta) * 25))
                handles = page.query_selector_all('.id-card')
                buf = io.BytesIO()
                ext = 'png' if image_format == 'png' else 'jpg'
                with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
                    for idx, h in enumerate(handles):
                        if image_format == 'png':
                            try:
                                shot = h.screenshot(type='png', scale='device')
                            except TypeError:
                                shot = h.screenshot(type='png')
                        else:
                            try:
                                shot = h.screenshot(type='jpeg', quality=92, scale='device')
                            except TypeError:
                                shot = h.screenshot(type='jpeg', quality=92)
                        student_idx = idx // 2
                        side = 'front' if (idx % 2 == 0) else 'back'
                        stu = cards_meta[student_idx]['student'] if student_idx < len(cards_meta) else None
                        sid_raw = (
                            (getattr(stu, 'student_id', None) or (str(stu.pk) if stu else 'card')).strip()
                        )
                        sid = re.sub(r'[^\w\-.]+', '_', sid_raw)[:80] or 'id'
                        zf.writestr(f'eid-{sid}-{side}.{ext}', shot)
            finally:
                try:
                    if context is not None:
                        context.close()
                except Exception:
                    pass
                browser.close()
    except Exception as e:
        _log.warning('student_eid_print_all: ZIP image export failed: %s', e)
        return HttpResponse(
            'Image ZIP export failed. Install Playwright Chromium: pip install playwright && playwright install chromium',
            status=503,
            content_type='text/plain; charset=utf-8',
        )

    payload = buf.getvalue()
    zip_name = 'student-eids-png.zip' if image_format == 'png' else 'student-eids-jpg.zip'
    response = HttpResponse(payload, content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{zip_name}"'
    _log.info(
        'student_eid_print_all: ZIP %s cards=%s bytes=%s elapsed=%.2fs',
        image_format,
        len(cards_meta),
        len(payload),
        time.monotonic() - started_at,
    )
    return response


def _students_for_eid_export(request):
    """Same GET filters as e-ID print-all / CSV export: status, course, year_level, section, sex, q."""
    status_filter = (request.GET.get('status') or '').strip()
    if status_filter == 'all':
        students = Student.objects.all()
    else:
        students = Student.objects.filter(account_status=Student.ACCOUNT_STATUS_APPROVED)
    filter_course = (request.GET.get('course') or '').strip()
    filter_year = (request.GET.get('year_level') or '').strip()
    filter_section = (request.GET.get('section') or '').strip()
    filter_sex = (request.GET.get('sex') or '').strip()
    search_q = (request.GET.get('q') or '').strip()
    if filter_course:
        students = students.filter(course=filter_course)
    if filter_year:
        students = students.filter(year_level=filter_year)
    if filter_section:
        students = students.filter(section__iexact=filter_section)
    if filter_sex:
        students = students.filter(sex=filter_sex)
    if search_q:
        students = students.filter(
            Q(first_name__icontains=search_q) |
            Q(last_name__icontains=search_q) |
            Q(middle_name__icontains=search_q) |
            Q(student_id__icontains=search_q)
        )
    return students.order_by('last_name', 'first_name')


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_eid_print_all(request):
    """Printable page with all (filtered) student e-ID cards. Same filters as student list. eid_zip_format=png|jpg → ZIP of card images."""
    students = list(_students_for_eid_export(request))
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')

    eid_zip_fmt = (request.GET.get('eid_zip_format') or '').strip().lower()
    if eid_zip_fmt == 'jpg':
        eid_zip_fmt = 'jpeg'
    if students and eid_zip_fmt in ('png', 'jpeg'):
        ctx = _build_student_eid_export_pdf_template_context(request, students)
        return _eid_student_cards_zip_response(request, ctx, eid_zip_fmt)

    # On Windows, weasyprint usually fails (GTK libs missing) and Playwright can be unstable,
    # which causes the browser "Server Closed Connection Suddenly" error.
    # Fallback: when pdf=1 on Windows, render the normal print HTML view instead.
    if request.GET.get('pdf') and students and platform.system() != 'Windows':
        import time

        ctx = _build_student_eid_export_pdf_template_context(request, students)
        html = render_to_string('gate/student_eid_export_pdf.html', ctx, request=request)

        base_url = (request.build_absolute_uri('/') or '/').rstrip('/') + '/'
        pdf_bytes = None
        _log = logging.getLogger(__name__)
        started_at = time.monotonic()
        pdf_method = None
        _log.info(
            "student_eid_print_all: start PDF generation for %s students (platform=%s)",
            len(students),
            platform.system(),
        )

        # Prefer WeasyPrint (fast, stable). If WeasyPrint can't run on this machine,
        # fallback to Playwright.
        try:
            from weasyprint import HTML
            pdf_bytes = HTML(string=html, base_url=base_url).write_pdf()
            pdf_method = 'weasyprint'
        except (ImportError, OSError) as e:
            _log.debug('student_eid_print_all: WeasyPrint unavailable: %s', e)
        except Exception as e:
            _log.warning('student_eid_print_all: WeasyPrint PDF failed: %s', e)

        # Playwright fallback: never use wait_until='networkidle' here — data-URI-only HTML
        # often never reaches "idle", which hangs the worker and drops the connection (IDM "server closed").
        if not pdf_bytes:
            try:
                from playwright.sync_api import sync_playwright
                pdf_method = 'playwright'
                with sync_playwright() as p:
                    browser = p.chromium.launch()
                    context = None
                    n_pdf = max(1, len(students))
                    vp_h_pdf = min(12000, 600 + n_pdf * 220)
                    try:
                        context = browser.new_context(
                            device_scale_factor=2,
                            viewport={'width': 1040, 'height': vp_h_pdf},
                        )
                        page = context.new_page()
                        page.set_default_timeout(20000)
                        # Ensure print-specific CSS applies.
                        try:
                            page.emulate_media(media='print')
                        except Exception:
                            pass
                        page.set_content(html, wait_until='domcontentloaded', timeout=20000)
                        # Give the browser a moment to lay out embedded images.
                        page.wait_for_timeout(800)
                        pdf_bytes = page.pdf(
                            width='3.377in',
                            height='2.127in',
                            print_background=True,
                            margin={'top': '0', 'right': '0', 'bottom': '0', 'left': '0'},
                            scale=2,
                        )
                    finally:
                        try:
                            if context is not None:
                                context.close()
                        except Exception:
                            pass
                        browser.close()
            except Exception as e:
                _log.warning('student_eid_print_all: Playwright PDF failed: %s', e)

        if pdf_bytes:
            response = HttpResponse(pdf_bytes, content_type='application/pdf')
            response['Content-Disposition'] = 'attachment; filename="student-eids.pdf"'
            _log.info(
                "student_eid_print_all: finished method=%s bytes=%s elapsed=%.2fs",
                pdf_method,
                len(pdf_bytes),
                time.monotonic() - started_at,
            )
            return response

        _log.warning(
            "student_eid_print_all: PDF export failed (method tried=%s) elapsed=%.2fs",
            pdf_method,
            time.monotonic() - started_at,
        )
        return HttpResponse(
            'PDF export failed. Install WeasyPrint dependencies on the server, or run '
            '"playwright install chromium". You can still use Print from the HTML view.',
            status=503,
            content_type='text/plain; charset=utf-8',
        )

    logo_url = None
    try:
        theme = SiteTheme.objects.first()
        if theme and theme.logo:
            logo_url = request.build_absolute_uri(theme.logo.url)
    except Exception:
        pass
    student_cards = []
    for s in students:
        qr_url = request.build_absolute_uri(reverse('gate-student-qr', kwargs={'pk': s.pk}))
        photo_url = None
        if getattr(s, 'photo', None) and s.photo:
            try:
                photo_url = request.build_absolute_uri(s.photo.url)
            except Exception:
                pass
        student_cards.append({'student': s, 'qr_url': qr_url, 'photo_url': photo_url})
    return render(request, 'gate/student_eid_print_all.html', {
        'student_cards': student_cards,
        'total': len(student_cards),
        'site_name': site_name,
        'logo_url': logo_url,
    })


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_qr_labels_pdf(request):
    """
    PDF download: one 1.5in×1.5in QR per student (student_id payload), full name + ID label below.
    Same GET filters as e-ID print-all (status, course, year_level, section, sex, q).
    """
    import time

    students = list(_students_for_eid_export(request))
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')
    qr_rows = []
    for student in students:
        qr_bytes = _student_qr_png_bytes(student, box_size=12)
        if qr_bytes:
            qr_url = f'data:image/png;base64,{base64.b64encode(qr_bytes).decode("ascii")}'
        else:
            qr_url = ''
        sid = (getattr(student, 'student_id', None) or '').strip() or str(student.pk)
        qr_rows.append({
            'qr_url': qr_url,
            'full_name': _fmt_student_name(student),
            'student_id': sid,
        })
    if not qr_rows:
        return HttpResponse(
            'No students match the selected filters.',
            status=404,
            content_type='text/plain; charset=utf-8',
        )
    ctx = {'qr_rows': qr_rows, 'site_name': site_name}
    html = render_to_string('gate/student_qr_labels_pdf.html', ctx, request=request)
    base_url = (request.build_absolute_uri('/') or '/').rstrip('/') + '/'
    pdf_bytes = None
    _log = logging.getLogger(__name__)
    started_at = time.monotonic()
    pdf_method = None

    try:
        from weasyprint import HTML
        pdf_bytes = HTML(string=html, base_url=base_url).write_pdf()
        pdf_method = 'weasyprint'
    except (ImportError, OSError) as e:
        _log.debug('student_qr_labels_pdf: WeasyPrint unavailable: %s', e)
    except Exception as e:
        _log.warning('student_qr_labels_pdf: WeasyPrint failed: %s', e)

    if not pdf_bytes:
        try:
            from playwright.sync_api import sync_playwright
            pdf_method = 'playwright'
            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page(viewport={'width': 816, 'height': 1056})
                try:
                    page.set_default_timeout(120000)
                    try:
                        page.emulate_media(media='print')
                    except Exception:
                        pass
                    page.set_content(html, wait_until='domcontentloaded', timeout=120000)
                    page.wait_for_timeout(800)
                    pdf_bytes = page.pdf(
                        format='Letter',
                        print_background=True,
                        margin={'top': '0.45in', 'right': '0.45in', 'bottom': '0.45in', 'left': '0.45in'},
                    )
                finally:
                    browser.close()
        except Exception as e:
            _log.warning('student_qr_labels_pdf: Playwright failed: %s', e)

    if pdf_bytes:
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        # preview=1 → inline so the browser opens the PDF in a tab instead of forcing download
        _preview = (request.GET.get('preview') or '').strip().lower() in ('1', 'true', 'yes')
        if _preview:
            response['Content-Disposition'] = 'inline; filename="student-qr-labels.pdf"'
        else:
            response['Content-Disposition'] = 'attachment; filename="student-qr-labels.pdf"'
        _log.info(
            'student_qr_labels_pdf: method=%s students=%s bytes=%s elapsed=%.2fs',
            pdf_method,
            len(qr_rows),
            len(pdf_bytes),
            time.monotonic() - started_at,
        )
        return response

    return HttpResponse(
        'PDF export failed. Install WeasyPrint (with GTK on Linux) or run '
        '"pip install playwright && playwright install chromium" for a browser fallback.',
        status=503,
        content_type='text/plain; charset=utf-8',
    )


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'student affairs')
def visitor_pass_qr_image(request, code):
    """Generate QR code image for a visitor pass code (e.g. VIS-001).

    Query params:
    - format: png|jpg|jpeg (default: png)
    - download: 1|true|yes -> force attachment filename
    """
    pass_obj = get_object_or_404(VisitorPass, code=code)
    try:
        image_format = (request.GET.get('format') or 'png').strip().lower()
        if image_format == 'jpg':
            image_format = 'jpeg'
        if image_format not in ('png', 'jpeg'):
            image_format = 'png'

        import qrcode
        qr = qrcode.QRCode(version=1, box_size=8, border=3)
        qr.add_data(pass_obj.code)
        qr.make(fit=True)
        img = qr.make_image(fill_color='#1a1a2e', back_color='white').convert('RGB')
        buffer = io.BytesIO()
        img.save(buffer, format='PNG' if image_format == 'png' else 'JPEG', quality=92)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type='image/png' if image_format == 'png' else 'image/jpeg'
        )
        should_download = (request.GET.get('download') or '').strip().lower() in ('1', 'true', 'yes')
        if should_download:
            ext = 'png' if image_format == 'png' else 'jpg'
            response['Content-Disposition'] = f'attachment; filename="visitor-qr-{pass_obj.code}.{ext}"'
        return response
    except Exception:
        from django.http import HttpResponseServerError
        return HttpResponseServerError(b'QR generation failed')


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'student affairs')
def visitor_pass_card(request, code):
    """Printable electronic ID card for a visitor pass (QR + code + branding)."""
    pass_obj = get_object_or_404(VisitorPass, code=code)
    should_download = (request.GET.get('download') or '').strip().lower() in ('1', 'true', 'yes')
    image_format = (request.GET.get('format') or 'png').strip().lower()
    if image_format == 'jpg':
        image_format = 'jpeg'
    if image_format not in ('png', 'jpeg'):
        image_format = 'png'

    qr_url = request.build_absolute_uri(
        reverse('visitor-qr-image', kwargs={'code': pass_obj.code})
    )
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')
    logo_url = None
    try:
        theme = SiteTheme.objects.first()
        if theme and theme.logo:
            logo_url = request.build_absolute_uri(theme.logo.url)
    except Exception:
        pass
    if not logo_url:
        # Fallback: university logo from project static file (place at static/gate/images/university-logo.png)
        try:
            logo_url = request.build_absolute_uri(static('gate/images/university-logo.png'))
        except Exception:
            pass

    if should_download:
        # Render the same HTML card layout and capture it as image.
        try:
            qr_bytes = _visitor_pass_qr_png_bytes(pass_obj)
            if qr_bytes:
                qr_url = f'data:image/png;base64,{base64.b64encode(qr_bytes).decode("ascii")}'

            html = render_to_string('gate/visitor_pass_card.html', {
                'pass_obj': pass_obj,
                'qr_url': qr_url,
                'site_name': site_name,
                'logo_url': logo_url,
            }, request=request)
            base_url = (request.build_absolute_uri('/') or '/').rstrip('/') + '/'
            html = _make_card_html_absolute_urls(html, base_url)

            shot = _screenshot_visitor_pass_card_html_bytes(html, image_format=image_format)
            if image_format == 'png':
                response = HttpResponse(shot, content_type='image/png')
                response['Content-Disposition'] = f'attachment; filename="visitor-eid-{pass_obj.code}.png"'
                return response
            response = HttpResponse(shot, content_type='image/jpeg')
            response['Content-Disposition'] = f'attachment; filename="visitor-eid-{pass_obj.code}.jpg"'
            return response
        except Exception:
            # Fallback path when Playwright is unavailable.
            card_png = _render_visitor_eid_card_png(pass_obj, site_name=site_name)
            if card_png:
                if image_format == 'png':
                    response = HttpResponse(card_png, content_type='image/png')
                    response['Content-Disposition'] = f'attachment; filename="visitor-eid-{pass_obj.code}.png"'
                    return response
                try:
                    from PIL import Image
                    source = Image.open(io.BytesIO(card_png)).convert('RGB')
                    out = io.BytesIO()
                    source.save(out, format='JPEG', quality=92)
                    out.seek(0)
                    response = HttpResponse(out.getvalue(), content_type='image/jpeg')
                    response['Content-Disposition'] = f'attachment; filename="visitor-eid-{pass_obj.code}.jpg"'
                    return response
                except Exception:
                    response = HttpResponse(card_png, content_type='image/png')
                    response['Content-Disposition'] = f'attachment; filename="visitor-eid-{pass_obj.code}.png"'
                    return response

    return render(request, 'gate/visitor_pass_card.html', {
        'pass_obj': pass_obj,
        'qr_url': qr_url,
        'site_name': site_name,
        'logo_url': logo_url,
    })


def _visitor_pass_qr_png_bytes(pass_obj):
    """Return QR code PNG bytes for a visitor pass (for embedding in PDF)."""
    try:
        import qrcode
        qr = qrcode.QRCode(version=1, box_size=8, border=3)
        qr.add_data(pass_obj.code)
        qr.make(fit=True)
        img = qr.make_image(fill_color='#1a1a2e', back_color='white')
        buf = io.BytesIO()
        img.save(buf, format='PNG')
        buf.seek(0)
        return buf.getvalue()
    except Exception:
        return None


def _make_card_html_absolute_urls(html, base_url):
    """Make static/media URLs in card HTML absolute so they load in Playwright set_content."""
    if base_url and not base_url.endswith('/'):
        base_url = base_url + '/'
    for old, new in (
        ('src="/static/', 'src="' + base_url + 'static/'),
        ('href="/static/', 'href="' + base_url + 'static/'),
        ('src="/media/', 'src="' + base_url + 'media/'),
        ('href="/media/', 'href="' + base_url + 'media/'),
    ):
        html = html.replace(old, new)
    return html


def _playwright_context_for_screenshots(browser):
    """Context whose requests skip ngrok's free-tier browser warning (logo/static via absolute tunnel URLs)."""
    # https://ngrok.com/docs/secure-tunnels/ngrok-agent/http — header bypasses interstitial for subresource loads.
    return browser.new_context(
        viewport={'width': 420, 'height': 560},
        extra_http_headers={'ngrok-skip-browser-warning': '1'},
    )


def _screenshot_visitor_pass_card_html_bytes(html, image_format='png'):
    """Capture visitor_pass_card.html as PNG/JPEG using the same Playwright pattern as single download.

    One browser per call — slower for batch, but matches modal layout reliably (esp. on Windows)
    and avoids long-lived sync_playwright().start() issues.

    Uses ngrok-skip-browser-warning on the Playwright context so when ``request.build_absolute_uri()``
    points at an ngrok tunnel, images (logo) load as real images instead of the interstitial HTML.
    """
    if image_format == 'jpg':
        image_format = 'jpeg'
    if image_format not in ('png', 'jpeg'):
        image_format = 'png'
    from playwright.sync_api import sync_playwright

    launch_args = [
        '--no-sandbox',
        '--disable-dev-shm-usage',
        '--disable-gpu',
    ]
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=True, args=launch_args)
        context = None
        try:
            context = _playwright_context_for_screenshots(browser)
            page = context.new_page()
            page.set_content(html, wait_until='domcontentloaded')
            page.wait_for_selector('.card-wrap', state='visible', timeout=15000)
            card = page.locator('.card-wrap').first
            if image_format == 'png':
                return card.screenshot(type='png')
            return card.screenshot(type='jpeg', quality=92)
        finally:
            try:
                if context is not None:
                    context.close()
            except Exception:
                pass
            browser.close()


def _render_card_html_to_pdf_with_page(page, html, height='4in', width='3in'):
    """Render one card HTML to PDF using an existing Playwright page. Returns PDF bytes or None."""
    try:
        page.set_content(html, wait_until='domcontentloaded')  # faster than networkidle
        return page.pdf(
            width=width,
            height=height,
            margin={'top': '0', 'right': '0', 'bottom': '0', 'left': '0'},
            print_background=True,
        )
    except Exception:
        return None


def _render_card_html_to_pdf_playwright(html, base_url):
    """Render card HTML to PDF using Playwright (Chromium). Returns PDF bytes or None. No GTK/Cairo needed."""
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        return None
    html = _make_card_html_absolute_urls(html, base_url)
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch()
            page = browser.new_page(viewport={'width': 240, 'height': 320})
            pdf_bytes = _render_card_html_to_pdf_with_page(page, html)
            browser.close()
        return pdf_bytes
    except Exception:
        return None


def _visitor_pass_list_for_user(request, no_limit=False):
    """Return passes queryset for visitor_pass_create and print-all.
    When a range (start/end) is in the URL, show all passes in that range so staff can view/print
    existing slots (regardless of who created them). Otherwise show recent passes for the whole campus.
    If no_limit=True, return full queryset (for pagination); otherwise default case is capped at 200 for print-all."""
    start_param = request.GET.get('start')
    end_param = request.GET.get('end')
    if start_param is not None and end_param is not None:
        try:
            s, e = int(start_param), int(end_param)
            if 1 <= s <= 999 and 1 <= e <= 999 and s <= e:
                code_min = VisitorPass.generate_reusable_code(s)
                code_max = VisitorPass.generate_reusable_code(e)
                return VisitorPass.objects.filter(
                    code__gte=code_min, code__lte=code_max
                ).order_by('code')
        except (TypeError, ValueError):
            pass
    qs = VisitorPass.objects.all().order_by('-created_at')
    if not no_limit:
        qs = qs[:200]
    return qs


def _render_visitor_eid_card_png(pass_obj, site_name='City College of Bayawan'):
    """Render a single visitor e-ID card as PNG bytes. Matches visitor_pass_card.html layout: green header + gray lines, yellow-orange bar, dark red Visitor Pass, white body, gold QR border, footer with gold line + shield/clock icons + divider. 288×384 px = 3×4 in at 96 DPI."""
    try:
        import qrcode
        from PIL import Image, ImageDraw, ImageFont
    except ImportError:
        return None
    card_w, card_h = 288, 384
    img = Image.new('RGB', (card_w, card_h), color='white')
    draw = ImageDraw.Draw(img)
    try:
        font_title = ImageFont.truetype("arial.ttf", 14)
        font_sub = ImageFont.truetype("arial.ttf", 10)
        font_bar = ImageFont.truetype("arial.ttf", 16)
        font_code = ImageFont.truetype("arial.ttf", 20)
        font_hint = ImageFont.truetype("arial.ttf", 9)
        font_footer = ImageFont.truetype("arial.ttf", 8)
    except (OSError, IOError):
        font_title = font_sub = font_bar = font_code = font_hint = font_footer = ImageFont.load_default()

    def _gradient_band(y1, y2, rgb_top, rgb_bottom):
        h = y2 - y1
        for i in range(h):
            t = i / h if h else 0
            r = int(rgb_top[0] + (rgb_bottom[0] - rgb_top[0]) * t)
            g = int(rgb_top[1] + (rgb_bottom[1] - rgb_top[1]) * t)
            b = int(rgb_top[2] + (rgb_bottom[2] - rgb_top[2]) * t)
            draw.rectangle([0, y1 + i, card_w, y1 + i + 1], fill=(r, g, b), outline=None)

    def _three_color_gradient(y1, y2, rgb_top, rgb_mid, rgb_bottom):
        """Vertical gradient with three stops to approximate card's 135deg header."""
        h = y2 - y1
        mid = h // 2
        for i in range(h):
            if i <= mid:
                t = i / mid if mid else 0
                r = int(rgb_top[0] + (rgb_mid[0] - rgb_top[0]) * t)
                g = int(rgb_top[1] + (rgb_mid[1] - rgb_top[1]) * t)
                b = int(rgb_top[2] + (rgb_mid[2] - rgb_top[2]) * t)
            else:
                t = (i - mid) / (h - mid) if (h - mid) else 0
                r = int(rgb_mid[0] + (rgb_bottom[0] - rgb_mid[0]) * t)
                g = int(rgb_mid[1] + (rgb_bottom[1] - rgb_mid[1]) * t)
                b = int(rgb_mid[2] + (rgb_bottom[2] - rgb_mid[2]) * t)
            draw.rectangle([0, y1 + i, card_w, y1 + i + 1], fill=(r, g, b), outline=None)

    y = 0
    top_h = 52
    # Header: card gradient 135deg approx as 3-stop vertical #8ddf00 -> #0e7001 -> #006b05
    _three_color_gradient(0, top_h, (140, 223, 0), (14, 112, 1), (0, 107, 5))
    # Gray diagonal lines (match card repeating-linear-gradient 45deg)
    for step in range(-top_h, card_w + top_h, 6):
        draw.line([(step, 0), (step + top_h + 10, top_h + 10)], fill=(25, 25, 25), width=1)
    # Logo circle (gold border)
    cx, cy = 30, top_h // 2
    draw.ellipse([cx - 18, cy - 18, cx + 18, cy + 18], outline=(212, 175, 55), fill=(80, 90, 50))
    draw.text((cx, cy), 'CCB', fill=(255, 255, 255), font=font_sub, anchor='mm')
    draw.text((card_w // 2 + 10, 14), site_name.upper(), fill=(255, 255, 255), font=font_title, anchor='mt')
    draw.text((card_w // 2 + 10, 34), 'SECURE CAMPUS ACCESS', fill=(230, 230, 230), font=font_sub, anchor='mt')
    y = top_h

    # Bar: yellow-orange gradient to match card #f5d236 -> #ffcd43 -> #fd9800
    bar_h = 32
    _gradient_band(y, y + bar_h, (245, 210, 54), (253, 152, 0))
    draw.text((card_w // 2, y + bar_h // 2), 'VISITOR PASS', fill=(97, 0, 0), font=font_bar, anchor='mm')
    y += bar_h

    # White body, QR with gold border (match card .qr-wrap 3px #ffd700, rounded)
    body_pad = 14
    qr_size = 110
    qr_border = 4
    qr_x = (card_w - qr_size - 2 * qr_border) // 2
    qr_y = y + body_pad
    draw.rectangle([qr_x, qr_y, qr_x + qr_size + 2 * qr_border, qr_y + qr_size + 2 * qr_border], outline=(255, 215, 0), fill=(255, 255, 255), width=qr_border)
    qr = qrcode.QRCode(version=1, box_size=3, border=1)
    qr.add_data(pass_obj.code)
    qr.make(fit=True)
    qr_img = qr.make_image(fill_color='#1a1a1a', back_color='white')
    qr_img = qr_img.resize((qr_size, qr_size))
    img.paste(qr_img, (qr_x + qr_border, qr_y + qr_border))
    y = qr_y + qr_size + 2 * qr_border + 10

    # Visitor code (green #27ae60)
    draw.text((card_w // 2, y), pass_obj.code, fill=(39, 174, 96), font=font_code, anchor='mt')
    y += 24
    draw.text((card_w // 2, y), 'SCAN AT GATE FOR ENTRY / EXIT', fill=(102, 102, 102), font=font_hint, anchor='mt')
    y += 20

    # Footer: gold top line 3px, dark green, shield + "AUTHORIZED VISITOR" | clock + "VALID TODAY ONLY"
    footer_h = 36
    fy = card_h - footer_h
    draw.rectangle([0, fy, card_w, fy + 3], fill=(255, 215, 0), outline=None)
    _gradient_band(fy + 3, card_h, (26, 77, 46), (22, 61, 36))
    # Vertical divider (white, center)
    div_x = card_w // 2
    draw.line([(div_x, fy + 6), (div_x, card_h - 6)], fill=(255, 255, 255), width=1)
    # Shield icon left side
    shield_cx = int(card_w * 0.28)
    shield_cy = fy + 3 + footer_h // 2
    draw.ellipse([shield_cx - 7, shield_cy - 9, shield_cx + 7, shield_cy + 9], outline=(255, 255, 255), fill=None, width=1)
    draw.line([(shield_cx - 4, shield_cy - 1), (shield_cx, shield_cy - 5), (shield_cx + 4, shield_cy - 1), (shield_cx, shield_cy + 3), (shield_cx - 4, shield_cy - 1)], fill=(255, 255, 255), width=1)
    # Clock icon right side
    clock_cx = int(card_w * 0.72)
    clock_cy = fy + 3 + footer_h // 2
    draw.ellipse([clock_cx - 7, clock_cy - 7, clock_cx + 7, clock_cy + 7], outline=(255, 255, 255), fill=None, width=1)
    draw.line([(clock_cx, clock_cy), (clock_cx, clock_cy - 5)], fill=(255, 255, 255), width=1)
    draw.line([(clock_cx, clock_cy), (clock_cx + 4, clock_cy + 3)], fill=(255, 255, 255), width=1)
    # Footer text (beside icons, not on top)
    draw.text((int(card_w * 0.42), shield_cy), 'AUTHORIZED VISITOR', fill=(255, 255, 255), font=font_footer, anchor='mm')
    draw.text((int(card_w * 0.85), clock_cy), 'VALID TODAY ONLY', fill=(255, 255, 255), font=font_footer, anchor='mm')

    buffer = io.BytesIO()
    img.save(buffer, format='PNG', dpi=(96, 96))
    buffer.seek(0)
    return buffer.getvalue()


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'student affairs')
def visitor_pass_print_all(request):
    """Printable page with all e-ID cards, or ZIP of card images when download=1."""
    passes = list(_visitor_pass_list_for_user(request))
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')

    if request.GET.get('download'):
        # ZIP of e-ID cards rendered from the same HTML layout shown in modal.
        image_format = (request.GET.get('format') or 'png').strip().lower()
        if image_format == 'jpg':
            image_format = 'jpeg'
        if image_format not in ('png', 'jpeg'):
            image_format = 'png'

        if not passes:
            return HttpResponse(
                'No visitor passes in this list to export. Generate slots or adjust filters, then try again.',
                status=400,
                content_type='text/plain; charset=utf-8',
            )

        base_url = (request.build_absolute_uri('/') or '/').rstrip('/') + '/'
        logo_url = None
        try:
            theme = SiteTheme.objects.first()
            if theme and theme.logo:
                logo_url = request.build_absolute_uri(theme.logo.url)
        except Exception:
            pass
        if not logo_url:
            try:
                logo_url = request.build_absolute_uri(static('gate/images/university-logo.png'))
            except Exception:
                pass

        # Same Playwright path as modal single-card download: one full browser session per card.
        zip_buffer = io.BytesIO()
        count = 0
        failed_cards = []
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            for p in passes:
                safe_code = "".join(c if c.isalnum() or c in '-_' else '_' for c in p.code)
                try:
                    qr_bytes = _visitor_pass_qr_png_bytes(p)
                    if qr_bytes:
                        qr_url = f'data:image/png;base64,{base64.b64encode(qr_bytes).decode("ascii")}'
                    else:
                        qr_url = request.build_absolute_uri(
                            reverse('visitor-qr-image', kwargs={'code': p.code})
                        )
                    html = render_to_string('gate/visitor_pass_card.html', {
                        'pass_obj': p,
                        'qr_url': qr_url,
                        'site_name': site_name,
                        'logo_url': logo_url,
                    }, request=request)
                    html = _make_card_html_absolute_urls(html, base_url)
                    shot = _screenshot_visitor_pass_card_html_bytes(html, image_format=image_format)
                    if image_format == 'png':
                        zf.writestr(f'visitor-eid-{safe_code}.png', shot)
                    else:
                        zf.writestr(f'visitor-eid-{safe_code}.jpg', shot)
                    count += 1
                except Exception as e:
                    logging.getLogger(__name__).warning(
                        'Visitor batch Playwright card export failed for %s: %s',
                        p.code, e
                    )
                    failed_cards.append(f'{p.code}: {str(e)}')
            if failed_cards:
                zf.writestr('failed-cards.txt', '\n'.join(failed_cards))

        if count <= 0:
            body = 'No visitor e-ID cards could be rendered.\n\n'
            if failed_cards:
                body += 'Details:\n' + '\n'.join(failed_cards[:20])
            else:
                body += 'Install browser automation if missing: pip install playwright && playwright install chromium'
            return HttpResponse(
                body,
                status=503,
                content_type='text/plain; charset=utf-8',
            )
        zip_buffer.seek(0)
        response = HttpResponse(zip_buffer.getvalue(), content_type='application/zip')
        suffix = 'png' if image_format == 'png' else 'jpg'
        response['Content-Disposition'] = f'attachment; filename="visitor-eids-{suffix}.zip"'
        response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response['Pragma'] = 'no-cache'
        return response

    # HTML print view
    logo_url = None
    try:
        theme = SiteTheme.objects.first()
        if theme and theme.logo:
            logo_url = request.build_absolute_uri(theme.logo.url)
    except Exception:
        pass
    if not logo_url:
        try:
            logo_url = request.build_absolute_uri(static('gate/images/university-logo.png'))
        except Exception:
            pass
    cards = []
    for p in passes:
        qr_url = request.build_absolute_uri(reverse('visitor-qr-image', kwargs={'code': p.code}))
        cards.append({'pass_obj': p, 'qr_url': qr_url, 'logo_url': logo_url})
    return render(request, 'gate/visitor_pass_print_all.html', {
        'site_name': site_name,
        'cards': cards,
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_list(request):
    """List students. Staff/faculty/student affairs get same registry actions as admin (import, export, e-ID, edit)."""
    from gate_analytics.roles import get_user_role
    role = get_user_role(request.user)
    can_edit_students = role in ('admin', 'staff', 'faculty', 'student affairs')
    is_student_affairs = role == 'student affairs'

    embed = (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes')
    filter_course = (request.GET.get('course') or '').strip()
    filter_year = (request.GET.get('year_level') or '').strip()
    filter_section = (request.GET.get('section') or '').strip()
    filter_sex = (request.GET.get('sex') or '').strip()
    search_q = (request.GET.get('q') or '').strip()
    highlight_raw = (request.GET.get('highlight') or '').strip()
    highlight_pk = None
    if highlight_raw.isdigit():
        highlight_pk = int(highlight_raw)

    # ---- Student Affairs: e-ID back-of-card signatories (modal on this page) ----
    eid_signatory_form = None
    eid_signatory_theme = None
    signatory_post_invalid = False
    eid_signatory_next_from_post = None
    if request.method == 'POST' and (request.POST.get('_eid_signatory_save') or '').strip() == '1':
        if not is_student_affairs:
            return HttpResponseForbidden(
                '<h1>403 Forbidden</h1><p>Only Student Affairs can update e-ID signatories.</p>'
            )
        eid_signatory_next_from_post = (request.POST.get('next') or '').strip()
        eid_signatory_theme = SiteTheme.objects.first()
        if eid_signatory_theme is None:
            eid_signatory_theme = SiteTheme.objects.create(
                site_name=getattr(settings, 'SITE_NAME', 'City College of Bayawan'),
            )
        eid_signatory_form = SiteThemeEidSignatoryForm(
            request.POST, request.FILES, instance=eid_signatory_theme,
        )
        if eid_signatory_form.is_valid():
            eid_signatory_form.save()
            cache.delete('site_theme_context_v2')
            messages.success(
                request,
                'E-ID signatory names and signatures saved. They appear on the back of printed student e-ID cards.',
            )
            next_path = (request.POST.get('next') or '').strip()
            if next_path.startswith('/') and not next_path.startswith('//'):
                return redirect(next_path)
            return redirect('gate-student-list')
        signatory_post_invalid = True

    students = Student.objects.all().order_by('last_name', 'first_name')

    if filter_course:
        students = students.filter(course=filter_course)
    if filter_year:
        students = students.filter(year_level=filter_year)
    if filter_section:
        students = students.filter(section__iexact=filter_section)
    if filter_sex:
        students = students.filter(sex=filter_sex)
    if search_q:
        students = students.filter(
            Q(first_name__icontains=search_q) |
            Q(last_name__icontains=search_q) |
            Q(middle_name__icontains=search_q) |
            Q(student_id__icontains=search_q)
        )

    # URL to clear only search (keeps filters)
    from urllib.parse import urlencode
    clear_params = {}
    if embed:
        clear_params['embed'] = '1'
    if filter_course:
        clear_params['course'] = filter_course
    if filter_year:
        clear_params['year_level'] = filter_year
    if filter_section:
        clear_params['section'] = filter_section
    if filter_sex:
        clear_params['sex'] = filter_sex
    from django.urls import reverse
    from django.core.paginator import Paginator
    clear_search_url = reverse('gate-student-list') + ('?' + urlencode(clear_params) if clear_params else '')

    if is_student_affairs:
        if eid_signatory_next_from_post is not None:
            n = eid_signatory_next_from_post
            eid_signatory_next_url = (
                n if (n.startswith('/') and not n.startswith('//')) else reverse('gate-student-list')
            )
        else:
            _next_pairs = [(k, v) for k, v in request.GET.items() if k not in ('eid_sig', 'partial')]
            eid_signatory_next_url = reverse('gate-student-list') + (
                '?' + urlencode(_next_pairs) if _next_pairs else ''
            )
    else:
        eid_signatory_next_url = ''

    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)

    def _strip_qs_param(qs, param):
        if not qs:
            return qs
        from urllib.parse import parse_qsl, urlencode
        return urlencode([(k, v) for k, v in parse_qsl(qs, keep_blank_values=True) if k != param])

    query_extra = _strip_qs_param(query_extra, 'partial')
    query_extra_base = _strip_qs_param(query_extra_base, 'partial')

    # Notification deep-link: ensure the student is visible and jump to their page.
    if highlight_pk:
        st_hi = Student.objects.filter(pk=highlight_pk).first()
        if st_hi:
            if not students.filter(pk=highlight_pk).exists():
                qparams = [
                    ('q', str(st_hi.student_id).strip()),
                    ('highlight', str(highlight_pk)),
                ]
                if embed:
                    qparams.insert(0, ('embed', '1'))
                return redirect(reverse('gate-student-list') + '?' + urlencode(qparams))
            pks = list(students.values_list('pk', flat=True))
            try:
                idx = pks.index(highlight_pk)
                need_page = idx // per_page + 1
            except ValueError:
                need_page = 1
            try:
                cur_page = int(request.GET.get('page') or 1)
            except (TypeError, ValueError):
                cur_page = 1
            if need_page != cur_page:
                qp = request.GET.copy()
                qp['page'] = str(need_page)
                return redirect(reverse('gate-student-list') + '?' + qp.urlencode())

    paginator = Paginator(students, per_page)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)

    if highlight_pk:
        AdminNotification.objects.filter(
            target_user=request.user,
            notification_type='student_registration',
            related_student_id=highlight_pk,
            is_read=False,
        ).update(is_read=True, read_at=timezone.now())

    # Student list sex filter: only show the two explicit options staff asked for.
    sex_choices = list(Student.SEX_CHOICES)

    if is_student_affairs and eid_signatory_form is None:
        eid_signatory_theme = SiteTheme.objects.first()
        if eid_signatory_theme is None:
            eid_signatory_theme = SiteTheme.objects.create(
                site_name=getattr(settings, 'SITE_NAME', 'City College of Bayawan'),
            )
        eid_signatory_form = SiteThemeEidSignatoryForm(instance=eid_signatory_theme)

    show_eid_signatory_modal = is_student_affairs and (
        signatory_post_invalid
        or (request.GET.get('eid_sig') or '').strip().lower() in ('1', 'true', 'yes', 'open')
    )

    list_ctx = {
        'site_name': 'City College of Bayawan',
        'can_edit_students': can_edit_students,
        'is_student_affairs': is_student_affairs,
        'eid_signatory_form': eid_signatory_form,
        'eid_signatory_theme': eid_signatory_theme,
        'show_eid_signatory_modal': show_eid_signatory_modal,
        'eid_signatory_next_url': eid_signatory_next_url,
        'embed': embed,
        'students': page_obj.object_list,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'filter_course': filter_course,
        'filter_year_level': filter_year,
        'filter_section': filter_section,
        'filter_sex': filter_sex,
        'search_q': search_q,
        'clear_search_url': clear_search_url,
        'course_choices': Student.COURSE_CHOICES,
        'year_level_choices': Student.YEAR_LEVEL_CHOICES,
        'sex_choices': sex_choices,
        'highlight_student_pk': highlight_pk,
        'student_qr_labels_pdf_url': reverse('gate-student-qr-labels-pdf'),
    }

    partial_table = (request.GET.get('partial') or '').strip().lower() in ('1', 'table', 'true')
    if partial_table:
        return render(request, 'gate/student_list_table_partial.html', list_ctx)

    return render(request, 'gate/student_list.html', list_ctx)


def _staff_personnel_registry_redirect_response(request):
    """After POST on staff registry, return to same list URL + query (hidden `next` field)."""
    next_path = (request.POST.get('next') or '').strip()
    if next_path.startswith('/') and not next_path.startswith('//'):
        return redirect(next_path)
    return redirect('pending-staff-personnel-list')


def _query_staff_registry_users(request):
    """Filter staff/faculty Users from GET (status, role, q, pending=1) — shared by list, export, print-all."""
    from django.contrib.auth import get_user_model

    User = get_user_model()
    if request.GET.get('pending') == '1':
        filter_status = 'pending'
    else:
        filter_status = (request.GET.get('status') or 'all').strip().lower()
    if filter_status not in ('all', 'active', 'pending'):
        filter_status = 'all'

    filter_role = (request.GET.get('role') or '').strip().lower()
    if filter_role not in ('', 'staff', 'faculty', 'student_affairs'):
        filter_role = ''

    search_q = (request.GET.get('q') or '').strip()

    id_qs = User.objects.filter(
        Q(groups__name__iexact='staff') |
        Q(groups__name__iexact='faculty') |
        Q(groups__name__iexact='Student Affairs')
    ).distinct().values_list('id', flat=True)
    qs = User.objects.filter(id__in=id_qs).select_related('staff_personnel_profile')

    if filter_status == 'pending':
        qs = qs.filter(is_active=False)
    elif filter_status == 'active':
        qs = qs.filter(is_active=True)

    if filter_role == 'staff':
        qs = qs.filter(groups__name__iexact='staff')
    elif filter_role == 'faculty':
        qs = qs.filter(groups__name__iexact='faculty')
    elif filter_role == 'student_affairs':
        qs = qs.filter(groups__name__iexact='Student Affairs')

    if search_q:
        qs = qs.filter(
            Q(username__icontains=search_q) |
            Q(first_name__icontains=search_q) |
            Q(last_name__icontains=search_q) |
            Q(email__icontains=search_q)
        ).distinct()

    return qs


def _user_is_staff_registry_account(u):
    """Staff, faculty, or Student Affairs (self-service registration targets)."""
    if not u:
        return False
    return (
        u.groups.filter(name__iexact='staff').exists()
        or u.groups.filter(name__iexact='faculty').exists()
        or u.groups.filter(name__iexact='Student Affairs').exists()
    )


@login_required(login_url='/login/')
@role_required('admin')
def pending_staff_personnel_list(request):
    """
    Staff/Faculty registry: filters, search, import/export, add.
    """
    from django.contrib.auth import get_user_model
    from django.contrib.auth.models import Group

    User = get_user_model()

    if request.GET.get('pending') == '1':
        filter_status = 'pending'
    else:
        filter_status = (request.GET.get('status') or 'all').strip().lower()
    if filter_status not in ('all', 'active', 'pending'):
        filter_status = 'all'

    filter_role = (request.GET.get('role') or '').strip().lower()
    if filter_role not in ('', 'staff', 'faculty', 'student_affairs'):
        filter_role = ''

    search_q = (request.GET.get('q') or '').strip()

    qs = _query_staff_registry_users(request).order_by('-date_joined')

    show_only_pending = filter_status == 'pending'

    clear_params = {}
    if filter_role:
        clear_params['role'] = filter_role
    if filter_status != 'all':
        if filter_status == 'pending':
            clear_params['pending'] = '1'
        else:
            clear_params['status'] = filter_status
    clear_search_url = reverse('pending-staff-personnel-list') + ('?' + urlencode(clear_params) if clear_params else '')

    # POST: approve, deactivate, or edit_user
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        user_id = request.POST.get('user_id')

        if action == 'edit_user' and user_id:
            try:
                u = User.objects.get(pk=user_id)
                if not _user_is_staff_registry_account(u):
                    messages.error(request, 'User is not a staff, faculty, or Student Affairs account.')
                    return _staff_personnel_registry_redirect_response(request)
                new_username = (request.POST.get('username') or '').strip()
                if new_username and User.objects.filter(username__iexact=new_username).exclude(pk=u.pk).exists():
                    messages.error(request, f'Username "{new_username}" is already taken.')
                    return _staff_personnel_registry_redirect_response(request)
                u.first_name = (request.POST.get('first_name') or '').strip()
                u.last_name = (request.POST.get('last_name') or '').strip()
                u.email = (request.POST.get('email') or '').strip()
                if new_username:
                    u.username = new_username
                u.save(update_fields=['first_name', 'last_name', 'email', 'username'])
                role = (request.POST.get('role') or '').strip().lower()
                if role in ('staff', 'faculty', 'student_affairs'):
                    role_groups = Group.objects.filter(
                        Q(name__iexact='staff')
                        | Q(name__iexact='faculty')
                        | Q(name__iexact='Student Affairs')
                    )
                    for g in role_groups:
                        u.groups.remove(g)
                    if role == 'student_affairs':
                        grp = Group.objects.filter(name__iexact='Student Affairs').first()
                    else:
                        grp = Group.objects.filter(name__iexact=role).first()
                    if grp:
                        u.groups.add(grp)
                if u.groups.filter(name__iexact='staff').exists() or u.groups.filter(name__iexact='faculty').exists():
                    profile, _ = StaffPersonnelProfile.objects.get_or_create(user=u, defaults={})
                    profile.middle_name = (request.POST.get('middle_name') or '')[:100]
                    profile.contact_number = (request.POST.get('contact_number') or '')[:20]
                    profile.department = (request.POST.get('department') or '')[:150]
                    profile.position = (request.POST.get('position') or '')[:150]
                    profile.address = (request.POST.get('address') or '')[:500]
                    profile.save(update_fields=['middle_name', 'contact_number', 'department', 'position', 'address'])
                messages.success(request, f'Account for {u.get_full_name() or u.username} has been updated.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')
            return _staff_personnel_registry_redirect_response(request)

        if action and user_id:
            try:
                u = User.objects.get(pk=user_id)
                if _user_is_staff_registry_account(u):
                    if action == 'approve':
                        u.is_active = True
                        u.save(update_fields=['is_active'])
                        messages.success(request, f'Account for {u.get_full_name() or u.username} has been approved. They can now log in.')
                    elif action == 'deactivate':
                        u.is_active = False
                        u.save(update_fields=['is_active'])
                        messages.success(request, f'Account for {u.get_full_name() or u.username} has been set to inactive.')
                    elif action == 'deny':
                        u.is_active = False
                        u.save(update_fields=['is_active'])
                        messages.warning(request, f'Access denied for {u.get_full_name() or u.username}. They cannot log in.')
                    elif action == 'delete':
                        if u.pk == request.user.pk:
                            messages.error(request, 'You cannot delete your own account.')
                        else:
                            name = u.get_full_name() or u.username
                            u.delete()
                            messages.success(request, f'Account "{name}" has been permanently deleted.')
            except User.DoesNotExist:
                messages.error(request, 'User not found.')
        return _staff_personnel_registry_redirect_response(request)

    def get_role(u):
        for g in u.groups.all():
            n = (g.name or '').lower()
            if n in ('staff', 'faculty', 'student affairs'):
                return g.name or n.title()
        return '—'

    def get_profile_dict(u):
        try:
            p = u.staff_personnel_profile
            return {
                'middle_name': p.middle_name or '',
                'contact_number': p.contact_number or '',
                'department': p.department or '',
                'position': p.position or '',
                'address': p.address or '',
            }
        except Exception:
            return {'middle_name': '', 'contact_number': '', 'department': '', 'position': '', 'address': ''}

    from django.core.paginator import Paginator
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    users_with_role = [(u, get_role(u), get_profile_dict(u)) for u in page_obj.object_list]

    # Pending count: same filter, DB-agnostic (MySQL and PostgreSQL)
    pending_count = User.objects.filter(
        is_active=False,
        id__in=User.objects.filter(
            Q(groups__name__iexact='staff') |
            Q(groups__name__iexact='faculty') |
            Q(groups__name__iexact='Student Affairs')
        ).distinct().values_list('id', flat=True)
    ).count()

    can_edit_staff_registry = get_user_role(request.user) in ('admin', 'staff', 'faculty')

    return render(request, 'gate/pending_staff_guard.html', {
        'site_name': 'City College of Bayawan',
        'users_with_role': users_with_role,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'show_only_pending': show_only_pending,
        'pending_count': pending_count,
        'filter_status': filter_status,
        'filter_role': filter_role,
        'search_q': search_q,
        'clear_search_url': clear_search_url,
        'can_edit_staff_registry': can_edit_staff_registry,
        'registry_current_path': request.get_full_path(),
    })


@require_GET
@login_required(login_url='/login/')
@role_required('admin')
def staff_personnel_eid_card(request, user_id):
    """Printable staff/faculty e-ID (same layout family as student e-ID)."""
    from django.contrib.auth import get_user_model

    User = get_user_model()
    user = get_object_or_404(User, pk=user_id)
    if not (user.groups.filter(name__iexact='staff').exists() or user.groups.filter(name__iexact='faculty').exists()):
        return HttpResponseForbidden('Not a staff/faculty account.')
    profile, _ = StaffPersonnelProfile.objects.get_or_create(user=user, defaults={})
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')
    logo_url = None
    try:
        theme = SiteTheme.objects.first()
        if theme and theme.logo:
            logo_url = request.build_absolute_uri(theme.logo.url)
    except Exception:
        pass
    if not logo_url:
        try:
            logo_url = request.build_absolute_uri(static('gate/images/university-logo.png'))
        except Exception:
            pass
    photo_url = None
    try:
        up = getattr(user, 'user_profile', None)
        if up and getattr(up, 'avatar', None) and up.avatar:
            url = up.avatar.url
            photo_url = url if url.startswith('http') else request.build_absolute_uri(url)
    except Exception:
        pass
    role_label = 'Faculty' if user.groups.filter(name__iexact='faculty').exists() else 'Staff'
    id_display = (profile.employee_id or '').strip() or user.username
    return render(request, 'gate/staff_eid_card.html', {
        'user': user,
        'profile': profile,
        'role_label': role_label,
        'qr_url': None,
        'site_name': site_name,
        'logo_url': logo_url,
        'photo_url': photo_url,
        'id_display': id_display,
    })


@require_GET
@login_required(login_url='/login/')
@role_required('admin')
def staff_personnel_eid_print_all(request):
    """Print all filtered staff/faculty e-ID cards (same GET filters as registry)."""
    from django.contrib.auth import get_user_model

    User = get_user_model()
    users = list(_query_staff_registry_users(request).order_by('last_name', 'first_name'))
    site_name = getattr(settings, 'SITE_NAME', 'City College of Bayawan')
    logo_url = None
    try:
        theme = SiteTheme.objects.first()
        if theme and theme.logo:
            logo_url = request.build_absolute_uri(theme.logo.url)
    except Exception:
        pass
    staff_cards = []
    for u in users:
        profile, _ = StaffPersonnelProfile.objects.get_or_create(user=u, defaults={})
        photo_url = None
        try:
            up = getattr(u, 'user_profile', None)
            if up and getattr(up, 'avatar', None) and up.avatar:
                url = up.avatar.url
                photo_url = url if url.startswith('http') else request.build_absolute_uri(url)
        except Exception:
            pass
        role_label = 'Faculty' if u.groups.filter(name__iexact='faculty').exists() else 'Staff'
        id_display = (profile.employee_id or '').strip() or u.username
        staff_cards.append({
            'user': u,
            'profile': profile,
            'role_label': role_label,
            'qr_url': None,
            'photo_url': photo_url,
            'id_display': id_display,
        })
    return render(request, 'gate/staff_eid_print_all.html', {
        'staff_cards': staff_cards,
        'total': len(staff_cards),
        'site_name': site_name,
        'logo_url': logo_url,
    })


@require_GET
@login_required(login_url='/login/')
@role_required('admin')
def staff_personnel_export_csv(request):
    """Export all staff/faculty users + profile fields to CSV (mirrors student export: full roster)."""
    from django.contrib.auth import get_user_model

    User = get_user_model()
    id_qs = User.objects.filter(
        Q(groups__name__iexact='staff') |
        Q(groups__name__iexact='faculty')
    ).distinct().values_list('id', flat=True)
    users = User.objects.filter(id__in=id_qs).order_by('last_name', 'first_name').select_related('staff_personnel_profile')
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="staff_faculty_export.csv"'
    response.write('\ufeff')
    writer = csv.writer(response)
    writer.writerow([
        'Username', 'Email', 'First name', 'Last name', 'Role', 'Employee ID',
        'Department', 'Position', 'Contact number', 'Middle Initial', 'Address',
        'Is active', 'Date joined',
    ])
    for u in users:
        profile, _ = StaffPersonnelProfile.objects.get_or_create(user=u, defaults={})
        role = ''
        for g in u.groups.all():
            n = (g.name or '').lower()
            if n in ('staff', 'faculty'):
                role = g.name or n.title()
                break
        writer.writerow([
            u.username,
            u.email or '',
            u.first_name,
            u.last_name,
            role,
            profile.employee_id or '',
            profile.department or '',
            profile.position or '',
            profile.contact_number or '',
            profile.middle_name or '',
            (profile.address or '').replace('\r\n', ' ').replace('\n', ' '),
            'Yes' if u.is_active else 'No',
            timezone.localtime(u.date_joined).strftime('%Y-%m-%d %H:%M:%S') if u.date_joined else '',
        ])
    return response


@login_required(login_url='/login/')
@role_required('admin')
def import_staff_personnel_csv(request):
    """Import staff/faculty from CSV: username, email, first_name, last_name, role, optional fields."""
    from django.contrib.auth import get_user_model
    import secrets

    if request.method != 'POST':
        return render(request, 'gate/import_staff_personnel_csv.html', {'site_name': 'City College of Bayawan'})

    csv_file = request.FILES.get('csv_file')
    if not csv_file or not csv_file.name.lower().endswith(('.csv', '.txt')):
        messages.warning(request, 'Please upload a CSV file.')
        return redirect('gate-staff-personnel-import-csv')

    try:
        content = csv_file.read().decode('utf-8-sig').strip()
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
    except Exception as e:
        messages.error(request, 'Could not read CSV: %s' % str(e))
        return redirect('gate-staff-personnel-import-csv')

    if not rows:
        messages.warning(request, 'CSV file is empty.')
        return redirect('gate-staff-personnel-import-csv')

    from django.contrib.auth.models import Group

    User = get_user_model()
    HEADER_ALIASES = {
        'username': 'username', 'user': 'username',
        'email': 'email',
        'first_name': 'first_name', 'first name': 'first_name',
        'last_name': 'last_name', 'last name': 'last_name',
        'middle_name': 'middle_name', 'middle name': 'middle_name', 'middle initial': 'middle_name',
        'role': 'role', 'type': 'role',
        'employee_id': 'employee_id', 'employee id': 'employee_id', 'id': 'employee_id',
        'department': 'department',
        'position': 'position',
        'contact_number': 'contact_number', 'contact number': 'contact_number', 'phone': 'contact_number',
        'password': 'password',
        'is_active': 'is_active', 'active': 'is_active',
    }

    first = rows[0]
    first_lower = [str(c or '').strip().lower() for c in first]
    has_header = 'username' in first_lower and 'email' in first_lower
    if has_header:
        col_map = {}
        for idx, cell in enumerate(first):
            key = str(cell or '').strip().lower()
            if key in HEADER_ALIASES:
                col_map[HEADER_ALIASES[key]] = idx
        data_rows = rows[1:]
    else:
        col_map = None
        data_rows = rows

    created = 0
    skipped = 0
    errors = []

    def _get(row, key, default=''):
        if col_map is None:
            return default
        idx = col_map.get(key)
        if idx is None or idx >= len(row):
            return default
        return (row[idx] or '').strip()

    for i, row in enumerate(data_rows):
        if not row:
            continue
        line_no = i + 2 if has_header else i + 1
        if col_map is not None:
            username = _get(row, 'username')
            email = _get(row, 'email')
            first_name = _get(row, 'first_name')
            last_name = _get(row, 'last_name')
            role_raw = (_get(row, 'role') or 'staff').lower()
            middle_name = _get(row, 'middle_name')
            employee_id = _get(row, 'employee_id')
            department = _get(row, 'department')
            position = _get(row, 'position')
            contact_number = _get(row, 'contact_number')
            password = _get(row, 'password')
            is_active_raw = (_get(row, 'is_active') or 'yes').lower()
        else:
            if len(row) < 4:
                continue
            username = (row[0] or '').strip()
            email = (row[1] or '').strip()
            first_name = (row[2] or '').strip()
            last_name = (row[3] or '').strip()
            role_raw = (row[4] or 'staff').strip().lower() if len(row) > 4 else 'staff'
            middle_name = ''
            employee_id = ''
            department = ''
            position = ''
            contact_number = ''
            password = ''
            is_active_raw = 'yes'

        if not username or not email or not first_name or not last_name:
            skipped += 1
            continue
        if User.objects.filter(username__iexact=username).exists():
            errors.append(f'Line {line_no}: username "{username}" already exists.')
            skipped += 1
            continue
        if role_raw not in ('staff', 'faculty'):
            errors.append(f'Line {line_no}: role must be staff or faculty.')
            skipped += 1
            continue
        is_active = is_active_raw in ('1', 'yes', 'true', 'active')

        pwd = password or secrets.token_urlsafe(14)
        try:
            with transaction.atomic():
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=pwd,
                    first_name=first_name,
                    last_name=last_name,
                    is_active=is_active,
                )
                grp = Group.objects.filter(name__iexact=role_raw).first()
                if grp:
                    user.groups.add(grp)
                profile, _ = StaffPersonnelProfile.objects.get_or_create(user=user, defaults={})
                profile.middle_name = (middle_name or '')[:100]
                profile.employee_id = (employee_id or '')[:50]
                profile.department = (department or '')[:150]
                profile.position = (position or '')[:150]
                profile.contact_number = (contact_number or '')[:20]
                profile.save()
                created += 1
        except Exception as ex:
            errors.append(f'Line {line_no}: {ex}')
            skipped += 1

    if created:
        messages.success(request, f'Imported {created} staff/faculty account(s).')
    if skipped and not created:
        messages.warning(request, 'No rows imported. Check errors below.')
    if errors:
        for err in errors[:15]:
            messages.warning(request, err)
        if len(errors) > 15:
            messages.warning(request, f'…and {len(errors) - 15} more errors.')
    return redirect('pending-staff-personnel-list')


@login_required(login_url='/login/')
@role_required('admin')
def staff_personnel_create(request):
    """Add staff/faculty user + profile (in-app, same idea as Add student)."""
    from django.contrib.auth import get_user_model
    from django.contrib.auth.models import Group

    User = get_user_model()
    form_modal = (request.GET.get('modal') == '1') or (request.POST.get('from_modal') == '1')
    form = StaffPersonnelCreateForm(request.POST or None)
    if form.is_valid():
        d = form.cleaned_data
        with transaction.atomic():
            user = User.objects.create_user(
                username=d['username'],
                email=d['email'],
                password=d['password'],
                first_name=d['first_name'],
                last_name=d['last_name'],
                is_active=d.get('is_active', True),
            )
            grp = Group.objects.filter(name__iexact=d['role']).first()
            if grp:
                user.groups.add(grp)
            profile, _ = StaffPersonnelProfile.objects.get_or_create(user=user, defaults={})
            profile.middle_name = (d.get('middle_name') or '')[:100]
            profile.department = (d.get('department') or '')[:150]
            profile.position = (d.get('position') or '')[:150]
            profile.contact_number = (d.get('contact_number') or '')[:20]
            profile.save()
        messages.success(request, f'Created account for {user.get_full_name() or user.username}.')
        if request.POST.get('from_modal') == '1':
            return redirect('gate-staff-personnel-create-modal-done')
        return redirect('pending-staff-personnel-list')
    return render(request, 'gate/staff_personnel_form.html', {
        'site_name': 'City College of Bayawan',
        'form': form,
        'title': 'Add staff / faculty',
        'form_modal': form_modal,
        'layout_template': 'gate/staff_personnel_form_modal_shell.html' if form_modal else 'base/base.html',
    })


@require_GET
@login_required(login_url='/login/')
@role_required('admin')
def staff_personnel_create_modal_done(request):
    """Minimal page loaded in iframe after create; tells parent window to close modal and refresh."""
    return render(request, 'gate/staff_personnel_create_modal_done.html', {
        'site_name': 'City College of Bayawan',
    })


@require_POST
@login_required(login_url='/login/')
@role_required('admin', 'student affairs')
def approve_all_pending_students(request):
    """Legacy URL: there are no pending student statuses; redirect to student list."""
    return redirect('gate-student-list')


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_list_export_csv(request):
    """Export student information to CSV (all fields). Honors GET filters: status, course, year_level, section, sex, q (same as e-ID print-all).
    If `status` is omitted from the query string (legacy direct link), all students are included; the export modal always sends `status`."""
    import time
    from django.db.models import Q
    started_at = time.monotonic()
    status_filter = (request.GET.get('status') or '').strip()
    if 'status' not in request.GET:
        students = Student.objects.all().order_by('last_name', 'first_name')
    elif status_filter == 'all':
        students = Student.objects.all().order_by('last_name', 'first_name')
    else:
        students = Student.objects.filter(account_status=Student.ACCOUNT_STATUS_APPROVED).order_by('last_name', 'first_name')
    filter_course = (request.GET.get('course') or '').strip()
    filter_year = (request.GET.get('year_level') or '').strip()
    filter_section = (request.GET.get('section') or '').strip()
    filter_sex = (request.GET.get('sex') or '').strip()
    search_q = (request.GET.get('q') or '').strip()
    if filter_course:
        students = students.filter(course=filter_course)
    if filter_year:
        students = students.filter(year_level=filter_year)
    if filter_section:
        students = students.filter(section__iexact=filter_section)
    if filter_sex:
        students = students.filter(sex=filter_sex)
    if search_q:
        students = students.filter(
            Q(first_name__icontains=search_q) |
            Q(last_name__icontains=search_q) |
            Q(middle_name__icontains=search_q) |
            Q(student_id__icontains=search_q)
        )
    # Build CSV fully first to make downloads reliable (some clients error on
    # chunked/streamed responses). Use UTF-8-SIG so Excel opens it correctly.
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Excel can convert long numeric IDs (student_id/contact numbers) into scientific
    # notation or truncate leading zeros. Prefix with TAB to force Excel to treat
    # them as text.
    def _excel_text(v):
        if v is None:
            return ''
        s = str(v)
        return '\t' + s

    writer.writerow([
        'Student ID', 'First name', 'Middle Initial', 'Last name', 'Email', 'Address',
        'Birthdate', 'Sex/Gender', 'Guardians / Parents', 'Guardian contact', 'Program', 'Year level', 'Section', 'Program or section (legacy)',
        'Your contact number', 'Account status', 'Is active', 'Approved by', 'Approved at',
        'Created at', 'Updated at',
    ])
    student_count = 0
    for s in students:
        student_count += 1
        writer.writerow([
            _excel_text(s.student_id),
            s.first_name,
            s.middle_name or '',
            s.last_name,
            s.email or '',
            (s.address or '').replace('\r\n', ' ').replace('\n', ' '),
            s.birthdate.isoformat() if s.birthdate else '',
            s.get_sex_display() if s.sex else '',
            s.guardians_parents or '',
            _excel_text(s.guardian_contact) if s.guardian_contact else '',
            s.course or '',
            s.get_year_level_display() if s.year_level else '',
            s.section or '',
            s.course_or_section or '',
            _excel_text(s.contact_number) if s.contact_number else '',
            s.account_status,
            'Yes' if s.is_active else 'No',
            s.approved_by.username if s.approved_by_id else '',
            timezone.localtime(s.approved_at).strftime('%Y-%m-%d %H:%M:%S') if s.approved_at else '',
            timezone.localtime(s.created_at).strftime('%Y-%m-%d %H:%M:%S') if s.created_at else '',
            timezone.localtime(s.updated_at).strftime('%Y-%m-%d %H:%M:%S') if s.updated_at else '',
        ])
    csv_bytes = buffer.getvalue().encode('utf-8-sig')
    elapsed = time.monotonic() - started_at
    response = HttpResponse(csv_bytes, content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="students_export.csv"'
    response['Content-Length'] = str(len(csv_bytes))
    response['Cache-Control'] = 'no-store'
    response['Pragma'] = 'no-cache'
    # Django/WGSI forbids hop-by-hop headers like 'Connection' and will throw 500.
    # Some proxies may try to add it; ensure it's removed.
    try:
        response.headers.pop('Connection', None)
    except Exception:
        pass
    logging.getLogger(__name__).info(
        "student_list_export_csv: students=%s bytes=%s elapsed=%.2fs",
        student_count,
        len(csv_bytes),
        elapsed,
    )
    # Also print to console so it's visible even if logging isn't configured.
    print(
        f"student_list_export_csv: students={student_count} bytes={len(csv_bytes)} elapsed={elapsed:.2f}s"
    )
    return response


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_create(request):
    """Create student. account_status drives gate access (APPROVED=active, INACTIVE=frozen)."""
    form_modal = (request.GET.get('modal') == '1') or (request.POST.get('from_modal') == '1')
    role = get_user_role(request.user)
    if form_modal:
        FormClass = StudentModalForm
    elif role == 'student affairs':
        FormClass = StudentStudentAffairsForm
    else:
        FormClass = StudentForm
    form = FormClass(request.POST or None, request.FILES or None)
    if form.is_valid():
        student = form.save()
        try:
            from .audit import log_action
            log_action(
                request,
                'student_created',
                'Student',
                object_id=student.pk,
                description=f'Student {student.student_id} created/updated by {get_user_role(request.user) or "user"}',
            )
        except Exception:
            pass
        # Sync is_active with status
        if student.account_status == Student.ACCOUNT_STATUS_APPROVED:
            student.is_active = True
            if not student.approved_at:
                student.approved_by = request.user
                student.approved_at = timezone.now()
            student.save(update_fields=['is_active', 'approved_by', 'approved_at'])
        else:
            student.is_active = False
            student.save(update_fields=['is_active'])
        # Always notify student about current status on creation (pending, approved, etc.).
        try:
            notify_student_status_change(student, new_status=student.account_status)
        except Exception:
            pass
        if student.account_status == Student.ACCOUNT_STATUS_APPROVED:
            try:
                from .audit import log_action
                log_action(request, 'student_created_approved', 'Student', object_id=student.pk, description=f'Student {student.student_id} created (approved)')
            except Exception:
                pass
        if request.POST.get('from_modal') == '1':
            return redirect('gate-student-create-modal-done')
        return redirect('gate-student-list')
    return render(request, 'gate/student_form.html', {
        'site_name': 'City College of Bayawan',
        'form': form,
        'title': 'Add Student',
        'form_modal': form_modal,
        'layout_template': 'gate/student_form_modal_shell.html' if form_modal else 'base/base.html',
    })


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_create_modal_done(request):
    """Minimal page loaded in iframe after create; tells parent window to close modal and refresh."""
    return render(request, 'gate/student_create_modal_done.html', {
        'site_name': 'City College of Bayawan',
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_edit(request, pk):
    """Edit student. account_status controls active/frozen access."""
    student = get_object_or_404(Student, pk=pk)
    role = get_user_role(request.user)
    FormClass = StudentStudentAffairsForm if role == 'student affairs' else StudentForm
    form = FormClass(request.POST or None, request.FILES or None, instance=student)
    if form.is_valid():
        old_photo_name = None
        if request.POST.get('photo-clear') and getattr(student, 'photo', None) and student.photo:
            old_photo_name = student.photo.name
        old_status = student.account_status
        form.save()
        try:
            from .audit import log_action
            log_action(
                request,
                'student_updated',
                'Student',
                object_id=student.pk,
                description=f'Student {student.student_id} profile fields updated',
            )
        except Exception:
            pass
        if old_photo_name:
            try:
                if default_storage.exists(old_photo_name):
                    default_storage.delete(old_photo_name)
            except Exception:
                pass
        new_status = form.cleaned_data.get('account_status')
        if new_status is None:
            new_status = old_status
        if new_status == Student.ACCOUNT_STATUS_APPROVED and old_status != Student.ACCOUNT_STATUS_APPROVED:
            student.refresh_from_db()
            if not student.approved_at:
                student.approved_by = request.user
                student.approved_at = timezone.now()
            student.is_active = True
            student.save(update_fields=['approved_by', 'approved_at', 'is_active'])
            from .audit import log_action
            log_action(request, 'student_approved', 'Student', object_id=student.pk, description=f'Student {student.student_id} approved')
        elif new_status == Student.ACCOUNT_STATUS_INACTIVE:
            student.is_active = False
            student.save(update_fields=['is_active'])
            from .audit import log_action
            log_action(request, 'student_rejected', 'Student', object_id=student.pk, description=f'Student {student.student_id} rejected/inactive')
        # If status actually changed, notify the student regardless of which value it changed to.
        if new_status != old_status:
            try:
                notify_student_status_change(student, new_status=new_status)
            except Exception:
                pass
        display_name = student.get_full_name() or student.student_id or str(student.pk)
        messages.success(
            request,
            f'Student "{display_name}" (ID {student.student_id}) has been updated successfully.',
        )
        return redirect('gate-student-list')
    return render(request, 'gate/student_form.html', {
        'site_name': 'City College of Bayawan',
        'form': form,
        'student': student,
        'title': 'Edit Student',
        'layout_template': 'base/base.html',
    })


@require_GET
@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def student_sample_csv(request):
    """Download sample students CSV: 50 students (20240001–20240050) with full details."""
    sample_path = os.path.join(os.path.dirname(__file__), 'sample_students_50.csv')
    if not os.path.isfile(sample_path):
        from django.http import HttpResponseNotFound
        return HttpResponseNotFound('Sample file not found.')
    with open(sample_path, 'rb') as f:
        content = f.read()
    response = HttpResponse(content, content_type='text/csv; charset=utf-8-sig')
    response['Content-Disposition'] = 'attachment; filename="sample_students_50.csv"'
    return response


@require_GET
@ensure_csrf_cookie
def entry_list(request):
    """List gate entries. Staff/Admin/Faculty can filter by date. Embed + guard_token for guard monitor iframe."""
    from django.contrib.auth.views import redirect_to_login
    from gate_analytics.roles import get_user_role
    if _guard_embed_query_token_ok(request):
        user_role = 'staff'
    else:
        if not request.user.is_authenticated:
            return redirect_to_login(next=request.get_full_path())
        user_role = get_user_role(request.user)
        if user_role not in ('admin', 'staff', 'faculty', 'student affairs'):
            return HttpResponseForbidden('Access denied')

    def _entry_list_tab_query(tab):
        p = request.GET.copy()
        p['tab'] = tab
        return p.urlencode()

    entries_tab = (request.GET.get('tab') or 'entries').strip().lower()
    if entries_tab not in ('entries', 'visitors', 'incidents'):
        entries_tab = 'entries'
    if request.GET.get('embed'):
        entries_tab = 'entries'

    q = (request.GET.get('q') or '').strip()
    f = (request.GET.get('filter') or '').strip().lower()

    from_date = (request.GET.get('from_date') or '').strip()
    to_date = (request.GET.get('to_date') or '').strip()
    from_time = (request.GET.get('from_time') or '').strip()
    if not from_date:
        from_date = timezone.localdate().isoformat()

    entries_qs = GateEntry.objects.select_related('student', 'event').order_by('-timestamp')
    if q:
        entries_qs = entries_qs.filter(
            Q(student__student_id__icontains=q) |
            Q(student__first_name__icontains=q) |
            Q(student__middle_name__icontains=q) |
            Q(student__last_name__icontains=q) |
            Q(notes__icontains=q) |
            Q(out_reason__icontains=q)
        )

    if f == 'granted':
        entries_qs = entries_qs.filter(granted=True)
    elif f == 'denied':
        entries_qs = entries_qs.filter(granted=False)
    elif f == 'in':
        entries_qs = entries_qs.filter(notes__iexact='IN')
    elif f == 'out':
        entries_qs = entries_qs.filter(notes__iexact='OUT')

    filter_date = None
    filter_to_date = None
    try:
        if from_date:
            filter_date = datetime.date.fromisoformat(from_date)
    except ValueError:
        filter_date = None
    try:
        if to_date:
            filter_to_date = datetime.date.fromisoformat(to_date)
    except ValueError:
        filter_to_date = None

    range_mode = bool(filter_date and filter_to_date and filter_to_date >= filter_date)
    range_start = range_end_exclusive = None

    # Use local-timezone day bounds so ranges match calendar days (fixes USE_TZ vs UTC)
    if filter_date:
        if range_mode:
            range_start = _local_day_bounds(filter_date)[0]
            range_end_exclusive = _local_day_bounds(filter_to_date)[1]
            entries_qs = entries_qs.filter(timestamp__gte=range_start, timestamp__lt=range_end_exclusive)
        else:
            day_start, day_end = _local_day_bounds(filter_date)
            range_start, range_end_exclusive = day_start, day_end
            entries_qs = entries_qs.filter(timestamp__gte=day_start, timestamp__lt=day_end)
            if from_time:
                tz = timezone.get_current_timezone()
                try:
                    start_t = datetime.time.fromisoformat(from_time)
                    start_dt = timezone.make_aware(datetime.datetime.combine(filter_date, start_t), tz)
                    entries_qs = entries_qs.filter(timestamp__gte=start_dt)
                except ValueError:
                    pass

    # Daily student gate only (matches analytics / yearly summary counts). Without this, pagination
    # mixes event + visitor rows so "View log" shows a handful of student rows vs hundreds counted.
    student_gate_only = (request.GET.get('student_gate_only') or '').strip().lower() in ('1', 'true', 'yes')
    if student_gate_only:
        entries_qs = entries_qs.filter(
            student_id__isnull=False,
            event__isnull=True,
            visitor_visit__isnull=True,
        )

    from django.core.paginator import Paginator
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    page_number = request.GET.get('page', 1)
    # When viewing daily student gate only, paginate merged visit rows (IN/OUT pairs) so
    # "Show N per page" matches the table. Fall back to entry-based pagination if very large.
    STUDENT_GATE_VISIT_PAGE_MAX_ENTRIES = 10000
    if student_gate_only and entries_qs.count() <= STUDENT_GATE_VISIT_PAGE_MAX_ENTRIES:
        entries_all = list(entries_qs.order_by('-timestamp'))
        _hydrate_gate_entry_students(entries_all)
        visits_all = _gate_entries_to_visits(entries_all)
        visits_all.sort(
            key=lambda v: (v[0].timestamp if v[0] else v[1].timestamp),
            reverse=True,
        )
        paginator_entries = Paginator(visits_all, per_page)
        page_obj_entries = paginator_entries.get_page(page_number)
        visits_page = list(page_obj_entries.object_list)
        visits_display = [(in_e, out_e, None, False) for in_e, out_e in visits_page]
        entries = []
        for in_e, out_e in visits_page:
            if in_e:
                entries.append(in_e)
            if out_e:
                entries.append(out_e)
        entries_visitor = []
        entries_event_only = []
        event_visits = []
    else:
        paginator_entries = Paginator(entries_qs, per_page)
        page_obj_entries = paginator_entries.get_page(page_number)
        entries = list(page_obj_entries.object_list)
        _hydrate_gate_entry_students(entries)
        entries_student_only = [e for e in entries if e.event_id is None and e.student_id is not None]
        entries_visitor = [e for e in entries if e.visitor_visit_id is not None]
        entries_event_only = [e for e in entries if e.event_id is not None]
        visits = _gate_entries_to_visits(entries_student_only)
        event_visits = _gate_entries_to_visits(entries_event_only)
        visits_display = [(in_e, out_e, None, False) for in_e, out_e in visits]

    # Visitors for same date or range (for tab inside Gate entries) — legacy VisitorEntry + reusable pass VisitorVisit
    visitors_qs = VisitorEntry.objects.select_related('recorded_by').order_by('-timestamp')
    if filter_date and range_start is not None:
        visitors_qs = visitors_qs.filter(timestamp__gte=range_start, timestamp__lt=range_end_exclusive)
    visitor_entries = list(visitors_qs[:200])
    # Reusable pass check-ins (VisitorVisit) — same shape as VisitorEntry for template
    visit_rows = []
    if filter_date and range_start is not None:
        visitor_visits_qs = VisitorVisit.objects.filter(
            checked_in_at__gte=range_start, checked_in_at__lt=range_end_exclusive
        ).select_related('checked_in_by').order_by('-checked_in_at')[:200]
        for vv in visitor_visits_qs:
            visit_rows.append(SimpleNamespace(
                timestamp=vv.checked_in_at,
                visitor_name=vv.full_name,
                purpose=vv.purpose or '',
                who_to_visit=vv.department or '',
                recorded_by=vv.checked_in_by,
                photo=vv.photo_in,
            ))
    visitors = sorted(visitor_entries + visit_rows, key=lambda x: x.timestamp, reverse=True)[:200]

    # Students currently inside — only meaningful for a single calendar day (not a multi-day range)
    if filter_date and not range_mode:
        currently_inside = _currently_inside_list(filter_date)
    else:
        currently_inside = None

    context = {
        'site_name': 'City College of Bayawan',
        'entries': entries,
        'visits': visits_display,
        'event_visits': event_visits,
        'entries_visitor': entries_visitor,
        'visitors': visitors,
        'currently_inside': currently_inside,
        'q': q,
        'filter': f,
        'from_date': from_date or timezone.localdate().isoformat(),
        'to_date': to_date,
        'from_time': from_time,
        'student_gate_only': student_gate_only,
        'user_role': user_role,
        'page_obj': page_obj_entries,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'guard_token': (request.GET.get('guard_token') or '').strip(),
        # Compact table CSS for guard live panel; filters stay visible (embed always).
        'compact_live_panel': (
            (request.GET.get('live_panel') or '').strip().lower() in ('1', 'true', 'yes')
        ),
        'entries_tab': entries_tab,
        'tab_q_entries': _entry_list_tab_query('entries'),
        'tab_q_visitors': _entry_list_tab_query('visitors'),
        'tab_q_incidents': _entry_list_tab_query('incidents'),
    }
    if entries_tab == 'incidents':
        inc_ctx, inc_early = _build_incident_list_context(request, embed_in_gate_entries=True)
        if inc_early is not None:
            return inc_early
        context['incident_page_obj'] = inc_ctx['page_obj']
        context['incident_query_extra'] = inc_ctx['query_extra']
        context['incident_query_extra_base'] = inc_ctx['query_extra_base']
        context['incident_per_page'] = inc_ctx['per_page']
        context['incident_per_page_options'] = inc_ctx['per_page_options']
        context['incidents'] = inc_ctx['incidents']
        context['can_sas_review'] = inc_ctx['can_sas_review']
        context['gate_incident_total_count'] = inc_ctx['gate_incident_total_count']
        context['sas_check_filter'] = inc_ctx['sas_check_filter']
        context['reason_filter'] = inc_ctx['reason_filter']
        context['reason_choices'] = inc_ctx['reason_choices']
        context['search_q'] = inc_ctx['search_q']
        context['filter_date'] = inc_ctx['filter_date']
        context['incident_url_chip_all'] = inc_ctx['incident_url_chip_all']
        context['incident_url_chip_to_check'] = inc_ctx['incident_url_chip_to_check']
        context['incident_url_chip_verified'] = inc_ctx['incident_url_chip_verified']
        context['highlight_incident_pk'] = inc_ctx['highlight_incident_pk']
        context['incident_embed_in_entries'] = True
        context['incident_list_clear_url'] = inc_ctx['incident_list_clear_url']
    else:
        context['incidents'] = []
    if request.GET.get('embed'):
        context['embed'] = True
        return render(request, 'gate/entry_list_embed.html', context)
    return render(request, 'gate/entry_list.html', context)


@require_GET
def event_attendees_embed(request):
    """Embeddable event attendees list for a single event (used beside event title on gate scan when tracking)."""
    from django.contrib.auth.views import redirect_to_login
    from gate_analytics.roles import get_user_role
    if not _guard_embed_get_token_ok(request):
        if not request.user.is_authenticated:
            return redirect_to_login(next=request.get_full_path())
        if get_user_role(request.user) not in ('admin', 'staff', 'faculty'):
            return HttpResponseForbidden('Access denied')
    event_id = (request.GET.get('event_id') or '').strip()
    if not event_id:
        return render(request, 'gate/event_attendees_embed.html', {
            'event': None,
            'event_visits': [],
            'from_date': timezone.localdate().isoformat(),
        })
    event = get_object_or_404(Event, id=int(event_id))
    from_date = (request.GET.get('from_date') or '').strip()
    if not from_date:
        from_date = timezone.localdate().isoformat()
    filter_date = None
    try:
        filter_date = datetime.date.fromisoformat(from_date)
    except ValueError:
        filter_date = timezone.localdate()
    day_start, day_end = _local_day_bounds(filter_date)
    if getattr(event, 'event_location', '') == 'field_trip':
        ft_att = (
            EventAttendance.objects.filter(event=event)
            .filter(
                Q(checked_in_at__gte=day_start, checked_in_at__lt=day_end)
                | Q(checked_out_at__gte=day_start, checked_out_at__lt=day_end)
            )
            .select_related('student')
            .distinct()
            .order_by('-checked_in_at', '-checked_out_at')[:200]
        )
        return render(request, 'gate/event_attendees_embed.html', {
            'event': event,
            'event_visits': [],
            'field_trip_attendances': list(ft_att),
            'from_date': from_date,
        })
    entries_qs = GateEntry.objects.filter(
        event_id=event.id,
        timestamp__gte=day_start, timestamp__lt=day_end,
    ).select_related('student').order_by('-timestamp')
    entries = list(entries_qs[:200])
    event_visits = _gate_entries_to_visits(entries)
    return render(request, 'gate/event_attendees_embed.html', {
        'event': event,
        'event_visits': event_visits,
        'field_trip_attendances': None,
        'from_date': from_date,
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def visitor_entry_list(request):
    """List visitor entries (manual log + reusable pass check-ins). Default to today. Supports from_date only (single day) or from_date+to_date (range), and search by name."""
    from_date = (request.GET.get('from_date') or '').strip()
    to_date = (request.GET.get('to_date') or '').strip()
    q = (request.GET.get('q') or '').strip()
    if not from_date:
        from_date = timezone.localdate().isoformat()

    qs = VisitorEntry.objects.select_related('recorded_by').order_by('-timestamp')
    if q:
        qs = qs.filter(visitor_name__icontains=q)
    filter_date = None
    filter_to_date = None
    try:
        if from_date:
            filter_date = datetime.date.fromisoformat(from_date)
        if to_date:
            filter_to_date = datetime.date.fromisoformat(to_date)
    except ValueError:
        filter_date = None
        filter_to_date = None

    if filter_date and filter_to_date and filter_to_date >= filter_date:
        # Date range: show all visitors from from_date through to_date (inclusive)
        range_start, _ = _local_day_bounds(filter_date)
        _, range_end = _local_day_bounds(filter_to_date)
        qs = qs.filter(timestamp__gte=range_start, timestamp__lt=range_end)
        visitor_entries = list(qs[:3000])
        visit_qs = VisitorVisit.objects.filter(
            checked_in_at__gte=range_start, checked_in_at__lt=range_end
        ).select_related('checked_in_by').order_by('-checked_in_at')
        if q:
            visit_qs = visit_qs.filter(full_name__icontains=q)
        visit_rows = []
        for vv in visit_qs[:3000]:
            visit_rows.append(SimpleNamespace(
                timestamp=vv.checked_in_at,
                visitor_name=vv.full_name,
                purpose=vv.purpose or '',
                who_to_visit=vv.department or '',
                recorded_by=vv.checked_in_by,
                photo=vv.photo_in,
            ))
    elif filter_date:
        # Single day
        day_start, day_end = _local_day_bounds(filter_date)
        qs = qs.filter(timestamp__gte=day_start, timestamp__lt=day_end)
        visitor_entries = list(qs[:3000])
        visit_qs = VisitorVisit.objects.filter(
            checked_in_at__gte=day_start, checked_in_at__lt=day_end
        ).select_related('checked_in_by').order_by('-checked_in_at')
        if q:
            visit_qs = visit_qs.filter(full_name__icontains=q)
        visit_rows = []
        for vv in visit_qs[:3000]:
            visit_rows.append(SimpleNamespace(
                timestamp=vv.checked_in_at,
                visitor_name=vv.full_name,
                purpose=vv.purpose or '',
                who_to_visit=vv.department or '',
                recorded_by=vv.checked_in_by,
                photo=vv.photo_in,
            ))
    else:
        visitor_entries = list(qs[:3000])
        visit_rows = []

    visitors_all = sorted(visitor_entries + visit_rows, key=lambda x: x.timestamp, reverse=True)
    from django.core.paginator import Paginator
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    paginator = Paginator(visitors_all, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    visitors = list(page_obj.object_list)

    context = {
        'site_name': 'City College of Bayawan',
        'visitors': visitors,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'from_date': from_date,
        'to_date': to_date or '',
        'filter_to_date': filter_to_date,
        'q': q,
    }
    if request.GET.get('embed'):
        context['embed'] = True
        return render(request, 'gate/visitor_list_embed.html', context)
    return render(request, 'gate/visitor_list.html', context)


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def visitor_yearly_summary(request):
    """Yearly visitor summary: visitors per month for a selected year. Uses local time bounds (no DB __year)."""
    embed = (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes')
    today = timezone.localdate()
    selected_year_raw = (request.GET.get('year') or '').strip()
    selected_year = int(selected_year_raw) if selected_year_raw.isdigit() else today.year

    year_start, year_end = _local_year_bounds(selected_year)

    # Use local datetime range (same approach as visitors_this_month) so MySQL/USE_TZ does not break counts
    entries_for_year = VisitorEntry.objects.filter(
        timestamp__gte=year_start, timestamp__lt=year_end,
    ).values_list('timestamp', flat=True)
    visits_for_year = VisitorVisit.objects.filter(
        checked_in_at__gte=year_start, checked_in_at__lt=year_end,
    ).values_list('checked_in_at', flat=True)

    # Bucket by month in Python (local time)
    entry_by_month = defaultdict(int)
    for ts in entries_for_year:
        if ts is not None:
            local_ts = timezone.localtime(ts)
            if 1 <= local_ts.month <= 12:
                entry_by_month[local_ts.month] += 1

    visit_by_month = defaultdict(int)
    for ts in visits_for_year:
        if ts is not None:
            local_ts = timezone.localtime(ts)
            if 1 <= local_ts.month <= 12:
                visit_by_month[local_ts.month] += 1

    # Build monthly list for template (include last_day_iso so "View log" shows whole month)
    monthly_stats = []
    total_year = 0
    for month_num in range(1, 13):
        count = entry_by_month.get(month_num, 0) + visit_by_month.get(month_num, 0)
        total_year += count
        first_day = datetime.date(selected_year, month_num, 1)
        _, last_day_num = calendar.monthrange(selected_year, month_num)
        last_day = datetime.date(selected_year, month_num, last_day_num)
        monthly_stats.append({
            'month': month_num,
            'month_name': calendar.month_name[month_num],
            'count': count,
            'first_day_iso': first_day.isoformat(),
            'last_day_iso': last_day.isoformat(),
        })

    # Available years from both models (exclude None from null timestamps)
    years_entries = set(
        VisitorEntry.objects.annotate(y=ExtractYear('timestamp')).values_list('y', flat=True).distinct()
    )
    years_visits = set(
        VisitorVisit.objects.annotate(y=ExtractYear('checked_in_at')).values_list('y', flat=True).distinct()
    )
    all_years = (years_entries | years_visits | {today.year, selected_year}) - {None}
    available_years = sorted(all_years, reverse=True)

    context = {
        'site_name': 'City College of Bayawan',
        'embed': embed,
        'selected_year': selected_year,
        'monthly_stats': monthly_stats,
        'total_year': total_year,
        'available_years': available_years,
    }
    return render(request, 'gate/visitor_yearly.html', context)


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def student_entries_yearly_summary(request):
    """Yearly student gate summary: granted daily-gate scans per month (same rules as analytics student entry counts)."""
    embed = (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes')
    today = timezone.localdate()
    selected_year_raw = (request.GET.get('year') or '').strip()
    selected_year = int(selected_year_raw) if selected_year_raw.isdigit() else today.year

    year_start, year_end = _local_year_bounds(selected_year)

    ts_list = GateEntry.objects.filter(
        timestamp__gte=year_start,
        timestamp__lt=year_end,
        granted=True,
        student_id__isnull=False,
        event__isnull=True,
        visitor_visit__isnull=True,
    ).values_list('timestamp', flat=True)

    entry_by_month = defaultdict(int)
    for ts in ts_list:
        if ts is not None:
            local_ts = timezone.localtime(ts)
            if 1 <= local_ts.month <= 12:
                entry_by_month[local_ts.month] += 1

    monthly_stats = []
    for month_num in range(1, 13):
        count = entry_by_month.get(month_num, 0)
        first_day = datetime.date(selected_year, month_num, 1)
        _, last_day_num = calendar.monthrange(selected_year, month_num)
        last_day = datetime.date(selected_year, month_num, last_day_num)
        monthly_stats.append({
            'month': month_num,
            'month_name': calendar.month_name[month_num],
            'count': count,
            'first_day_iso': first_day.isoformat(),
            'last_day_iso': last_day.isoformat(),
        })

    total_year = sum(s['count'] for s in monthly_stats)

    years_qs = GateEntry.objects.filter(
        granted=True,
        student_id__isnull=False,
        event__isnull=True,
        visitor_visit__isnull=True,
    ).annotate(y=ExtractYear('timestamp')).values_list('y', flat=True).distinct()
    data_years = {y for y in list(years_qs) if y}
    all_years = (data_years | {today.year, selected_year}) - {None}
    available_years = sorted(all_years, reverse=True)

    context = {
        'site_name': 'City College of Bayawan',
        'embed': embed,
        'selected_year': selected_year,
        'monthly_stats': monthly_stats,
        'total_year': total_year,
        'available_years': available_years,
    }
    return render(request, 'gate/student_entries_yearly.html', context)


def _build_incident_list_context(request, embed_in_gate_entries=False):
    """
    Shared context for gate incidents table (standalone /gate/incidents/ or Gate entries ?tab=incidents).
    """
    highlight_raw = (request.GET.get('highlight') or '').strip()
    highlight_pk = int(highlight_raw) if highlight_raw.isdigit() else None

    qs = GateIncident.objects.select_related('student', 'sas_checked_by').order_by('-timestamp')
    sas_check_filter = (request.GET.get('sas_check') or '').strip()
    if sas_check_filter in ('to_check', 'verified'):
        qs = qs.filter(sas_review_status=sas_check_filter)
    reason_filter = (request.GET.get('reason') or '').strip()
    if reason_filter and reason_filter in dict(GateIncident.REASON_CHOICES):
        qs = qs.filter(reason=reason_filter)
    search_q = (request.GET.get('q') or '').strip()
    if search_q:
        qs = qs.filter(
            Q(details__icontains=search_q)
            | Q(scanned_id__icontains=search_q)
            | Q(student__first_name__icontains=search_q)
            | Q(student__last_name__icontains=search_q)
            | Q(student__student_id__icontains=search_q)
        ).distinct()

    date_filter_raw = (request.GET.get('date') or '').strip()
    filter_date = None
    if date_filter_raw:
        try:
            filter_date = datetime.datetime.strptime(date_filter_raw, '%Y-%m-%d').date()
        except ValueError:
            filter_date = None
        if filter_date:
            tz = timezone.get_current_timezone()
            day_start = timezone.make_aware(
                datetime.datetime.combine(filter_date, datetime.time.min),
                tz,
            )
            day_end = day_start + datetime.timedelta(days=1)
            qs = qs.filter(timestamp__gte=day_start, timestamp__lt=day_end)

    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)

    list_name = 'gate-entry-list' if embed_in_gate_entries else 'gate-incident-list'

    if highlight_pk and not embed_in_gate_entries:
        inc_hi = GateIncident.objects.filter(pk=highlight_pk).first()
        if inc_hi:
            if not qs.filter(pk=highlight_pk).exists():
                return None, redirect(f'{reverse("gate-incident-list")}?highlight={highlight_pk}')
            pks = list(qs.values_list('pk', flat=True))
            try:
                idx = pks.index(highlight_pk)
                need_page = idx // per_page + 1
            except ValueError:
                need_page = 1
            try:
                cur_page = int(request.GET.get('page') or 1)
            except (TypeError, ValueError):
                cur_page = 1
            if need_page != cur_page:
                qp = request.GET.copy()
                qp['page'] = str(need_page)
                return None, redirect(reverse('gate-incident-list') + '?' + qp.urlencode())

    from django.core.paginator import Paginator
    paginator = Paginator(qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    if highlight_pk and request.user.is_authenticated:
        AdminNotification.objects.filter(
            target_user=request.user,
            notification_type='incident',
            related_incident_id=highlight_pk,
            is_read=False,
        ).update(is_read=True, read_at=timezone.now())
    incidents = list(page_obj.object_list)
    missing_student_ids = {
        (inc.scanned_id or '').strip()
        for inc in incidents
        if not inc.student and (inc.scanned_id or '').strip()
    }
    resolved_students = {}
    if missing_student_ids:
        sid_q = Q()
        for sid in missing_student_ids:
            sid_q |= Q(student_id__iexact=sid)
        resolved_students = {
            s.student_id.upper(): s
            for s in Student.objects.filter(sid_q)
        }
    for inc in incidents:
        inc.display_student = inc.student
        if not inc.display_student:
            sid = (inc.scanned_id or '').strip().upper()
            if sid:
                inc.display_student = resolved_students.get(sid)
    role = get_user_role(request.user)
    # SAS-only: mark checked, clear all, and related incident actions (not admins/staff on this workflow).
    can_sas_review = role == 'student affairs'

    def _incident_chip_url(sas_val):
        qp = request.GET.copy()
        qp.pop('page', None)
        if embed_in_gate_entries:
            qp['tab'] = 'incidents'
        if filter_date is None and date_filter_raw:
            qp.pop('date', None)
        if sas_val is None:
            qp.pop('sas_check', None)
        else:
            qp['sas_check'] = sas_val
        tail = qp.urlencode()
        return reverse(list_name) + ('?' + tail if tail else '')

    incident_list_clear_url = (
        reverse('gate-entry-list') + '?tab=incidents'
        if embed_in_gate_entries else reverse('gate-incident-list')
    )

    ctx = {
        'site_name': 'City College of Bayawan',
        'page_obj': page_obj,
        'incidents': incidents,
        'can_sas_review': can_sas_review,
        'gate_incident_total_count': GateIncident.objects.count(),
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'sas_check_filter': sas_check_filter,
        'reason_filter': reason_filter,
        'reason_choices': GateIncident.REASON_CHOICES,
        'search_q': search_q,
        'filter_date': filter_date,
        'incident_url_chip_all': _incident_chip_url(None),
        'incident_url_chip_to_check': _incident_chip_url('to_check'),
        'incident_url_chip_verified': _incident_chip_url('verified'),
        'highlight_incident_pk': highlight_pk,
        'incident_embed_in_entries': embed_in_gate_entries,
        'incident_list_clear_url': incident_list_clear_url,
    }
    return ctx, None


@ensure_csrf_cookie
@login_required(login_url='/login/')
@role_required('admin', 'student affairs', 'staff', 'faculty')
def incident_list(request):
    """List gate incidents (ID mismatch, proxy, etc.) — SAS follow-up, admin, and gate staff."""
    ctx, early = _build_incident_list_context(request, embed_in_gate_entries=False)
    if early is not None:
        return early
    return render(request, 'gate/incident_list.html', ctx)


def _gate_incident_resolved_student(incident):
    """Student linked on the incident, or matched by scanned_id (same logic as incident list display)."""
    if incident.student_id:
        return incident.student
    sid = (incident.scanned_id or '').strip()
    if not sid:
        return None
    return Student.objects.filter(student_id__iexact=sid).first()


def _pending_sibling_incidents(incident):
    """
    Other incidents still awaiting SAS check for the same person (student FK or matching scanned_id).
    Excludes the given incident; caller should only use while those rows are still to_check.
    """
    qs = GateIncident.objects.filter(sas_review_status='to_check').exclude(pk=incident.pk)
    st = _gate_incident_resolved_student(incident)
    if st:
        sid = (st.student_id or '').strip()
        if sid:
            return qs.filter(Q(student=st) | Q(scanned_id__iexact=sid))
        return qs.filter(student=st)
    sid = (incident.scanned_id or '').strip()
    if not sid:
        return qs.none()
    return qs.filter(Q(scanned_id__iexact=sid) | Q(student__student_id__iexact=sid))


@require_POST
@login_required(login_url='/login/')
@transaction.atomic
def incident_sas_verify(request, incident_id):
    """Student Affairs (SAS) marks an incident as verified/checked. Admins can view the list but cannot confirm."""
    if get_user_role(request.user) != 'student affairs':
        messages.error(
            request,
            'Only Student Affairs (SAS) can mark incidents as checked. Students are verified in person at the office.',
        )
        return redirect(request.META.get('HTTP_REFERER') or 'gate-incident-list')
    incident = get_object_or_404(GateIncident, pk=incident_id)
    if incident.sas_review_status == 'verified':
        messages.info(request, 'Incident already verified.')
    else:
        now = timezone.now()
        if not incident.sas_check_notes:
            incident.sas_check_notes = 'Verified by Student Affairs.'
        incident.sas_review_status = 'verified'
        incident.sas_checked_by = request.user
        incident.sas_checked_at = now
        incident.save(
            update_fields=[
                'sas_review_status',
                'sas_checked_by',
                'sas_checked_at',
                'sas_check_notes',
            ]
        )
        note = (incident.sas_check_notes or '')[:255]
        n_extra = _pending_sibling_incidents(incident).update(
            sas_review_status='verified',
            sas_checked_by=request.user,
            sas_checked_at=now,
            sas_check_notes=note,
        )
        st = _gate_incident_resolved_student(incident)
        if st:
            _clear_student_office_hold_if_no_pending(st)
        # When SAS marks checked: notify admins (activate if inactive; confirm resolved if already active)
        if get_user_role(request.user) == 'student affairs':
            if st:
                from .admin_notification_service import AdminNotificationService

                AdminNotificationService.notify_admins_sas_verified_incident(
                    incident, st, request.user
                )
        if n_extra:
            messages.success(
                request,
                f'Incident marked as verified. {n_extra} related incident(s) for the same student '
                f'were also marked checked.',
            )
        else:
            messages.success(request, 'Incident marked as verified.')
    return redirect(request.META.get('HTTP_REFERER') or 'gate-incident-list')


@require_POST
@login_required(login_url='/login/')
@role_required('student affairs')
def incident_clear_all(request):
    """Delete every gate incident row and clear office holds (Student Affairs only)."""
    qs = GateIncident.objects.exclude(student_id__isnull=True)
    student_ids = list(qs.values_list('student_id', flat=True).distinct())
    n = GateIncident.objects.count()
    if n == 0:
        messages.info(request, 'There are no gate incidents to remove.')
        return redirect('gate-incident-list')
    with transaction.atomic():
        GateIncident.objects.all().delete()
        if student_ids:
            Student.objects.filter(pk__in=student_ids, office_clearance_hold=True).update(
                office_clearance_hold=False,
                office_clearance_note='',
            )
    messages.success(
        request,
        f'Removed {n} gate incident record(s). Student accounts were not deleted.',
    )
    return redirect('gate-incident-list')


@login_required(login_url='/login/')
def reports_incidents_overrides_redirect(request):
    """Redirect to reports hub (incidents feature removed)."""
    return redirect('reports-hub')


def _allocate_next_import_student_id():
    """Next unique numeric student_id (8 digits) for CSV import when the file has no ID column."""
    max_n = 0
    for sid in Student.objects.values_list('student_id', flat=True):
        if not sid:
            continue
        s = str(sid).strip()
        if s.isdigit() and len(s) <= 8:
            max_n = max(max_n, int(s))
    nxt = max_n + 1
    if max_n == 0:
        nxt = 10000001
    while nxt <= 99999999:
        cand = str(nxt).zfill(8)
        if not Student.objects.filter(student_id=cand).exists():
            return cand
        nxt += 1
    return 'IMP' + str(int(timezone.now().timestamp()))[-12:]


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def import_students_csv(request):
    """Import students from CSV (from registrar).
    student_id is optional: omit the column or leave cells blank — IDs are assigned automatically (8-digit).
    Full header row must include sex/gender (column sex, gender, or sex/gender) with MALE or FEMALE per row.
    Other columns may include: first_name, last_name, email, address, birthdate,
    contact_number, guardians_parents, guardian_contact, course, year_level, section, and optionally student_id.
    Legacy rows without a header: optional sex as last column (after email): student_id, first, last, email, sex
    or first, last, email, sex — sex is required for every imported row.
    """
    if request.method != 'POST':
        return render(request, 'gate/import_students_csv.html', {
            'site_name': 'City College of Bayawan',
        })
    sas_import = get_user_role(request.user) == 'student affairs'
    csv_file = request.FILES.get('csv_file')
    if not csv_file or not csv_file.name.lower().endswith(('.csv', '.txt')):
        messages.warning(request, 'Please upload a CSV file.')
        return redirect('gate-student-import')
    try:
        content = csv_file.read().decode('utf-8-sig').strip()
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
    except Exception as e:
        messages.error(request, 'Could not read CSV: %s' % str(e))
        return redirect('gate-student-import')
    if not rows:
        messages.warning(request, 'CSV file is empty.')
        return redirect('gate-student-import')

    def _normalize_student_id(val):
        if val is None:
            return ''
        s = str(val).strip()
        try:
            f = float(s)
            if f == int(f):
                return str(int(f))
        except (TypeError, ValueError):
            pass
        return s

    def _parse_birthdate(s):
        if not s or not str(s).strip():
            return None
        s = str(s).strip()
        for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%d/%m/%Y', '%m-%d-%Y'):
            try:
                return datetime.datetime.strptime(s, fmt).date()
            except ValueError:
                continue
        return None

    def _normalize_year_level(val):
        if not val:
            return ''
        s = str(val).strip().lower()
        if s in ('1', '1st', '1st year', 'first'):
            return '1'
        if s in ('2', '2nd', '2nd year', 'second'):
            return '2'
        if s in ('3', '3rd', '3rd year', 'third'):
            return '3'
        if s in ('4', '4th', '4th year', 'fourth'):
            return '4'
        return val.strip()

    def _normalize_sex(val):
        if not val:
            return ''
        s = str(val).strip().lower()
        if s in ('m', 'male'):
            return Student.SEX_MALE
        if s in ('f', 'female'):
            return Student.SEX_FEMALE
        return ''

    # Column name -> index (case-insensitive); used when first row is a header
    HEADER_ALIASES = {
        'student_id': 'student_id', 'student id': 'student_id', 'id': 'student_id',
        'first_name': 'first_name', 'first name': 'first_name',
        'middle_name': 'middle_name', 'middle name': 'middle_name', 'middle initial': 'middle_name',
        'last_name': 'last_name', 'last name': 'last_name',
        'email': 'email',
        'address': 'address',
        'birthdate': 'birthdate', 'birth date': 'birthdate',
        'sex': 'sex', 'gender': 'sex', 'sex/gender': 'sex', 'sex_gender': 'sex',
        'contact_number': 'contact_number', 'contact number': 'contact_number',
        'guardians_parents': 'guardians_parents', 'guardians parents': 'guardians_parents', 'guardian': 'guardians_parents',
        'guardian_contact': 'guardian_contact', 'guardian contact': 'guardian_contact',
        'course': 'course',
        'program': 'course',
        'year_level': 'year_level', 'year level': 'year_level',
        'section': 'section',
    }
    # Normalize "contact number" key
    HEADER_ALIASES['contact number'] = 'contact_number'

    data_rows = []
    col_map = None  # None = legacy (no header mapping)

    first = rows[0]
    first_lower = [str(c or '').strip().lower() for c in first]
    has_sid = any(c in ('student_id', 'student id', 'id') for c in first_lower)
    has_first = any(c in ('first_name', 'first name') for c in first_lower)
    has_last = any(c in ('last_name', 'last name') for c in first_lower)
    if (has_first and has_last) or (has_sid and has_first):
        # Build column map from header
        col_map = {}
        for idx, cell in enumerate(first):
            key = str(cell or '').strip().lower()
            if key in HEADER_ALIASES:
                col_map[HEADER_ALIASES[key]] = idx
            else:
                col_map[key] = idx
        data_rows = rows[1:]
    else:
        # Legacy: skip single header line if it looks like "student_id" or "id"
        if rows and len(rows) > 1 and (first[0] or '').strip().lower() in ('student_id', 'student id', 'id'):
            data_rows = rows[1:]
        else:
            data_rows = rows

    if col_map is not None and 'sex' not in col_map:
        messages.error(
            request,
            'CSV must include a sex/gender column (header: sex, gender, or sex/gender).',
        )
        return redirect('gate-student-import')

    created = 0
    skipped = 0
    errors = []
    for i, row in enumerate(data_rows):
        if not row:
            continue
        if col_map is not None:
            def _get(key, default=''):
                idx = col_map.get(key)
                if idx is None:
                    return default
                if idx >= len(row):
                    return default
                return (row[idx] or '').strip()
            student_id = _normalize_student_id(_get('student_id'))
            first_name = _get('first_name')
            middle_name = _get('middle_name')
            last_name = _get('last_name')
            email = _get('email')
            address = _get('address')
            birthdate = _parse_birthdate(_get('birthdate'))
            sex = _normalize_sex(_get('sex'))
            contact_number = _get('contact_number')
            guardians_parents = _get('guardians_parents')
            guardian_contact = _get('guardian_contact')
            course = _get('course')
            year_level = _normalize_year_level(_get('year_level'))
            section = _get('section')
        else:
            fc = (row[0] or '').strip()
            fc_digits = fc.isdigit() and len(fc) <= 8
            if fc_digits and len(row) >= 3:
                student_id = _normalize_student_id(fc)
                first_name = (row[1] or '').strip()
                last_name = (row[2] or '').strip()
                email = (row[3] or '').strip() if len(row) > 3 else ''
                sex = _normalize_sex(row[4]) if len(row) > 4 else ''
            elif len(row) >= 2:
                student_id = ''
                first_name = fc
                last_name = (row[1] or '').strip()
                email = (row[2] or '').strip() if len(row) > 2 else ''
                sex = _normalize_sex(row[3]) if len(row) > 3 else ''
            else:
                continue
            middle_name = ''
            address = ''
            birthdate = None
            contact_number = ''
            guardians_parents = ''
            guardian_contact = ''
            course = ''
            year_level = ''
            section = ''

        if not first_name or not last_name:
            continue
        if not sex:
            skipped += 1
            _file_row = i + 2 if col_map is not None else i + 1
            errors.append(
                'Row %s: sex/gender is required (MALE/FEMALE or M/F).' % _file_row
            )
            continue
        if not student_id:
            student_id = _allocate_next_import_student_id()
        elif Student.objects.filter(student_id=student_id).exists():
            skipped += 1
            continue
        try:
            Student.objects.create(
                student_id=student_id,
                first_name=first_name,
                middle_name=middle_name or '',
                last_name=last_name,
                email=email or '',
                address=address or '',
                birthdate=birthdate,
                sex=sex or '',
                contact_number=contact_number or '',
                guardians_parents=guardians_parents or '',
                guardian_contact=guardian_contact or '',
                course=course or '',
                year_level=year_level or '',
                section=section or '',
                is_active=True,
                account_status=Student.ACCOUNT_STATUS_APPROVED,
            )
            created += 1
        except Exception as e:
            _file_row = i + 2 if col_map is not None else i + 1
            errors.append('Row %s: %s' % (_file_row, str(e)))
    if created:
        from .audit import log_action
        log_action(request, 'bulk_import', 'Student', object_id='', description=f'CSV import: {created} student(s) created')
        messages.success(request, 'Imported %s student(s). QR codes are available in the student list.' % created)
    if skipped:
        messages.info(
            request,
            'Skipped %s row(s) (duplicate student_id, or missing sex/gender).' % skipped,
        )
    if errors:
        for err in errors[:5]:
            messages.error(request, err)
        if len(errors) > 5:
            messages.error(request, '… and %s more errors.' % (len(errors) - 5))
    return redirect('gate-student-list')



@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def gate_today_report(request):
    """Today's report at the gate: today's entries and visitors. Staff at gate use this. HTML or CSV export."""
    today = timezone.localdate()
    day_start, day_end = _local_day_bounds(today)

    entries_qs = GateEntry.objects.filter(
        timestamp__gte=day_start, timestamp__lt=day_end,
        event_id=None, student_id__isnull=False
    ).select_related('student', 'recorded_by').order_by('-timestamp')
    entries = list(entries_qs[:500])

    incidents = []

    visitor_entries = list(VisitorEntry.objects.filter(
        timestamp__gte=day_start, timestamp__lt=day_end
    ).select_related('recorded_by').order_by('-timestamp')[:500])
    visitor_visits_qs = VisitorVisit.objects.filter(
        checked_in_at__gte=day_start, checked_in_at__lt=day_end
    ).select_related('checked_in_by').order_by('-checked_in_at')[:500]
    visit_rows = []
    for vv in visitor_visits_qs:
        visit_rows.append(SimpleNamespace(
            timestamp=vv.checked_in_at,
            visitor_name=vv.full_name,
            purpose=vv.purpose or '',
            who_to_visit=vv.department or '',
            recorded_by=vv.checked_in_by,
        ))
    visitors = sorted(visitor_entries + visit_rows, key=lambda x: x.timestamp, reverse=True)[:500]

    if request.GET.get('format') == 'csv':
        buf = io.StringIO()
        w = csv.writer(buf)
        w.writerow(['Today\'s report', today.isoformat()])
        w.writerow([])
        w.writerow(['Student entries', 'Time', 'Student ID', 'Name', 'Result', 'Recorded by'])
        for e in entries:
            name = e.student.get_full_name() if e.student else '—'
            rec = (e.recorded_by.get_full_name() or e.recorded_by.username) if e.recorded_by else '—'
            w.writerow(['', timezone.localtime(e.timestamp).strftime('%Y-%m-%d %I:%M %p'), e.student.student_id if e.student else '—', name, 'Granted' if e.granted else 'Denied', rec])
        w.writerow([])
        w.writerow(['Visitors', 'Time', 'Name', 'Purpose', 'Who to visit', 'Recorded by'])
        for v in visitors:
            rec = (getattr(v.recorded_by, 'get_full_name', lambda: '')() or getattr(v.recorded_by, 'username', '—')) if getattr(v, 'recorded_by', None) else '—'
            w.writerow(['', timezone.localtime(v.timestamp).strftime('%Y-%m-%d %I:%M %p'), getattr(v, 'visitor_name', getattr(v, 'full_name', '—')), getattr(v, 'purpose', '—'), getattr(v, 'who_to_visit', getattr(v, 'department', '—')), rec])
        response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
        response['Content-Disposition'] = f'attachment; filename="gate_report_{today.isoformat()}.csv"'
        return response

    return render(request, 'gate/gate_today_report.html', {
        'site_name': 'City College of Bayawan',
        'today': today,
        'entries': entries,
        'incidents': incidents,
        'visitors': visitors,
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_manage_registrations(request, event_id):
    """Manage event registrations and generate QR tokens (admin/staff/faculty)."""
    event = get_object_or_404(Event, id=event_id)
    
    if request.method == 'POST':
        action = request.POST.get('action', '').strip()
        
        if action == 'register_all':
            # Register all active students
            students = Student.objects.filter(is_active=True)
            created_count = 0
            for student in students:
                reg, created = EventRegistration.objects.get_or_create(
                    event=event,
                    student=student
                )
                if created or not reg.token:
                    reg.token = EventRegistration.generate_token()
                    reg.save()
                    created_count += 1
            messages.success(request, f'Registered {created_count} students for {event.name}.')
            return redirect('event-manage-registrations', event_id=event.id)
        
        elif action == 'register_csv':
            # Import from CSV (student_id column)
            csv_file = request.FILES.get('csv_file')
            if not csv_file:
                messages.error(request, 'Please upload a CSV file.')
                return redirect('event-manage-registrations', event_id=event.id)
            
            try:
                decoded_file = csv_file.read().decode('utf-8')
                reader = csv.DictReader(io.StringIO(decoded_file))
                created_count = 0
                for row in reader:
                    student_id = row.get('student_id', '').strip()
                    if not student_id:
                        continue
                    try:
                        student = Student.objects.get(student_id=student_id, is_active=True)
                        reg, created = EventRegistration.objects.get_or_create(
                            event=event,
                            student=student
                        )
                        if created or not reg.token:
                            reg.token = EventRegistration.generate_token()
                            reg.save()
                            created_count += 1
                    except Student.DoesNotExist:
                        pass
                from .audit import log_action
                log_action(request, 'bulk_import', 'EventRegistration', object_id=event.id, description=f'Event "{event.name}": {created_count} registration(s) from CSV')
                messages.success(request, f'Imported {created_count} registrations from CSV.')
            except Exception as e:
                messages.error(request, f'Error importing CSV: {str(e)}')
            return redirect('event-manage-registrations', event_id=event.id)
    
    # Get all registrations (for stats use full counts)
    total_registered = EventRegistration.objects.filter(event=event).count()
    checked_in = EventRegistration.objects.filter(event=event, checked_in_at__isnull=False).count()
    checked_out = EventRegistration.objects.filter(event=event, checked_out_at__isnull=False).count()

    registrations_qs = EventRegistration.objects.filter(event=event).select_related('student').order_by('student__last_name', 'student__first_name')
    search_q = (request.GET.get('q') or '').strip()
    if search_q:
        registrations_qs = registrations_qs.filter(
            Q(student__student_id__icontains=search_q) |
            Q(student__first_name__icontains=search_q) |
            Q(student__last_name__icontains=search_q)
        )
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    from django.core.paginator import Paginator
    paginator = Paginator(registrations_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    registrations = list(page_obj.object_list)

    from django.utils import timezone
    today = timezone.localdate()
    event_ended = event.end_date < today

    context = {
        'site_name': 'City College of Bayawan',
        'event': event,
        'today': today,
        'event_ended': event_ended,
        'registrations': registrations,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'total_registered': total_registered,
        'checked_in': checked_in,
        'checked_out': checked_out,
        'search_q': search_q,
    }
    if (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes'):
        return render(request, 'gate/event_registrations_embed.html', context)
    return render(request, 'gate/event_registrations.html', context)


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_registrations_export_csv(request, event_id):
    """Export event registrations as CSV."""
    event = get_object_or_404(Event, id=event_id)
    regs = EventRegistration.objects.filter(event=event).select_related('student').order_by('student__last_name', 'student__first_name')
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(['student_id', 'name', 'token', 'status', 'issued_at', 'checked_in_at', 'checked_out_at'])
    for r in regs:
        name = r.student.get_full_name() if r.student else ''
        student_id = r.student.student_id if r.student else ''
        writer.writerow([
            student_id,
            name,
            r.token or '',
            r.status or '',
            r.issued_at.strftime('%Y-%m-%d %H:%M') if r.issued_at else '',
            r.checked_in_at.strftime('%Y-%m-%d %H:%M') if r.checked_in_at else '',
            r.checked_out_at.strftime('%Y-%m-%d %H:%M') if r.checked_out_at else '',
        ])
    buf.seek(0)
    response = HttpResponse(buf.getvalue(), content_type='text/csv; charset=utf-8')
    safe_name = "".join(c if c.isalnum() or c in ' -_' else '_' for c in event.name)[:80]
    response['Content-Disposition'] = f'attachment; filename="event_registrations_{safe_name}.csv"'
    return response


def _guard_display_scanner_url_for_event(request, event_id):
    """
    Same UX as the wall scanner dashboard (/gate/guard-display/?token=…) with event pre-selected.
    Returns None if GATE_GUARD_DISPLAY_TOKEN is not set (caller falls back to /gate/?event=).
    """
    token = (getattr(settings, 'GATE_GUARD_DISPLAY_TOKEN', '') or '').strip()
    if not token:
        return None
    params = {'token': token, 'event': str(event_id)}
    kiosk = (request.GET.get('kiosk') or '').strip().lower() in ('1', 'true', 'yes', 'on')
    if kiosk:
        params['kiosk'] = '1'
    return f"{reverse('gate-guard-display')}?{urlencode(params)}"


def _redirect_gate_scan_with_event(request, event_id):
    """Prefer guard scanner dashboard when token is configured; else /gate/?event=…"""
    url = _guard_display_scanner_url_for_event(request, event_id)
    if url:
        return redirect(url)
    return redirect(f"{reverse('gate-scan')}?{urlencode({'event': str(event_id)})}")


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_attendance_scanner(request):
    """Event attendance scanner hub: lists today's active/scheduled events so staff can open the scanner for each."""
    today = timezone.localdate()
    events = Event.objects.filter(
        status__in=('active', 'scheduled'),
        start_date__lte=today,
        end_date__gte=today,
    ).order_by('start_date', 'name')
    return render(request, 'gate/event_attendance_scanner.html', {
        'site_name': getattr(settings, 'SITE_NAME', 'City College of Bayawan'),
        'events': events,
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_attendance_report(request, event_id):
    """View attendance logs and analytics for an event. Optional ?q= for quick search by student ID or name."""
    event = get_object_or_404(Event, id=event_id)
    
    logs_qs = AttendanceLog.objects.filter(event=event, voided=False).select_related('student', 'registration').order_by('-scan_time')
    q = (request.GET.get('q') or '').strip()
    if q:
        from django.db.models import Q
        logs_qs = logs_qs.filter(
            Q(student__student_id__icontains=q) |
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q)
        )
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    from django.core.paginator import Paginator
    paginator = Paginator(logs_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    logs = list(page_obj.object_list)
    
    # Stats (on full set for this event, not filtered)
    stats_qs = AttendanceLog.objects.filter(event=event, voided=False)
    total_scans = stats_qs.count()
    successful = stats_qs.filter(result='SUCCESS').count()
    duplicates = stats_qs.filter(result='DUPLICATE').count()
    invalid = stats_qs.filter(result__in=['INVALID', 'REVOKED', 'WRONG_EVENT', 'OUTSIDE_WINDOW', 'SECURE_EVENT_REQUIRES_TOKEN']).count()
    overrides = stats_qs.filter(remarks='AUDIENCE_OVERRIDE').count()
    
    # EventAttendance summary (covers both token and student-ID scans)
    attendances = EventAttendance.objects.filter(event=event)
    checked_in = attendances.filter(checked_in_at__isnull=False).count()
    checked_out = attendances.filter(checked_out_at__isnull=False).count()
    currently_inside = attendances.filter(checked_in_at__isnull=False, checked_out_at__isnull=True).count()
    total_registered = EventRegistration.objects.filter(event=event).count()
    
    from django.utils import timezone
    today = timezone.localdate()
    event_ended = event.end_date < today

    context = {
        'site_name': 'City College of Bayawan',
        'event': event,
        'today': today,
        'event_ended': event_ended,
        'logs': logs,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'search_query': q,
        'total_scans': total_scans,
        'successful': successful,
        'duplicates': duplicates,
        'invalid': invalid,
        'overrides': overrides,
        'total_registered': total_registered,
        'checked_in': checked_in,
        'checked_out': checked_out,
        'currently_inside': currently_inside,
        'attendance_rate': round((checked_in / total_registered * 100) if total_registered else 0, 1),
    }
    if (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes'):
        return render(request, 'gate/event_attendance_report_embed.html', context)
    return render(request, 'gate/event_attendance_report.html', context)


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_attendance_report_export_csv(request, event_id):
    """Export event attendance summary as CSV. Sorted by checked-in time (first to enter first). 12-hour time; name as Family, Given Middle."""
    event = get_object_or_404(Event, id=event_id)
    attendances = (
        EventAttendance.objects.filter(event=event)
        .select_related('student')
        .order_by(F('checked_in_at').asc(nulls_last=True))
    )

    def name_family_first(student):
        given_middle = ' '.join(filter(None, [student.first_name or '', student.middle_name or ''])).strip()
        if student.last_name:
            return f"{student.last_name}, {given_middle}".strip()
        return given_middle or student.get_full_name()
    
    def get_course_section_year(student):
        """Build course/section/year display for export."""
        course_section = (student.course_or_section or '').strip()
        if not course_section:
            parts = []
            if getattr(student, 'course', None):
                parts.append(student.get_course_display() if hasattr(student, 'get_course_display') else student.course)
            if getattr(student, 'section', None) and (student.section or '').strip():
                parts.append((student.section or '').strip())
            course_section = ' - '.join(parts) if parts else ''
        
        # Add year level
        year_level = getattr(student, 'year_level', None) or ''
        if course_section and year_level:
            return f"{course_section} - {year_level}"
        elif course_section:
            return course_section
        elif year_level:
            return f"Year {year_level}"
        return '—'

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{event.id}_{event.name[:30].replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Student ID', 'Name', 'Program/Section/Year', 'Checked In', 'Checked Out'])
    for a in attendances:
        # Use 12-hour format; prefix with apostrophe so Excel displays as text (not military time)
        checked_in = _format_event_time_12h(a.checked_in_at)
        checked_out = _format_event_time_12h(a.checked_out_at)
        if checked_in:
            checked_in = "'" + checked_in
        if checked_out:
            checked_out = "'" + checked_out
        name = name_family_first(a.student)
        course_section_year = get_course_section_year(a.student)
        writer.writerow([a.student.student_id, name, course_section_year, checked_in, checked_out])
    return response


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_scan_logs_export_csv(request, event_id):
    """Export event scan logs as CSV."""
    event = get_object_or_404(Event, id=event_id)
    logs = AttendanceLog.objects.filter(event=event, voided=False).select_related('student').order_by('-scan_time')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="scan_logs_{event.id}_{event.name[:30].replace(" ", "_")}.csv"'
    writer = csv.writer(response)
    writer.writerow(['Scan Time', 'Student ID', 'Name', 'Scan Type', 'Result', 'Device ID', 'Recorded By', 'Remarks'])
    for log in logs:
        student_id = log.student.student_id if log.student else ''
        name = log.student.get_full_name() if log.student else ''
        scan_time = timezone.localtime(log.scan_time).strftime('%Y-%m-%d %H:%M:%S') if log.scan_time else ''
        recorded_by = (log.recorded_by.username if log.recorded_by else '') or ''
        writer.writerow([scan_time, student_id, name, log.scan_type, log.result, log.device_id or '', recorded_by, log.remarks or ''])
    return response


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_attendance_report_export_xlsx(request, event_id):
    """Export event attendance summary as Excel (.xlsx). Sorted by checked-in time; 12-hour time; name as Family, Given Middle."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment
    except ImportError:
        return HttpResponse('Excel export requires openpyxl. Install: pip install openpyxl', status=501)
    event = get_object_or_404(Event, id=event_id)
    attendances = (
        EventAttendance.objects.filter(event=event)
        .select_related('student')
        .order_by(F('checked_in_at').asc(nulls_last=True))
    )

    def name_family_first(student):
        given_middle = ' '.join(filter(None, [student.first_name or '', student.middle_name or ''])).strip()
        if student.last_name:
            return f"{student.last_name}, {given_middle}".strip()
        return given_middle or student.get_full_name()
    
    def get_course_section_year(student):
        """Build course/section/year display for export."""
        course_section = (student.course_or_section or '').strip()
        if not course_section:
            parts = []
            if getattr(student, 'course', None):
                parts.append(student.get_course_display() if hasattr(student, 'get_course_display') else student.course)
            if getattr(student, 'section', None) and (student.section or '').strip():
                parts.append((student.section or '').strip())
            course_section = ' - '.join(parts) if parts else ''
        
        # Add year level
        year_level = getattr(student, 'year_level', None) or ''
        if course_section and year_level:
            return f"{course_section} - {year_level}"
        elif course_section:
            return course_section
        elif year_level:
            return f"Year {year_level}"
        return '—'

    wb = Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    headers = ['Student ID', 'Name', 'Program/Section/Year', 'Checked In', 'Checked Out']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True)
    for row, a in enumerate(attendances, 2):
        checked_in = _format_event_time_12h(a.checked_in_at)
        checked_out = _format_event_time_12h(a.checked_out_at)
        name = name_family_first(a.student)
        course_section_year = get_course_section_year(a.student)
        ws.cell(row=row, column=1, value=a.student.student_id)
        ws.cell(row=row, column=2, value=name)
        ws.cell(row=row, column=3, value=course_section_year)
        c4 = ws.cell(row=row, column=4, value=checked_in)
        c5 = ws.cell(row=row, column=5, value=checked_out)
        c4.number_format = '@'
        c5.number_format = '@'
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{event.id}_{event.name[:30].replace(" ", "_")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_attendance_report_export_pdf(request, event_id):
    """Export event attendance summary as PDF."""
    try:
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
    except ImportError:
        return HttpResponse('PDF export requires reportlab. Install: pip install reportlab', status=501)
    event = get_object_or_404(Event, id=event_id)
    attendances = (
        EventAttendance.objects.filter(event=event)
        .select_related('student')
        .order_by(F('checked_in_at').asc(nulls_last=True))
    )

    def name_family_first(student):
        given_middle = ' '.join(filter(None, [student.first_name or '', student.middle_name or ''])).strip()
        if student.last_name:
            return f"{student.last_name}, {given_middle}".strip()
        return given_middle or student.get_full_name()
    
    def get_course_section_year(student):
        """Build course/section/year display for export."""
        course_section = (student.course_or_section or '').strip()
        if not course_section:
            parts = []
            if getattr(student, 'course', None):
                parts.append(student.get_course_display() if hasattr(student, 'get_course_display') else student.course)
            if getattr(student, 'section', None) and (student.section or '').strip():
                parts.append((student.section or '').strip())
            course_section = ' - '.join(parts) if parts else ''
        
        # Add year level
        year_level = getattr(student, 'year_level', None) or ''
        if course_section and year_level:
            return f"{course_section} - {year_level}"
        elif course_section:
            return course_section
        elif year_level:
            return f"Year {year_level}"
        return '—'

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="attendance_report_{event.id}_{event.name[:30].replace(" ", "_")}.pdf"'
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=36, leftMargin=36, topMargin=36, bottomMargin=36)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph(f'Attendance Report: {event.name}', styles['Title']))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f'Event dates: {event.start_date} to {event.end_date}', styles['Normal']))
    elements.append(Spacer(1, 16))

    def _pdf_cell_txt(val):
        s = (val if val is not None else '') or ''
        s = str(s)
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    # Header: Paragraphs wrap; body: Paragraphs with fixed width from colWidths so long text does not bleed.
    th_style = ParagraphStyle(
        'eath', parent=styles['Normal'], fontName='Helvetica-Bold', fontSize=7.5, leading=9,
        textColor=colors.whitesmoke, alignment=TA_CENTER, spaceAfter=0, spaceBefore=0,
    )
    td_style = ParagraphStyle(
        'eatd', parent=styles['Normal'], fontName='Helvetica', fontSize=7.5, leading=9.5,
        textColor=colors.black, alignment=TA_LEFT, spaceAfter=0, spaceBefore=0,
    )
    # Five columns: drop "Recorded at" (auto_now_add) — usually duplicates check-in and wasted space.
    data = [[
        Paragraph('Student ID', th_style),
        Paragraph('Name', th_style),
        Paragraph('Program / section / year', th_style),
        Paragraph('Checked In', th_style),
        Paragraph('Checked Out', th_style),
    ]]
    for a in attendances:
        checked_in = _format_event_time_12h(a.checked_in_at) or '—'
        checked_out = _format_event_time_12h(a.checked_out_at) or '—'
        name = name_family_first(a.student)
        course_section_year = get_course_section_year(a.student)
        data.append([
            Paragraph(_pdf_cell_txt(a.student.student_id), td_style),
            Paragraph(_pdf_cell_txt(name), td_style),
            Paragraph(_pdf_cell_txt(course_section_year), td_style),
            Paragraph(_pdf_cell_txt(checked_in), td_style),
            Paragraph(_pdf_cell_txt(checked_out), td_style),
        ])
    # ~7.27" content width: wider time columns for "MM/DD/YYYY h:mm AM"; program column matches typical data.
    t = Table(
        data,
        colWidths=[0.8*inch, 1.55*inch, 1.5*inch, 1.38*inch, 1.38*inch],
    )
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
        ('TOPPADDING', (0, 0), (-1, 0), 4),
        ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
        ('VALIGN', (0, 1), (-1, -1), 'TOP'),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('LEFTPADDING', (0, 0), (-1, -1), 3),
        ('RIGHTPADDING', (0, 0), (-1, -1), 3),
        ('TOPPADDING', (0, 1), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(t)
    doc.build(elements)
    return response


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_scan_logs_export_xlsx(request, event_id):
    """Export event scan logs as Excel (.xlsx)."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return HttpResponse('Excel export requires openpyxl. Install: pip install openpyxl', status=501)
    event = get_object_or_404(Event, id=event_id)
    logs = AttendanceLog.objects.filter(event=event, voided=False).select_related('student', 'recorded_by').order_by('-scan_time')
    wb = Workbook()
    ws = wb.active
    ws.title = 'Scan logs'
    headers = ['Scan Time', 'Student ID', 'Name', 'Scan Type', 'Result', 'Device ID', 'Recorded By', 'Remarks']
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h).font = Font(bold=True)
    for row, log in enumerate(logs, 2):
        student_id = log.student.student_id if log.student else ''
        name = log.student.get_full_name() if log.student else ''
        scan_time = timezone.localtime(log.scan_time).strftime('%Y-%m-%d %H:%M:%S') if log.scan_time else ''
        recorded_by = (log.recorded_by.username if log.recorded_by else '') or ''
        ws.cell(row=row, column=1, value=scan_time)
        ws.cell(row=row, column=2, value=student_id)
        ws.cell(row=row, column=3, value=name)
        ws.cell(row=row, column=4, value=log.scan_type)
        ws.cell(row=row, column=5, value=log.result)
        ws.cell(row=row, column=6, value=log.device_id or '')
        ws.cell(row=row, column=7, value=recorded_by)
        ws.cell(row=row, column=8, value=log.remarks or '')
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="scan_logs_{event.id}_{event.name[:30].replace(" ", "_")}.xlsx"'
    wb.save(response)
    return response


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_currently_inside(request, event_id):
    """List students currently inside event (checked in, not checked out)."""
    event = get_object_or_404(Event, id=event_id)
    inside = EventAttendance.objects.filter(
        event=event,
        checked_in_at__isnull=False,
        checked_out_at__isnull=True,
    ).select_related('student').order_by('checked_in_at')
    
    embed = request.GET.get('embed') == '1'
    template = 'gate/event_currently_inside_embed.html' if embed else 'gate/event_currently_inside.html'
    
    return render(request, template, {
        'site_name': 'City College of Bayawan',
        'event': event,
        'inside': inside,
        'count': inside.count(),
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_live_dashboard(request, event_id):
    """Live dashboard for event: real-time counts, optional auto-refresh.
    When ?embed=1 render compact table-only view. When event ended, no auto-refresh by default."""
    event = get_object_or_404(Event, id=event_id)
    from django.utils import timezone
    today = timezone.localdate()
    event_ended = event.end_date < today

    attendances = EventAttendance.objects.filter(event=event)
    checked_in = attendances.filter(checked_in_at__isnull=False).count()
    checked_out = attendances.filter(checked_out_at__isnull=False).count()
    currently_inside = attendances.filter(checked_in_at__isnull=False, checked_out_at__isnull=True).count()

    logs_qs = AttendanceLog.objects.filter(event=event, voided=False).select_related('student').order_by('-scan_time')
    scan_type_filter = (request.GET.get('type') or '').strip().upper()
    if scan_type_filter in ('IN', 'OUT'):
        logs_qs = logs_qs.filter(scan_type=scan_type_filter)
    search_q = (request.GET.get('q') or '').strip()
    if search_q:
        logs_qs = logs_qs.filter(
            Q(student__student_id__icontains=search_q) |
            Q(student__first_name__icontains=search_q) |
            Q(student__last_name__icontains=search_q)
        )
    recent_limit = min(int(request.GET.get('limit', 25)), 100)
    recent = list(logs_qs[:recent_limit])

    refresh_seconds = None if event_ended else (5 if request.GET.get('refresh', '1') != '0' else None)
    if not event_ended and request.GET.get('refresh') == '10':
        refresh_seconds = 10
    if not event_ended and request.GET.get('refresh') == '30':
        refresh_seconds = 30

    context = {
        'site_name': 'City College of Bayawan',
        'event': event,
        'today': today,
        'event_ended': event_ended,
        'checked_in': checked_in,
        'checked_out': checked_out,
        'currently_inside': currently_inside,
        'recent': recent,
        'recent_limit': recent_limit,
        'scan_type_filter': scan_type_filter,
        'search_q': search_q,
        'refresh_seconds': refresh_seconds,
    }
    if (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes'):
        return render(request, 'gate/event_live_dashboard_embed.html', context)
    return render(request, 'gate/event_live_dashboard.html', context)


@login_required(login_url='/login/')
@role_required('student', 'admin', 'staff')
def student_portal(request):
    """Student view: my gate logs, my event attendance, my points (if linked to a student)."""
    student = None
    # Link: username can match student_id (admin creates user with username=student_id and role Student)
    if request.user.is_authenticated:
        student = Student.objects.filter(student_id=request.user.username.strip(), is_active=True).first()
    
    if not student:
        return render(request, 'gate/student_portal.html', {
            'site_name': 'City College of Bayawan',
            'student': None,
            'gate_entries': [],
            'event_attendances': [],
            'total_points': 0,
        })
    
    gate_entries = GateEntry.objects.filter(student=student).order_by('-timestamp')[:50]
    event_attendances = EventAttendance.objects.filter(student=student).select_related('event').order_by('-recorded_at')[:50]
    from django.db.models import Sum
    points_earned = EventAttendance.objects.filter(student=student, participated=True).aggregate(
        total=Sum('event__points')
    )
    total_points = points_earned['total'] or 0
    
    return render(request, 'gate/student_portal.html', {
        'site_name': 'City College of Bayawan',
        'student': student,
        'gate_entries': gate_entries,
        'event_attendances': event_attendances,
        'total_points': total_points,
    })


# ------------------------- Report Generation -------------------------

# ----- New Reports Menu (Overview, Daily Gate, Event Attendance, Incidents & Overrides, Exports) -----

@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def reports_overview(request):
    """Reports Overview.

    This menu now defaults to the Export/Print flow (reports-exports) so the
    Reports sidebar stays simple: filter + preview + export.
    """
    from django.urls import reverse
    base_url = reverse('reports-exports')
    query_string = request.META.get('QUERY_STRING') or ''
    if query_string:
        base_url = f'{base_url}?{query_string}'
    return redirect(base_url)
    filter_date, day_start, day_end, from_time, to_time, search_q, event_id, date_range_label, from_date_str, to_date_str, time_error, report_timestamp_q, report_checked_in_at_q, report_checked_out_at_q, report_recorded_at_q = _report_filter_from_request(request)
    # Student daily gate only (no event, no visitor)
    entries_qs = GateEntry.objects.filter(
        timestamp__gte=day_start, timestamp__lt=day_end,
        event__isnull=True, visitor_visit__isnull=True, student_id__isnull=False,
    ).select_related('student', 'recorded_by')
    if search_q:
        entries_qs = entries_qs.filter(
            Q(student__student_id__icontains=search_q) |
            Q(student__first_name__icontains=search_q) |
            Q(student__last_name__icontains=search_q)
        )
    entries = list(entries_qs.order_by('-timestamp')[:500])
    visits = _gate_entries_to_visits(entries)
    total_visits = len(visits)
    inside_now = _currently_inside_count(filter_date)
    denied_entries = GateEntry.objects.filter(
        timestamp__gte=day_start, timestamp__lt=day_end,
        granted=False, event__isnull=True, visitor_visit__isnull=True,
    ).count()
    incidents_count = 0  # Incidents feature removed
    # Forced OUT (no IN): visits where in_entry is None, out_entry is not None
    forced_out_no_in = sum(1 for in_e, out_e in visits if in_e is None and out_e is not None)
    # Top early-out reasons (out_reason_code)
    from django.db.models import Count
    early_out_qs = GateEntry.objects.filter(
        timestamp__gte=day_start, timestamp__lt=day_end,
        scan_type='OUT', granted=True, event__isnull=True, visitor_visit__isnull=True,
        out_reason_code__isnull=False,
    ).exclude(out_reason_code='').values('out_reason_code').annotate(c=Count('id')).order_by('-c')[:10]
    top_early_out = list(early_out_qs)
    # PH time window for reports: 6:00 AM–10:00 PM (class day + allowance)
    report_hours = list(range(6, 23))  # 6..22 inclusive
    hour_to_idx = {h: i for i, h in enumerate(report_hours)}

    # Hourly buckets for chart (IN and OUT by hour)
    hourly_in = defaultdict(int)
    hourly_out = defaultdict(int)
    for e in entries:
        h = timezone.localtime(e.timestamp).hour if timezone.is_aware(e.timestamp) else e.timestamp.hour
        if h not in hour_to_idx:
            continue
        if getattr(e, 'scan_type', None) == 'OUT' or (e.notes or '').strip().upper() == 'OUT':
            hourly_out[h] += 1
        else:
            hourly_in[h] += 1
    hourly_chart = [{'hour': h, 'in': hourly_in[h], 'out': hourly_out[h]} for h in report_hours]
    hourly_chart_json = json.dumps(hourly_chart)
    # Recent activity (last 20)
    recent = list(entries_qs.order_by('-timestamp')[:20])
    entries_list_url = reverse('gate-entry-list') + f'?from_date={filter_date.isoformat()}'
    incidents_list_url = reverse('gate-entry-list') + f'?from_date={filter_date.isoformat()}&filter=denied'

    # Heatmap: day × hour (last 7 days, 6 AM–10 PM PH time)
    heatmap_days = []
    heatmap_data = []
    heatmap_max = 1
    for i in range(7):
        d = filter_date - datetime.timedelta(days=(6 - i))
        heatmap_days.append(d)
        ds, de = _local_day_bounds(d)
        hour_counts = [0] * len(report_hours)
        entries_hm = GateEntry.objects.filter(
            timestamp__gte=ds, timestamp__lt=de,
            event__isnull=True, visitor_visit__isnull=True,
        )
        for e in entries_hm:
            h = timezone.localtime(e.timestamp).hour if timezone.is_aware(e.timestamp) else e.timestamp.hour
            idx = hour_to_idx.get(h)
            if idx is not None:
                hour_counts[idx] += 1
        heatmap_data.append(hour_counts)
        heatmap_max = max(heatmap_max, max(hour_counts))
    heatmap_display = []
    for row in heatmap_data:
        heatmap_display.append([{'count': c, 'opacity': round(c / max(heatmap_max, 1), 2)} for c in row])
    heatmap_rows = list(zip(heatmap_days, heatmap_display))
    heatmap_json = json.dumps({'days': [d.isoformat() for d in heatmap_days], 'data': heatmap_data, 'max': heatmap_max})

    return render(request, 'gate/reports/overview.html', {
        'site_name': 'City College of Bayawan',
        'filter_date': filter_date,
        'day_start': day_start,
        'day_end': day_end,
        'date_range_label': date_range_label,
        'from_time': from_time,
        'to_time': to_time,
        'search_q': search_q,
        'event_id': event_id,
        'total_visits': total_visits,
        'inside_now': inside_now,
        'denied_attempts': denied_entries,
        'incidents_created': incidents_count,
        'forced_out_no_in': forced_out_no_in,
        'top_early_out': top_early_out,
        'hourly_chart': hourly_chart,
        'hourly_chart_json': hourly_chart_json,
        'recent_activity': recent,
        'entries_list_url': entries_list_url,
        'incidents_list_url': incidents_list_url,
        'show_event_filter': False,
        'heatmap_days': heatmap_days,
        'heatmap_data': heatmap_data,
        'heatmap_rows': heatmap_rows,
        'heatmap_max': heatmap_max,
        'heatmap_json': heatmap_json,
        'heatmap_hours': report_hours,
        'from_date_str': from_date_str,
        'to_date_str': to_date_str,
        'time_error': time_error,
        'timezone_str': getattr(settings, 'TIME_ZONE', 'Asia/Manila'),
        'applied_filter_chips': _report_applied_filter_chips(request, date_range_label, from_time, to_time, search_q, event_id, active_events=None),
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def reports_daily_gate(request):
    """Daily Gate: Visits, Inside Now, Trends. Sub-tab via ?tab=visits|inside_now|trends."""
    filter_date, day_start, day_end, from_time, to_time, search_q, event_id, date_range_label, from_date_str, to_date_str, time_error, report_timestamp_q, report_checked_in_at_q, report_checked_out_at_q, report_recorded_at_q = _report_filter_from_request(request)
    tab = (request.GET.get('tab') or 'visits').strip().lower()
    if tab not in ('visits', 'inside_now', 'trends'):
        tab = 'visits'
    entries_qs = GateEntry.objects.filter(
        event__isnull=True, visitor_visit__isnull=True, student_id__isnull=False,
    ).select_related('student', 'recorded_by')
    entries_qs = _apply_report_timestamp_filter(entries_qs, report_timestamp_q, day_start, day_end)
    if search_q:
        entries_qs = entries_qs.filter(
            Q(student__student_id__icontains=search_q) |
            Q(student__first_name__icontains=search_q) |
            Q(student__last_name__icontains=search_q)
        )
    entries = list(entries_qs.order_by('-timestamp')[:500])
    visits = _gate_entries_to_visits(entries)
    # Build visit rows with status badge
    visit_rows = []
    for in_e, out_e in visits:
        entry = in_e or out_e
        sid = entry.student.student_id if entry.student else ''
        name = (entry.student.get_full_name() or entry.student.username or '') if entry.student else ''
        in_ts = in_e.timestamp if in_e else None
        out_ts = out_e.timestamp if out_e else None
        if in_e and out_e:
            status = 'Completed'
        elif in_e:
            status = 'Inside'
        else:
            status = 'Forced OUT'
        duration = None
        if in_ts and out_ts:
            duration = (out_ts - in_ts).total_seconds() / 60  # minutes
        visit_rows.append({
            'student_id': sid,
            'name': name,
            'in_time': in_ts,
            'out_time': out_ts,
            'duration_min': duration,
            'duration_hm': _format_duration_hm(duration),
            'status': status,
            'out_reason_code': getattr(out_e, 'out_reason_code', '') if out_e else '',
        })
    inside_now_list = _currently_inside_list(filter_date)
    # Early outs: group by out_reason_code
    early_out_entries = GateEntry.objects.filter(
        scan_type='OUT', granted=True, event__isnull=True, visitor_visit__isnull=True,
    ).select_related('student')
    early_out_entries = _apply_report_timestamp_filter(early_out_entries, report_timestamp_q, day_start, day_end)
    early_out_entries = early_out_entries.order_by('-timestamp')
    if search_q:
        early_out_entries = early_out_entries.filter(
            Q(student__student_id__icontains=search_q) |
            Q(student__first_name__icontains=search_q) |
            Q(student__last_name__icontains=search_q)
        )
    early_out_entries = list(early_out_entries[:200])
    forced_out_no_in_list = [r for r in visit_rows if r['status'] == 'Forced OUT']
    # Trends: simple daily counts for the week (for chart or table)
    trend_days = []
    for i in range(7):
        d = filter_date - datetime.timedelta(days=(6 - i))
        ds, de = _local_day_bounds(d)
        eq = GateEntry.objects.filter(
            timestamp__gte=ds, timestamp__lt=de,
            granted=True, event__isnull=True, visitor_visit__isnull=True, student_id__isnull=False,
        )
        cnt = len(_gate_entries_to_visits(list(eq[:300])))
        trend_days.append({'date': d, 'visits': cnt})
    return render(request, 'gate/reports/daily_gate.html', {
        'site_name': 'City College of Bayawan',
        'filter_date': filter_date,
        'date_range_label': date_range_label,
        'from_time': from_time,
        'to_time': to_time,
        'search_q': search_q,
        'event_id': event_id,
        'tab': tab,
        'visit_rows': visit_rows,
        'inside_now_list': inside_now_list,
        'early_out_entries': early_out_entries,
        'forced_out_no_in_list': forced_out_no_in_list,
        'trend_days': trend_days,
        'show_event_filter': False,
        'from_date_str': from_date_str,
        'to_date_str': to_date_str,
        'time_error': time_error,
        'timezone_str': getattr(settings, 'TIME_ZONE', 'Asia/Manila'),
        'applied_filter_chips': _report_applied_filter_chips(request, date_range_label, from_time, to_time, search_q, event_id, active_events=None),
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def reports_event_attendance(request):
    """Event Attendance: Summary, Attendees, Timeline. Requires event_id in filter."""
    filter_date, day_start, day_end, from_time, to_time, search_q, event_id, date_range_label, from_date_str, to_date_str, time_error, report_timestamp_q, report_checked_in_at_q, report_checked_out_at_q, report_recorded_at_q = _report_filter_from_request(request)
    tab = (request.GET.get('tab') or 'summary').strip().lower()
    if tab not in ('summary', 'attendees', 'timeline'):
        tab = 'summary'
    active_events = Event.objects.filter(status__in=('active', 'scheduled', 'completed')).order_by('-start_date')[:50]
    event = None
    summary = {}
    attendees = []
    timeline = []
    if event_id:
        event = Event.objects.filter(pk=event_id).first()
    if not event:
        event = active_events.first()
        if event:
            event_id = event.id
    if event:
        from django.db.models import Count
        
        # If no explicit date filter is set, show all attendance for the event (not just today)
        # This makes the report more useful by default
        if not request.GET.get('date_range') and not request.GET.get('from_date'):
            # No date filter applied - show all attendance for this event
            checkins = EventAttendance.objects.filter(event=event, checked_in_at__isnull=False)
        else:
            # Date filter applied - respect it
            checkins = EventAttendance.objects.filter(event=event, checked_in_at__isnull=False)
            checkins = _apply_report_checked_in_at_filter(checkins, report_checked_in_at_q, day_start, day_end)
        
        if search_q:
            checkins = checkins.filter(
                Q(student__student_id__icontains=search_q) |
                Q(student__first_name__icontains=search_q) |
                Q(student__last_name__icontains=search_q)
            )
        checkins = checkins.select_related('student').order_by('-checked_in_at')
        total = checkins.count()
        checked_out = checkins.exclude(checked_out_at__isnull=True).count()
        summary = {'total': total, 'checked_in': total, 'checked_out': checked_out}
        
        # Process attendees and add formatted course/section/year
        attendees_list = list(checkins[:200])
        for att in attendees_list:
            # Build course_section_year display: use course_or_section if set, else derive from course + section + year
            s = att.student
            course_section = (s.course_or_section or '').strip()
            if not course_section:
                parts = []
                if getattr(s, 'course', None):
                    parts.append(s.get_course_display() if hasattr(s, 'get_course_display') else s.course)
                if getattr(s, 'section', None) and (s.section or '').strip():
                    parts.append((s.section or '').strip())
                course_section = ' - '.join(parts) if parts else ''
            
            # Add year level to the display
            year_level = getattr(s, 'year_level', None) or ''
            if course_section and year_level:
                course_section_year = f"{course_section} - {year_level}"
            elif course_section:
                course_section_year = course_section
            elif year_level:
                course_section_year = f"Year {year_level}"
            else:
                course_section_year = '—'
            
            # Add as attribute for template access
            att.course_section_display = course_section_year
        attendees = attendees_list
        
        # Timeline: by 10-min bucket (using Python to avoid MySQL timezone issues)
        if not request.GET.get('date_range') and not request.GET.get('from_date'):
            # No date filter - show all timeline data for the event
            timeline_data = EventAttendance.objects.filter(
                event=event, checked_in_at__isnull=False,
            ).values_list('checked_in_at', flat=True).order_by('checked_in_at')
        else:
            # Date filter applied - respect it
            tl = EventAttendance.objects.filter(event=event, checked_in_at__isnull=False)
            tl = _apply_report_checked_in_at_filter(tl, report_checked_in_at_q, day_start, day_end)
            timeline_data = tl.values_list('checked_in_at', flat=True).order_by('checked_in_at')
        
        # Group by 10-minute buckets in Python
        from collections import defaultdict
        bucket_counts = defaultdict(int)
        for dt in timeline_data:
            if dt:
                # Round down to 10-minute bucket
                local_dt = timezone.localtime(dt)
                bucket_minute = (local_dt.minute // 10) * 10
                bucket_time = local_dt.replace(minute=bucket_minute, second=0, microsecond=0)
                bucket_counts[bucket_time] += 1
        
        # Convert to list format expected by template
        timeline = [{'bucket': k, 'c': v} for k, v in sorted(bucket_counts.items())][:100]
    return render(request, 'gate/reports/event_attendance.html', {
        'site_name': 'City College of Bayawan',
        'filter_date': filter_date,
        'date_range_label': date_range_label,
        'from_time': from_time,
        'to_time': to_time,
        'search_q': search_q,
        'event_id': event_id,
        'active_events': active_events,
        'event': event,
        'tab': tab,
        'summary': summary,
        'attendees': attendees,
        'timeline': timeline,
        'show_event_filter': True,
        'from_date_str': from_date_str,
        'to_date_str': to_date_str,
        'time_error': time_error,
        'timezone_str': getattr(settings, 'TIME_ZONE', 'Asia/Manila'),
        'applied_filter_chips': _report_applied_filter_chips(request, date_range_label, from_time, to_time, search_q, event_id, active_events=active_events),
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def reports_incidents_overrides(request):
    """Redirect to reports hub (incidents feature removed)."""
    return redirect('reports-hub')


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def reports_exports(request):
    """Exports: all report types in one menu; choose one specific export or Export all (ZIP)."""
    if request.method == 'POST' and (request.POST.get('_save_report_signatories') or '').strip() == '1':
        theme = SiteTheme.objects.first()
        if theme is None:
            theme = SiteTheme.objects.create(
                site_name=getattr(settings, 'SITE_NAME', 'City College of Bayawan'),
            )
        form_post = SiteThemeReportSignatoryForm(request.POST, request.FILES, instance=theme)
        if form_post.is_valid():
            form_post.save()
            messages.success(
                request,
                'Report signatories saved. They appear on CSV, Excel, PDF exports and when you print.',
            )
            pq = (request.POST.get('preserve_query') or '').strip()
            if pq:
                return redirect(f'{request.path}?{pq}')
            return redirect('reports-exports')
        messages.error(
            request,
            'Could not save report signatories. Use PNG/JPEG images under 3 MB each.',
        )
        pq = (request.POST.get('preserve_query') or '').strip()
        if pq:
            return redirect(f'{request.path}?{pq}')
        return redirect('reports-exports')

    filter_date, day_start, day_end, from_time, to_time, search_q, event_id, date_range_label, from_date_str, to_date_str, time_error, report_timestamp_q, report_checked_in_at_q, report_checked_out_at_q, report_recorded_at_q = _report_filter_from_request(request)
    VALID = (
        'export_all', 'overview_summary', 'recent_activity',
        'daily_gate_visits', 'inside_now', 'raw_entries',
        'event_attendance', 'by_personnel',
    )
    template_type = (request.GET.get('template') or 'daily_gate_visits').strip()
    if template_type not in VALID:
        template_type = 'raw_entries'
    audience_filter = (request.GET.get('audience') or 'all').strip().lower()
    if audience_filter not in ('all', 'students', 'visitors'):
        audience_filter = 'all'
    if template_type != 'daily_gate_visits':
        audience_filter = 'all'

    program_course = (request.GET.get('program') or '').strip().upper()
    year_level = (request.GET.get('year_level') or '').strip()
    section = (request.GET.get('section') or '').strip()
    event_io = (request.GET.get('event_io') or '').strip().upper()
    if event_io not in ('IN', 'OUT'):
        event_io = None
    valid_year_codes = {code for code, _label in Student.YEAR_LEVEL_CHOICES}
    if year_level not in valid_year_codes:
        year_level = None
    if not section:
        section = None
    preview_rows = []
    visitor_preview_rows = []
    # Pagination context (used by exports_inner for the event_attendance dashboard preview)
    page_obj = None
    per_page = None
    per_page_options = None
    query_extra = ''
    query_extra_base = ''
    if template_type == 'export_all':
        preview_rows = []
    else:
        if template_type == 'event_attendance':
            # Event attendance preview: paginate so the dashboard can show all attendees.
            from django.core.paginator import Paginator

            per_page = 50  # fixed: matches your request ("preview of 50 attendees")
            per_page_options = PER_PAGE_OPTIONS

            # Pagination links should preserve every filter except `page` (and the fixed per_page).
            q_extra = request.GET.copy()
            q_extra.pop('page', None)
            q_extra.pop('per_page', None)
            q_extra.pop('partial', None)
            query_extra = q_extra.urlencode()
            query_extra_base = query_extra

            try:
                page_num = int((request.GET.get('page') or '1').strip())
            except (TypeError, ValueError):
                page_num = 1
            page_num = max(1, page_num)

            if event_io == 'IN':
                qs = EventAttendance.objects.all()
                qs = _apply_report_checked_in_at_filter(qs, report_checked_in_at_q, day_start, day_end)
            elif event_io == 'OUT':
                qs = EventAttendance.objects.all()
                qs = _apply_report_checked_out_at_filter(qs, report_checked_out_at_q, day_start, day_end)
            else:
                from django.db.models import Q
                if report_checked_in_at_q is not None:
                    qs = EventAttendance.objects.filter(
                        report_checked_in_at_q | report_checked_out_at_q | report_recorded_at_q
                    )
                else:
                    qs = EventAttendance.objects.filter(
                        Q(checked_in_at__gte=day_start, checked_in_at__lt=day_end) |
                        Q(checked_out_at__gte=day_start, checked_out_at__lt=day_end) |
                        Q(recorded_at__gte=day_start, recorded_at__lt=day_end)
                    )

            if event_id:
                qs = qs.filter(event_id=event_id)

            qs = qs.select_related('student', 'event').order_by('checked_in_at')
            paginator = Paginator(qs, per_page)
            page_obj = paginator.get_page(page_num)

            preview_rows = []
            attendees = list(page_obj.object_list)
            for a in attendees:
                cs = ''
                if a.student:
                    course = (a.student.course or '').strip()
                    section_val = (a.student.section or '').strip()
                    year_level_val = (a.student.year_level or '').strip()

                    # Build course/section display (best-effort, mirrors older preview logic)
                    if course and section_val:
                        cs = f"{course} {section_val}"
                    elif course:
                        cs = course
                    elif section_val:
                        cs = section_val
                    elif getattr(a.student, 'course_or_section', ''):
                        cs = a.student.course_or_section

                    if cs and year_level_val:
                        cs = f"{cs} - {year_level_val}"
                    elif year_level_val:
                        cs = f"Year {year_level_val}"

                preview_rows.append({
                    'event': a.event.name if a.event else '',
                    'student_id': a.student.student_id if a.student else '',
                    'name': _fmt_student_name(a.student) if a.student else '',
                    'Program/Section': cs,
                    'checked_in_at': a.checked_in_at,
                    'checked_out_at': getattr(a, 'checked_out_at', None),
                })
        else:
            preview_rows = _reports_export_preview(
                filter_date, day_start, day_end,
                template_type, search_q, event_id,
                program_course=program_course or None,
                event_io=event_io,
                year_level=year_level,
                section=section,
                audience_filter=audience_filter,
                report_timestamp_q=report_timestamp_q,
                report_checked_in_at_q=report_checked_in_at_q,
                report_checked_out_at_q=report_checked_out_at_q,
                report_recorded_at_q=report_recorded_at_q,
            )
    if template_type == 'daily_gate_visits':
        visitor_preview_rows = _reports_export_visitor_preview(
            day_start, day_end, report_timestamp_q, report_checked_in_at_q,
        )
        if audience_filter == 'students':
            visitor_preview_rows = []
        elif audience_filter == 'visitors':
            preview_rows = []
    # Always provide events for the "specific event" dropdown.
    # Some campuses keep events in statuses other than (active/scheduled/completed),
    # so filtering here caused the dropdown to not render at all.
    active_events = Event.objects.order_by('-start_date', 'name')[:30]
    year_level_choices = Student.YEAR_LEVEL_CHOICES
    _section_cache_key = 'reports_exports_section_choices_v1'
    section_choices = cache.get(_section_cache_key)
    if section_choices is None:
        section_choices = list(
            Student.objects.exclude(section='').values_list('section', flat=True).distinct().order_by('section')
        )
        cache.set(_section_cache_key, section_choices, 120)
    ctx = {
        'site_name': 'City College of Bayawan',
        'filter_date': filter_date,
        'date_range_label': date_range_label,
        'from_time': from_time,
        'to_time': to_time,
        'search_q': search_q,
        'event_id': event_id,
        'template_type': template_type,
        'audience_filter': audience_filter,
        'preview_rows': preview_rows,
        'visitor_preview_rows': visitor_preview_rows,
        'active_events': active_events,
        'program_choices': Student.COURSE_CHOICES,
        'show_event_filter': True,
        'show_report_type_filter': True,
        'from_date_str': from_date_str,
        'to_date_str': to_date_str,
        'time_error': time_error,
        'timezone_str': getattr(settings, 'TIME_ZONE', 'Asia/Manila'),
        'applied_filter_chips': _report_applied_filter_chips(request, date_range_label, from_time, to_time, search_q, event_id, active_events=active_events),
        'year_level_choices': year_level_choices,
        'section_choices': section_choices,
        # Pagination for the dashboard preview table
        'page_obj': page_obj,
        'per_page': per_page,
        'per_page_options': per_page_options,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'reports_exports_ajax': True,
        'report_signatories_form': SiteThemeReportSignatoryForm(instance=SiteTheme.objects.first()),
        'preserve_query_string': request.GET.urlencode(),
    }
    if request.GET.get('partial') == '1':
        if request.headers.get('X-Requested-With') != 'XMLHttpRequest':
            q = request.GET.copy()
            q.pop('partial', None)
            next_url = request.path + ('?' + q.urlencode() if q else '')
            return redirect(next_url)
        return render(request, 'gate/reports/exports_inner.html', ctx)
    return render(request, 'gate/reports/exports.html', ctx)


def _reports_export_visitor_preview(day_start, day_end, report_timestamp_q=None, report_checked_in_at_q=None):
    """Visitor preview rows for the daily_gate_visits report type."""
    visitor_rows = []
    ve = VisitorEntry.objects.all()
    ve = _apply_report_timestamp_filter(ve, report_timestamp_q, day_start, day_end)
    visitor_entries = ve.order_by('-timestamp')[:10]
    vv = VisitorVisit.objects.all()
    vv = _apply_report_checked_in_at_filter(vv, report_checked_in_at_q, day_start, day_end)
    visitor_visits_qs = vv.order_by('-checked_in_at')[:10]
    for v in visitor_entries:
        visitor_rows.append({
            'Name': v.visitor_name,
            'Purpose': getattr(v, 'purpose', '') or '',
            'In time': v.timestamp,
            'Out time': '',
        })
    for vv in visitor_visits_qs:
        visitor_rows.append({
            'Name': vv.full_name,
            'Purpose': getattr(vv, 'purpose', '') or '',
            'In time': vv.checked_in_at,
            'Out time': getattr(vv, 'checked_out_at', None) or '',
        })
    return visitor_rows


def _reports_export_preview(
    filter_date, day_start, day_end, template_type, search_q, event_id,
    program_course=None, event_io=None, year_level=None, section=None,
    audience_filter='all',
    report_timestamp_q=None, report_checked_in_at_q=None, report_checked_out_at_q=None, report_recorded_at_q=None,
):
    """First 10 rows for preview."""
    preview = []
    if template_type == 'daily_gate_visits' and audience_filter != 'visitors':
        entries_qs = GateEntry.objects.filter(
            event__isnull=True, visitor_visit__isnull=True, student_id__isnull=False,
        ).select_related('student')
        entries_qs = _apply_report_timestamp_filter(entries_qs, report_timestamp_q, day_start, day_end)
        if program_course:
            entries_qs = entries_qs.filter(student__course=program_course)
        if year_level:
            entries_qs = entries_qs.filter(student__year_level=year_level)
        if section:
            entries_qs = entries_qs.filter(student__section__iexact=section)
        entries = list(entries_qs.order_by('-timestamp')[:300])
        visits = _gate_entries_to_visits(entries)
        for in_e, out_e in visits[:10]:
            e = in_e or out_e
            cs = ''
            if e.student:
                course = (e.student.course or '').strip()
                section = (e.student.section or '').strip()
                cs = (course + ' ' + section).strip()
                if not cs and getattr(e.student, 'course_or_section', ''):
                    cs = e.student.course_or_section
            gender = ''
            if e.student and getattr(e.student, 'sex', ''):
                gender = e.student.get_sex_display()
            preview.append({
                'ID': e.student.student_id if e.student else '',
                'Name': _fmt_student_name(e.student) if e.student else '',
                'Gender': gender,
                'Program/Section': cs,
                'In time': in_e.timestamp if in_e else None,
                'Out time': out_e.timestamp if out_e else None,
            })
    elif template_type == 'raw_entries':
        qs = GateEntry.objects.filter().select_related('student')
        qs = _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end)
        if program_course:
            qs = qs.filter(student__course=program_course)
        if year_level:
            qs = qs.filter(student__year_level=year_level)
        if section:
            qs = qs.filter(student__section__iexact=section)
        if event_io == 'IN':
            qs = qs.filter(scan_type='IN')
        elif event_io == 'OUT':
            qs = qs.filter(scan_type='OUT')
        preview = list(qs.order_by('-timestamp')[:10].values('timestamp', 'student__student_id', 'scan_type', 'result', 'granted'))
    elif template_type == 'incidents_log':
        iq = GateIncident.objects.filter().select_related('student')
        iq = _apply_report_timestamp_filter(iq, report_timestamp_q, day_start, day_end)
        preview = list(iq.order_by('-timestamp')[:10].values('timestamp', 'scanned_id', 'reason', 'details'))
    elif template_type == 'inside_now':
        inside = _currently_inside_list(filter_date)
        for e in inside[:10]:
            preview.append({
                'student_id': e.student.student_id if e.student else '',
                'name': (e.student.get_full_name() or '') if e.student else '',
                'in_time': e.timestamp,
            })
    elif template_type == 'event_attendance':
        # when event_id is provided, filter by that event; otherwise show all events
        if event_io == 'IN':
            qs = EventAttendance.objects.all()
            qs = _apply_report_checked_in_at_filter(qs, report_checked_in_at_q, day_start, day_end)
        elif event_io == 'OUT':
            qs = EventAttendance.objects.all()
            qs = _apply_report_checked_out_at_filter(qs, report_checked_out_at_q, day_start, day_end)
        else:
            from django.db.models import Q
            if report_checked_in_at_q is not None:
                qs = EventAttendance.objects.filter(
                    report_checked_in_at_q | report_checked_out_at_q | report_recorded_at_q
                )
            else:
                qs = EventAttendance.objects.filter(
                    Q(checked_in_at__gte=day_start, checked_in_at__lt=day_end) |
                    Q(checked_out_at__gte=day_start, checked_out_at__lt=day_end) |
                    Q(recorded_at__gte=day_start, recorded_at__lt=day_end)
                )

        qs = qs.select_related('student', 'event').order_by('checked_in_at')
        if event_id:
            qs = qs.filter(event_id=event_id)
        # preview is limited to avoid huge tables, but 50 rows should cover most needs
        att = list(qs[:50])
        for a in att:
            # combine course/section/year for student
            cs = ''
            if a.student:
                course = (a.student.course or '').strip()
                section = (a.student.section or '').strip()
                year_level = (a.student.year_level or '').strip()
                
                # Build course + section
                if course and section:
                    cs = f"{course} {section}"
                elif course:
                    cs = course
                elif section:
                    cs = section
                elif getattr(a.student, 'course_or_section', ''):
                    cs = a.student.course_or_section
                
                # Add year level
                if cs and year_level:
                    cs = f"{cs} - {year_level}"
                elif year_level:
                    cs = f"Year {year_level}"
            
            preview.append({
                'event': a.event.name if a.event else '',
                'student_id': a.student.student_id if a.student else '',
                'name': _fmt_student_name(a.student) if a.student else '',
                'Program/Section': cs,
                'checked_in_at': a.checked_in_at,
                'checked_out_at': getattr(a, 'checked_out_at', None),
            })
    elif template_type == 'overview_summary':
        # total visits needs to respect the full day_start/day_end range (not just
        # the start date). Likewise, "inside now" should be the current count independent
        # of the selected filter range.
        total_visits = _granted_visits_count_for_bounds(day_start, day_end, daily_gate_only=True, report_timestamp_q=report_timestamp_q)
        inside_now = _currently_inside_count()  # always show current inside count
        dq = GateEntry.objects.filter(granted=False, event__isnull=True, visitor_visit__isnull=True)
        dq = _apply_report_timestamp_filter(dq, report_timestamp_q, day_start, day_end)
        denied = dq.count()
        iq = GateIncident.objects.filter()
        iq = _apply_report_timestamp_filter(iq, report_timestamp_q, day_start, day_end)
        incidents = iq.count()
        eq = GateEntry.objects.filter(event__isnull=True, visitor_visit__isnull=True, student_id__isnull=False)
        eq = _apply_report_timestamp_filter(eq, report_timestamp_q, day_start, day_end)
        entries = list(eq.order_by('-timestamp')[:500])
        visits = _gate_entries_to_visits(entries)
        forced_out = sum(1 for in_e, out_e in visits if in_e is None and out_e is not None)
        preview = [
            {'Metric': 'Total visits', 'Value': total_visits},
            {'Metric': 'Inside now', 'Value': inside_now},
            {'Metric': 'Denied attempts', 'Value': denied},
            {'Metric': 'Incidents created', 'Value': incidents},
            {'Metric': 'Forced OUT (no IN)', 'Value': forced_out},
        ]
    elif template_type == 'recent_activity':
        ra = GateEntry.objects.filter(event__isnull=True, visitor_visit__isnull=True).select_related('student')
        ra = _apply_report_timestamp_filter(ra, report_timestamp_q, day_start, day_end)
        entries = ra.order_by('-timestamp')[:10]
        for e in entries:
            preview.append({
                'timestamp': e.timestamp,
                'student_id': e.student.student_id if e.student else '',
                'name': (e.student.get_full_name() or '') if e.student else '',
                'scan_type': e.scan_type or '',
                'result': 'Granted' if e.granted else 'Denied',
            })
    elif template_type == 'incidents_proxy':
        ip = GateIncident.objects.filter(reason='proxy_attendance').select_related('student')
        ip = _apply_report_timestamp_filter(ip, report_timestamp_q, day_start, day_end)
        qs = ip.order_by('-timestamp')[:10]
        for i in qs:
            preview.append({'timestamp': i.timestamp, 'scanned_id': i.scanned_id or '', 'details': (i.details or '')[:100]})
    elif template_type == 'incidents_overrides':
        io = GateEntry.objects.filter(granted=True, event__isnull=True, visitor_visit__isnull=True, out_reason_code='OVERRIDE_BY_PERSONNEL').select_related('student', 'recorded_by')
        io = _apply_report_timestamp_filter(io, report_timestamp_q, day_start, day_end)
        qs = io.order_by('-timestamp')[:10]
        for e in qs:
            preview.append({
                'timestamp': e.timestamp, 'student_id': e.student.student_id if e.student else '',
                'name': (e.student.get_full_name() or '') if e.student else '',
                'recorded_by': e.recorded_by.username if e.recorded_by else '',
                'out_reason': (e.out_reason or '')[:80],
            })
    elif template_type == 'by_personnel':
        from django.db.models import Count
        from django.contrib.auth import get_user_model
        User = get_user_model()
        bp = GateEntry.objects.filter(event__isnull=True, visitor_visit__isnull=True, recorded_by_id__isnull=False)
        bp = _apply_report_timestamp_filter(bp, report_timestamp_q, day_start, day_end)
        qs = bp.values('recorded_by_id').annotate(scans=Count('id'), denied=Count('id', filter=Q(granted=False))).order_by('-scans')[:10]
        users_map = {u.pk: (u.get_full_name() or u.username) for u in User.objects.filter(pk__in=[r['recorded_by_id'] for r in qs])}
        for r in qs:
            preview.append({'Recorder': users_map.get(r['recorded_by_id'], '—'), 'Scans': r['scans'], 'Denied': r['denied']})
    return preview


def _format_duration_hm(total_minutes):
    """Format a stay length in minutes as e.g. '1h 45m', '38m', or '2h' (no seconds)."""
    if total_minutes is None:
        return ''
    try:
        m = int(round(float(total_minutes)))
    except (TypeError, ValueError):
        return ''
    if m < 0:
        m = 0
    if m == 0:
        return '0m'
    h, rem = divmod(m, 60)
    if h and rem:
        return f'{h}h {rem}m'
    if h:
        return f'{h}h'
    return f'{rem}m'


def _reports_export_build_data(
    filter_date, day_start, day_end, template_type, search_q, event_id,
    program_course=None, event_io=None, year_level=None, section=None,
    audience_filter='all',
    report_timestamp_q=None, report_checked_in_at_q=None, report_checked_out_at_q=None, report_recorded_at_q=None,
):
    """Build full export rows for the given template and filter."""
    headers = []
    rows = []
    if template_type == 'daily_gate_visits':
        entries_qs = GateEntry.objects.filter(
            event__isnull=True, visitor_visit__isnull=True, student_id__isnull=False,
        ).select_related('student')
        entries_qs = _apply_report_timestamp_filter(entries_qs, report_timestamp_q, day_start, day_end)
        if program_course:
            entries_qs = entries_qs.filter(student__course=program_course)
        if year_level:
            entries_qs = entries_qs.filter(student__year_level=year_level)
        if section:
            entries_qs = entries_qs.filter(student__section__iexact=section)
        entries = list(entries_qs.order_by('-timestamp')[:5000])
        if search_q:
            entries = [e for e in entries if e.student and (
                search_q.lower() in (e.student.student_id or '').lower() or
                search_q.lower() in (e.student.get_full_name() or '').lower()
            )]
        if audience_filter != 'visitors':
            visits = _gate_entries_to_visits(entries)
            headers = ['Student ID', 'Name', 'Gender', 'Program/Section', 'IN time', 'OUT time', 'Duration', 'Status']
            for in_e, out_e in visits:
                e = in_e or out_e
                sid = e.student.student_id if e.student else ''
                name = _fmt_student_name(e.student) if e.student else ''
                gender = ''
                if e.student and getattr(e.student, 'sex', ''):
                    gender = e.student.get_sex_display()
                # compute course/section string
                cs = ''
                if e.student:
                    course = (e.student.course or '').strip()
                    section = (e.student.section or '').strip()
                    cs = (course + ' ' + section).strip()
                    if not cs and getattr(e.student, 'course_or_section', ''):
                        cs = e.student.course_or_section
                in_ts = in_e.timestamp if in_e else None
                out_ts = out_e.timestamp if out_e else None
                in_str = timezone.localtime(in_ts).strftime('%Y-%m-%d %I:%M %p') if in_ts else ''
                out_str = timezone.localtime(out_ts).strftime('%Y-%m-%d %I:%M %p') if out_ts else ''
                duration = ''
                if in_ts and out_ts:
                    mins = (out_ts - in_ts).total_seconds() / 60.0
                    duration = _format_duration_hm(mins)
                status = 'Completed' if (in_e and out_e) else ('Inside' if in_e else 'Forced OUT')
                rows.append([sid, name, gender, cs, in_str, out_str, duration, status])
        # append visitors section
        ve = VisitorEntry.objects.select_related('recorded_by')
        ve = _apply_report_timestamp_filter(ve, report_timestamp_q, day_start, day_end)
        visitor_entries = list(ve.order_by('-timestamp'))
        vv = VisitorVisit.objects.select_related('checked_in_by')
        vv = _apply_report_checked_in_at_filter(vv, report_checked_in_at_q, day_start, day_end)
        visitor_visits_qs = list(vv.order_by('-checked_in_at'))
        if audience_filter != 'students' and (visitor_entries or visitor_visits_qs):
            if not headers:
                headers = ['Visitor Name', 'IN time', 'OUT time', 'Purpose', 'Who to visit', 'Recorded by']
            # blank line then header for visitor block
            if rows:
                rows.append([])
                rows.append(['Visitor Name', 'IN time', 'OUT time', 'Purpose', 'Who to visit', 'Recorded by'])
            for v in visitor_entries:
                rec = (getattr(v.recorded_by, 'get_full_name', lambda: '')() or getattr(v.recorded_by, 'username', '—')) if getattr(v, 'recorded_by', None) else '—'
                rows.append([
                    v.visitor_name,
                    timezone.localtime(v.timestamp).strftime('%Y-%m-%d %I:%M %p'),
                    '',
                    v.purpose or '',
                    v.who_to_visit or '',
                    rec,
                ])
            for vv in visitor_visits_qs:
                rec = (getattr(vv.checked_in_by, 'get_full_name', lambda: '')() or getattr(vv.checked_in_by, 'username', '—')) if getattr(vv, 'checked_in_by', None) else '—'
                in_str = timezone.localtime(vv.checked_in_at).strftime('%Y-%m-%d %I:%M %p') if vv.checked_in_at else ''
                out_str = timezone.localtime(vv.checked_out_at).strftime('%Y-%m-%d %I:%M %p') if getattr(vv, 'checked_out_at', None) else ''
                rows.append([
                    vv.full_name,
                    in_str,
                    out_str,
                    vv.purpose or '',
                    vv.department or '',
                    rec,
                ])
    elif template_type == 'overview_summary':
        total_visits = _granted_visits_count_for_bounds(day_start, day_end, daily_gate_only=True, report_timestamp_q=report_timestamp_q)
        inside_now = _currently_inside_count()
        dq = GateEntry.objects.filter(granted=False, event__isnull=True, visitor_visit__isnull=True)
        dq = _apply_report_timestamp_filter(dq, report_timestamp_q, day_start, day_end)
        denied = dq.count()
        iq = GateIncident.objects.filter()
        iq = _apply_report_timestamp_filter(iq, report_timestamp_q, day_start, day_end)
        incidents = iq.count()
        eq = GateEntry.objects.filter(event__isnull=True, visitor_visit__isnull=True, student_id__isnull=False)
        eq = _apply_report_timestamp_filter(eq, report_timestamp_q, day_start, day_end)
        entries = list(eq.order_by('-timestamp')[:500])
        visits = _gate_entries_to_visits(entries)
        forced_out = sum(1 for in_e, out_e in visits if in_e is None and out_e is not None)
        headers = ['Metric', 'Value']
        rows = [
            ['Total visits', total_visits],
            ['Inside now', inside_now],
            ['Denied attempts', denied],
            ['Incidents created', incidents],
            ['Forced OUT (no IN)', forced_out],
        ]
    elif template_type == 'recent_activity':
        qs = GateEntry.objects.filter(event__isnull=True, visitor_visit__isnull=True).select_related('student')
        qs = _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end)
        qs = qs.order_by('-timestamp')[:500]
        if search_q:
            qs = qs.filter(
                Q(student__student_id__icontains=search_q) |
                Q(student__first_name__icontains=search_q) |
                Q(student__last_name__icontains=search_q)
            )
        headers = ['Timestamp', 'Student ID', 'Name', 'Scan type', 'Result']
        for e in qs:
            ts = timezone.localtime(e.timestamp).strftime('%Y-%m-%d %I:%M:%S %p') if e.timestamp else ''
            sid = e.student.student_id if e.student else ''
            name = (e.student.get_full_name() or '') if e.student else ''
            rows.append([ts, sid, name, e.scan_type or '', 'Granted' if e.granted else 'Denied'])
    elif template_type == 'raw_entries':
        qs = GateEntry.objects.filter().select_related('student')
        qs = _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end)
        if program_course:
            qs = qs.filter(student__course=program_course)
        if year_level:
            qs = qs.filter(student__year_level=year_level)
        if section:
            qs = qs.filter(student__section__iexact=section)
        if event_io == 'IN':
            qs = qs.filter(scan_type='IN')
        elif event_io == 'OUT':
            qs = qs.filter(scan_type='OUT')
        if search_q:
            qs = qs.filter(
                Q(student__student_id__icontains=search_q) |
                Q(student__first_name__icontains=search_q) |
                Q(student__last_name__icontains=search_q)
            )
        qs = qs.order_by('-timestamp')[:5000]
        headers = ['Timestamp', 'Student ID', 'Scan type', 'Result', 'Granted']
        for e in qs:
            ts = timezone.localtime(e.timestamp).strftime('%Y-%m-%d %H:%M:%S') if e.timestamp else ''
            sid = e.student.student_id if e.student else ''
            rows.append([ts, sid, e.scan_type or '', e.result or '', 'Yes' if e.granted else 'No'])
    elif template_type == 'incidents_log':
        qs = GateIncident.objects.filter().select_related('student')
        qs = _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end)
        qs = qs.order_by('-timestamp')[:5000]
        if search_q:
            qs = qs.filter(
                Q(scanned_id__icontains=search_q) |
                Q(student__student_id__icontains=search_q) |
                Q(details__icontains=search_q)
            )
        headers = ['Timestamp', 'Scanned ID', 'Reason', 'Details']
        for i in qs:
            ts = timezone.localtime(i.timestamp).strftime('%Y-%m-%d %H:%M:%S') if i.timestamp else ''
            rows.append([ts, i.scanned_id or '', i.get_reason_display() if i.reason else '', (i.details or '')[:200]])
    elif template_type == 'incidents_proxy':
        qs = GateIncident.objects.filter(reason='proxy_attendance').select_related('student')
        qs = _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end)
        qs = qs.order_by('-timestamp')[:5000]
        if search_q:
            qs = qs.filter(Q(scanned_id__icontains=search_q) | Q(details__icontains=search_q))
        headers = ['Timestamp', 'Scanned ID', 'Details']
        for i in qs:
            ts = timezone.localtime(i.timestamp).strftime('%Y-%m-%d %H:%M:%S') if i.timestamp else ''
            rows.append([ts, i.scanned_id or '', (i.details or '')[:200]])
    elif template_type == 'incidents_overrides':
        qs = GateEntry.objects.filter(granted=True, event__isnull=True, visitor_visit__isnull=True, out_reason_code='OVERRIDE_BY_PERSONNEL').select_related('student', 'recorded_by')
        qs = _apply_report_timestamp_filter(qs, report_timestamp_q, day_start, day_end)
        qs = qs.order_by('-timestamp')
        if search_q:
            qs = qs.filter(
                Q(student__student_id__icontains=search_q) |
                Q(student__first_name__icontains=search_q) |
                Q(student__last_name__icontains=search_q)
            )
        headers = ['Timestamp', 'Student ID', 'Name', 'Recorded by', 'OUT note']
        for e in qs[:2000]:
            ts = timezone.localtime(e.timestamp).strftime('%Y-%m-%d %H:%M:%S') if e.timestamp else ''
            sid = e.student.student_id if e.student else ''
            name = (e.student.get_full_name() or '') if e.student else ''
            rows.append([ts, sid, name, e.recorded_by.username if e.recorded_by else '', (e.out_reason or '')[:200]])
    elif template_type == 'by_personnel':
        from django.db.models import Count
        from django.contrib.auth import get_user_model
        User = get_user_model()
        bp = GateEntry.objects.filter(event__isnull=True, visitor_visit__isnull=True, recorded_by_id__isnull=False)
        bp = _apply_report_timestamp_filter(bp, report_timestamp_q, day_start, day_end)
        qs = bp.values('recorded_by_id').annotate(scans=Count('id'), denied=Count('id', filter=Q(granted=False))).order_by('-scans')
        by_personnel_raw = list(qs[:100])
        user_ids = [r['recorded_by_id'] for r in by_personnel_raw]
        users_map = {u.pk: (u.get_full_name() or u.username) for u in User.objects.filter(pk__in=user_ids)}
        headers = ['Recorder', 'Scans recorded', 'Denied']
        for r in by_personnel_raw:
            rows.append([users_map.get(r['recorded_by_id'], '—'), r['scans'], r.get('denied') or 0])
    elif template_type == 'inside_now':
        inside = _currently_inside_list(filter_date)
        headers = ['Student ID', 'Name', 'IN time']
        for e in inside:
            sid = e.student.student_id if e.student else ''
            name = (e.student.get_full_name() or '') if e.student else ''
            ts = timezone.localtime(e.timestamp).strftime('%Y-%m-%d %I:%M %p') if e.timestamp else ''
            rows.append([sid, name, ts])
    elif template_type == 'event_attendance':
        # support both specific event and all events, and include records where
        # check-in, check-out or creation timestamp falls within the selected interval
        if event_io == 'IN':
            qs = EventAttendance.objects.all()
            qs = _apply_report_checked_in_at_filter(qs, report_checked_in_at_q, day_start, day_end)
        elif event_io == 'OUT':
            qs = EventAttendance.objects.all()
            qs = _apply_report_checked_out_at_filter(qs, report_checked_out_at_q, day_start, day_end)
        else:
            from django.db.models import Q
            if report_checked_in_at_q is not None:
                qs = EventAttendance.objects.filter(
                    report_checked_in_at_q | report_checked_out_at_q | report_recorded_at_q
                )
            else:
                qs = EventAttendance.objects.filter(
                    Q(checked_in_at__gte=day_start, checked_in_at__lt=day_end) |
                    Q(checked_out_at__gte=day_start, checked_out_at__lt=day_end) |
                    Q(recorded_at__gte=day_start, recorded_at__lt=day_end)
                )

        qs = qs.select_related('student', 'event').order_by('checked_in_at')
        if event_id:
            qs = qs.filter(event_id=event_id)
        if search_q:
            qs = qs.filter(
                Q(student__student_id__icontains=search_q) |
                Q(student__first_name__icontains=search_q) |
                Q(student__last_name__icontains=search_q)
            )
        att = list(qs[:5000])
        headers = ['Event', 'Student ID', 'Name', 'Program/Section/Year', 'Checked in', 'Checked out']
        for a in att:
            sid = a.student.student_id if a.student else ''
            name = _fmt_student_name(a.student) if a.student else ''
            # compute course/section/year for event row
            csy = ''
            if a.student:
                course = (a.student.course or '').strip()
                section = (a.student.section or '').strip()
                year = (a.student.year_level or '').strip()
                cs = (course + ' ' + section).strip()
                if not cs and getattr(a.student, 'course_or_section', ''):
                    cs = a.student.course_or_section
                # Build course/section/year display
                if cs and year:
                    csy = f"{cs} - {year}"
                elif cs:
                    csy = cs
                elif year:
                    csy = year
            evname = a.event.name if a.event else ''
            ci = timezone.localtime(a.checked_in_at).strftime('%Y-%m-%d %I:%M %p') if a.checked_in_at else ''
            co = timezone.localtime(a.checked_out_at).strftime('%Y-%m-%d %I:%M %p') if getattr(a, 'checked_out_at', None) else ''
            rows.append([evname, sid, name, csy, ci, co])
    return headers, rows


def _report_signatory_theme_fields(theme):
    """Return (name, title) from SiteTheme report-only fields (not e-ID signatories)."""
    if not theme:
        return '', ''
    n1 = (getattr(theme, 'report_first_signatory_name', '') or '').strip()
    t1 = (getattr(theme, 'report_first_signatory_title', '') or '').strip()
    return n1, t1


def _report_signatories_nonempty(theme):
    n1, t1 = _report_signatory_theme_fields(theme)
    return bool(n1 or t1)


def _append_report_signatories_csv(writer, num_cols, theme=None):
    """Append a signatory block to a CSV export (uses first column; pads row width)."""
    theme = theme if theme is not None else SiteTheme.objects.first()
    if not _report_signatories_nonempty(theme):
        return
    n1, t1 = _report_signatory_theme_fields(theme)
    pad = [''] * max(0, num_cols - 1)

    def _line(name, title):
        parts = []
        if name:
            parts.append(name)
        if title:
            parts.append(f'({title})')
        return ' '.join(parts) if parts else ''

    writer.writerow([])
    writer.writerow(['— Signatory —'] + pad)
    writer.writerow([_line(n1, t1)] + pad)


def _append_report_signatories_xlsx(ws, start_row, theme=None):
    """Write signatory rows starting at start_row; returns next free row index."""
    theme = theme if theme is not None else SiteTheme.objects.first()
    if not _report_signatories_nonempty(theme):
        return start_row
    n1, t1 = _report_signatory_theme_fields(theme)

    def _line(name, title):
        parts = []
        if name:
            parts.append(name)
        if title:
            parts.append(f'({title})')
        return ' '.join(parts) if parts else ''

    r = start_row
    ws.cell(row=r, column=1, value='— Signatory —')
    r += 1
    ws.cell(row=r, column=1, value=_line(n1, t1))
    return r + 1


def _report_signatories_context_for_template(request):
    """Context dict for print preview (single report signatory)."""
    theme = SiteTheme.objects.first()
    n1, t1 = _report_signatory_theme_fields(theme)

    def _abs_url(f):
        if not f or not getattr(f, 'name', None):
            return ''
        try:
            url = f.url
        except Exception:
            return ''
        if not url:
            return ''
        if request:
            try:
                return request.build_absolute_uri(url)
            except Exception:
                pass
        return url

    sig1 = _abs_url(getattr(theme, 'report_first_signatory_signature', None)) if theme else ''
    return {
        'has_signatories': bool(n1 or t1),
        'first_name': n1,
        'first_title': t1,
        'first_signature_url': sig1,
    }


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def reports_export_download(request):
    """Download export as CSV, Excel, or PDF. GET: format=csv|xlsx|pdf|zip, same filter/template params. template=export_all → ZIP of all report types."""
    fmt = (request.GET.get('format') or 'csv').strip().lower()
    if fmt not in ('csv', 'xlsx', 'pdf', 'zip'):
        fmt = 'csv'
    filter_date, day_start, day_end, from_time, to_time, search_q, event_id, date_range_label, from_date_str, to_date_str, time_error, report_timestamp_q, report_checked_in_at_q, report_checked_out_at_q, report_recorded_at_q = _report_filter_from_request(request)
    VALID = (
        'export_all', 'overview_summary', 'recent_activity',
        'daily_gate_visits', 'inside_now', 'raw_entries',
        'event_attendance', 'by_personnel',
    )
    template_type = (request.GET.get('template') or 'daily_gate_visits').strip()
    if template_type not in VALID:
        template_type = 'raw_entries'
    audience_filter = (request.GET.get('audience') or 'all').strip().lower()
    if audience_filter not in ('all', 'students', 'visitors'):
        audience_filter = 'all'
    if template_type != 'daily_gate_visits':
        audience_filter = 'all'

    program_course = (request.GET.get('program') or '').strip().upper()
    year_level = (request.GET.get('year_level') or '').strip()
    section = (request.GET.get('section') or '').strip()
    event_io = (request.GET.get('event_io') or '').strip().upper()
    if event_io not in ('IN', 'OUT'):
        event_io = None
    valid_year_codes = {code for code, _label in Student.YEAR_LEVEL_CHOICES}
    if year_level not in valid_year_codes:
        year_level = None
    if not section:
        section = None

    export_theme = SiteTheme.objects.first()

    if template_type == 'export_all':
        safe_date = filter_date.strftime('%Y-%m-%d')
        buffer = io.BytesIO()
        with zipfile.ZipFile(buffer, 'w', zipfile.ZIP_DEFLATED) as zf:
            templates_to_export = [
                'overview_summary', 'recent_activity', 'daily_gate_visits', 'inside_now',
                'raw_entries', 'by_personnel',
            ]
            if event_id:
                templates_to_export.append('event_attendance')
            for t in templates_to_export:
                headers, rows = _reports_export_build_data(
                    filter_date, day_start, day_end, t, search_q, event_id,
                    program_course=program_course or None, event_io=event_io,
                    year_level=year_level, section=section,
                    audience_filter=audience_filter,
                    report_timestamp_q=report_timestamp_q,
                    report_checked_in_at_q=report_checked_in_at_q,
                    report_checked_out_at_q=report_checked_out_at_q,
                    report_recorded_at_q=report_recorded_at_q,
                )
                buf = io.StringIO()
                w = csv.writer(buf)
                w.writerow(headers)
                w.writerows(rows)
                _append_report_signatories_csv(w, len(headers), theme=export_theme)
                zf.writestr(f'{t}_{safe_date}.csv', buf.getvalue())
        response = HttpResponse(buffer.getvalue(), content_type='application/zip')
        response['Content-Disposition'] = f'attachment; filename="reports_export_all_{safe_date}.zip"'
        return response

    headers, rows = _reports_export_build_data(
        filter_date, day_start, day_end, template_type, search_q, event_id,
        program_course=program_course or None, event_io=event_io,
        year_level=year_level, section=section,
        audience_filter=audience_filter,
        report_timestamp_q=report_timestamp_q,
        report_checked_in_at_q=report_checked_in_at_q,
        report_checked_out_at_q=report_checked_out_at_q,
        report_recorded_at_q=report_recorded_at_q,
    )
    safe_date = filter_date.strftime('%Y-%m-%d')
    base_name = f'report_{template_type}_{safe_date}'

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{base_name}.csv"'
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(rows)
        _append_report_signatories_csv(writer, len(headers), theme=export_theme)
        return response

    if fmt == 'xlsx':
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            return HttpResponse('Excel export requires openpyxl. Install: pip install openpyxl', status=501)
        wb = Workbook()
        ws = wb.active
        ws.title = 'Export'
        for col, h in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=h).font = Font(bold=True)
        for row_idx, row in enumerate(rows, 2):
            for col_idx, val in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)
        _append_report_signatories_xlsx(ws, len(rows) + 3, theme=export_theme)
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{base_name}.xlsx"'
        wb.save(response)
        return response

    if fmt == 'pdf':
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4, landscape
            from reportlab.lib.styles import getSampleStyleSheet
            from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, KeepTogether
            from reportlab.lib.units import cm
            from reportlab.lib.utils import ImageReader
        except ImportError:
            return HttpResponse('PDF export requires reportlab. Install: pip install reportlab', status=501)
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{base_name}.pdf"'
        # Landscape PDF so tables match the on-screen preview columns.
        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            rightMargin=1.2 * cm,
            leftMargin=1.2 * cm,
            topMargin=3.0 * cm,  # leave room for logo/title header
            bottomMargin=1.0 * cm,
        )

        # ----- Header (logo + title) drawn directly onto the PDF canvas -----
        def _find_logo_path():
            try:
                from django.contrib.staticfiles import finders
                p = finders.find('img/logo.png')
                return p
            except Exception:
                return None

        logo_path = _find_logo_path()
        logo_img = None
        if logo_path:
            try:
                logo_img = ImageReader(logo_path)
            except Exception:
                logo_img = None

        def _draw_header(canvas, doc_obj):
            canvas.saveState()
            page_w, page_h = doc_obj.pagesize
            left = doc_obj.leftMargin
            right = page_w - doc_obj.rightMargin
            top = page_h - doc_obj.topMargin + (2.1 * cm)  # header sits in the reserved top margin

            # Logo (optional)
            logo_size = 1.45 * cm
            # Title + subtitle
            title = getattr(settings, 'SITE_NAME', None) or 'City College of Bayawan'
            report_title = template_type.replace('_', ' ').title()
            subtitle = 'Gate Entry Report' if template_type == 'daily_gate_visits' else (report_title + ' Report')

            # Meta line (date + generated)
            generated = timezone.localtime(timezone.now()).strftime('%Y-%m-%d %H:%M:%S')
            meta = f"Date: {safe_date}  |  Generated: {generated}"

            # Center the whole header block (logo + text) for a cleaner look.
            gap = 0.35 * cm
            try:
                title_w = canvas.stringWidth(str(title), 'Helvetica-Bold', 13)
                sub_w = canvas.stringWidth(str(subtitle), 'Helvetica', 10)
                meta_w = canvas.stringWidth(str(meta), 'Helvetica', 8.5)
            except Exception:
                title_w = sub_w = meta_w = 0
            text_w = max(title_w, sub_w, meta_w)
            group_w = (logo_size + gap + text_w) if logo_img else text_w
            x0 = (page_w - group_w) / 2.0
            if x0 < left:
                x0 = left
            tx = x0 + (logo_size + gap if logo_img else 0)

            if logo_img:
                try:
                    canvas.drawImage(logo_img, x0, top - logo_size, width=logo_size, height=logo_size, preserveAspectRatio=True, mask='auto')
                except Exception:
                    pass

            canvas.setFillColor(colors.HexColor('#14532d'))
            canvas.setFont('Helvetica-Bold', 13)
            canvas.drawString(tx, top - 2, str(title))
            canvas.setFillColor(colors.HexColor('#334155'))
            canvas.setFont('Helvetica', 10)
            canvas.drawString(tx, top - 16, str(subtitle))

            canvas.setFillColor(colors.HexColor('#475569'))
            canvas.setFont('Helvetica', 8.5)
            canvas.drawString(tx, top - 30, meta)

            # Accent rule (leave extra breathing room under header)
            canvas.setStrokeColor(colors.HexColor('#16a34a'))
            canvas.setLineWidth(1)
            canvas.line(left, top - 48, right, top - 48)

            canvas.restoreState()

        # Build readable, full-width table.
        styles = getSampleStyleSheet()
        cell_style = styles['BodyText']
        cell_style.fontName = 'Helvetica'
        cell_style.fontSize = 8.5
        cell_style.leading = 10.5
        head_style = styles['BodyText']
        head_style.fontName = 'Helvetica-Bold'
        head_style.fontSize = 9
        head_style.leading = 11

        def _para(val, st):
            if val is None:
                val = ''
            txt = str(val)
            # Basic XML escaping for Paragraph
            txt = txt.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            return Paragraph(txt, st)

        data = [[_para(h, head_style) for h in headers]]
        for r in rows:
            data.append([_para(v, cell_style) for v in r])

        # Smart column widths: allocate more space to Name/Program/Time columns.
        weights = []
        for h in headers:
            hu = str(h).upper()
            if 'NAME' in hu:
                w = 2.6
            elif 'PROGRAM' in hu or 'SECTION' in hu:
                w = 2.2
            elif 'TIME' in hu or 'DATE' in hu or 'TIMESTAMP' in hu:
                w = 2.0
            elif hu.strip() == 'ID' or 'STUDENT ID' in hu:
                w = 1.4
            else:
                w = 1.6
            weights.append(w)
        total_w = sum(weights) or 1.0
        col_widths = [doc.width * (w / total_w) for w in weights]

        t = Table(data, colWidths=col_widths, repeatRows=1)
        t.setStyle(TableStyle([
            # Clean, print-friendly palette (high contrast but not too dark)
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#eef2f7')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('ALIGN', (0, 0), (-1, 0), 'LEFT'),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 7),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            ('LINEBELOW', (0, 0), (-1, 0), 0.8, colors.HexColor('#94a3b8')),
            ('GRID', (0, 0), (-1, -1), 0.3, colors.HexColor('#cbd5e1')),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8fafc')]),
        ]))

        title = f"{template_type.replace('_', ' ').title()} — {safe_date}"
        subtitle = f"{date_range_label} · Audience: {audience_filter.title() if audience_filter else 'All'}"
        story = [
            # Header is drawn on the canvas; keep a small spacer so the table starts cleanly.
            Spacer(1, 4),
            Paragraph(subtitle.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;'), styles['Normal']),
            Spacer(1, 8),
            t,
        ]
        n1, t1 = _report_signatory_theme_fields(export_theme)
        if n1 or t1:
            from reportlab.platypus import Image as RLImage

            sig_h = styles['Heading4']
            sig_h.fontName = 'Helvetica-Bold'
            sig_h.fontSize = 11
            sig_h.textColor = colors.HexColor('#14532d')
            sig_h.leading = 14
            sig_n = styles['Normal']
            sig_n.fontSize = 9.5
            sig_n.textColor = colors.HexColor('#334155')
            name, title = n1, t1
            img_path = None
            if export_theme:
                f = getattr(export_theme, 'report_first_signatory_signature', None)
                if f and getattr(f, 'name', None):
                    try:
                        pth = f.path
                        if os.path.exists(pth):
                            img_path = pth
                    except Exception:
                        pass
            nm = (name or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            tl = (title or '').replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')
            if nm and tl:
                txt = f'<b>{nm}</b><br/><font size="8" color="#64748b">{tl}</font>'
            elif nm:
                txt = f'<b>{nm}</b>'
            else:
                txt = f'<font size="8" color="#64748b">{tl}</font>'

            # Keep heading + image + name together so page breaks do not split "Signatory" from the name.
            sig_flowables = [
                Spacer(1, 18),
                Paragraph('Signatory', sig_h),
                Spacer(1, 8),
            ]
            if img_path:
                try:
                    sig_flowables.append(RLImage(img_path, width=3.4 * cm, height=1.5 * cm))
                    sig_flowables.append(Spacer(1, 6))
                except Exception:
                    pass
            sig_flowables.append(Paragraph(txt, sig_n))
            sig_flowables.append(Spacer(1, 14))
            story.append(KeepTogether(sig_flowables))
        doc.build(story, onFirstPage=_draw_header, onLaterPages=_draw_header)
        return response

    return redirect('reports-exports')


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def reports_hub(request):
    """Real-time: live occupancy, today's entries, gate status."""
    today = timezone.localdate()
    day_start, day_end = _local_day_bounds(today)
    # Today's gate entries (count visits so IN+OUT = 1)
    entries_today = GateEntry.objects.filter(timestamp__gte=day_start, timestamp__lt=day_end)
    granted_today = _granted_visits_count_for_date(today, daily_gate_only=True)
    denied_entries_count = entries_today.filter(granted=False).count()
    denied_today = denied_entries_count  # Incidents feature removed
    # Today's event scans (AttendanceLog, non-voided)
    scans_today = AttendanceLog.objects.filter(scan_time__gte=day_start, scan_time__lt=day_end, voided=False)
    success_today = scans_today.filter(result='SUCCESS').count()
    # Live occupancy: events active today with currently inside count
    active_events = Event.objects.filter(status='active', start_date__lte=today, end_date__gte=today)
    live_occupancy = []
    for evt in active_events[:10]:
        inside = EventAttendance.objects.filter(
            event=evt, checked_in_at__isnull=False, checked_out_at__isnull=True
        ).count()
        live_occupancy.append({'event': evt, 'currently_inside': inside})
    # Gate status: registered devices with online/offline (last_seen within 10 min = online)
    device_list = []
    from datetime import timedelta
    threshold = timezone.now() - timedelta(minutes=10)
    for d in ScannerDevice.objects.filter(is_active=True).order_by('name', 'device_id'):
        device_list.append({
            'device': d,
            'last_seen_at': d.last_seen_at,
            'is_online': d.last_seen_at >= threshold if d.last_seen_at else False,
        })
    # Latest generated reports (so hub and Generated Reports page show the same data)
    reports = GeneratedReport.objects.select_related('generated_by').order_by('-generated_at')[:50]
    return render(request, 'gate/reports_hub.html', {
        'site_name': 'City College of Bayawan',
        'today': today,
        'granted_today': granted_today,
        'denied_today': denied_today,
        'success_today': success_today,
        'live_occupancy': live_occupancy,
        'devices': device_list,
        'reports': reports,
    })


@login_required(login_url='/login/')
def api_notification_count(request):
    """JSON endpoint for polling unread notification count (used for notification sound)."""
    from gate_analytics.context_processors import notifications_context
    ctx = notifications_context(request)
    return JsonResponse({'unread': ctx.get('unread_notifications_count', 0)})


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
@require_POST
def notifications_mark_all_read(request):
    """AJAX endpoint: mark all navbar notifications as read for admin/staff/faculty/SAS."""
    from gate.models import NotificationRead, Student, Event
    from gate_analytics.roles import get_user_role
    from datetime import timedelta

    role = get_user_role(request.user)
    keys = []

    AdminNotification.objects.filter(
        target_user=request.user,
        is_read=False,
    ).update(is_read=True, read_at=timezone.now())

    if role == 'admin':
        keys.append('notif_pending_students')
        for pk in Student.objects.filter(
            account_status=Student.ACCOUNT_STATUS_APPROVED
        ).values_list('pk', flat=True)[:50]:
            keys.append(f'notif_student_{pk}')

    if role in ('admin', 'staff'):
        from django.contrib.auth import get_user_model
        from django.db.models import Q
        User = get_user_model()
        keys.append('notif_pending_staff_personnel')
        for pk in User.objects.filter(is_active=False).filter(
            Q(groups__name__iexact='staff') |
            Q(groups__name__iexact='faculty')
        ).distinct().values_list('pk', flat=True)[:50]:
            keys.append(f'notif_staff_personnel_{pk}')

    if role in ('admin', 'staff', 'faculty'):
        today = timezone.localdate()
        until = today + timedelta(days=30)
        keys.append('notif_upcoming_events')
        keys.append('notif_new_events')
        now = timezone.now()
        new_since = now - timedelta(days=7)
        event_pks = set(
            Event.objects.filter(
                start_date__gte=today, start_date__lte=until,
                status__in=('scheduled', 'active'),
            ).values_list('pk', flat=True)[:50]
        ) | set(
            Event.objects.filter(
                created_date__gte=new_since
            ).values_list('pk', flat=True)[:50]
        )
        for pk in event_pks:
            keys.append(f'notif_event_{pk}')

    created = 0
    for key in keys:
        _, was_created = NotificationRead.objects.get_or_create(
            user=request.user, notification_key=key,
        )
        if was_created:
            created += 1

    return JsonResponse({'success': True, 'marked': created})


def _notification_message_for_nav_dropdown(message):
    """
    Compact body text for the bell history list: remove path/URL lines and other
    noise so each row fits without an inner scrollbar (row link covers navigation).
    """
    if not message or not str(message).strip():
        return ''
    kept = []
    for raw in str(message).splitlines():
        ln = raw.strip()
        if not ln:
            continue
        low = ln.lower()
        if ln.startswith('/'):
            continue
        if re.match(r'^https?://\S+$', ln, re.I):
            continue
        if re.search(r':\s*/gate/', low):
            continue
        if re.search(r':\s*/\S*/edit', low):
            continue
        if 'student (app)' in low and '/gate' in ln:
            continue
        if re.match(r'^incidents?\s*:', low) and '/' in ln:
            continue
        if low.startswith('recorded by:'):
            continue
        if re.fullmatch(r'^\([\w\-]+\)$', ln):
            continue
        kept.append(ln)
    if not kept:
        collapsed = ' '.join(str(message).split())
        return (collapsed[:200] + '…') if len(collapsed) > 200 else collapsed
    text = ' '.join(kept)
    if len(text) > 220:
        text = text[:217].rsplit(' ', 1)[0] + '…'
    return text


def _admin_notification_history_queryset(user):
    """In-app AdminNotification rows for notification history (navbar + API)."""
    role = get_user_role(user)
    qs = (
        AdminNotification.objects.filter(target_user=user)
        .filter(Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now()))
        .select_related('related_student', 'related_incident')
        .order_by('-created_at')
    )
    if role in ('staff', 'faculty'):
        qs = qs.exclude(notification_type='incident').exclude(
            notification_type__in=(
                'sas_inactive_ready_activation',
                'sas_verified_gate_followup',
                'gate_manual_referral',
            )
        )
    elif role == 'student affairs':
        qs = qs.exclude(notification_type='staff_personnel_registration')
    return qs


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
@require_GET
def notifications_history_api(request):
    """JSON pages of AdminNotification history for the navbar (no full page load)."""
    from django.core.paginator import Paginator
    from django.utils.formats import date_format
    from django.utils.translation import gettext as _

    qs = _admin_notification_history_queryset(request.user)
    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    items = []
    for n in page_obj:
        href = ''
        link_label = ''
        if n.related_student_id:
            try:
                href = reverse('gate-student-edit', kwargs={'pk': n.related_student_id})
                link_label = _('Student')
            except Exception:
                pass
        if not href and n.related_incident_id:
            try:
                href = f"{reverse('gate-incident-list')}?highlight={n.related_incident_id}"
                link_label = _('Incident')
            except Exception:
                pass
        raw_msg = n.message or ''
        items.append(
            {
                'id': n.pk,
                'created_display': date_format(
                    timezone.localtime(n.created_at), 'SHORT_DATETIME_FORMAT'
                ),
                'type_label': n.get_notification_type_display(),
                'title': n.title,
                'message': _notification_message_for_nav_dropdown(raw_msg),
                'message_tooltip': raw_msg[:800] + ('…' if len(raw_msg) > 800 else ''),
                'is_read': n.is_read,
                'href': href,
                'link_label': link_label,
            }
        )
    return JsonResponse(
        {
            'items': items,
            'page': page_obj.number,
            'num_pages': paginator.num_pages,
            'has_next': page_obj.has_next(),
        }
    )


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty', 'student affairs')
def check_new_admin_notifications_api_view(request):
    """
    AJAX endpoint to check for new admin notifications since a given timestamp.
    """
    from gate.admin_notification_service import AdminNotificationService
    from datetime import timedelta
    
    since_str = request.GET.get('since', '')
    
    try:
        # Parse the since timestamp
        if since_str:
            since = datetime.datetime.fromisoformat(since_str.replace('Z', '+00:00'))
        else:
            # Default to last 1 minute
            since = timezone.now() - timedelta(minutes=1)
    except (ValueError, AttributeError):
        since = timezone.now() - timedelta(minutes=1)
    
    # Get unread notifications created after 'since'
    notifications = AdminNotification.objects.filter(
        target_user=request.user,
        is_read=False,
        created_at__gt=since
    ).filter(
        Q(expires_at__isnull=True) | Q(expires_at__gt=timezone.now())
    )
    _api_role = get_user_role(request.user)
    if _api_role in ('staff', 'faculty'):
        notifications = notifications.exclude(notification_type='incident')
    if _api_role != 'admin':
        notifications = notifications.exclude(notification_type='sas_inactive_ready_activation')
        notifications = notifications.exclude(notification_type='sas_verified_gate_followup')
        notifications = notifications.exclude(notification_type='gate_manual_referral')
    if _api_role == 'student affairs':
        notifications = notifications.exclude(notification_type='staff_personnel_registration')
    notifications = notifications.order_by('-created_at')[:5]

    # Format notifications for JSON
    notification_list = []
    for notif in notifications:
        notification_list.append({
            'id': notif.id,
            'title': notif.title,
            'message': notif.message,
            'priority': notif.priority,
            'notification_type': notif.notification_type,
            'created_at': notif.created_at.isoformat(),
        })
    
    return JsonResponse({
        'success': True,
        'notifications': notification_list
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def report_list(request):
    """List generated reports (daily/weekly/monthly) with download."""
    reports_qs = GeneratedReport.objects.select_related('generated_by').order_by('-generated_at')
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    from django.core.paginator import Paginator
    paginator = Paginator(reports_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    reports = list(page_obj.object_list)
    return render(request, 'gate/report_list.html', {
        'site_name': 'City College of Bayawan',
        'reports': reports,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def report_download(request, pk):
    """Download a generated report file."""
    report = get_object_or_404(GeneratedReport, pk=pk)
    if not report.file:
        messages.warning(request, 'No file attached for this report.')
        return redirect('report-list')
    try:
        path = report.file.path
    except Exception:
        path = None
    if path and os.path.exists(path):
        with open(path, 'rb') as f:
            response = HttpResponse(f.read(), content_type='application/octet-stream')
            response['Content-Disposition'] = f'attachment; filename="{os.path.basename(path)}"'
            return response
    # Fallback: redirect to file URL if storage is remote
    if report.file:
        from django.http import HttpResponseRedirect
        return HttpResponseRedirect(report.file.url)
    messages.error(request, 'File not found.')
    return redirect('report-list')


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def report_on_demand_view(request, pk):
    """Re-display an on-demand report that has no file (was viewed as HTML)."""
    import json
    report = get_object_or_404(GeneratedReport, pk=pk)
    if report.report_type != 'on_demand' or not report.period_start or not report.period_end:
        messages.warning(request, 'Cannot view this report.')
        return redirect('report-list')
    try:
        summary = json.loads(report.summary) if report.summary else {}
    except (ValueError, TypeError):
        summary = {}
    group_by = summary.get('group_by', 'student')
    date_from = report.period_start
    date_to = report.period_end
    data = _build_on_demand_data(date_from, date_to, group_by, request)
    row_keys = data.get('row_keys') or []
    row_values = [[r.get(k, '') for k in row_keys] for r in data.get('rows', [])]
    return render(request, 'gate/report_on_demand_result.html', {
        'site_name': 'City College of Bayawan',
        'date_from': date_from,
        'date_to': date_to,
        'group_by': group_by,
        'data': data,
        'row_values': row_values,
    })


def _build_on_demand_data(date_from, date_to, group_by, request):
    """Build report data for on-demand: by student, gate, time_window, or course_section."""
    from django.db.models import Count
    date_from = timezone.make_aware(datetime.datetime.combine(date_from, datetime.time.min), timezone.get_current_timezone())
    date_to = timezone.make_aware(datetime.datetime.combine(date_to, datetime.time.max), timezone.get_current_timezone())
    logs = AttendanceLog.objects.filter(
        scan_time__gte=date_from, scan_time__lte=date_to, voided=False
    ).select_related('student', 'event')
    entries = GateEntry.objects.filter(
        timestamp__gte=date_from, timestamp__lte=date_to
    ).select_related('student')
    if group_by == 'student':
        # By student: count scans and gate entries
        student_ids = set()
        for log in logs:
            if log.student_id:
                student_ids.add(log.student_id)
        for e in entries:
            student_ids.add(e.student_id)
        from .models import Student
        students = Student.objects.filter(id__in=student_ids).order_by('last_name', 'first_name')
        rows = []
        for s in students:
            scan_count = logs.filter(student=s).count()
            entry_count = entries.filter(student=s).count()
            # Program/Section: use course_or_section if set, else derive from course + section
            course_section = (s.course_or_section or '').strip()
            if not course_section:
                parts = []
                if getattr(s, 'course', None):
                    parts.append(s.get_course_display() if hasattr(s, 'get_course_display') else s.course)
                if getattr(s, 'section', None) and (s.section or '').strip():
                    parts.append((s.section or '').strip())
                course_section = ' - '.join(parts) if parts else '—'
            rows.append({
                'student_id': s.student_id,
                'name': s.get_full_name(),
                'sex': s.get_sex_display() if getattr(s, 'sex', '') else '—',
                'year_level': s.year_level or '—',
                'course_or_section': course_section,
                'scan_count': scan_count,
                'entry_count': entry_count,
            })
        return {'group_by': 'student', 'rows': rows, 'headers': ['Student ID', 'Name', 'Sex/Gender', 'Year Level', 'Program/Section', 'Event scans', 'Gate entries'],
                'row_keys': ['student_id', 'name', 'sex', 'year_level', 'course_or_section', 'scan_count', 'entry_count']}
    if group_by == 'gate':
        # By device_id
        gate_rows = logs.values('device_id').annotate(count=Count('id')).order_by('-count')
        rows = [{'device_id': r['device_id'] or '(unknown)', 'count': r['count']} for r in gate_rows]
        return {'group_by': 'gate', 'rows': rows, 'headers': ['Device / Gate', 'Scan count'], 'row_keys': ['device_id', 'count']}
    if group_by == 'time_window':
        # By hour (using Python to avoid MySQL timezone issues)
        from collections import defaultdict
        hour_counts_dict = defaultdict(int)
        for log in logs:
            if log.scan_time:
                local_dt = timezone.localtime(log.scan_time)
                hour_counts_dict[local_dt.hour] += 1
        rows = [{'hour': f"{h}:00", 'count': c} for h, c in sorted(hour_counts_dict.items())]
        return {'group_by': 'time_window', 'rows': rows, 'headers': ['Hour', 'Scan count'], 'row_keys': ['hour', 'count']}
    if group_by == 'course_section':
        from .models import Student
        section_counts = {}
        for log in logs:
            if not log.student_id:
                continue
            try:
                s = Student.objects.get(pk=log.student_id)
            except Student.DoesNotExist:
                continue
            course_section = (s.course_or_section or '').strip()
            if not course_section:
                parts = []
                if getattr(s, 'course', None):
                    parts.append(s.get_course_display() if hasattr(s, 'get_course_display') else s.course)
                if getattr(s, 'section', None) and (s.section or '').strip():
                    parts.append((s.section or '').strip())
                course_section = ' - '.join(parts) if parts else '—'
            key = (course_section, s.year_level or '—')
            section_counts[key] = section_counts.get(key, 0) + 1
        rows = [{'course_section': k[0], 'year_level': k[1], 'count': v} for k, v in sorted(section_counts.items())]
        return {'group_by': 'course_section', 'rows': rows, 'headers': ['Program/Section', 'Year Level', 'Scan count'], 'row_keys': ['course_section', 'year_level', 'count']}
    return {'group_by': None, 'rows': [], 'headers': [], 'row_keys': []}


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def report_on_demand(request):
    """On-demand report: admin chooses date range and group by (student, gate, time window, course/section)."""
    if request.method == 'POST':
        import json
        from django.utils.dateparse import parse_date
        from django.core.files.base import ContentFile
        date_from = parse_date(request.POST.get('date_from') or '')
        date_to = parse_date(request.POST.get('date_to') or '')
        group_by = request.POST.get('group_by') or 'student'
        if not date_from or not date_to or date_from > date_to:
            messages.error(request, 'Please select a valid date range.')
            return redirect('report-on-demand')
        data = _build_on_demand_data(date_from, date_to, group_by, request)
        format = request.POST.get('format') or 'html'
        row_keys = data.get('row_keys') or []
        rows = data.get('rows', [])
        # Create GeneratedReport so it appears in Generated Reports list
        group_by_label = {'student': 'student', 'gate': 'gate', 'time_window': 'time window', 'course_section': 'program/section'}.get(group_by, group_by)
        title = f"On-demand: {date_from.strftime('%b %d, %Y')} – {date_to.strftime('%b %d, %Y')} (by {group_by_label})"
        summary_dict = {'group_by': group_by, 'row_count': len(rows)}
        report = None
        try:
            report = GeneratedReport.objects.create(
                report_type='on_demand',
                period_start=date_from,
                period_end=date_to,
                title=title,
                summary=json.dumps(summary_dict, default=str),
                generated_by=request.user,
            )
        except Exception as e:
            import logging
            logging.getLogger(__name__).exception('Failed to save on-demand report to GeneratedReport: %s', e)
        if format == 'csv':
            from io import StringIO
            buf = StringIO()
            w = csv.writer(buf)
            w.writerow(data['headers'])
            for row in rows:
                w.writerow([row.get(k, '') for k in row_keys])
            csv_content = buf.getvalue().encode('utf-8')
            if report:
                report.file.save(
                    f'report_{date_from}_{date_to}_{group_by}.csv',
                    ContentFile(csv_content),
                    save=True,
                )
            response = HttpResponse(csv_content, content_type='text/csv')
            response['Content-Disposition'] = f'attachment; filename="report_{date_from}_{date_to}_{group_by}.csv"'
            return response
        if format == 'xlsx':
            try:
                from openpyxl import Workbook
                from openpyxl.styles import Font
                from io import BytesIO
            except ImportError:
                messages.error(request, 'Excel export requires openpyxl.')
                return redirect('report-on-demand')
            wb = Workbook()
            ws = wb.active
            ws.title = 'Report'
            for col, h in enumerate(data['headers'], 1):
                ws.cell(row=1, column=col, value=h).font = Font(bold=True)
            for row_idx, row in enumerate(rows, 2):
                for col_idx, k in enumerate(row_keys, 1):
                    ws.cell(row=row_idx, column=col_idx, value=row.get(k, ''))
            xlsx_io = BytesIO()
            wb.save(xlsx_io)
            xlsx_io.seek(0)
            if report:
                report.file.save(
                f'report_{date_from}_{date_to}_{group_by}.xlsx',
                ContentFile(xlsx_io.read()),
                save=True,
            )
            xlsx_io.seek(0)
            response = HttpResponse(xlsx_io.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = f'attachment; filename="report_{date_from}_{date_to}_{group_by}.xlsx"'
            return response
        if report:
            messages.success(request, 'Report saved to Generated Reports. You can open it from Reports hub or Generated reports.')
        row_values = [[r.get(k, '') for k in row_keys] for r in rows]
        return render(request, 'gate/report_on_demand_result.html', {
            'site_name': 'City College of Bayawan',
            'date_from': date_from,
            'date_to': date_to,
            'group_by': group_by,
            'data': data,
            'row_values': row_values,
        })
    today = timezone.localdate()
    return render(request, 'gate/report_on_demand.html', {
        'site_name': 'City College of Bayawan',
        'default_from': today - datetime.timedelta(days=7),
        'default_to': today,
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_expected_today(request, event_id):
    """List target-audience students who have not checked in yet (expected today)."""
    event = get_object_or_404(Event, id=event_id)
    eligible_students = _event_audience_students_qs(event)
    checked_in_ids = set(
        EventAttendance.objects.filter(
            event=event, checked_in_at__isnull=False
        ).values_list('student_id', flat=True)
    )
    expected_students = eligible_students.exclude(id__in=checked_in_ids).order_by('last_name', 'first_name')[:500]
    # Keep template compatibility (expects each row to have .student)
    expected = [SimpleNamespace(student=s) for s in expected_students]
    return render(request, 'gate/event_expected_today.html', {
        'site_name': 'City College of Bayawan',
        'event': event,
        'expected': expected,
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def event_manual_checkin(request, event_id):
    """Manual check-in: search by student ID or name and mark present."""
    event = get_object_or_404(Event, id=event_id)
    if request.method == 'POST':
        student_id = (request.POST.get('student_id') or '').strip()
        if not student_id:
            messages.error(request, 'Enter student ID.')
            return redirect('event-manual-checkin', event_id=event_id)
        try:
            student = Student.objects.get(student_id=student_id, is_active=True)
        except Student.DoesNotExist:
            messages.error(request, f'Student {student_id} not found or inactive.')
            return redirect('event-manual-checkin', event_id=event_id)

        allowed = _is_student_allowed_for_event(event, student)
        override = _should_override_audience(request)
        audience_overridden = False
        if not allowed:
            if override and _can_override_audience(request.user):
                audience_overridden = True
            else:
                messages.error(
                    request,
                    f'{student.get_full_name()} is not included in this event audience ({event.audience_summary()}).'
                )
                return redirect('event-manual-checkin', event_id=event_id)

        att, _ = EventAttendance.objects.get_or_create(student=student, event=event, defaults={'participated': False})
        was_already_checked_in = att.checked_in_at is not None
        if att.checked_in_at is None:
            att.checked_in_at = timezone.now()
            att.save(update_fields=['checked_in_at'])
            from .audit import log_action
            desc = f'{student_id} @ {event.name}'
            if audience_overridden:
                desc += ' (audience override)'
            log_action(request, 'manual_checkin', 'EventAttendance', object_id=att.pk, description=desc)
            if audience_overridden:
                messages.warning(request, f'{student.get_full_name()} marked checked in (override: outside target audience).')
            else:
                messages.success(request, f'{student.get_full_name()} marked checked in.')
        else:
            messages.info(request, f'{student.get_full_name()} was already checked in.')

        # Log into AttendanceLog so overrides appear in event attendance report
        result_code = 'DUPLICATE' if was_already_checked_in else 'SUCCESS'
        _create_event_log_single_duplicate(
            event=event,
            student=student,
            scan_type='IN',
            result=result_code,
            registration=None,
            token='',
            device_id='manual_checkin',
            remarks='AUDIENCE_OVERRIDE' if audience_overridden else '',
            recorded_by=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
        )
        return redirect('event-manual-checkin', event_id=event_id)
    q = (request.GET.get('q') or '').strip()
    attendances = EventAttendance.objects.filter(event=event).select_related('student').order_by('student__last_name', 'student__first_name')
    if q:
        attendances = attendances.filter(
            Q(student__student_id__icontains=q) |
            Q(student__first_name__icontains=q) |
            Q(student__last_name__icontains=q)
        )
    return render(request, 'gate/event_manual_checkin.html', {
        'site_name': 'City College of Bayawan',
        'event': event,
        'attendances': attendances[:100],
        'search_query': q,
    })


@require_GET
@ensure_csrf_cookie
def event_attendance_live_embed(request, event_id):
    """Embeddable list of event attendance scan logs (AttendanceLog), not daily GateEntry rows."""
    from django.contrib.auth.views import redirect_to_login
    from gate_analytics.roles import get_user_role
    if _guard_embed_query_token_ok(request):
        user_role = 'staff'
    else:
        if not request.user.is_authenticated:
            return redirect_to_login(next=request.get_full_path())
        user_role = get_user_role(request.user)
        if user_role not in ('admin', 'staff', 'faculty', 'student affairs'):
            return HttpResponseForbidden('Access denied')
    event = get_object_or_404(Event, id=event_id)
    from_date = (request.GET.get('from_date') or '').strip()
    if not from_date:
        from_date = timezone.localdate().isoformat()
    try:
        filter_date = datetime.date.fromisoformat(from_date)
    except ValueError:
        filter_date = timezone.localdate()
    day_start, day_end = _local_day_bounds(filter_date)
    logs = (
        AttendanceLog.objects.filter(
            event=event,
            voided=False,
            scan_time__gte=day_start,
            scan_time__lt=day_end,
        )
        .select_related('student')
        .order_by('-scan_time')[:200]
    )
    return render(request, 'gate/event_attendance_live_embed.html', {
        'event': event,
        'filter_date': filter_date,
        'logs': logs,
        'embed': (request.GET.get('embed') or '').strip().lower() in ('1', 'true', 'yes'),
    })


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
@ensure_csrf_cookie
def field_trip_event_scan(request, event_id):
    """
    Event attendance POSTs: scan student permanent QR (eEID) for field-trip / off-campus events.
    GET serves the same scanner UI as /gate/ with this event locked (no daily gate entries).
    Records EventAttendance only (no GateEntry).
    """
    event = get_object_or_404(Event, id=event_id)
    if request.method == 'GET':
        if getattr(event, 'event_location', '') != 'field_trip':
            return _redirect_gate_scan_with_event(request, event_id)
        user_role = get_user_role(request.user)
        kiosk_mode = (request.GET.get('kiosk') or '').strip().lower() in ('1', 'true', 'yes', 'on')
        return render(request, 'gate/gate_scan.html', {
            'site_name': getattr(settings, 'SITE_NAME', 'City College of Bayawan'),
            'page_title': f'Event attendance – {event.name}',
            'active_events': [],
            'event_attendance_scanner_event': event,
            'event_attendance_field_trip': True,
            'campus_departments': CAMPUS_DEPARTMENT_CHOICES,
            'user_role': user_role,
            'kiosk_mode': kiosk_mode,
            'guard_student_popup_style': getattr(settings, 'GATE_GUARD_STUDENT_POPUP_STYLE', 'split'),
        })

    today = timezone.localdate()
    if event.start_date > today or event.end_date < today:
        messages.warning(request, f'Event is not scheduled for today ({today}). You can still record attendance.')

    if request.method == 'POST':
        student_id = (request.POST.get('student_id') or request.POST.get('qr') or '').strip()
        requested_scan_type = (request.POST.get('scan_type') or 'IN').strip().upper()
        if requested_scan_type not in ('IN', 'OUT'):
            requested_scan_type = 'IN'
        if not student_id:
            if _http_request_wants_json(request):
                return JsonResponse({'success': False, 'message': 'No student ID provided.', 'color': 'error'})
            messages.error(request, 'Enter or scan student ID.')
            return redirect('event-field-trip-scan', event_id=event_id)

        student = Student.objects.filter(student_id=student_id, is_active=True).first()
        if not student:
            if _http_request_wants_json(request):
                return JsonResponse({
                    'success': False,
                    'message': 'Student not found or inactive.',
                    'student_id': student_id,
                    'color': 'error',
                })
            messages.error(request, f'Student {student_id} not found or inactive.')
            return redirect('event-field-trip-scan', event_id=event_id)

        allowed = _is_student_allowed_for_event(event, student)
        override = _should_override_audience(request)
        audience_overridden = False
        if not allowed:
            if override and _can_override_audience(request.user):
                audience_overridden = True
            else:
                message = f'{student.get_full_name()} is not included in this event audience ({event.audience_summary()}).'
                if _http_request_wants_json(request):
                    return JsonResponse({
                        'success': False,
                        'message': message,
                        'student_id': student.student_id,
                        'not_allowed': True,
                        'color': 'warning',
                    })
                messages.error(request, message)
                return redirect('event-field-trip-scan', event_id=event_id)

        # Determine current attendance record (if any); create placeholder when first scanned
        att, _ = EventAttendance.objects.get_or_create(
            student=student,
            event=event,
            defaults={'participated': False},
        )

        # Determine effective scan type based on existing attendance:
        #  - First-ever scan: always IN (check-in), even if OUT was selected.
        #  - Already checked in but not yet checked out: treat as OUT (check-out),
        #    even if staff accidentally left mode on IN.
        #  - Already checked out: keep OUT and simply inform "already checked out".
        is_first_scan = att.checked_in_at is None and att.checked_out_at is None
        if is_first_scan:
            scan_type = 'IN'
        elif att.checked_in_at is not None and att.checked_out_at is None:
            scan_type = 'OUT'
        else:
            scan_type = 'OUT'

        # For OUT scans, deny exit while event is not finished — but only when
        # the student hasn't checked out yet. If already checked out, fall
        # through so the "already checked out" message is shown instead.
        if scan_type == 'OUT' and att.checked_out_at is None and event.status in ('draft', 'scheduled', 'active'):
            msg = 'Event is not yet finished. Check-out will be allowed once the event is completed.'
            if _http_request_wants_json(request):
                return JsonResponse({
                    'success': False,
                    'message': msg,
                    'event_status': event.status,
                    'color': 'warning',
                }, status=400)
            messages.error(request, msg)
            return redirect('event-field-trip-scan', event_id=event_id)

        # OUT: require an existing check-in before allowing check-out
        if scan_type == 'OUT' and att.checked_in_at is None:
            # Log as NOT_CHECKED_IN so report shows failed attempt
            AttendanceLog.objects.create(
                event=event,
                student=student,
                registration=None,
                scan_type='OUT',
                result='NOT_CHECKED_IN',
                token='',
                device_id=request.POST.get('device_id', '')[:64],
                remarks='CHECKOUT_BEFORE_CHECKIN',
                recorded_by=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
            )
            msg = f'{student.get_full_name()} has not checked in yet for this event.'
            if _http_request_wants_json(request):
                return JsonResponse({
                    'success': False,
                    'message': msg,
                    'student_id': student.student_id,
                    'color': 'warning',
                })
            messages.error(request, msg)
            return redirect('event-field-trip-scan', event_id=event_id)

        already_checked_in = att.checked_in_at is not None
        already_checked_out = att.checked_out_at is not None

        if scan_type == 'IN':
            if not att.checked_in_at:
                att.checked_in_at = timezone.now()
            att.participated = True
            att.save(update_fields=['checked_in_at', 'participated'])
        else:  # OUT
            if not att.checked_out_at:
                att.checked_out_at = timezone.now()
            att.participated = True
            att.save(update_fields=['checked_out_at', 'participated'])

        if audience_overridden:
            from .audit import log_action
            log_action(
                request,
                'event_attendance_override',
                'EventAttendance',
                object_id=att.pk,
                description=f'{student.student_id} @ {event.name} (audience override)'
            )

        # Log into AttendanceLog so overrides appear in event attendance report.
        # For event scanner we do NOT record DUPLICATE rows – re-scans either
        # become OUT attempts or are blocked with a message.
        result_code = 'SUCCESS'
        _create_event_log_single_duplicate(
            event=event,
            student=student,
            scan_type=scan_type,
            result=result_code,
            registration=None,
            token='',
            device_id=request.POST.get('device_id', '')[:64],
            remarks='AUDIENCE_OVERRIDE' if audience_overridden else '',
            recorded_by=request.user if getattr(request, 'user', None) and request.user.is_authenticated else None,
        )

        if _http_request_wants_json(request):
            photo_url = _scan_ui_photo_url(request, student)
            if scan_type == 'IN':
                msg = (
                    f'{student.get_full_name()} checked in.'
                    if not already_checked_in else
                    f'{student.get_full_name()} already checked in.'
                )
            else:
                msg = (
                    f'{student.get_full_name()} checked out.'
                    if not already_checked_out else
                    f'{student.get_full_name()} already checked out.'
                )
            if audience_overridden:
                msg += ' (Override: outside target audience.)'
            time_str = timezone.localtime(timezone.now()).strftime('%I:%M %p')
            return JsonResponse({
                'success': True,
                'ok': True,
                'result': 'ALLOWED',
                'message': msg,
                'color': 'success',
                'status': scan_type,
                'scan_type': scan_type,
                'time': time_str,
                'student_name': student.get_full_name(),
                'student_id': student.student_id,
                'already_checked_in': already_checked_in,
                'already_checked_out': already_checked_out,
                'photo_url': photo_url or '',
                'audience_overridden': audience_overridden,
                'student': {
                    'student_id': student.student_id,
                    'name': student.get_full_name(),
                    'first_name': student.first_name,
                    'middle_name': student.middle_name or '',
                    'last_name': student.last_name,
                    'email': student.email or '',
                    'photo_url': photo_url or '',
                    'course_or_section': getattr(student, 'course_or_section', '') or '',
                    'year_level': getattr(student, 'year_level', '') or '',
                },
            })

        if scan_type == 'IN':
            if already_checked_in:
                messages.info(request, f'{student.get_full_name()} was already checked in.')
            else:
                if audience_overridden:
                    messages.warning(request, f'{student.get_full_name()} checked in (override: outside target audience).')
                else:
                    messages.success(request, f'{student.get_full_name()} checked in.')
        else:
            if already_checked_out:
                messages.info(request, f'{student.get_full_name()} was already checked out.')
            else:
                messages.success(request, f'{student.get_full_name()} checked out.')
        return redirect('event-field-trip-scan', event_id=event_id)


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'faculty')
def report_compare_events(request):
    """Compare two events side-by-side (attendance, check-in count, etc.)."""
    event_a_id = request.GET.get('event_a')
    event_b_id = request.GET.get('event_b')
    event_a = Event.objects.filter(id=event_a_id).first() if event_a_id else None
    event_b = Event.objects.filter(id=event_b_id).first() if event_b_id else None
    stats_a = stats_b = None
    if event_a:
        regs_a = EventRegistration.objects.filter(event=event_a).count()
        checked_a = EventAttendance.objects.filter(event=event_a, checked_in_at__isnull=False).count()
        inside_a = EventAttendance.objects.filter(event=event_a, checked_in_at__isnull=False, checked_out_at__isnull=True).count()
        stats_a = {'registered': regs_a, 'checked_in': checked_a, 'currently_inside': inside_a}
    if event_b:
        regs_b = EventRegistration.objects.filter(event=event_b).count()
        checked_b = EventAttendance.objects.filter(event=event_b, checked_in_at__isnull=False).count()
        inside_b = EventAttendance.objects.filter(event=event_b, checked_in_at__isnull=False, checked_out_at__isnull=True).count()
        stats_b = {'registered': regs_b, 'checked_in': checked_b, 'currently_inside': inside_b}
    events_list = Event.objects.filter(status__in=('active', 'completed', 'scheduled')).order_by('-start_date')[:200]
    return render(request, 'gate/report_compare_events.html', {
        'site_name': 'City College of Bayawan',
        'event_a': event_a,
        'event_b': event_b,
        'stats_a': stats_a,
        'stats_b': stats_b,
        'events_list': events_list,
    })


@login_required(login_url='/login/')
@role_required('admin')
def audit_log_viewer(request):
    """Admin view: list audit logs (logins and other actions). Filter by a single date for day-to-day view."""
    from .models import AuditLog
    from django.utils.dateparse import parse_date
    logs_qs = AuditLog.objects.select_related('user').order_by('-created_at')
    action_filter = (request.GET.get('action') or '').strip().lower()
    if action_filter == 'login':
        logs_qs = logs_qs.filter(action='login')
    # Single date filter (YYYY-MM-DD) – show logs for that day only
    date_param = request.GET.get('date', '').strip()
    date_parsed = parse_date(date_param) if date_param else None
    if date_parsed:
        day_start, day_end = _local_day_bounds(date_parsed)
        logs_qs = logs_qs.filter(created_at__gte=day_start, created_at__lt=day_end)
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    from django.core.paginator import Paginator
    paginator = Paginator(logs_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    logs = list(page_obj.object_list)
    all_qs = AuditLog.objects.select_related('user')
    today_start, today_end = _local_day_bounds(timezone.localdate())
    today_qs = all_qs.filter(created_at__gte=today_start, created_at__lt=today_end)
    return render(request, 'gate/audit_log_viewer.html', {
        'site_name': 'City College of Bayawan',
        'logs': logs,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'action_filter': action_filter,
        'filter_date': date_param,
        'total_logs': all_qs.count(),
        'today_logs': today_qs.count(),
        'today_logins': today_qs.filter(action='login').count(),
        'unique_users_today': today_qs.values('user').distinct().count(),
        'blocked_ips': list(BlockedIP.objects.order_by('-blocked_at')[:100]),
        'blocked_count': BlockedIP.objects.filter(is_active=True).count(),
    })


@login_required(login_url='/login/')
@role_required('admin')
@require_POST
def block_ip(request):
    """Block an IP address from the audit log."""
    from .models import BlockedIP
    ip = request.POST.get('ip_address', '').strip()
    reason = request.POST.get('reason', '').strip()
    if not ip:
        messages.error(request, 'No IP address provided.')
        return redirect('audit-log-viewer')
    if ip in ('127.0.0.1', '::1', 'localhost'):
        messages.warning(request, 'Cannot block localhost.')
        return redirect('audit-log-viewer')
    obj, created = BlockedIP.objects.get_or_create(
        ip_address=ip,
        defaults={'reason': reason or 'Blocked from audit log', 'blocked_by': request.user, 'is_active': True},
    )
    if not created and not obj.is_active:
        obj.is_active = True
        obj.reason = reason or obj.reason or 'Re-blocked from audit log'
        obj.blocked_by = request.user
        obj.save(update_fields=['is_active', 'reason', 'blocked_by'])
    AuditLog.objects.create(
        user=request.user,
        action='block_ip',
        description=f'Blocked IP {ip}: {reason or "no reason given"}',
        ip_address=_get_client_ip(request),
    )
    from gate_analytics.middleware import BlockedIPMiddleware
    BlockedIPMiddleware.clear_cache()
    messages.success(request, f'IP {ip} has been blocked.')
    return redirect('audit-log-viewer')


@login_required(login_url='/login/')
@role_required('admin')
@require_POST
def unblock_ip(request):
    """Unblock an IP address."""
    from .models import BlockedIP
    ip_id = request.POST.get('ip_id', '').strip()
    try:
        obj = BlockedIP.objects.get(pk=ip_id)
        obj.is_active = False
        obj.save(update_fields=['is_active'])
        AuditLog.objects.create(
            user=request.user,
            action='unblock_ip',
            description=f'Unblocked IP {obj.ip_address}',
            ip_address=_get_client_ip(request),
        )
        from gate_analytics.middleware import BlockedIPMiddleware
        BlockedIPMiddleware.clear_cache()
        messages.success(request, f'IP {obj.ip_address} has been unblocked.')
    except BlockedIP.DoesNotExist:
        messages.error(request, 'Blocked IP record not found.')
    return redirect('audit-log-viewer')


@login_required(login_url='/login/')
@role_required('admin')
@require_POST
def delete_blocked_ip(request):
    """Permanently delete a blocked IP record."""
    from .models import BlockedIP
    ip_id = request.POST.get('ip_id', '').strip()
    try:
        obj = BlockedIP.objects.get(pk=ip_id)
        ip_addr = obj.ip_address
        obj.delete()
        from gate_analytics.middleware import BlockedIPMiddleware
        BlockedIPMiddleware.clear_cache()
        messages.success(request, f'Blocked IP record for {ip_addr} deleted.')
    except BlockedIP.DoesNotExist:
        messages.error(request, 'Record not found.')
    return redirect('audit-log-viewer')


def _get_client_ip(request):
    xff = request.META.get('HTTP_X_FORWARDED_FOR', '').split(',')[0].strip()
    return xff or request.META.get('REMOTE_ADDR', '')


def _department_label(value):
    """Return display label for a department value, or the value itself."""
    if not value:
        return ''
    for v, label in CAMPUS_DEPARTMENT_CHOICES:
        if v == value and v:
            return label
    return value


@login_required(login_url='/login/')
@role_required('admin', 'staff', 'student affairs')
def visitor_pass_create(request):
    """Create a one-time visitor pass or generate bulk reusable slots (VIS-001...VIS-N)."""
    if request.method == 'POST':
        action = (request.POST.get('action') or '').strip()
        # Bulk generate reusable slots (VIS-001, VIS-002, ...) – From/To range from user
        if action == 'generate_slots':
            # Read slot range from POST; use only provided values to avoid defaulting to 100
            raw_start = (request.POST.get('slot_start') or '').strip()
            raw_end = (request.POST.get('slot_end') or '').strip()
            try:
                start = int(raw_start) if raw_start else 1
            except (TypeError, ValueError):
                start = 1
            try:
                end = int(raw_end) if raw_end else 100
            except (TypeError, ValueError):
                end = 100
            if start < 1:
                start = 1
            if end > 999:
                end = 999
            if start > end:
                start, end = end, start
            created = 0
            for i in range(start, end + 1):
                code = VisitorPass.generate_reusable_code(i)
                if not VisitorPass.objects.filter(code=code).exists():
                    VisitorPass.objects.create(
                        code=code,
                        status=VisitorPass.STATUS_AVAILABLE,
                        created_by=request.user,
                    )
                    created += 1
            if created:
                messages.success(request, f'Created {created} reusable visitor pass slot(s) (VIS-{start:03d} to VIS-{end:03d}).')
            else:
                messages.info(request, f'Slots VIS-{start:03d} to VIS-{end:03d} already exist. You can view and print them below.')
            return redirect(reverse('visitor-pass-create') + f'?start={start}&end={end}')

        guest_name = (request.POST.get('guest_name') or '').strip()
        purpose = (request.POST.get('purpose') or '').strip()
        department_value = (request.POST.get('department') or '').strip()
        department_label = _department_label(department_value) if department_value else ''
        valid_from = request.POST.get('valid_from')
        valid_until = request.POST.get('valid_until')
        if not guest_name or not valid_from or not valid_until:
            messages.error(request, 'Guest name and valid from/until are required.')
            return redirect('visitor-pass-create')
        from django.utils.dateparse import parse_date
        try:
            d_from = parse_date(valid_from[:10]) if valid_from else None
            d_until = parse_date(valid_until[:10]) if valid_until else None
            if not d_from or not d_until:
                raise ValueError('Invalid date')
            tz = timezone.get_current_timezone()
            dt_from = timezone.make_aware(datetime.datetime.combine(d_from, datetime.time.min), tz)
            dt_until = timezone.make_aware(datetime.datetime.combine(d_until, datetime.time.max), tz)
        except Exception:
            messages.error(request, 'Invalid date format. Use YYYY-MM-DD.')
            return redirect('visitor-pass-create')
        code = VisitorPass.generate_code()
        VisitorPass.objects.create(
            code=code,
            guest_name=guest_name,
            purpose=purpose,
            department=department_label,
            valid_from=dt_from,
            valid_until=dt_until,
            created_by=request.user,
            notes=request.POST.get('notes', ''),
        )
        messages.success(request, f'Visitor pass created: {code}')
        return redirect('visitor-pass-create')
    # List: paginate passes (when start/end in URL show that range, else all recent)
    passes_qs = _visitor_pass_list_for_user(request, no_limit=True)
    per_page, query_extra, query_extra_base = _get_per_page_and_query(request)
    from django.core.paginator import Paginator
    paginator = Paginator(passes_qs, per_page)
    page_obj = paginator.get_page(request.GET.get('page', 1))
    passes = list(page_obj.object_list)
    display_start = None
    display_end = None
    start_param = request.GET.get('start')
    end_param = request.GET.get('end')
    if start_param is not None and end_param is not None:
        try:
            s, e = int(start_param), int(end_param)
            if 1 <= s <= 999 and 1 <= e <= 999 and s <= e:
                display_start = s
                display_end = e
        except (TypeError, ValueError):
            pass
    return render(request, 'gate/visitor_pass_create.html', {
        'site_name': 'City College of Bayawan',
        'passes': passes,
        'page_obj': page_obj,
        'query_extra': query_extra,
        'query_extra_base': query_extra_base,
        'per_page': per_page,
        'per_page_options': PER_PAGE_OPTIONS,
        'campus_departments': CAMPUS_DEPARTMENT_CHOICES,
        'display_start': display_start,
        'display_end': display_end,
    })


@require_GET
@login_required(login_url='/login/')
def calendar_ics(request):
    """Export events as .ics (iCalendar) for Google Calendar / Outlook."""
    from django.http import HttpResponse
    events = Event.objects.filter(status__in=('scheduled', 'active')).order_by('start_date')[:100]
    lines = ['BEGIN:VCALENDAR', 'VERSION:2.0', 'PRODID:-//CCB Gate//EN']
    for e in events:
        lines.append('BEGIN:VEVENT')
        lines.append(f'UID:event-{e.id}@ccb')
        lines.append(f'DTSTART;VALUE=DATE:{e.start_date.strftime("%Y%m%d")}')
        lines.append(f'DTEND;VALUE=DATE:{(e.end_date + datetime.timedelta(days=1)).strftime("%Y%m%d")}')
        lines.append(f'SUMMARY:{e.name[:200].replace(chr(10), " ").replace(chr(13), " ")}')
        lines.append(f'LOCATION:{e.venue[:200].replace(chr(10), " ")}')
        lines.append('END:VEVENT')
    lines.append('END:VCALENDAR')
    response = HttpResponse('\r\n'.join(lines), content_type='text/calendar; charset=utf-8')
    response['Content-Disposition'] = 'attachment; filename="events.ics"'
    return response


# ------------------------- Optional API for integration -------------------------

@require_GET
def api_attendance(request):
    """
    Read-only API for gate/attendance integration. Token-based auth.
    Preferred: Authorization: Bearer <token> header.
    Legacy (deprecated): ?api_key=<token> query param — tokens in URLs leak via logs/Referer.
    Query params: date_from (YYYY-MM-DD), date_to (YYYY-MM-DD).
    Returns: gate_entries_count, event_scans_count, and sample rows (minimal fields).
    Set API_ATTENDANCE_TOKEN in environment (or settings) to enable.
    """
    header_token = (request.headers.get('Authorization') or '').replace('Bearer ', '').strip()
    query_token = request.GET.get('api_key', '')
    if query_token and not header_token:
        import logging
        logging.getLogger('gate').warning(
            'api_attendance called with api_key in query string (deprecated) — '
            'use Authorization: Bearer header instead. client=%s',
            request.META.get('REMOTE_ADDR', '?'),
        )
    token = header_token or query_token
    expected = getattr(settings, 'API_ATTENDANCE_TOKEN', '')
    if not expected or token != expected:
        return JsonResponse({'error': 'Unauthorized', 'detail': 'Invalid or missing API token.'}, status=401)
    from django.utils.dateparse import parse_date
    today = timezone.localdate()
    date_from = parse_date(request.GET.get('date_from') or '') or (today - datetime.timedelta(days=7))
    date_to = parse_date(request.GET.get('date_to') or '') or today
    if date_from > date_to:
        date_from, date_to = date_to, date_from
    tz = timezone.get_current_timezone()
    start = timezone.make_aware(datetime.datetime.combine(date_from, datetime.time.min), tz)
    end = timezone.make_aware(datetime.datetime.combine(date_to, datetime.time.max), tz)
    gate_qs = GateEntry.objects.filter(timestamp__gte=start, timestamp__lte=end).select_related('student').order_by('-timestamp')[:500]
    gate_entries = [
        {'student_id': e.student.student_id if e.student else None, 'timestamp': e.timestamp.isoformat(), 'granted': e.granted}
        for e in gate_qs
    ]
    logs_qs = AttendanceLog.objects.filter(scan_time__gte=start, scan_time__lte=end, voided=False).order_by('-scan_time')[:500]
    event_scans = [
        {'event_id': l.event_id, 'student_id': l.student.student_id if l.student else None, 'scan_time': l.scan_time.isoformat(), 'result': l.result}
        for l in logs_qs.select_related('student', 'event')
    ]
    return JsonResponse({
        'date_from': date_from.isoformat(),
        'date_to': date_to.isoformat(),
        'gate_entries_count': GateEntry.objects.filter(timestamp__gte=start, timestamp__lte=end).count(),
        'event_scans_count': AttendanceLog.objects.filter(scan_time__gte=start, scan_time__lte=end, voided=False).count(),
        'gate_entries_sample': gate_entries,
        'event_scans_sample': event_scans,
    })
