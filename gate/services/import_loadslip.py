"""
Import load slip data from CSV or Excel (many students in one file).
Matches City College of Bayawan enrollment load slip format:
CODE, COURSE TITLE, SECTION, UNITS, SCHEDULE (e.g. TTH/ 10:00-11:30 AM), ROOM.
Validates: student exists, semester/day, start < end time, units > 0,
no duplicate subject+day+time, no overlapping times per day.
"""
import csv
import io
import re
from datetime import datetime, time
from django.db import transaction
from django.core.exceptions import ValidationError

from ..models import Student, StudentLoadSlip, LoadSlipSubject

ALLOWED_SEM = {'1st', '2nd', 'summer'}
ALLOWED_DAYS = {'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun'}

# Standard columns (one row per day)
REQUIRED_COLS_STANDARD = [
    'student_id', 'school_year', 'semester', 'subject_code', 'subject_title',
    'section', 'units', 'day', 'start_time', 'end_time', 'room', 'instructor',
]
# Slip-style: same but SCHEDULE instead of day/start_time/end_time (one row per subject; we expand to one per day)
REQUIRED_COLS_SLIP = [
    'student_id', 'school_year', 'semester', 'subject_code', 'subject_title',
    'section', 'units', 'schedule', 'room', 'instructor',
]

# Map slip header names to internal keys (case-insensitive match after strip)
COLUMN_ALIASES = {
    'code': 'subject_code',
    'course title': 'subject_title',
    'course_title': 'subject_title',
    'section': 'section',
    'units': 'units',
    'schedule': 'schedule',
    'room': 'room',
    'instructor': 'instructor',
    'student_id': 'student_id',
    'student id': 'student_id',
    'school_year': 'school_year',
    'school year': 'school_year',
    'semester': 'semester',
}

# Schedule day abbreviations (as on CCB load slip): M, T, W, TH, F, S/SAT, etc.
DAY_ABBREV_TO_FULL = {
    'M': 'Mon', 'MO': 'Mon', 'MON': 'Mon',
    'T': 'Tue', 'TU': 'Tue', 'TUE': 'Tue', 'TUES': 'Tue',
    'W': 'Wed', 'WE': 'Wed', 'WED': 'Wed',
    'TH': 'Thu', 'R': 'Thu', 'THU': 'Thu', 'THUR': 'Thu',
    'F': 'Fri', 'FR': 'Fri', 'FRI': 'Fri',
    'S': 'Sat', 'SA': 'Sat', 'SAT': 'Sat',
    'U': 'Sun', 'SU': 'Sun', 'SUN': 'Sun',
}


def _parse_time(val):
    """Accept '07:30', '10:00-11:30 AM', datetime.time, or datetime.datetime."""
    if val is None or (isinstance(val, str) and not str(val).strip()):
        raise ValidationError('Time is required.')
    if hasattr(val, 'hour'):
        if hasattr(val, 'year') and getattr(val, 'year', None):
            return val.time()
        return val
    s = str(val).strip()
    for fmt in ('%H:%M', '%H:%M:%S', '%I:%M %p', '%I:%M:%S %p', '%I:%M %p', '%H:%M'):
        try:
            return datetime.strptime(s, fmt).time()
        except ValueError:
            continue
    raise ValidationError(f"Invalid time format: '{val}'")


def _parse_schedule(schedule_str):
    """
    Parse CCB load slip SCHEDULE column e.g. 'TTH/ 10:00-11:30 AM', 'MWF/ 11:00-12:00 AM', 'SAT/ 1:00-5:00 PM'.
    Returns list of (day, start_time, end_time) with day in ALLOWED_DAYS.
    """
    if not schedule_str or not str(schedule_str).strip():
        raise ValidationError('Schedule is required.')
    s = str(schedule_str).strip().upper()
    part = s.split('/')
    if len(part) < 2:
        raise ValidationError(f"Schedule should be like 'TTH/ 10:00-11:30 AM'. Got: '{schedule_str}'")
    days_part = part[0].strip().replace(' ', '')
    time_part = part[1].strip() if len(part) > 1 else ''
    if not time_part:
        raise ValidationError(f"Schedule missing time part. Got: '{schedule_str}'")
    # Accept: "10:00-11:30 AM", "10:00 AM-11:30 AM", "02:30 PM-04:00 PM" (AM/PM after each time or once at end)
    time_match = re.match(
        r'(\d{1,2}:\d{2}(?::\d{2})?)\s*(AM|PM)?\s*-\s*(\d{1,2}:\d{2}(?::\d{2})?)\s*(AM|PM)?',
        time_part, re.I
    )
    if not time_match:
        time_match = re.match(
            r'(\d{1,2}:\d{2})\s*(AM|PM)?\s*-\s*(\d{1,2}:\d{2})\s*(AM|PM)?',
            time_part, re.I
        )
    if not time_match:
        raise ValidationError(f"Schedule time should be like '10:00-11:30 AM' or '02:30 PM-04:00 PM'. Got: '{time_part}'")
    start_s = time_match.group(1)
    start_ampm = (time_match.group(2) or '').upper()
    end_s = time_match.group(3)
    end_ampm = (time_match.group(4) or '').upper()
    # If only one AM/PM given (at end), it applies to both times
    if start_ampm and not end_ampm:
        end_ampm = start_ampm
    if end_ampm and not start_ampm:
        start_ampm = end_ampm
    # When no AM/PM: assume PM for hours 1-11 (common for afternoon classes e.g. "TTH/5:00-6:01")
    if not start_ampm and not end_ampm:
        try:
            start_hr = int(start_s.split(':')[0])
            end_hr = int(end_s.split(':')[0])
            if 1 <= start_hr <= 11 or 1 <= end_hr <= 11:
                start_ampm = 'PM'
                end_ampm = 'PM'
        except (ValueError, IndexError):
            pass
    start_str = f"{start_s} {start_ampm}".strip() if start_ampm else start_s
    end_str = f"{end_s} {end_ampm}".strip() if end_ampm else end_s
    # "11:00-12:00 PM" means 11 AM to 12 PM (noon), not 11 PM to 12 PM. If only one PM at end and start is 1-11, end is 12, treat start as AM.
    if start_ampm == 'PM' and end_ampm == 'PM':
        try:
            start_hr = int(start_s.split(':')[0])
            end_hr = int(end_s.split(':')[0])
            if 1 <= start_hr <= 11 and end_hr == 12:
                start_str = f"{start_s} AM".strip()
        except (ValueError, IndexError):
            pass
    start_time = _parse_time(start_str)
    end_time = _parse_time(end_str)
    # Common slip typo: "11:00-12:00 AM" means 11 AM–12 PM (noon). 12:00 AM is midnight, so start > end.
    if end_time == time(0, 0) and start_time.hour < 12:
        end_time = time(12, 0)
    if start_time >= end_time:
        raise ValidationError(f"Start time must be before end time in schedule: '{schedule_str}'")
    day_tokens = []
    i = 0
    while i < len(days_part):
        two = days_part[i:i + 2]
        one = days_part[i]
        if two == 'TH':
            day_tokens.append('TH')
            i += 2
        elif two == 'SA' and (i + 3 <= len(days_part) and days_part[i:i + 3] == 'SAT'):
            day_tokens.append('SAT')
            i += 3
        elif two == 'SU':
            day_tokens.append('SU')
            i += 2
        elif one in 'MWF':
            day_tokens.append(one)
            i += 1
        elif one == 'T':
            day_tokens.append('T')
            i += 1
        elif one == 'S' and (i + 1 >= len(days_part) or days_part[i + 1] != 'U'):
            day_tokens.append('S')
            i += 1
        elif one == 'U':
            day_tokens.append('U')
            i += 1
        else:
            i += 1
    days = []
    for t in day_tokens:
        d = DAY_ABBREV_TO_FULL.get(t) or DAY_ABBREV_TO_FULL.get(t.upper())
        if d and d not in days:
            days.append(d)
    if not days:
        raise ValidationError(f"Could not parse days from schedule: '{schedule_str}'. Use e.g. MWF, TTH, SAT.")
    return [(d, start_time, end_time) for d in days]


def parse_schedule_safe(schedule_str):
    """Return list of (day, start_time, end_time) or [] if invalid. For gate logic (no ValidationError)."""
    if not schedule_str or not str(schedule_str).strip():
        return []
    try:
        return _parse_schedule(schedule_str)
    except ValidationError:
        return []


def _normalize_student_id(val):
    """Normalize so Excel numeric IDs (e.g. 20240001.0) match DB (20240001)."""
    if val is None:
        return ''
    s = str(val).strip()
    try:
        f = float(s)
        if f == int(f):
            return str(int(f))
    except (TypeError, ValueError):
        pass
    return s


def _normalize_columns(rows):
    """Normalize column names using aliases (slip-style headers → internal names)."""
    if not rows:
        return rows
    first = rows[0]
    key_map = {}
    for raw_key in first.keys():
        k = str(raw_key or '').strip().lower()
        key_map[raw_key] = COLUMN_ALIASES.get(k) or raw_key
    out = []
    for row in rows:
        out.append({key_map.get(k, k): row.get(k) for k in row})
    return out


def _read_rows(file_obj, filename):
    """Return list of dicts (one per row). Supports .csv and .xlsx."""
    name = (filename or '').lower()
    if name.endswith('.csv'):
        try:
            data = file_obj.read().decode('utf-8-sig')
        except Exception as e:
            raise ValidationError(f'Could not read CSV: {e}')
        reader = csv.DictReader(io.StringIO(data))
        return list(reader)
    if name.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ValidationError('Excel support requires openpyxl. Install: pip install openpyxl')
        try:
            wb = load_workbook(file_obj, data_only=True)
        except Exception as e:
            raise ValidationError(f'Could not read Excel file: {e}')
        ws = wb.active
        header_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not header_row:
            raise ValidationError('Excel file has no header row.')
        headers = [str(c.value or '').strip() or f'Col{i}' for i, c in enumerate(header_row)]
        rows = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            r = list(r) if r else []
            row = {h: (r[i] if i < len(r) else None) for i, h in enumerate(headers)}
            if any(v is not None and str(v).strip() for v in row.values()):
                rows.append(row)
        return rows
    raise ValidationError('Unsupported file type. Upload CSV or XLSX.')


def import_loadslip(file_obj, filename, uploaded_by=None, replace_existing=True):
    """
    Parse file and create/update load slips and subjects.
    Supports (1) standard columns: day, start_time, end_time; (2) slip-style: SCHEDULE (e.g. TTH/ 10:00-11:30 AM).
    Column aliases: CODE→subject_code, COURSE TITLE→subject_title, SECTION, UNITS, ROOM, SCHEDULE.
    replace_existing=True: delete existing subjects for each (student, SY, semester) then insert.
    Returns: dict with created_subjects, updated_slips, errors (list of strings).
    """
    rows = _read_rows(file_obj, filename)
    if not rows:
        raise ValidationError('File is empty.')
    rows = _normalize_columns(rows)
    first = rows[0]
    has_schedule = 'schedule' in first
    has_day_time = 'day' in first and 'start_time' in first and 'end_time' in first
    if has_schedule:
        required = ['student_id', 'school_year', 'semester', 'subject_code', 'subject_title', 'section', 'units', 'schedule']
    else:
        required = REQUIRED_COLS_STANDARD
    missing = [c for c in required if c not in first]
    if missing:
        raise ValidationError(
            f"Missing columns: {', '.join(missing)}. "
            "Use standard (day, start_time, end_time) or slip-style (schedule e.g. TTH/ 10:00-11:30 AM)."
        )

    errors = []
    created_subjects = 0
    updated_slips = 0
    input_course_rows = 0
    # Which (student_id, school_year, semester) were updated — only these are touched; other semesters left unchanged
    updated_slip_keys = []

    # Group rows by (student_id, school_year, semester). Only these combinations are updated.
    grouped = {}
    for idx, row in enumerate(rows, start=2):
        sid = _normalize_student_id(row.get('student_id'))
        sy = str(row.get('school_year') or '').strip()
        sem = str(row.get('semester') or '').strip()
        if not sid and not sy and not sem:
            continue
        key = (sid, sy, sem)
        grouped.setdefault(key, []).append((idx, row))

    input_course_rows = sum(len(items) for _, items in grouped.items())

    with transaction.atomic():
        for (student_id, school_year, semester), items in grouped.items():
            sem = semester.strip().lower()
            if sem not in ALLOWED_SEM:
                errors.append(f"[{student_id}] Invalid semester '{semester}'. Use 1st, 2nd, or summer.")
                continue
            if not student_id:
                errors.append('Row group missing student_id.')
                continue

            try:
                student = Student.objects.get(student_id=student_id)
            except Student.DoesNotExist:
                errors.append(f"Student ID '{student_id}' not found.")
                continue

            slip, slip_created = StudentLoadSlip.objects.get_or_create(
                student=student, school_year=school_year, semester=sem,
                defaults={}
            )
            if not slip_created:
                updated_slips += 1
            updated_slip_keys.append((student_id, school_year, semester))

            # Replace this semester's subjects with the file (registrar is source of truth). Other semesters untouched.
            if replace_existing:
                slip.subjects.all().delete()

            schedule_map = {d: [] for d in ALLOWED_DAYS}
            seen_keys = set()
            # Slip-style: one row per course (avoid duplicate course in same file)
            seen_schedule_key = set()

            for row_num, row in items:
                try:
                    try:
                        units = float(row.get('units') or 0)
                    except (TypeError, ValueError):
                        raise ValidationError('units must be a number.')
                    if units <= 0:
                        raise ValidationError('units must be > 0.')

                    subject_code = str(row.get('subject_code') or '').strip()
                    subject_title = str(row.get('subject_title') or '').strip()
                    section = str(row.get('section') or '').strip()
                    if not subject_code:
                        raise ValidationError('subject_code (or CODE) is required.')
                    room = str(row.get('room') or '').strip() if row.get('room') is not None else ''
                    instructor = str(row.get('instructor') or '').strip() if row.get('instructor') is not None else ''

                    if has_schedule:
                        schedule_str = str(row.get('schedule') or '').strip()
                        schedule_entries = _parse_schedule(schedule_str)
                        # Validate overlaps (same as before)
                        for day, start_time, end_time in schedule_entries:
                            for (s, e, code) in schedule_map[day]:
                                if start_time < e and end_time > s:
                                    raise ValidationError(
                                        f"Schedule overlap on {day}: {subject_code} conflicts with {code}."
                                    )
                            schedule_map[day].append((start_time, end_time, subject_code))
                        # One row per course: store schedule string; day/start/end = first occurrence
                        dedup_key = (subject_code, section, schedule_str)
                        if dedup_key in seen_schedule_key:
                            continue
                        seen_schedule_key.add(dedup_key)
                        first_day, first_start, first_end = schedule_entries[0]
                        if LoadSlipSubject.objects.filter(
                            load_slip=slip,
                            subject_code=subject_code,
                            section=section,
                            schedule=schedule_str,
                        ).exists():
                            continue
                        LoadSlipSubject.objects.create(
                            load_slip=slip,
                            subject_code=subject_code,
                            subject_title=subject_title,
                            section=section,
                            units=units,
                            schedule=schedule_str,
                            day=first_day,
                            start_time=first_start,
                            end_time=first_end,
                            room=room,
                            instructor=instructor,
                        )
                        created_subjects += 1
                    else:
                        # Standard: one row per day (legacy)
                        day = str(row.get('day') or '').strip()
                        if day not in ALLOWED_DAYS:
                            raise ValidationError(f"Invalid day '{day}'. Use Mon, Tue, Wed, Thu, Fri, Sat, Sun.")
                        start_time = _parse_time(row.get('start_time'))
                        end_time = _parse_time(row.get('end_time'))
                        if start_time >= end_time:
                            raise ValidationError('start_time must be earlier than end_time.')
                        schedule_entries = [(day, start_time, end_time)]

                        for day, start_time, end_time in schedule_entries:
                            dedup_key = (subject_code, day, start_time, end_time, section)
                            if dedup_key in seen_keys:
                                continue
                            seen_keys.add(dedup_key)

                            for (s, e, code) in schedule_map[day]:
                                if start_time < e and end_time > s:
                                    raise ValidationError(
                                        f"Schedule overlap on {day}: {subject_code} conflicts with {code}."
                                    )
                            schedule_map[day].append((start_time, end_time, subject_code))

                            if LoadSlipSubject.objects.filter(
                                load_slip=slip,
                                subject_code=subject_code,
                                day=day,
                                start_time=start_time,
                                end_time=end_time,
                            ).exists():
                                continue

                            LoadSlipSubject.objects.create(
                                load_slip=slip,
                                subject_code=subject_code,
                                subject_title=subject_title,
                                section=section,
                                units=units,
                                day=day,
                                start_time=start_time,
                                end_time=end_time,
                                room=room,
                                instructor=instructor,
                            )
                            created_subjects += 1

                except ValidationError as e:
                    msg = e.messages[0] if getattr(e, 'messages', None) else str(e)
                    errors.append(f"Row {row_num} [{student_id}]: {msg}")
                except Exception as e:
                    errors.append(f"Row {row_num} [{student_id}]: {str(e)}")

    return {
        'updated_slips': updated_slips,
        'created_subjects': created_subjects,
        'input_course_rows': input_course_rows,
        'updated_slip_keys': updated_slip_keys,
        'errors': errors,
    }


def import_loadslip_by_filter(
    file_obj, filename,
    school_year, semester, course, year_level, sections_list,
    replace_existing=True,
):
    """
    Apply one load slip (from file) to all students matching course + year + section(s).
    File has NO student_id: slip-style only (CODE, COURSE TITLE, SECTION, UNITS, SCHEDULE, ROOM).
    sections_list: None or [] = all sections for that course/year; else e.g. ['A','B','C'].
    Returns same dict as import_loadslip.
    """
    rows = _read_rows(file_obj, filename)
    if not rows:
        raise ValidationError('File is empty.')
    rows = _normalize_columns(rows)
    first = rows[0]
    if 'schedule' not in first:
        raise ValidationError(
            "Import by filter requires slip-style format with SCHEDULE column. "
            "File must have: CODE, COURSE TITLE, SECTION, UNITS, SCHEDULE, ROOM (no student_id)."
        )
    required = ['subject_code', 'subject_title', 'section', 'units', 'schedule']
    missing = [c for c in required if c not in first]
    if missing:
        raise ValidationError(f"Missing columns: {', '.join(missing)}. Filter mode requires slip-style (no student_id).")

    if not course or not year_level:
        raise ValidationError('Course and year level are required when importing by filter.')
    sem = str(semester or '').strip().lower()
    if sem not in ALLOWED_SEM:
        raise ValidationError(f"Invalid semester '{semester}'. Use 1st, 2nd, or summer.")
    sy = str(school_year or '').strip()
    if not sy:
        raise ValidationError('School year is required (e.g. 2025-2026).')

    # Resolve sections: None or [] = all sections; otherwise filter by these sections
    section_filter = None
    if sections_list:
        section_filter = [s.strip() for s in sections_list if str(s).strip()]
        if not section_filter:
            section_filter = None

    students_qs = Student.objects.filter(course=course, year_level=year_level, is_active=True)
    if section_filter is not None:
        # Match section exactly (case-sensitive as stored); allow comma list "A,B,C"
        from django.db.models import Q
        q = Q()
        for sec in section_filter:
            q = q | Q(section=sec)
        students_qs = students_qs.filter(q)
    students = list(students_qs)
    if not students:
        return {
            'updated_slips': 0,
            'created_subjects': 0,
            'input_course_rows': len(rows),
            'updated_slip_keys': [],
            'errors': ['No students found matching Course=%s, Year=%s%s.' % (
                course, year_level,
                ', Sections=%s' % (section_filter or 'all') if section_filter else ' (all sections)',
            )],
        }

    # Parse all subject rows once (validate and build list of subject dicts)
    errors = []
    subject_rows = []
    schedule_map = {d: [] for d in ALLOWED_DAYS}
    seen_schedule_key = set()
    for idx, row in enumerate(rows, start=2):
        try:
            units = float(row.get('units') or 0)
        except (TypeError, ValueError):
            errors.append(f"Row {idx}: units must be a number.")
            continue
        if units <= 0:
            errors.append(f"Row {idx}: units must be > 0.")
            continue
        subject_code = str(row.get('subject_code') or '').strip()
        subject_title = str(row.get('subject_title') or '').strip()
        section = str(row.get('section') or '').strip()
        if not subject_code:
            errors.append(f"Row {idx}: subject code (CODE) is required.")
            continue
        room = str(row.get('room') or '').strip() if row.get('room') is not None else ''
        instructor = str(row.get('instructor') or '').strip() if row.get('instructor') is not None else ''
        schedule_str = str(row.get('schedule') or '').strip()
        if not schedule_str:
            errors.append(f"Row {idx}: SCHEDULE is required.")
            continue
        try:
            schedule_entries = _parse_schedule(schedule_str)
        except ValidationError as e:
            errors.append(f"Row {idx}: {e}")
            continue
        for day, start_time, end_time in schedule_entries:
            for (s, e, code) in schedule_map[day]:
                if start_time < e and end_time > s:
                    errors.append(f"Row {idx}: Schedule overlap on {day}: {subject_code} conflicts with {code}.")
                    break
            else:
                schedule_map[day].append((start_time, end_time, subject_code))
        dedup_key = (subject_code, section, schedule_str)
        if dedup_key in seen_schedule_key:
            continue
        seen_schedule_key.add(dedup_key)
        first_day, first_start, first_end = schedule_entries[0]
        subject_rows.append({
            'subject_code': subject_code,
            'subject_title': subject_title,
            'section': section,
            'units': units,
            'schedule': schedule_str,
            'day': first_day,
            'start_time': first_start,
            'end_time': first_end,
            'room': room,
            'instructor': instructor,
        })
    if errors:
        return {
            'updated_slips': 0,
            'created_subjects': 0,
            'input_course_rows': len(rows),
            'updated_slip_keys': [],
            'errors': errors,
        }

    created_subjects = 0
    updated_slips = 0
    updated_slip_keys = []
    with transaction.atomic():
        for student in students:
            slip, slip_created = StudentLoadSlip.objects.get_or_create(
                student=student, school_year=sy, semester=sem, defaults={}
            )
            if not slip_created:
                updated_slips += 1
            updated_slip_keys.append((student.student_id, sy, sem))
            if replace_existing:
                slip.subjects.all().delete()
            for sub in subject_rows:
                if LoadSlipSubject.objects.filter(
                    load_slip=slip,
                    subject_code=sub['subject_code'],
                    section=sub['section'],
                    schedule=sub['schedule'],
                ).exists():
                    continue
                LoadSlipSubject.objects.create(
                    load_slip=slip,
                    subject_code=sub['subject_code'],
                    subject_title=sub['subject_title'],
                    section=sub['section'],
                    units=sub['units'],
                    schedule=sub['schedule'],
                    day=sub['day'],
                    start_time=sub['start_time'],
                    end_time=sub['end_time'],
                    room=sub['room'],
                    instructor=sub['instructor'],
                )
                created_subjects += 1

    return {
        'updated_slips': updated_slips,
        'created_subjects': created_subjects,
        'input_course_rows': len(rows) * len(students),
        'updated_slip_keys': updated_slip_keys,
        'errors': [],
    }
